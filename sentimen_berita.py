#!/usr/bin/env python3
"""
Analisa sentimen berita Indonesia.
Scrape dari berbagai sumber, analisa sentimen, simpan ke Excel.
"""
import urllib.request
import re
import json
import os
import sys
from datetime import datetime
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

EXCEL_PATH = os.path.expanduser("~/sembako/sentimen_berita.xlsx")
HISTORY_PATH = os.path.expanduser("~/sembako/sentimen_history.json")

# Kata kunci sentimen bahasa Indonesia
POSITIVE_ID = [
    "positif", "optimis", "berhasil", "sukses", "naik", "meningkat", "tumbuh",
    "untung", "profit", "bagus", "baik", "stabil", "aman", "sejahtera",
    "kemajuan", "berkah", "puas", "senang", "bangga", "pulih", "rebound",
    "bullish", "melonjak", "melambung", "cemerlang", "modal", "perkuat",
    "sentuh", "ciptakan", "bangkit", "kejutan"
]

NEGATIVE_ID = [
    "negatif", "gagal", "turun", "menurun", "anjlok", "merosot", "rugi",
    "krisis", "resesi", "inflasi", "korupsi", "bencana", "gempa", "banjir",
    "kecelakaan", "konflik", "perang", "demo", "protes", "ancaman",
    "bearish", "merah", "darurat", "wabah", "pandemi", "PHK", "kemiskinan",
    "kelangkaan", "keluhan", "masalah", "buruk", "pelemahan", "melambat",
    "miskin", "ilegal", "ditutup", "diragukan", "dampak", "eskalasi"
]

# Outdated/seasonal keywords to exclude (not relevant to current date)
# Use word boundaries to avoid partial matches
OUTDATED_PATTERNS = [
    r'\bramadhan\b', r'\bramadan\b', r'\blebaran\b', r'\bmudik\b',
    r'\bidul\s*fitri\b', r'\bpuasa\b', r'\bnatal\b', r'\btahun\s*baru\b',
    r'\bimlek\b', r'\bkenaikan\s*gaji\b', r'\bthr\b', r'\bhari\s*raya\b',
    r'\bmenjelang\s*ramadhan\b', r'\bjelang\s*ramadhan\b',
    r'\bjelang\s*ramadan\b', r'\bjelang\s*lebaran\b',
    r'\bmenjelang\s*natal\b'
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml",
    "Accept-Language": "id-ID,id;q=0.9",
}

def fetch_url(url, timeout=15):
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read().decode("utf-8", errors="ignore")
    except Exception as e:
        print(f"  Error fetching: {e}")
        return ""

def get_article_date(article_url):
    """Fetch article page and extract publish date."""
    html = fetch_url(article_url, timeout=10)
    if not html:
        return None

    # Try multiple date patterns
    patterns = [
        r'<time[^>]*datetime="([^"]+)"',  # datetime="2026-06-25T07:01:24+07:00"
        r'"datePublished"\s*:\s*"([^"]+)"',  # "datePublished": "2026-06-25T07:01:24+07:00"
        r'<meta[^>]*name="date"[^>]*content="([^"]+)"',  # meta date
        r'<time[^>]*>([^<]+)</time>',  # <time>Kamis, 25 Jun 2026 07:01 WIB</time>
    ]

    for pattern in patterns:
        match = re.search(pattern, html)
        if match:
            date_str = match.group(1)
            # Parse ISO date
            iso_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', date_str)
            if iso_match:
                return datetime.fromisoformat(date_str.replace('+07:00', ''))
            # Parse Indonesian format
            indo_match = re.search(r'(\d{1,2})\s+(\w+)\s+(\d{4})', date_str)
            if indo_match:
                MONTH_MAP = {'jan':1,'feb':2,'mar':3,'apr':4,'mei':5,'jun':6,'jul':7,'agu':8,'sep':9,'okt':10,'nov':11,'des':12}
                day = int(indo_match.group(1))
                month = MONTH_MAP.get(indo_match.group(2).lower()[:3], 6)
                year = int(indo_match.group(3))
                return datetime(year, month, day)

    return None

def scrape_detik(query, limit=5):
    """Scrape headlines from detik.com search - verified current date only."""
    search_url = f"https://www.detik.com/search/searchall?query={query.replace(' ', '+')}&siteid=2&result_type=latest"
    html = fetch_url(search_url)
    if not html:
        return []

    # Find article URLs
    article_urls = re.findall(r'href="(https?://[^"]*detik[^"]*-\d+[^"]*)"', html)
    # Deduplicate
    article_urls = list(dict.fromkeys(article_urls))

    today = datetime.now().date()
    results = []

    skip_patterns = ['detikcom', 'detikNews', 'detikFinance', 'detikInet',
                    'detikHot', 'detikSport', 'MENU', 'Beranda', 'Login',
                    'Terpopuler', 'Koleksi', 'Selengkapnya']

    for article_url in article_urls:
        if len(results) >= limit:
            break

        # Fetch article to get title and date
        article_html = fetch_url(article_url, timeout=10)
        if not article_html:
            continue

        # Extract title
        title_match = re.search(r'<meta[^>]*property="og:title"[^>]*content="([^"]+)"', article_html, re.IGNORECASE)
        if not title_match:
            title_match = re.search(r'<title>([^<]+)</title>', article_html)
        if not title_match:
            continue

        title = title_match.group(1).strip()
        if len(title) < 15:
            continue
        if any(p.lower() in title.lower() for p in skip_patterns):
            continue

        # Get publish date
        pub_date = get_article_date(article_url)
        if pub_date is None:
            continue
        if pub_date.date() != today:
            print(f"  SKIP (not today): {title[:50]}... ({pub_date.strftime('%d %b %Y')})")
            continue

        # Skip outdated/seasonal
        title_lower = title.lower()
        if any(re.search(pattern, title_lower) for pattern in OUTDATED_PATTERNS):
            print(f"  SKIP (outdated): {title[:50]}...")
            continue

        results.append(title)
        print(f"  ✓ {title[:60]}...")

    return results[:limit]

def analyze_sentiment_id(text):
    """Analyze sentiment for Indonesian text."""
    text_lower = text.lower()
    
    pos_count = sum(1 for w in POSITIVE_ID if w in text_lower)
    neg_count = sum(1 for w in NEGATIVE_ID if w in text_lower)
    
    analyzer = SentimentIntensityAnalyzer()
    vader = analyzer.polarity_scores(text)
    
    if pos_count > neg_count:
        label = "POSITIF"
        score = min(0.5 + (pos_count - neg_count) * 0.15, 1.0)
    elif neg_count > pos_count:
        label = "NEGATIF"
        score = max(-0.5 - (neg_count - pos_count) * 0.15, -1.0)
    else:
        compound = vader['compound']
        if compound > 0.1:
            label = "POSITIF"
            score = compound
        elif compound < -0.1:
            label = "NEGATIF"
            score = compound
        else:
            label = "NETRAL"
            score = compound
    
    return {"label": label, "score": round(score, 3)}

def analyze_news(keywords=None):
    if keywords is None:
        keywords = [
            "ekonomi indonesia",
            "harga pangan",
            "inflasi indonesia",
            "rupiah dollar",
            "ihsg bursa"
        ]
    
    all_results = []
    
    for kw in keywords:
        print(f"\n🔍 Mencari: {kw}")
        titles = scrape_detik(kw, limit=5)
        print(f"  Ditemukan {len(titles)} berita")
        
        for title in titles:
            sentiment = analyze_sentiment_id(title)
            result = {
                "keyword": kw,
                "headline": title,
                "sentiment": sentiment["label"],
                "score": sentiment["score"],
                "source": "detik.com",
                "timestamp": datetime.now().isoformat()
            }
            all_results.append(result)
            emoji = {"POSITIF": "🟢", "NEGATIF": "🔴", "NETRAL": "⚪"}[sentiment["label"]]
            print(f"  {emoji} [{sentiment['label']:>7}] {title[:60]}")
    
    return all_results

def save_to_history(results):
    history = []
    if os.path.exists(HISTORY_PATH):
        with open(HISTORY_PATH) as f:
            history = json.load(f)
    history.extend(results)
    history = history[-500:]
    with open(HISTORY_PATH, "w") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)

def save_to_excel(results):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, Reference
    
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    pos_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    neg_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    net_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Detail"]
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detail"
        headers = ["Tanggal", "Waktu", "Keyword", "Headline", "Sentimen", "Skor", "Sumber"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
        ws.row_dimensions[1].height = 30
        widths = [13, 10, 20, 55, 12, 10, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
    
    row = ws.max_row + 1
    for r in results:
        dt = datetime.fromisoformat(r["timestamp"])
        fill = {"POSITIF": pos_fill, "NEGATIF": neg_fill, "NETRAL": net_fill}[r["sentiment"]]
        data = [dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M"), r["keyword"],
                r["headline"], r["sentiment"], r["score"], r["source"]]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin; c.fill = fill
            c.alignment = la if col == 4 else ca
        row += 1
    
    # Summary sheet
    if "Ringkasan" in wb.sheetnames:
        del wb["Ringkasan"]
    ws_sum = wb.create_sheet("Ringkasan")
    
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    pos = sum(1 for r in all_rows if r[4] == "POSITIF")
    neg = sum(1 for r in all_rows if r[4] == "NEGATIF")
    net = sum(1 for r in all_rows if r[4] == "NETRAL")
    total = pos + neg + net
    
    for col, h in enumerate(["Metrik", "Jumlah", "Persentase"], 1):
        c = ws_sum.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
    
    scores = [r[5] for r in all_rows if r[5] is not None]
    avg_score = sum(scores) / len(scores) if scores else 0
    
    sum_data = [
        ("Total Berita", total, "100%"),
        ("🟢 Positif", pos, f"{pos/total*100:.1f}%" if total else "0%"),
        ("🔴 Negatif", neg, f"{neg/total*100:.1f}%" if total else "0%"),
        ("⚪ Netral", net, f"{net/total*100:.1f}%" if total else "0%"),
        ("Rata-rata Skor", round(avg_score, 3), ""),
    ]
    for i, rd in enumerate(sum_data, 2):
        for col, val in enumerate(rd, 1):
            c = ws_sum.cell(row=i, column=col, value=val)
            c.border = thin; c.alignment = ca
    
    for i in range(1, 4):
        ws_sum.column_dimensions[get_column_letter(i)].width = 18
    
    # Per keyword
    row_k = len(sum_data) + 3
    ws_sum.cell(row=row_k, column=1, value="Per Keyword").font = Font(bold=True, size=12)
    row_k += 1
    for col, h in enumerate(["Keyword", "Total", "Positif", "Negatif", "Netral", "Skor Avg"], 1):
        c = ws_sum.cell(row=row_k, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
    
    keywords = {}
    for r in all_rows:
        kw = r[2]
        if kw not in keywords:
            keywords[kw] = {"pos": 0, "neg": 0, "net": 0, "scores": []}
        if r[4] == "POSITIF": keywords[kw]["pos"] += 1
        elif r[4] == "NEGATIF": keywords[kw]["neg"] += 1
        else: keywords[kw]["net"] += 1
        if r[5] is not None: keywords[kw]["scores"].append(r[5])
    
    row_k += 1
    for kw, data in keywords.items():
        t = data["pos"] + data["neg"] + data["net"]
        avg = sum(data["scores"]) / len(data["scores"]) if data["scores"] else 0
        for col, val in enumerate([kw, t, data["pos"], data["neg"], data["net"], round(avg, 3)], 1):
            c = ws_sum.cell(row=row_k, column=col, value=val)
            c.border = thin; c.alignment = ca
        row_k += 1
    
    # Pie chart
    if total > 0:
        pie = PieChart()
        pie.title = "Distribusi Sentimen"
        pie.style = 10
        pie.add_data(Reference(ws_sum, min_col=2, min_row=1, max_row=4), titles_from_data=True)
        pie.set_categories(Reference(ws_sum, min_col=1, min_row=2, max_row=4))
        pie.width = 16; pie.height = 12
        ws_sum.add_chart(pie, "A" + str(row_k + 2))
    
    wb.save(EXCEL_PATH)
    print(f"\n✅ Disimpan ke {EXCEL_PATH}")

def main():
    print("=" * 50)
    print("📰 ANALISA SENTIMEN BERITA INDONESIA")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M WIB')}")
    print("=" * 50)
    
    results = analyze_news()
    
    if results:
        save_to_history(results)
        save_to_excel(results)
        
        pos = sum(1 for r in results if r["sentiment"] == "POSITIF")
        neg = sum(1 for r in results if r["sentiment"] == "NEGATIF")
        net = sum(1 for r in results if r["sentiment"] == "NETRAL")
        total = len(results)
        avg = sum(r["score"] for r in results) / total if total else 0
        
        msg = f"📊 *Analisa Sentimen Hari Ini*\n"
        msg += f"📅 {datetime.now().strftime('%d %B %Y')}\n\n"
        msg += f"📰 Total berita: *{total}*\n"
        msg += f"🟢 Positif: *{pos}* ({pos/total*100:.0f}%)\n"
        msg += f"🔴 Negatif: *{neg}* ({neg/total*100:.0f}%)\n"
        msg += f"⚪ Netral: *{net}* ({net/total*100:.0f}%)\n"
        msg += f"📈 Skor rata-rata: *{avg:.2f}*\n\n"
        
        if neg > pos:
            msg += "⚠️ Sentimen cenderung *NEGATIF* hari ini"
        elif pos > neg:
            msg += "✅ Sentimen cenderung *POSITIF* hari ini"
        else:
            msg += "⚖️ Sentimen *SEIMBANG* hari ini"
        
        print(f"\n{msg}")
    else:
        print("⚠️ Tidak ada berita ditemukan")

if __name__ == "__main__":
    main()
