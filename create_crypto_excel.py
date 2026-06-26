import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from datetime import datetime
import os

EXCEL_PATH = os.path.expanduser("~/sembako/crypto_monitor.xlsx")

COINS = ["bitcoin", "ethereum", "solana", "cardano", "dogecoin", "ripple"]
COIN_NAMES = {"bitcoin": "BTC", "ethereum": "ETH", "solana": "SOL", 
              "cardano": "ADA", "dogecoin": "DOGE", "ripple": "XRP"}

def create_or_load_workbook():
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        return wb, wb["Harga"]
    return create_new_workbook()

def create_new_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Harga"
    
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    
    headers = ["Tanggal", "Waktu",
               "BTC (USD)", "BTC (IDR)", "BTC 24h %",
               "ETH (USD)", "ETH (IDR)", "ETH 24h %",
               "SOL (USD)", "SOL (IDR)", "SOL 24h %",
               "ADA (USD)", "ADA (IDR)", "ADA 24h %",
               "DOGE (USD)", "DOGE (IDR)", "DOGE 24h %",
               "XRP (USD)", "XRP (IDR)", "XRP 24h %",
               "Total Market Cap (T)", "Sentimen"]
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
    
    # Color per coin
    colors = {"BTC": "F7931A", "ETH": "627EEA", "SOL": "9945FF", 
              "ADA": "0033AD", "DOGE": "C2A633", "XRP": "23292F"}
    for i, coin in enumerate(["BTC", "ETH", "SOL", "ADA", "DOGE", "XRP"]):
        base_col = 3 + i * 3
        fill = PatternFill(start_color=colors[coin], end_color=colors[coin], fill_type="solid")
        for c in range(base_col, base_col + 3):
            ws.cell(row=1, column=c).fill = fill
    
    widths = [13, 10] + [14, 18, 10] * 6 + [18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "C2"
    
    # Sheet Sentimen
    ws2 = wb.create_sheet("Sentimen")
    for col, h in enumerate(["Tanggal", "Koin", "Harga (USD)", "Perubahan 24h", 
                              "Sentimen", "Skor", "Sumber Berita"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
    widths2 = [13, 8, 14, 12, 12, 10, 40]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    
    wb.save(EXCEL_PATH)
    return wb, ws

def add_price_row(tanggal, waktu, data, sentimen=""):
    """
    data: dict with keys like 'btc_usd', 'btc_idr', 'btc_change', etc.
    """
    wb, ws = create_or_load_workbook()
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    neg_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    pos_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=tanggal).alignment = ca
    ws.cell(row=row, column=1).border = thin
    ws.cell(row=row, column=2, value=waktu).alignment = ca
    ws.cell(row=row, column=2).border = thin
    
    coins = ["btc", "eth", "sol", "ada", "doge", "xrp"]
    for i, coin in enumerate(coins):
        base_col = 3 + i * 3
        usd = data.get(f"{coin}_usd", 0)
        idr = data.get(f"{coin}_idr", 0)
        change = data.get(f"{coin}_change", 0)
        
        for col, val in [(base_col, usd), (base_col+1, idr), (base_col+2, round(change, 2))]:
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = ca; c.border = thin
            if col in [base_col, base_col+1] and val:
                c.number_format = '#,##0'
            if col == base_col + 2:
                c.number_format = '0.00'
                if val < 0: c.fill = neg_fill
                elif val > 0: c.fill = pos_fill
    
    total_mcap = data.get("total_mcap", 0)
    c = ws.cell(row=row, column=21, value=round(total_mcap / 1e12, 2) if total_mcap else 0)
    c.alignment = ca; c.border = thin; c.number_format = '0.00'
    
    c = ws.cell(row=row, column=22, value=sentimen)
    c.alignment = ca; c.border = thin
    
    wb.save(EXCEL_PATH)
    print(f"Price row added: {tanggal} {waktu}")

def add_sentimen_row(tanggal, coin, harga, change, sentimen, skor, sumber):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # Ensure Sentimen sheet exists
    if "Sentimen" not in wb.sheetnames:
        ws = wb.create_sheet("Sentimen")
        hf = Font(bold=True, color="FFFFFF", size=10)
        hfill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                      top=Side(style="thin"), bottom=Side(style="thin"))
        
        headers = ["Tanggal", "Koin", "Harga (USD)", "Perubahan 24h", "Sentimen", "Skor", "Sumber Berita"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
        widths2 = [13, 8, 14, 12, 12, 10, 40]
        for i, w in enumerate(widths2, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(EXCEL_PATH)
    
    ws = wb["Sentimen"]
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    neg_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    pos_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    
    row = ws.max_row + 1
    data = [tanggal, coin, harga, change, sentimen, skor, sumber]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = thin
        c.alignment = la if col == 7 else ca
        if col == 3 and val: c.number_format = '#,##0.00'
        if col == 4:
            c.number_format = '0.00'
            if val < 0: c.fill = neg_fill
            elif val > 0: c.fill = pos_fill
    
    wb.save(EXCEL_PATH)

if __name__ == "__main__":
    # Test data
    add_price_row("2026-06-23", "23:30", {
        "btc_usd": 62607, "btc_idr": 1123571395, "btc_change": -1.88,
        "eth_usd": 1664.67, "eth_idr": 29874814, "eth_change": -3.28,
        "sol_usd": 69.56, "sol_idr": 1248406, "sol_change": -2.70,
        "ada_usd": 0.1525, "ada_idr": 2737, "ada_change": -3.58,
        "doge_usd": 0.0791, "doge_idr": 1420, "doge_change": -3.03,
        "xrp_usd": 1.11, "xrp_idr": 19839, "xrp_change": -1.40,
        "total_mcap": 2.25e12
    }, "BEARISH")
    print(f"File saved: {EXCEL_PATH}")
