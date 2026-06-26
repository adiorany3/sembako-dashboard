import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, LineChart, Reference, BarChart
from datetime import datetime
import os

EXCEL_PATH = os.path.expanduser("~/sembako/keuangan.xlsx")

KATEGORI_PENGELUARAN = [
    "Makanan & Minuman", "Transportasi", "Belanja Rumah Tangga",
    "Tagihan & Utilitas", "Kesehatan", "Pendidikan", "Hiburan",
    "Pakaian", "Pertanian & Ternak", "Investasi", "Lainnya"
]

KATEGORI_PEMASUKAN = [
    "Gaji", "Usaha", "Investasi", "Hadiah", "Lainnya"
]

def create_or_load():
    if os.path.exists(EXCEL_PATH):
        return openpyxl.load_workbook(EXCEL_PATH)
    return create_new()

def create_new():
    wb = openpyxl.Workbook()
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    hfill2 = PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid")
    hfill3 = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(horizontal="center", vertical="center")
    la = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    
    # === Sheet 1: Transaksi ===
    ws = wb.active
    ws.title = "Transaksi"
    headers = ["Tanggal", "Jenis", "Kategori", "Deskripsi", "Jumlah (Rp)", 
               "Metode Bayar", "Sumber Nota"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf
        c.fill = hfill3
        c.alignment = ha; c.border = thin
    
    widths = [13, 12, 22, 40, 18, 15, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    
    # === Sheet 2: Ringkasan Bulanan ===
    ws2 = wb.create_sheet("Ringkasan Bulanan")
    headers2 = ["Bulan", "Total Pemasukan", "Total Pengeluaran", "Saldo", "Savings Rate (%)"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill3; c.alignment = ha; c.border = thin
    widths2 = [14, 20, 20, 20, 16]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    
    # === Sheet 3: Per Kategori ===
    ws3 = wb.create_sheet("Per Kategori")
    headers3 = ["Kategori", "Jumlah Transaksi", "Total (Rp)", "Rata-rata", "Persentase"]
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill3; c.alignment = ha; c.border = thin
    for i in range(1, 6):
        ws3.column_dimensions[get_column_letter(i)].width = 20
    ws3.freeze_panes = "A2"
    
    # === Sheet 4: Anggaran ===
    ws4 = wb.create_sheet("Anggaran")
    headers4 = ["Kategori", "Anggaran Bulanan", "Terpakai", "Sisa", "Persentase"]
    for col, h in enumerate(headers4, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill3; c.alignment = ha; c.border = thin
    for i in range(1, 6):
        ws4.column_dimensions[get_column_letter(i)].width = 20
    
    # Default anggaran
    default_budget = {
        "Makanan & Minuman": 2000000, "Transportasi": 500000,
        "Belanja Rumah Tangga": 1000000, "Tagihan & Utilitas": 500000,
        "Kesehatan": 300000, "Hiburan": 200000, "Lainnya": 500000
    }
    for i, (kat, budget) in enumerate(default_budget.items(), 2):
        ws4.cell(row=i, column=1, value=kat).border = thin
        c = ws4.cell(row=i, column=2, value=budget)
        c.border = thin; c.number_format = '#,##0'
        for col in [3, 4, 5]:
            ws4.cell(row=i, column=col).border = thin
    ws4.freeze_panes = "A2"
    
    wb.save(EXCEL_PATH)
    return wb

def add_transaksi(tanggal, jenis, kategori, deskripsi, jumlah, metode="", nota=""):
    """jenis: 'Pemasukan' atau 'Pengeluaran'"""
    wb = create_or_load()
    ws = wb["Transaksi"]
    ca = Alignment(horizontal="center", vertical="center")
    la = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    pos_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    neg_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    row = ws.max_row + 1
    data = [tanggal, jenis, kategori, deskripsi, jumlah, metode, nota]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = thin
        c.alignment = la if col in [4, 7] else ca
        if col == 5:
            c.number_format = '#,##0'
            c.fill = pos_fill if jenis == "Pemasukan" else neg_fill
    
    wb.save(EXCEL_PATH)
    print(f"✅ {jenis}: Rp {jumlah:,} - {deskripsi}")

def update_ringkasan():
    """Update summary sheets."""
    wb = create_or_load()
    ws = wb["Transaksi"]
    
    # Read all transactions
    transaksi = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            transaksi.append(row)
    
    # Monthly summary
    if "Ringkasan Bulanan" in wb.sheetnames:
        del wb["Ringkasan Bulanan"]
    ws2 = wb.create_sheet("Ringkasan Bulanan")
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    
    for col, h in enumerate(["Bulan", "Pemasukan", "Pengeluaran", "Saldo", "Savings %"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
    
    # Group by month
    months = {}
    for t in transaksi:
        tanggal = str(t[0])
        bulan = tanggal[:7]  # YYYY-MM
        if bulan not in months:
            months[bulan] = {"masuk": 0, "keluar": 0}
        if t[1] == "Pemasukan":
            months[bulan]["masuk"] += t[4] or 0
        else:
            months[bulan]["keluar"] += t[4] or 0
    
    row = 2
    for bulan in sorted(months.keys()):
        m = months[bulan]
        saldo = m["masuk"] - m["keluar"]
        savings = round(saldo / m["masuk"] * 100, 1) if m["masuk"] else 0
        
        ws2.cell(row=row, column=1, value=bulan).border = thin
        for col, val in [(2, m["masuk"]), (3, m["keluar"]), (4, saldo)]:
            c = ws2.cell(row=row, column=col, value=val)
            c.border = thin; c.alignment = ha; c.number_format = '#,##0'
        c = ws2.cell(row=row, column=5, value=savings)
        c.border = thin; c.alignment = ha; c.number_format = '0.0'
        row += 1
    
    # Category summary
    if "Per Kategori" in wb.sheetnames:
        del wb["Per Kategori"]
    ws3 = wb.create_sheet("Per Kategori")
    for col, h in enumerate(["Kategori", "Jumlah", "Total (Rp)", "Rata-rata", "Persen"], 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
    
    categories = {}
    for t in transaksi:
        if t[1] == "Pengeluaran":
            kat = t[2] or "Lainnya"
            if kat not in categories:
                categories[kat] = {"count": 0, "total": 0}
            categories[kat]["count"] += 1
            categories[kat]["total"] += t[4] or 0
    
    total_all = sum(c["total"] for c in categories.values())
    row = 2
    for kat in sorted(categories.keys(), key=lambda k: categories[k]["total"], reverse=True):
        c = categories[kat]
        avg = c["total"] // c["count"] if c["count"] else 0
        pct = round(c["total"] / total_all * 100, 1) if total_all else 0
        
        ws3.cell(row=row, column=1, value=kat).border = thin
        ws3.cell(row=row, column=2, value=c["count"]).border = thin
        for col, val in [(3, c["total"]), (4, avg)]:
            c2 = ws3.cell(row=row, column=col, value=val)
            c2.border = thin; c2.number_format = '#,##0'
        ws3.cell(row=row, column=5, value=pct).border = thin
        row += 1
    
    wb.save(EXCEL_PATH)
    print("✅ Ringkasan diperbarui")

if __name__ == "__main__":
    # Sample data
    add_transaksi("2026-06-20", "Pengeluaran", "Makanan & Minuman", "Nasi goreng + es teh", 25000, "Tunai")
    add_transaksi("2026-06-20", "Pengeluaran", "Transportasi", "Grab ke kantor", 35000, "GoPay")
    add_transaksi("2026-06-21", "Pemasukan", "Gaji", "Gaji Juni 2026", 8000000, "Transfer BCA")
    add_transaksi("2026-06-21", "Pengeluaran", "Belanja Rumah Tangga", "Belanja bulanan Indomaret", 350000, "Debit BCA")
    add_transaksi("2026-06-22", "Pengeluaran", "Pertanian & Ternak", "Pakan ayam 50kg", 400000, "Tunai")
    add_transaksi("2026-06-22", "Pengeluaran", "Tagihan & Utilitas", "Listrik PLN", 250000, "GoPay")
    add_transaksi("2026-06-23", "Pemasukan", "Usaha", "Jual telur ayam", 750000, "Tunai")
    add_transaksi("2026-06-23", "Pengeluaran", "Makanan & Minuman", "Warung makan", 45000, "QRIS")
    
    update_ringkasan()
    print(f"\nFile: {EXCEL_PATH}")
