"""Test dedup utilities."""
import os
import sys
import pytest
import openpyxl
from pathlib import Path

# Import from project
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))


class TestDedupExcelFile:
    def test_dedup_removes_duplicates(self, temp_excel_with_dupes):
        """Dedup should remove duplicate rows."""
        from utils.dedup import dedup_excel_file
        report = dedup_excel_file(str(temp_excel_with_dupes), dry_run=False, verbose=False)
        assert report['status'] == 'ok'
        assert report['duplicates_removed'] > 0
        assert report['rows_after'] < report['rows_before']

    def test_dedup_dry_run_no_changes(self, temp_excel_with_dupes):
        """Dry run should not modify file."""
        from utils.dedup import dedup_excel_file
        # Read original row count
        wb = openpyxl.load_workbook(str(temp_excel_with_dupes))
        original_rows = wb['TestSheet'].max_row
        wb.close()

        report = dedup_excel_file(str(temp_excel_with_dupes), dry_run=True, verbose=False)
        assert report['duplicates_removed'] > 0

        # File should be unchanged
        wb = openpyxl.load_workbook(str(temp_excel_with_dupes))
        assert wb['TestSheet'].max_row == original_rows
        wb.close()

    def test_dedup_no_duplicates(self, temp_excel):
        """No duplicates should be reported as zero."""
        from utils.dedup import dedup_excel_file
        report = dedup_excel_file(str(temp_excel), dry_run=False, verbose=False)
        assert report['duplicates_removed'] == 0

    def test_dedup_missing_file(self, tmp_path):
        """Missing file should return ok with zero changes."""
        from utils.dedup import dedup_excel_file
        report = dedup_excel_file(str(tmp_path / "nonexistent.xlsx"), dry_run=False, verbose=False)
        assert report['status'] == 'error'
