#!/usr/bin/env python3
"""
Master Update Script - Update all dashboard data
- Generates 30 days historical data for new installations
- Updates daily data for all categories
- Saves to ~/sembako/ AND commits to GitHub

Usage:
    python3 master_update.py --init    # Initial setup (30 days back)
    python3 master_update.py            # Daily update
"""

import os
import sys
import json
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Base paths
SEMBAKO_DIR = os.path.expanduser('~/sembako')
GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN', '')
GIT_REMOTE = 'https://github.com/adiorany3/sembako-dashboard.git'

# ==========================================
# STYLE DEFINITIONS
# ==========================================
header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
date_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
date_font = Font(bold=True, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def git_commit_push(message):
    """Commit and push changes to GitHub"""
    if not os.path.exists(f'{SEMBAKO_DIR}/.git'):
        print("⚠️ Not a git repo, skipping push")
        return False
    
    os.system(f'cd {SEMBAKO_DIR} && git add . > /dev/null 2>&1')
    result = os.system(f'cd {SEMBAKO_DIR} && git commit -m "{message}" > /dev/null 2>&1')
    if result != 0:
        print("⚠️ Nothing to commit")
        return False
    
    result = os.system(f'cd {SEMBAKO_DIR} && git push origin main > /dev/null 2>&1')
    if result == 0:
        print("  ✅ Pushed to GitHub")
        return True
    else:
        print("  ⚠️ Push failed (will retry next time)")
        return False

# ==========================================
# DATA DEFINITIONS
# ==========================================

# NOTE: Real data is fetched by individual scraping scripts.
# These definitions are kept only for --init fallback mode.

# ==========================================
# EXCEL CREATION FUNCTIONS
# ==========================================

def create_sembako_excel(days_back=30):
    """Create/update harga_sembako.xlsx with historical data"""
    print("\n📊 Updating Sembako data...")
    filepath = f'{SEMBAKO_DIR}/harga_sembako.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Harga'
    
    # Headers
    headers = ['Tanggal', 'Beras Premium', 'Beras Medium', 'Minyak Goreng', 'Gula Pasir',
               'Garam', 'Tepung Terigu', 'Cabai Merah', 'Cabai Rawit', 'Bawang Merah',
               'Bawang Putih', 'Minyak Tanah', 'Telur Ras', 'Telur Kampung', 'Ayam Ras',
               'Ayam Kampung', 'Daging Sapi', 'Gas Elpiji', 'Garam Bata', 'Garam Halus',
               'Susu KM Bendera', 'Susu KM Indomilk', 'Susu Bubuk Bendera', 'Susu Bubuk Indomilk']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Generate historical data
    today = datetime.now()
    for day_offset in range(days_back, -1, -1):
        date = today - timedelta(days=day_offset)
        row = [date.strftime('%Y-%m-%d')]
        
        for product, base_price, satuan in SEMBAKO_PRODUCTS:
            # Add slight variation
            variation = base_price * random.uniform(-0.05, 0.05)
            price = round((base_price + variation) / 100) * 100
            row.append(int(price))
        
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=days_back - day_offset + 2, column=col, value=value)
            if col == 1:
                cell.fill = date_fill
                cell.font = date_font
            cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    wb.save(filepath)
    print(f"  ✅ {filepath} - {days_back + 1} days, {len(SEMBAKO_PRODUCTS)} products")

def create_crypto_excel(days_back=30):
    """Create/update crypto_monitor.xlsx"""
    print("\n🪙 Updating Crypto data...")
    filepath = f'{SEMBAKO_DIR}/crypto_monitor.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    
    headers = ['Tanggal', 'Waktu', 'BTC USD', 'BTC IDR', 'BTC 24h',
               'ETH USD', 'ETH IDR', 'ETH 24h', 'SOL USD', 'SOL IDR', 'SOL 24h',
               'ADA USD', 'ADA IDR', 'ADA 24h', 'DOGE USD', 'DOGE IDR', 'DOGE 24h',
               'XRP USD', 'XRP IDR', 'XRP 24h', 'Market Cap', 'Sentimen']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    today = datetime.now()
    for day_offset in range(days_back, -1, -1):
        date = today - timedelta(days=day_offset)
        row = [date.strftime('%Y-%m-%d'), '08:00']
        
        for coin, data in CRYPTO_BASE.items():
            usd_price = data['price'] / 15000  # Approx USD
            idr_price = data['price']
            change = data['change'] + random.uniform(-2, 2)
            row.extend([round(usd_price, 2), idr_price, round(change, 2)])
        
        row.extend(['High', 'Netral'])
        
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=days_back - day_offset + 2, column=col, value=value)
            if col == 1:
                cell.fill = date_fill
                cell.font = date_font
            cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12
    
    wb.save(filepath)
    print(f"  ✅ {filepath} - {days_back + 1} days")

def create_emas_excel(days_back=30):
    """Create/update harga_emas.xlsx"""
    print("\n💰 Updating Emas data...")
    filepath = f'{SEMBAKO_DIR}/harga_emas.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Harian'
    
    headers = ['Tanggal', 'Antam Beli', 'Antam Buyback', 'Antam Pegadaian', 
               'Galeri 24', 'UBS Beli', 'Selisih', 'Spread %']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    today = datetime.now()
    for day_offset in range(days_back, -1, -1):
        date = today - timedelta(days=day_offset)
        row = [date.strftime('%Y-%m-%d')]
        
        for name, base_price in GOLD_BASE.items():
            variation = base_price * random.uniform(-0.02, 0.02)
            price = round((base_price + variation) / 1000) * 1000
            row.append(int(price))
        
        selisih = row[-1] - row[1]
        spread = round((selisih / row[1]) * 100, 2)
        row.extend([selisih, spread])
        
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=days_back - day_offset + 2, column=col, value=value)
            if col == 1:
                cell.fill = date_fill
                cell.font = date_font
            cell.border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    wb.save(filepath)
    print(f"  ✅ {filepath} - {days_back + 1} days")

def create_pertanian_excel(days_back=30):
    """Create/update harga_pertanian_ternak.xlsx"""
    print("\n🌾 Updating Pertanian data...")
    filepath = f'{SEMBAKO_DIR}/harga_pertanian_ternak.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    headers = ['Tanggal', 'Jagung Pipil', 'Jagung Pakan', 'Kedelai Impor', 'Kedelai Lokal',
               'Pakan Broiler', 'Pakan Layer', 'Bungkil Kedelai', 'Jagung Giling']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    today = datetime.now()
    for day_offset in range(days_back, -1, -1):
        date = today - timedelta(days=day_offset)
        row = [date.strftime('%Y-%m-%d')]
        
        for product, base_price, satuan in PERTANIAN_PRODUCTS:
            variation = base_price * random.uniform(-0.05, 0.05)
            price = round((base_price + variation) / 100) * 100
            row.append(int(price))
        
        for col, value in enumerate(row, 1):
            cell = ws.cell(row=days_back - day_offset + 2, column=col, value=value)
            if col == 1:
                cell.fill = date_fill
                cell.font = date_font
            cell.border = thin_border
    
    ws.column_dimensions['A'].width = 12
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    wb.save(filepath)
    print(f"  ✅ {filepath} - {days_back + 1} days")

def daily_update_all():
    """Call real scraping scripts to get actual data, then dedup."""
    print("\n📝 Running real scraping scripts...")
    
    today = datetime.now().strftime('%Y-%m-%d')
    scripts_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Real scraping scripts that fetch actual prices from web
    scrapers = [
        ("Sembako", "update_harga.py"),
        ("Crypto", "update_crypto.py"),
        ("Emas", "update_emas.py"),
        ("Pertanian", "update_pertanian.py"),
        ("Peternakan", "update_peternakan.py"),
        ("Pakan", "update_pakan_nutrisi.py"),
        ("Saham/IHSG", "update_saham.py"),
        ("Kurs Valuta", "update_kurs.py"),
        ("BI Rate", "update_bi_rate.py"),
        ("Cuaca", "fetch_cuaca_pagi.py"),
        ("CPO", "update_cpo.py"),
        ("Oil", "update_oil.py"),
    ]
    
    success = 0
    failed = 0
    for name, script in scrapers:
        script_path = os.path.join(scripts_dir, script)
        if os.path.exists(script_path):
            print(f"  🔍 {name}...")
            result = os.system(f'cd {SEMBAKO_DIR} && python3 {script_path} > /dev/null 2>&1')
            if result == 0:
                print(f"    ✅ {name} updated")
                success += 1
            else:
                print(f"    ⚠️ {name} failed (exit {result})")
                failed += 1
        else:
            print(f"    ❌ {script} not found")
            failed += 1
    
    # Auto-dedup all files after update
    _parent = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _parent not in sys.path:
        sys.path.insert(0, _parent)
    from utils.dedup import dedup_all
    reports = dedup_all(dry_run=False, verbose=False)
    removed = sum(r['duplicates_removed'] for r in reports)
    if removed > 0:
        print(f"\n🧹 Auto-dedup: removed {removed} duplicate rows")
    
    # Recovery: reload JSON history → Excel if data missing
    try:
        _scripts = os.path.dirname(os.path.abspath(__file__))
        if _scripts not in sys.path:
            sys.path.insert(0, _scripts)
        from recovery_json_to_excel import load_crypto_history_to_excel, load_emas_history_to_excel
        load_crypto_history_to_excel()
        load_emas_history_to_excel()
    except Exception as e:
        print(f"  ⚠️ Recovery loader error: {e}")
    
    print(f"\n📊 Results: {success} succeeded, {failed} failed")

# ==========================================
# MAIN
# ==========================================

if __name__ == '__main__':
    init_mode = len(sys.argv) > 1 and sys.argv[1] == '--init'
    
    if init_mode:
        print("=" * 50)
        print("🚀 INITIAL SETUP - Fetching real data from all scrapers")
        print("=" * 50)
        
        # Init mode also uses real scrapers (same as daily)
        daily_update_all()
        
        print("\n" + "=" * 50)
        print("✅ ALL DATA INITIALIZED!")
        print("=" * 50)
        
        # Commit and push
        git_commit_push("Initial data setup - real scraped data")
        
    else:
        print("=" * 50)
        print("📅 DAILY UPDATE")
        print("=" * 50)
        
        daily_update_all()
        
        print("\n" + "=" * 50)
        print("✅ DAILY UPDATE COMPLETE!")
        print("=" * 50)
        
        # Commit and push
        git_commit_push(f"Daily update - {datetime.now().strftime('%Y-%m-%d')}")
