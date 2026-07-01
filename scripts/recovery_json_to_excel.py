#!/usr/bin/env python3
"""
Recovery: load JSON history → Excel for crypto & emas.
Run when Excel data is missing/stale but JSON history has data.
"""
import json
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.expanduser("~/sembako"))

DATA_DIR = os.path.expanduser("~/sembako/data")


def load_crypto_history_to_excel():
    """Load crypto_history.json → crypto_monitor.xlsx (fill missing dates)."""
    json_path = os.path.join(DATA_DIR, "crypto_history.json")
    excel_path = os.path.join(DATA_DIR, "crypto_monitor.xlsx")
    
    if not os.path.exists(json_path):
        print("  ⚠️ crypto_history.json not found")
        return 0
    
    with open(json_path) as f:
        history = json.load(f)
    
    if not history:
        print("  ⚠️ crypto_history.json is empty")
        return 0
    
    # Import and use add_price_row
    from scripts.create_crypto_excel import add_price_row
    
    import openpyxl
    # Get existing dates from Excel
    existing_dates = set()
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        if "Harga" in wb.sheetnames:
            ws = wb["Harga"]
            for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
                if row[0]:
                    existing_dates.add(str(row[0])[:10])
        wb.close()
    
    loaded = 0
    # Dedup JSON: keep latest entry per date
    seen = {}
    for entry in history:
        date_str = str(entry.get("date", ""))[:10]
        seen[date_str] = entry
    for entry in seen.values():
        date_str = str(entry.get("date", ""))[:10]
        if date_str in existing_dates:
            continue
        
        prices = entry.get("prices", {})
        if not prices:
            continue
        
        waktu = entry.get("time", "08:00")
        sentimen = entry.get("sentiment", "")
        add_price_row(date_str, waktu, prices, sentimen)
        loaded += 1
    
    print(f"  ✅ Crypto: loaded {loaded} rows from JSON history ({len(history)} total, {len(existing_dates)} already in Excel)")
    return loaded


def load_emas_history_to_excel():
    """Load emas_history.json → harga_emas.xlsx (fill missing dates)."""
    json_path = os.path.join(DATA_DIR, "emas_history.json")
    excel_path = os.path.join(DATA_DIR, "harga_emas.xlsx")
    
    if not os.path.exists(json_path):
        print("  ⚠️ emas_history.json not found")
        return 0
    
    with open(json_path) as f:
        content = f.read().strip()
        if not content:
            print("  ⚠️ emas_history.json is empty")
            return 0
        history = json.loads(content)
    
    if not history:
        print("  ⚠️ emas_history.json is empty")
        return 0
    
    from scripts.create_emas_excel import add_daily_row
    
    import openpyxl
    # Get existing dates from Excel
    existing_dates = set()
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        if "Harian" in wb.sheetnames:
            ws = wb["Harian"]
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0]:
                    existing_dates.add(str(row[0])[:10])
        wb.close()
    
    loaded = 0
    for entry in history:
        date_str = str(entry.get("date", ""))[:10]
        if date_str in existing_dates:
            continue
        
        antam_beli = entry.get("antam_beli", 0)
        antam_buyback = entry.get("antam_buyback", 0)
        pegadaian = entry.get("antam_pegadaian", 0)
        galeri24 = entry.get("galeri24", 0)
        ubs = entry.get("ubs", 0)
        
        if antam_beli:
            add_daily_row(date_str, antam_beli, antam_buyback, pegadaian, galeri24, ubs)
            loaded += 1
    
    print(f"  ✅ Emas: loaded {loaded} rows from JSON history ({len(history)} total, {len(existing_dates)} already in Excel)")
    return loaded


if __name__ == "__main__":
    print("=" * 50)
    print("🔄 RECOVERY: JSON → Excel")
    print("=" * 50)
    c = load_crypto_history_to_excel()
    e = load_emas_history_to_excel()
    print(f"\nTotal: {c + e} rows recovered")
