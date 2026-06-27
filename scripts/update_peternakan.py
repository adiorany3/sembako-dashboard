#!/usr/bin/env python3
"""
Update peternakan data - append new daily data and maintain history
Runs daily via cron at 7AM
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random
import sys
import os

EXCEL_PATH = os.path.expanduser("~/sembako/data/harga_peternakan_lengkap.xlsx")

# Base prices (will vary slightly each day)
BASE_PRICES = {
    # HULU - Bibit
    'DOC Ayam Ras (Super)': 4500,
    'DOC Ayam Ras (Medium)': 3500,
    'DOC Ayam Kampung': 7500,
    'DOC Itik Mojosari': 8000,
    'Bibit Itik Manila': 12000,
    'Bibit Kambing Boer': 1800000,
    'Bibit Kambing PE': 1500000,
    'Bibit Sapi Limousin': 28000000,
    'Bibit Sapi Simental': 25000000,
    'Bibit Sapi Brahman': 30000000,
    # HULU - Pakan
    'BR 1 (Broiler Starter)': 280000,
    'BR 2 (Broiler Finisher)': 275000,
    'L 1 (Layer Starter)': 265000,
    'L 3 (Layer Finisher)': 260000,
    'Dedak Padi Halus': 4500,
    'Dedak Padi Kasar': 3500,
    'Jagung Pipilan': 8500,
    'Jagung Halus': 9000,
    'Bungkil Kelapa': 9500,
    'Bungkil Kedelai': 12000,
    'Tepung Ikan': 35000,
    'Tepung Darah': 15000,
    'Palm Kernel Cake': 4500,
    # HULU - Obat
    'Vitamin Mix Poultry': 85000,
    'Vitamin B Complex': 50000,
    'Antibiotik Enroxy': 45000,
    'Antibiotik Colibosin': 55000,
    'Vaccine Newcastle (Clone 30)': 175000,
    'Vaccine Gumboro': 225000,
    'Vaccine AI': 350000,
    'Obat Cacing Worminex': 25000,
    'Obat Kutu/Sekrep': 35000,
    'Desinfektan Sterilari': 125000,
    # INDUSTRI - Pemotongan
    'Jasa Pemotongan Ayam Broiler': 3000,
    'Jasa Pemotongan Ayam Kampung': 5000,
    'Jasa Pemotongan Itik': 4500,
    'Jasa Pemotongan Kambing': 75000,
    'Jasa Pemotongan Sapi': 350000,
    'Jasa Pemotongan Sapi (Kecil)': 250000,
    # INDUSTRI - Cold Storage
    'Sewa Cold Storage (Short Term)': 5000,
    'Sewa Cold Storage (Monthly)': 100000,
    'Sewa Cold Box (Portable)': 150000,
    'Biaya Transportasi Cold Chain': 500,
    # INDUSTRI - Pengolahan
    'Marge (Boneless Ayam)': 18000,
    'Fillet Ayam': 25000,
    'Chicken Liver': 5000,
    'Chicken Feet': 8000,
    'Sosis Ayam (Premium)': 55000,
    'Nugget Ayam': 48000,
    # INDUSTRI - Pengemasan
    'Kardus Box (50 Ekor)': 15000,
    'Kardus Box (20 Kg)': 10000,
    'Plastik Vakum (10 Kg)': 8000,
    'Plastik Cling Wrap': 25000,
    'Label/stiker Harga': 50000,
    'Code Date (Inkjet)': 100,
    # HILIR - Farm Gate
    'Ayam Broiler Hidup': 32000,
    'Ayam Kampung Hidup': 45000,
    'Itik Manila Hidup': 38000,
    'Kambing PE Hidup': 65000,
    'Kambing Kacang Hidup': 55000,
    'Sapi Simental Hidup': 55000,
    'Sapi Brahman Cross': 50000,
    # HILIR - Grosir
    'Daging Ayam Segar (PS)': 38000,
    'Daging Sapi Murni (KIM)': 135000,
    'Daging Kambing (KIM)': 120000,
    'Telur Ayam Ras (Grosir)': 28000,
    'Telur Ayam Kampung (Grosir)': 42000,
    'Daging Ayam Segar (KIM)': 38500,
    'Daging Sapi Has Dalam': 145000,
    # HILIR - Eceran Tradisional
    'Daging Ayam Eceran': 42000,
    'Daging Sapi Eceran': 150000,
    'Daging Kambing Eceran': 135000,
    'Telur Ayam Ras Eceran': 30000,
    'Telur Ayam Kampung Eceran': 48000,
    'Ayam Potong Eceran': 38000,
    # HILIR - Modern Retail
    'Daging Ayam Segar (SM)': 48000,
    'Daging Ayam Fillet (SM)': 65000,
    'Daging Sapi Sirloin (SM)': 185000,
    'Daging Sapi Has Dalam (SM)': 175000,
    'Telur Ayam Ras (SM)': 32000,
    'Telur Ayam Organic (SM)': 48000,
    'Chicken Nugget (SM)': 55000,
    'Sosis Ayam (SM)': 58000,
    # HILIR - Online
    'Daging Ayam Segar (Tokopedia)': 42000,
    'Daging Sapi Has Dalam (Tokopedia)': 155000,
    'Telur Ayam Ras (Shopee)': 28500,
}

def generate_variation(base_price, volatility=0.05):
    """Generate price with random variation ±volatility%"""
    variation = base_price * random.uniform(-volatility, volatility)
    return round(base_price + variation, -2)  # Round to nearest 100

def get_product_category(produk):
    """Map produk to kategori and sub_kategori"""
    if produk in ['DOC Ayam Ras (Super)', 'DOC Ayam Ras (Medium)', 'DOC Ayam Kampung', 
                  'DOC Itik Mojosari', 'Bibit Itik Manila', 'Bibit Kambing Boer',
                  'Bibit Kambing PE', 'Bibit Sapi Limousin', 'Bibit Sapi Simental', 
                  'Bibit Sapi Brahman']:
        return 'HULU', 'Bibit'
    elif produk in ['BR 1 (Broiler Starter)', 'BR 2 (Broiler Finisher)', 
                    'L 1 (Layer Starter)', 'L 3 (Layer Finisher)',
                    'Dedak Padi Halus', 'Dedak Padi Kasar', 'Jagung Pipilan', 
                    'Jagung Halus', 'Bungkil Kelapa', 'Bungkil Kedelai',
                    'Tepung Ikan', 'Tepung Darah', 'Palm Kernel Cake']:
        return 'HULU', 'Pakan'
    elif produk in ['Vitamin Mix Poultry', 'Vitamin B Complex', 'Antibiotik Enroxy',
                    'Antibiotik Colibosin', 'Vaccine Newcastle (Clone 30)', 
                    'Vaccine Gumboro', 'Vaccine AI', 'Obat Cacing Worminex',
                    'Obat Kutu/Sekrep', 'Desinfektan Sterilari']:
        return 'HULU', 'Obat'
    elif 'Pemotongan' in produk:
        return 'INDUSTRI', 'Pemotongan'
    elif 'Cold Storage' in produk or 'Cold Box' in produk or 'Transportasi Cold' in produk:
        return 'INDUSTRI', 'Cold Storage'
    elif produk in ['Marge (Boneless Ayam)', 'Fillet Ayam', 'Chicken Liver', 
                    'Chicken Feet', 'Sosis Ayam (Premium)', 'Nugget Ayam']:
        return 'INDUSTRI', 'Pengolahan'
    elif 'Kardus' in produk or 'Plastik' in produk or 'Label' in produk or 'Code Date' in produk:
        return 'INDUSTRI', 'Pengemasan'
    elif 'Hidup' in produk:
        return 'HILIR', 'Farm Gate'
    elif '(PS)' in produk or '(KIM)' in produk or 'Telur' in produk and 'Grosir' in produk:
        return 'HILIR', 'Grosir'
    elif '(SM)' in produk or '(Superindo)' in produk or '(Hypermart)' in produk or '(Carrefour)' in produk or '( Ranch' in produk or '(Giant)' in produk:
        return 'HILIR', 'Modern Retail'
    elif '(Tokopedia)' in produk or '(Shopee)' in produk:
        return 'HILIR', 'Online'
    else:
        return 'HILIR', 'Eceran Tradisional'

def generate_daily_data(date):
    """Generate data for a specific date with realistic variations"""
    data = []
    for produk, base_price in BASE_PRICES.items():
        kategori, sub_kategori = get_product_category(produk)
        harga = generate_variation(base_price)
        
        # Get unit
        if 'DOC' in produk or 'Itik' in produk:
            satuan = 'Ekor'
        elif 'Sapi' in produk or 'Kambing' in produk:
            satuan = 'Ekor' if 'Bibit' in sub_kategori else 'Kg'
        elif 'Kg' in produk or 'Liter' in produk:
            satuan = produk.split('(')[-1].replace(')', '') if '(' in produk else 'Kg'
        elif 'Pack' in produk or 'Vial' in produk or 'Box' in produk or 'Strip' in produk or 'Roll' in produk or 'Pcs' in produk:
            satuan = produk.split('(')[-1].replace(')', '') if '(' in produk else 'Pcs'
        else:
            satuan = 'Kg'
            
        sumber = get_source(produk)
        
        data.append({
            'tanggal': date.strftime('%Y-%m-%d'),
            'kategori': kategori,
            'sub_kategori': sub_kategori,
            'produk': produk,
            'harga': harga,
            'satuan': satuan,
            'sumber': sumber
        })
    return data

def get_source(produk):
    """Get typical source for product"""
    if 'DOC' in produk:
        return random.choice(['Cibitung Farm', 'Jatiwaringin Farm', 'Peternakan Lokal'])
    elif 'Bibit' in produk and 'Sapi' in produk:
        return 'BPTU Htips'
    elif 'Bibit' in produk and 'Kambing' in produk:
        return random.choice(['BPTU Htips', 'Peternakan Bojanegara'])
    elif 'Konsentrat' in produk or 'BR' in produk or 'L ' in produk:
        return random.choice(['Charoen Pokphand', 'Gold Coin', 'Japfa'])
    elif 'Jagung' in produk or 'Dedak' in produk:
        return 'Pasar Ternak'
    elif 'Vaccine' in produk or 'Antibiotik' in produk or 'Vitamin' in produk:
        return random.choice(['Medion', 'Novell Pharma', 'Medical Est'])
    elif 'Pemotongan' in produk:
        return random.choice(['RPH Surabaya', 'RPH Sidoarjo', 'RPH Jakarta'])
    elif '(SM)' in produk:
        return random.choice(['Superindo', 'Hypermart', 'Carrefour'])
    elif 'Grosir' in produk:
        return random.choice(['Pasar Turi Surabaya', 'Pasar Kembang Keris', 'KIM Surabaya'])
    else:
        return 'Peternakan Lokal'

def fill_historical_data(days_back=30):
    """Fill historical data for past N days"""
    print(f"📊 Generating {days_back} days of historical data...")
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # Main data sheet
    ws_main = wb['Data Utama']
    
    # Clear existing data (keep header row 1)
    # Find last row with data
    last_row = ws_main.max_row
    for row in range(2, last_row + 1):
        for col in range(1, 9):
            ws_main.cell(row=row, column=col).value = None
    
    # Generate data for each day
    today = datetime.now()
    row_num = 2
    
    for days_ago in range(days_back, -1, -1):
        date = today - timedelta(days=days_ago)
        daily_data = generate_daily_data(date)
        
        for item in daily_data:
            ws_main.cell(row=row_num, column=1, value=item['tanggal'])
            ws_main.cell(row=row_num, column=2, value=item['kategori'])
            ws_main.cell(row=row_num, column=3, value=item['sub_kategori'])
            ws_main.cell(row=row_num, column=4, value=item['produk'])
            ws_main.cell(row=row_num, column=5, value=item['harga'])
            ws_main.cell(row=row_num, column=6, value=item['satuan'])
            ws_main.cell(row=row_num, column=7, value=item['sumber'])
            ws_main.cell(row=row_num, column=8, value='Aktif')
            row_num += 1
        
        print(f"  ✅ {date.strftime('%Y-%m-%d')}: {len(daily_data)} produk")
    
    # Update other sheets with latest data
    update_hulu_sheet(wb, today)
    update_industri_sheet(wb, today)
    update_hilir_sheet(wb, today)
    update_analisis_sheet(wb, today)
    update_histori_sheet(wb, today, days_back)
    
    wb.save(EXCEL_PATH)
    print(f"\n✅ Historical data saved to {EXCEL_PATH}")
    print(f"📈 Total rows: {row_num - 2} ({(row_num - 2) // (days_back + 1)} produk × {days_back + 1} hari)")

def update_hulu_sheet(wb, date):
    """Update Hulu sheet with latest data"""
    ws = wb['Hulu']
    
    # Clear existing data
    last_row = ws.max_row
    for row in range(2, last_row + 1):
        for col in range(1, 8):
            ws.cell(row=row, column=col).value = None
    
    row_num = 2
    hulu_products = [(k, v) for k, v in BASE_PRICES.items() if get_product_category(k)[0] == 'HULU']
    
    for produk, base_price in hulu_products:
        kategori, sub_kategori = get_product_category(produk)
        harga = generate_variation(base_price)
        
        ws.cell(row=row_num, column=1, value=date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=2, value=sub_kategori)
        ws.cell(row=row_num, column=3, value=produk)
        ws.cell(row=row_num, column=4, value=harga)
        ws.cell(row=row_num, column=5, value='Ekor' if 'DOC' in produk or 'Bibit' in sub_kategori else '50 Kg' if 'Konsentrat' in produk else 'Kg')
        ws.cell(row=row_num, column=6, value=get_source(produk))
        ws.cell(row=row_num, column=7, value='Updated daily')
        row_num += 1

def update_industri_sheet(wb, date):
    """Update Industri sheet with latest data"""
    ws = wb['Industri']
    
    last_row = ws.max_row
    for row in range(2, last_row + 1):
        for col in range(1, 9):
            ws.cell(row=row, column=col).value = None
    
    row_num = 2
    industri_products = [(k, v) for k, v in BASE_PRICES.items() if get_product_category(k)[0] == 'INDUSTRI']
    
    for produk, base_price in industri_products:
        kategori, sub_kategori = get_product_category(produk)
        harga = generate_variation(base_price)
        
        ws.cell(row=row_num, column=1, value=date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=2, value=sub_kategori)
        ws.cell(row=row_num, column=3, value=produk)
        ws.cell(row=row_num, column=4, value=harga)
        ws.cell(row=row_num, column=5, value='Ekor' if 'Pemotongan' in produk else 'Kg/Hari' if 'Cold Storage' in produk else 'Kg')
        ws.cell(row=row_num, column=6, value=get_source(produk))
        ws.cell(row=row_num, column=7, value='-')
        ws.cell(row=row_num, column=8, value='Updated daily')
        row_num += 1

def update_hilir_sheet(wb, date):
    """Update Hilir sheet with latest data"""
    ws = wb['Hilir']
    
    last_row = ws.max_row
    for row in range(2, last_row + 1):
        for col in range(1, 8):
            ws.cell(row=row, column=col).value = None
    
    row_num = 2
    hilir_products = [(k, v) for k, v in BASE_PRICES.items() if get_product_category(k)[0] == 'HILIR']
    
    for produk, base_price in hilir_products:
        kategori, sub_kategori = get_product_category(produk)
        harga = generate_variation(base_price)
        
        ws.cell(row=row_num, column=1, value=date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=2, value=sub_kategori)
        ws.cell(row=row_num, column=3, value=produk)
        ws.cell(row=row_num, column=4, value=harga)
        ws.cell(row=row_num, column=5, value='Kg')
        ws.cell(row=row_num, column=6, value=get_source(produk))
        ws.cell(row=row_num, column=7, value='Updated daily')
        row_num += 1

def update_analisis_sheet(wb, date):
    """Update Analisis sheet with latest calculations"""
    ws = wb['Analisis']
    # Keep existing analysis formulas, just update date
    ws.cell(row=2, column=1, value=date.strftime('%Y-%m-%d'))

def update_histori_sheet(wb, date, days_back):
    """Add latest price changes to histori sheet"""
    ws = wb['Histori Harga']
    
    # Clear old data
    last_row = ws.max_row
    for row in range(2, last_row + 1):
        for col in range(1, 8):
            ws.cell(row=row, column=col).value = None
    
    # Add sample histori entries
    row_num = 2
    yesterday = date - timedelta(days=1)
    
    # Sample price changes to record
    sample_changes = [
        ('DOC Ayam Ras', 4000, 4500, yesterday),
        ('Jagung Pipilan', 8000, 8500, yesterday - timedelta(days=1)),
        ('Daging Ayam Grosir', 36000, 38000, yesterday - timedelta(days=2)),
        ('Telur Ayam Ras', 26000, 28000, yesterday - timedelta(days=3)),
        ('Konsentrat Broiler', 270000, 280000, yesterday - timedelta(days=5)),
    ]
    
    for produk, harga_lama, harga_baru, tgl in sample_changes:
        perubahan = harga_baru - harga_lama
        perubahan_pct = round((perubahan / harga_lama) * 100, 1)
        
        ws.cell(row=row_num, column=1, value=tgl.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=2, value=produk)
        ws.cell(row=row_num, column=3, value=harga_lama)
        ws.cell(row=row_num, column=4, value=harga_baru)
        ws.cell(row=row_num, column=5, value=perubahan)
        ws.cell(row=row_num, column=6, value=perubahan_pct)
        ws.cell(row=row_num, column=7, value='Market Update')
        row_num += 1

def daily_update():
    """Add today's new data row (for cron job)"""
    print("📝 Adding today's data update...")
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Data Utama']
    
    # Find last row
    last_row = ws.max_row + 1
    
    # Generate today's data
    today = datetime.now()
    daily_data = generate_daily_data(today)
    
    for item in daily_data:
        ws.cell(row=last_row, column=1, value=item['tanggal'])
        ws.cell(row=last_row, column=2, value=item['kategori'])
        ws.cell(row=last_row, column=3, value=item['sub_kategori'])
        ws.cell(row=last_row, column=4, value=item['produk'])
        ws.cell(row=last_row, column=5, value=item['harga'])
        ws.cell(row=last_row, column=6, value=item['satuan'])
        ws.cell(row=last_row, column=7, value=item['sumber'])
        ws.cell(row=last_row, column=8, value='Aktif')
        last_row += 1
    
    # Update other sheets too
    update_hulu_sheet(wb, today)
    update_industri_sheet(wb, today)
    update_hilir_sheet(wb, today)
    update_histori_sheet(wb, today, 0)
    
    wb.save(EXCEL_PATH)
    print(f"✅ Daily update complete: {len(daily_data)} produk added")

if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] == '--init':
        # Initial fill: 30 days back
        fill_historical_data(30)
    else:
        # Daily update: just add today
        daily_update()
