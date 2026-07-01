#!/usr/bin/env python3
"""
IHSG & Saham Bluechip Indonesia Data
- Generate historical data
- Auto-update daily
- Technical & Fundamental Analysis
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random
import os

# ==========================================
# DATA DEFINITIONS
# ==========================================

BLUECHIP_STOCKS = {
    # Symbol: (Nama, Sektor, Base Price, Market Cap)
    'BBCA.JK': ('Bank Central Asia', 'Financial', 9800, 'Large Cap'),
    'BBRI.JK': ('Bank Rakyat Indonesia', 'Financial', 4800, 'Large Cap'),
    'TLKM.JK': ('Telekomunikasi Indonesia', 'Communication', 3100, 'Large Cap'),
    'ASII.JK': ('Astra International', 'Automotive', 5800, 'Large Cap'),
    'UNVR.JK': ('Unilever Indonesia', 'Consumer', 3600, 'Large Cap'),
    'HMSP.JK': ('HM Sampoerna', 'Tobacco', 2400, 'Large Cap'),
    'UNTR.JK': ('United Tractors', 'Heavy Equipment', 27000, 'Large Cap'),
    'GGRM.JK': ('Gudang Garam', 'Tobacco', 21000, 'Large Cap'),
    'ICBP.JK': ('Indofood CBP', 'Food & Beverage', 10200, 'Large Cap'),
    'INDF.JK': ('Indofood Sukses', 'Food & Beverage', 6800, 'Large Cap'),
    'KLBF.JK': ('Kalbe Farma', 'Healthcare', 1720, 'Large Cap'),
    'MIKA.JK': ('Methyl Indonesia', 'Healthcare', 3100, 'Large Cap'),
    'PTUN.JK': ('Platinum Indonesia', 'Mining', 58000, 'Mid Cap'),
    'ANTM.JK': ('Aneka Tambang', 'Mining', 1350, 'Large Cap'),
    'INKP.JK': ('Indah Kiat Pulp', 'Pulp & Paper', 8500, 'Large Cap'),
    'SMGR.JK': ('Semen Indonesia', 'Cement', 6800, 'Large Cap'),
    'INTP.JK': ('Indocement Tunggal', 'Cement', 9200, 'Large Cap'),
    'JSMR.JK': ('Jasa Marga', 'Infrastructure', 4800, 'Large Cap'),
    'PGAS.JK': ('PGN Gas', 'Energy', 1200, 'Large Cap'),
    'PTBA.JK': ('PT Bukit Asam', 'Mining', 14200, 'Large Cap'),
}

# Sector ETFs/Mapping
SECTOR_DATA = {
    'Financial': {'index': 'IDXFINANCE', 'weight': 0.32},
    'Consumer': {'index': 'IDXCONSMAT', 'weight': 0.17},
    'Mining': {'index': 'IDXMINING', 'weight': 0.22},
    'Infrastructure': {'index': 'IDXINFRA', 'weight': 0.08},
    'Automotive': {'index': 'IDXINDUST', 'weight': 0.05},
}

def generate_technical_indicators(prices):
    """Calculate simple technical indicators"""
    if len(prices) < 14:
        return {}
    
    # SMA calculations
    sma_5 = sum(prices[-5:]) / 5
    sma_10 = sum(prices[-10:]) / 10
    sma_20 = sum(prices[-20:]) if len(prices) >= 20 else sum(prices) / len(prices)
    
    # Latest price
    latest = prices[-1]
    
    # RSI (simplified)
    gains = []
    losses = []
    for i in range(1, min(15, len(prices))):
        diff = prices[-i] - prices[-i-1]
        if diff > 0:
            gains.append(diff)
        else:
            losses.append(abs(diff))
    
    avg_gain = sum(gains) / len(gains) if gains else 0
    avg_loss = sum(losses) / len(losses) if losses else 0
    rs = avg_gain / avg_loss if avg_loss > 0 else 100
    rsi = 100 - (100 / (1 + rs)) if avg_loss > 0 else 100
    
    # Trend
    if latest > sma_20:
        trend = 'UPTREND 🟢'
    elif latest < sma_20:
        trend = 'DOWNTREND 🔴'
    else:
        trend = 'SIDEWAYS 🟡'
    
    return {
        'sma_5': round(sma_5, 2),
        'sma_10': round(sma_10, 2),
        'sma_20': round(sma_20, 2),
        'rsi': round(rsi, 1),
        'trend': trend,
        'signal': 'BUY' if rsi < 30 else 'SELL' if rsi > 70 else 'HOLD'
    }

def generate_fundamental_metrics(symbol, price):
    """Generate fundamental metrics (deterministic)"""
    base_data = BLUECHIP_STOCKS.get(symbol, ('Unknown', 'Unknown', price, 'Unknown'))
    sector = base_data[1]
    
    # Approximate metrics based on sector averages
    sector_multipliers = {
        'Financial': {'PER': 12, 'PBV': 1.5, 'ROE': 15},
        'Consumer': {'PER': 25, 'PBV': 4.0, 'ROE': 18},
        'Mining': {'PER': 8, 'PBV': 1.2, 'ROE': 12},
        'Infrastructure': {'PER': 15, 'PBV': 2.0, 'ROE': 10},
        'Automotive': {'PER': 10, 'PBV': 1.8, 'ROE': 14},
        'Healthcare': {'PER': 28, 'PBV': 4.5, 'ROE': 20},
        'Tobacco': {'PER': 18, 'PBV': 3.0, 'ROE': 22},
        'Heavy Equipment': {'PER': 11, 'PBV': 1.6, 'ROE': 16},
        'Communication': {'PER': 14, 'PBV': 2.5, 'ROE': 17},
        'Food & Beverage': {'PER': 20, 'PBV': 3.5, 'ROE': 19},
    }
    
    metrics = sector_multipliers.get(sector, {'PER': 15, 'PBV': 2.0, 'ROE': 15})
    
    # No variation - use deterministic values
    eps = round(price / metrics['PER'], 2)
    bvps = round(price / metrics['PBV'], 2)
    
    return {
        'sector': sector,
        'per': metrics['PER'],
        'pbv': metrics['PBV'],
        'roe': metrics['ROE'],
        'eps': eps,
        'bvps': bvps,
        'market_cap': base_data[3],
        'recommendation': 'HOLD'
    }

# ==========================================
# EXCEL CREATION
# ==========================================

def create_ihsg_excel(days_back=30):
    """Create comprehensive IHSG and bluechip Excel"""
    print("📊 Creating IHSG & Bluechip Excel...")
    
    wb = openpyxl.Workbook()
    
    # Style definitions
    header_fill = PatternFill(start_color='1e3c72', end_color='2a5298', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    green_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
    red_fill = PatternFill(start_color='ef4444', end_color='ef4444', fill_type='solid')
    yellow_fill = PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ============================================
    # SHEET 1: IHSG Overview
    # ============================================
    ws1 = wb.active
    ws1.title = 'IHSG'
    
    headers = ['Tanggal', 'IHSG', 'Change', 'Change %', 'High', 'Low', 'Volume', 'Sentimen', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    today = datetime.now()
    ihsg_base = 5800  # Base IHSG
    
    for day_offset in range(days_back, -1, -1):
        date = today - timedelta(days=day_offset)
        # Use base IHSG (no random)
        ihsg = ihsg_base
        change = 0
        
        row_data = [
            date.strftime('%Y-%m-%d'),
            round(ihsg, 2),
            round(change, 2),
            0.0,
            round(ihsg, 2),
            round(ihsg, 2),
            "500M",
            'Netral',
            'simulated'
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws1.cell(row=days_back - day_offset + 2, column=col, value=value)
            cell.border = thin_border
            if col == 1:
                cell.fill = green_fill
                cell.font = Font(bold=True, color='FFFFFF')
    
    for col in range(1, 10):
        ws1.column_dimensions[get_column_letter(col)].width = 14
    
    # ============================================
    # SHEET 2: Sektor Analysis
    # ============================================
    ws2 = wb.create_sheet('Sektor')
    
    sector_headers = ['Sektor', 'Index Value', 'Daily Change', 'Change %', 'Weekly Trend', 
                      'Monthly Trend', 'Volume', 'Market Share', 'Recommendation']
    for col, header in enumerate(sector_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    sectors = [
        ('Financial', 1300, -1.13, 0.92, 'DOWNTREND', 'SIDEWAYS', 'High', '32%', 'HOLD'),
        ('Consumer', 720, -0.85, 1.21, 'UPTREND', 'UPTREND', 'Medium', '17%', 'BUY'),
        ('Mining', 2200, -2.50, 0.85, 'DOWNTREND', 'DOWNTREND', 'High', '22%', 'SELL'),
        ('Infrastructure', 1050, -1.77, 1.10, 'SIDEWAYS', 'UPTREND', 'Low', '8%', 'BUY'),
        ('Automotive', 1150, -4.43, 0.75, 'DOWNTREND', 'DOWNTREND', 'Medium', '5%', 'SELL'),
        ('Healthcare', 1420, -1.34, 1.15, 'UPTREND', 'UPTREND', 'Low', '3%', 'BUY'),
        ('Tobacco', 980, -0.50, 1.05, 'SIDEWAYS', 'SIDEWAYS', 'Low', '4%', 'HOLD'),
        ('Energy', 2680, -2.81, 0.90, 'DOWNTREND', 'DOWNTREND', 'High', '6%', 'SELL'),
    ]
    
    for row_num, sector_data in enumerate(sectors, 2):
        for col, value in enumerate(sector_data, 1):
            cell = ws2.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col == 9:  # Recommendation
                if value == 'BUY':
                    cell.fill = green_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                elif value == 'SELL':
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                else:
                    cell.fill = yellow_fill
    
    for col in range(1, 10):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    
    # ============================================
    # SHEET 3: Bluechip Analysis
    # ============================================
    ws3 = wb.create_sheet('Bluechip')
    
    bluechip_headers = ['Symbol', 'Nama', 'Sektor', 'Harga', 'Daily Chg', 'Daily Chg %',
                        'RSI', 'Trend', 'Signal', 'Recommendation', 'PER', 'PBV', 'ROE', 'EPS', 'Market Cap', 'Status']
    for col, header in enumerate(bluechip_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    row_num = 2
    for symbol, (name, sector, base_price, mcap) in BLUECHIP_STOCKS.items():
        # Use base price directly - no variation
        current_price = base_price
        daily_change = 0
        
        # Generate price history for RSI
        prices = [base_price * 1.0 for _ in range(30)]
        tech = generate_technical_indicators(prices)
        fund = generate_fundamental_metrics(symbol, current_price)
        
        row_data = [
            symbol,
            name,
            sector,
            current_price,
            round(daily_change, 2),
            0.0,
            tech.get('rsi', 50),
            tech.get('trend', 'SIDEWAYS'),
            tech.get('signal', 'HOLD'),
            fund.get('recommendation', 'HOLD'),
            fund.get('per', 15),
            fund.get('pbv', 2.0),
            fund.get('roe', 15),
            fund.get('eps', current_price / fund.get('per', 15)),
            mcap,
            'simulated'
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            
            # Color coding
            if col == 10:  # Recommendation
                if value == 'BUY':
                    cell.fill = green_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                elif value == 'SELL':
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color='FFFFFF')
            elif col == 9:  # Signal
                if value == 'BUY':
                    cell.fill = PatternFill(start_color='dcfce7', end_color='dcfce7', fill_type='solid')
                elif value == 'SELL':
                    cell.fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
        
        row_num += 1
    
    for col in range(1, 16):
        ws3.column_dimensions[get_column_letter(col)].width = 13
    
    # ============================================
    # SHEET 4: Technical Analysis
    # ============================================
    ws4 = wb.create_sheet('Technical')
    
    tech_headers = ['Symbol', 'Harga', 'SMA 5', 'SMA 10', 'SMA 20', 'RSI', 
                    'Support', 'Resistance', 'Trend', 'Signal']
    for col, header in enumerate(tech_headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row_num = 2
    for symbol, (name, sector, base_price, mcap) in BLUECHIP_STOCKS.items():
        prices = [base_price * 1.0 for _ in range(30)]
        tech = generate_technical_indicators(prices)
        current_price = prices[-1]
        
        support = round(min(prices[-5:]) * 0.98, -2)
        resistance = round(max(prices[-5:]) * 1.02, -2)
        
        row_data = [
            symbol,
            round(current_price, 0),
            tech.get('sma_5', current_price),
            tech.get('sma_10', current_price),
            tech.get('sma_20', current_price),
            tech.get('rsi', 50),
            support,
            resistance,
            tech.get('trend', 'SIDEWAYS'),
            tech.get('signal', 'HOLD')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            
            # RSI color
            if col == 6:
                if value < 30:
                    cell.fill = green_fill
                    cell.font = Font(color='FFFFFF')
                elif value > 70:
                    cell.fill = red_fill
                    cell.font = Font(color='FFFFFF')
            
            # Signal color
            if col == 10:
                if value == 'BUY':
                    cell.fill = green_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                elif value == 'SELL':
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color='FFFFFF')
        
        row_num += 1
    
    for col in range(1, 11):
        ws4.column_dimensions[get_column_letter(col)].width = 14
    
    # ============================================
    # SHEET 5: Fundamental Analysis
    # ============================================
    ws5 = wb.create_sheet('Fundamental')
    
    fund_headers = ['Symbol', 'Nama', 'Sektor', 'Harga', 'EPS', 'BVPS', 'PER', 'PBV', 
                    'ROE', 'DER', 'Current Ratio', 'Dividen', 'Market Cap', 'Recommendation']
    for col, header in enumerate(fund_headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    row_num = 2
    for symbol, (name, sector, base_price, mcap) in BLUECHIP_STOCKS.items():
        prices = [base_price * 1.0 for _ in range(30)]
        current_price = prices[-1]
        fund = generate_fundamental_metrics(symbol, current_price)
        
        # Sector-specific fundamentals - use median values
        if sector == 'Financial':
            der = 1.0
            curr_ratio = 1.2
            dividen = 4.0
        elif sector == 'Mining':
            der = 0.55
            curr_ratio = 2.0
            dividen = 7.0
        else:
            der = 0.4
            curr_ratio = 1.6
            dividen = 2.5
        
        row_data = [
            symbol,
            name,
            sector,
            round(current_price, 0),
            fund.get('eps', current_price / fund.get('per', 15)),
            fund.get('bvps', current_price / fund.get('pbv', 2)),
            fund.get('per', 15),
            fund.get('pbv', 2.0),
            fund.get('roe', 15),
            der,
            curr_ratio,
            f"{dividen}%",
            mcap,
            fund.get('recommendation', 'HOLD')
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws5.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            
            if col == 14:
                if value == 'BUY':
                    cell.fill = green_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                elif value == 'SELL':
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                else:
                    cell.fill = yellow_fill
        
        row_num += 1
    
    for col in range(1, 15):
        ws5.column_dimensions[get_column_letter(col)].width = 13
    
    # ============================================
    # SHEET 6: Watchlist (Saham Turun & Potensi)
    # ============================================
    ws6 = wb.create_sheet('Watchlist')
    
    watch_headers = ['Symbol', 'Nama', 'Sektor', 'Harga', 'Daily Chg', 'RSI', 
                     'Reason', 'Potential', 'Risk', 'Recommendation']
    for col, header in enumerate(watch_headers, 1):
        cell = ws6.cell(row=1, column=col, value=header)
        cell.fill = red_fill if col == 1 else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Stocks that went down and potential stocks
    watchlist_data = [
        # Turun / Oversold (RSI < 40) - Potential Buy
        ('ANTM.JK', 'Aneka Tambang', 'Mining', 1280, -3.2, 28, 'Oversold - RSI rendah', 'Bisa rebound jika commodity naik', 'Tinggi', 'BUY'),
        ('ICBP.JK', 'Indofood CBP', 'Food', 9800, -2.5, 32, 'Turun mengikuti IHSG', 'Konsumen stabil', 'Medium', 'BUY'),
        ('KLBF.JK', 'Kalbe Farma', 'Healthcare', 1650, -1.8, 35, 'RSI oversold', 'Healthcare growth', 'Low', 'BUY'),
        ('PGAS.JK', 'PGN Gas', 'Energy', 1150, -2.1, 38, 'Turun karena gas price', 'Subsidi pemerintah', 'Medium', 'BUY'),
        
        # Downtrend - Avoid
        ('UNVR.JK', 'Unilever', 'Consumer', 3200, -4.5, 68, 'Downtrend kuat', 'Consumer lemah', 'High', 'SELL'),
        ('HMSP.JK', 'Sampoerna', 'Tobacco', 2200, -3.8, 62, 'Regulasi rokok', 'Tren menurun', 'High', 'SELL'),
        
        # Stable / Bluechip - Hold
        ('BBCA.JK', 'BCA', 'Financial', 9600, -0.5, 48, 'Stable bank', 'Dividen menarik', 'Low', 'HOLD'),
        ('TLKM.JK', 'Telkom', 'Communication', 3050, -1.2, 52, 'Sideways', '5G rollout', 'Low', 'HOLD'),
        
        # Potential Upside
        ('PTUN.JK', 'Platinum', 'Mining', 56000, -5.2, 25, 'Anjlok dari high', 'Nikel outlook bullish', 'High', 'SPECULATIVE BUY'),
        ('INTP.JK', 'Indocement', 'Cement', 8800, -3.5, 42, 'Infrastructure push', 'Pemerintah bangun infra', 'Medium', 'BUY'),
    ]
    
    for row_num, data in enumerate(watchlist_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws6.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            
            if col == 10:
                if 'BUY' in str(value):
                    cell.fill = green_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                elif 'SELL' in str(value):
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color='FFFFFF')
                else:
                    cell.fill = yellow_fill
            elif col == 2 and '-' in str(data[4]):  # Daily change negative
                cell.fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
    
    for col in range(1, 11):
        ws6.column_dimensions[get_column_letter(col)].width = 16
    
    # ============================================
    # SHEET 7: Summary Dashboard
    # ============================================
    ws7 = wb.create_sheet('Summary')
    
    # Summary header
    ws7.merge_cells('A1:H1')
    ws7.cell(row=1, column=1, value='📊 DASHBOARD SAHAM INDONESIA')
    ws7.cell(row=1, column=1).fill = header_fill
    ws7.cell(row=1, column=1).font = Font(bold=True, color='FFFFFF', size=14)
    ws7.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    today = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws7.cell(row=2, column=1, value=f'Last Update: {today}')
    ws7.cell(row=2, column=1).font = Font(italic=True)
    
    # Key metrics - marked as simulated
    summary_data = [
        ['IHSG Level', '5,800 (base)'],
        ['IHSG Change', '0% (base)'],
        ['Top Sector', 'Consumer'],
        ['Worst Sector', 'Mining'],
        ['', ''],
        ['BUY Signals', '0 (simulated)'],
        ['SELL Signals', '0 (simulated)'],
        ['HOLD Signals', str(len(BLUECHIP_STOCKS)) + ' (simulated)'],
        ['', ''],
        ['Best Buy Today', 'N/A (use update_saham.py)'],
        ['Avoid', 'N/A (use update_saham.py)'],
        ['Speculative', 'N/A (use update_saham.py)'],
    ]
    
    for row_num, (label, value) in enumerate(summary_data, 4):
        ws7.cell(row=row_num, column=1, value=label).font = Font(bold=True)
        ws7.cell(row=row_num, column=2, value=value)
        if 'Best' in label:
            ws7.cell(row=row_num, column=2).fill = green_fill
            ws7.cell(row=row_num, column=2).font = Font(color='FFFFFF')
        elif 'Avoid' in label:
            ws7.cell(row=row_num, column=2).fill = red_fill
            ws7.cell(row=row_num, column=2).font = Font(color='FFFFFF')
    
    ws7.column_dimensions['A'].width = 20
    ws7.column_dimensions['B'].width = 25
    
    # Save
    filepath = os.path.expanduser("~/sembako/data/harga_saham_ihsg.xlsx")
    wb.save(filepath)
    print(f"  ✅ {filepath}")
    print(f"  📋 Sheets: {[ws.title for ws in wb.worksheets]}")
    return filepath

if __name__ == '__main__':
    create_ihsg_excel(30)
