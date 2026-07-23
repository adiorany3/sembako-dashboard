import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from datetime import datetime
import os

EXCEL_PATH = os.path.expanduser("~/sembako/data/harga_sembako.xlsx")

def create_or_load_workbook():
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        return wb, ws
    return create_new_workbook()

def create_new_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Harga Sembako"
    
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = [
        "Tanggal",
        "Beras Premium\n(Rp/kg)", "Beras Medium\n(Rp/kg)",
        "Minyak Goreng\nCurah (Rp/kg)", "Gula Putih\n(Rp/kg)",
        "Garam Halus\n(Rp/kg)", "Tepung Terigu\n(Rp/kg)",
        "Telur Ayam Ras\n(Rp/kg)", "Telur Ayam Kampung\n(Rp/kg)",
        "Daging Ayam Ras\n(Rp/kg)", "Daging Ayam Kampung\n(Rp/kg)",
        "Daging Sapi\n(Rp/kg)",
        "Cabai Merah\nKeriting (Rp/kg)", "Cabai Rawit\nMerah (Rp/kg)",
        "Bawang Merah\n(Rp/kg)", "Bawang Putih\n(Rp/kg)",
        "Minyak Tanah\n(Rp/kg)", "Gas Elpiji\n(Rp)",
        "Sumber"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
        cell.border = thin
    
    # Category sub-headers coloring
    cat_colors = {
        range(2, 4): "4472C4",    # Beras - biru
        range(4, 7): "548235",    # Minyak/Gula - hijau
        range(7, 9): "ED7D31",    # Telur - oranye
        range(9, 11): "C00000",   # Daging ayam - merah
        range(11, 12): "7030A0",  # Daging sapi - ungu
        range(12, 16): "BF8F00",  # Bumbu - kuning
        range(16, 18): "808080",  # Lainnya - abu
    }
    for cols, color in cat_colors.items():
        for col in cols:
            ws.cell(row=1, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    widths = [13, 16, 16, 14, 18, 20, 16, 18, 16, 18, 14, 18, 18, 16, 16, 14, 12, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "B2"
    return wb, ws

def add_row(tanggal, data_dict, sumber="Siskaperbapo Jatim"):
    """
    data_dict keys from update_harga.py:
      beras_premium, beras_medium, gula, minyak_curah, minyak_kemasan,
      telur_ras, telur_kampung, ayam_ras, ayam_kampung, sapi,
      cabai_keriting, cabai_rawit, bawang_merah, bawang_putih, garam, elpiji

    24-col layout (col → header):
      1=Tanggal, 2=Beras Premium, 3=Beras Medium, 4=Minyak Goreng,
      5=Gula Pasir, 6=Garam, 7=Tepung Terigu, 8=Cabai Merah,
      9=Cabai Rawit, 10=Bawang Merah, 11=Bawang Putih, 12=Minyak Tanah,
      13=Telur Ras, 14=Telur Kampung, 15=Ayam Ras, 16=Ayam Kampung,
      17=Daging Sapi, 18=Gas Elpiji, 19=Garam Bata, 20=Garam Halus,
      21=Susu KM Bendera, 22=Susu KM Indomilk, 23=Susu Bubuk Bendera,
      24=Susu Bubuk Indomilk

    Mapping: scrape_key → col, extra cols stay blank (None).
    Returns: True if row added, False if date already exists
    """
    try:
        wb, ws = create_or_load_workbook()
        ca = Alignment(horizontal="center", vertical="center")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Check if date already exists — skip duplicate
        for r in range(2, ws.max_row + 1):
            existing = str(ws.cell(row=r, column=1).value or '')[:10]
            if existing == str(tanggal)[:10]:
                wb.close()
                print(f"  ⏭️ Date {tanggal} already exists, skipping")
                return False

        row = ws.max_row + 1

        # Map scrape keys → 24-col layout columns
        KEY_TO_COL = {
            "beras_premium":  2,   # Beras Premium
            "beras_medium":   3,   # Beras Medium
            "minyak_curah":   4,   # Minyak Goreng (use curah value as representative)
            "gula":           5,   # Gula Pasir
            "garam":          6,   # Garam
            # col 7 = Tepung Terigu — no scraper, leave None
            "cabai_keriting": 8,   # Cabai Merah (keriting is subcategory)
            "cabai_rawit":    9,   # Cabai Rawit
            "bawang_merah":  10,   # Bawang Merah
            "bawang_putih":  11,   # Bawang Putih
            # col 12 = Minyak Tanah — no scraper, leave None
            "telur_ras":     13,   # Telur Ras
            "telur_kampung": 14,   # Telur Kampung
            "ayam_ras":      15,   # Ayam Ras
            "ayam_kampung":  16,   # Ayam Kampung
            "sapi":          17,   # Daging Sapi
            "elpiji":        18,   # Gas Elpiji
            "gas_elpiji":    18,   # alias from cache
            # cols 19-24: Garam Bata, Garam Halus, Susu KM/Bubuk — no scraper
        }

        c = ws.cell(row=row, column=1, value=tanggal)
        c.alignment = ca; c.border = thin

        # Get last known values for NULL columns (fallback)
        last_row = ws.max_row - 1 if ws.max_row > 1 else None
        
        for key, col in KEY_TO_COL.items():
            val = data_dict.get(key)
            # Fallback to last known if genuinely missing (not provided)
            if (val is None or val == '') and last_row:
                last_val = ws.cell(row=last_row, column=col).value
                if last_val is not None and last_val != 0:
                    val = last_val
            c = ws.cell(row=row, column=col, value=val if val is not None else None)
            c.alignment = ca; c.border = thin
            if val is not None and val != '': c.number_format = '#,##0'

        # Cols 19-24: copy from last row if available (no scraper for these)
        # Use hardcoded default for gas_elpiji if missing (HET government price)
        if last_row:
            for col in range(19, 25):
                last_val = ws.cell(row=last_row, column=col).value
                c = ws.cell(row=row, column=col, value=last_val)
                c.alignment = ca; c.border = thin
                if last_val: c.number_format = '#,##0'

        wb.save(EXCEL_PATH)
        print(f"Row added: {tanggal}")
        return EXCEL_PATH
    except Exception as e:
        print(f"  ❌ Error adding row: {e}")
        return None

if __name__ == "__main__":
    # Data 21 Juni 2026 (dari Siskaperbapo Jatim via detik.com)
    add_row("2026-06-21", {
        "beras_premium": 15010, "beras_medium": 12953,
        "gula": 17166, "minyak_curah": 20462, "minyak_kemasan": 21708,
        "telur_ras": 25198, "telur_kampung": 46663,
        "ayam_ras": 32256, "ayam_kampung": 66888,
        "sapi": 126568,
        "cabai_keriting": 37870, "cabai_rawit": 57232,
        "bawang_merah": 42529, "bawang_putih": 35265,
        "garam": 9662, "elpiji": 20092
    })
    
    # Data 23 Juni 2026 (sama, belum ada update baru)
    add_row("2026-06-23", {
        "beras_premium": 15010, "beras_medium": 12953,
        "gula": 17166, "minyak_curah": 20462, "minyak_kemasan": 21708,
        "telur_ras": 25198, "telur_kampung": 46663,
        "ayam_ras": 32256, "ayam_kampung": 66888,
        "sapi": 126568,
        "cabai_keriting": 37870, "cabai_rawit": 57232,
        "bawang_merah": 42529, "bawang_putih": 35265,
        "garam": 9662, "elpiji": 20092
    })
    
    print(f"File saved: {EXCEL_PATH}")
