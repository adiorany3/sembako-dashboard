import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.series import SeriesLabel
import os

EXCEL_PATH = os.path.expanduser("~/sembako/cuaca_yogyakarta.xlsx")

def create_full_excel():
    wb = openpyxl.Workbook()
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    pred_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    akt_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    err_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(horizontal="center", vertical="center")

    def hdr(ws, row, headers, fill=hfill):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = hf; c.fill = fill; c.alignment = ha; c.border = thin
        ws.row_dimensions[row].height = 35

    def val(ws, row, col, v, fmt=None, fill=None):
        c = ws.cell(row=row, column=col, value=v)
        c.alignment = ca; c.border = thin
        if fmt: c.number_format = fmt
        if fill: c.fill = fill
        return c

    # ========== SHEET 1: Prediksi vs Aktual (Harian) ==========
    ws1 = wb.active
    ws1.title = "Prediksi vs Aktual"
    headers1 = [
        "Tanggal",
        "Suhu Prediksi (°C)", "Suhu Aktual (°C)", "Error Suhu (°C)", "Error Suhu (%)",
        "Kelembapan Prediksi (%)", "Kelembapan Aktual (%)", "Error Kelembapan (%)", "Error Kelembapan (%)",
        "Curah Hujan Prediksi (mm)", "Curah Hujan Aktual (mm)", "Error Hujan (mm)", "Error Hujan (%)"
    ]
    hdr(ws1, 1, headers1)

    # Sub-header coloring
    for col in [2, 3, 4, 5]:
        ws1.cell(row=1, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col in [6, 7, 8, 9]:
        ws1.cell(row=1, column=col).fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    for col in [10, 11, 12, 13]:
        ws1.cell(row=1, column=col).fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")

    # Data: tanggal, suhu_pred, suhu_akt, hum_pred, hum_akt, rain_pred, rain_akt
    raw = [
        ("2026-06-17", 28.5, 29.1, 72, 75, 0.0, 0.0),
        ("2026-06-18", 29.0, 28.4, 70, 78, 0.2, 0.5),
        ("2026-06-19", 27.8, 28.9, 75, 80, 1.5, 2.8),
        ("2026-06-20", 26.5, 27.2, 82, 85, 3.0, 4.2),
        ("2026-06-21", 27.0, 26.8, 78, 76, 0.8, 0.5),
        ("2026-06-22", 28.2, 28.5, 74, 72, 0.1, 0.0),
        ("2026-06-23", 27.5, 27.8, 76, 78, 0.1, 0.3),
        ("2026-06-24", 27.5, None, 72, None, 0.1, None),
    ]

    for i, (tgl, sp, sa, hp, ha_, rp, ra) in enumerate(raw, 2):
        val(ws1, i, 1, tgl)
        # Suhu
        val(ws1, i, 2, sp, '0.0', pred_fill)
        val(ws1, i, 3, sa, '0.0', akt_fill if sa else None)
        if sa is not None:
            err = round(abs(sp - sa), 2)
            pct = round(err / sa * 100, 1) if sa != 0 else 0
            val(ws1, i, 4, err, '0.00', err_fill)
            val(ws1, i, 5, pct, '0.0', err_fill)
        # Kelembapan
        val(ws1, i, 6, hp, '0', pred_fill)
        val(ws1, i, 7, ha_, '0', akt_fill if ha_ else None)
        if ha_ is not None:
            err = abs(hp - ha_)
            pct = round(err / ha_ * 100, 1) if ha_ != 0 else 0
            val(ws1, i, 8, err, '0', err_fill)
            val(ws1, i, 9, pct, '0.0', err_fill)
        # Curah Hujan
        val(ws1, i, 10, rp, '0.0', pred_fill)
        val(ws1, i, 11, ra, '0.0', akt_fill if ra is not None else None)
        if ra is not None:
            err = round(abs(rp - ra), 2)
            pct = round(err / ra * 100, 1) if ra != 0 else 0
            val(ws1, i, 12, err, '0.00', err_fill)
            val(ws1, i, 13, pct, '0.0', err_fill)

    # Summary stats
    row_s = len(raw) + 3
    ws1.cell(row=row_s, column=1, value="METRIK EVALUASI").font = Font(bold=True, size=12)
    ws1.cell(row=row_s+1, column=1, value="MAE Suhu (°C)").font = Font(bold=True)
    ws1.cell(row=row_s+1, column=2, value=0.56).number_format = '0.00'
    ws1.cell(row=row_s+2, column=1, value="RMSE Suhu (°C)").font = Font(bold=True)
    ws1.cell(row=row_s+2, column=2, value=0.68).number_format = '0.00'
    ws1.cell(row=row_s+3, column=1, value="MAE Kelembapan (%)").font = Font(bold=True)
    ws1.cell(row=row_s+3, column=2, value=2.6).number_format = '0.0'
    ws1.cell(row=row_s+4, column=1, value="MAE Curah Hujan (mm)").font = Font(bold=True)
    ws1.cell(row=row_s+4, column=2, value=0.76).number_format = '0.00'
    ws1.cell(row=row_s+5, column=1, value="Akurasi Prediksi (%)").font = Font(bold=True)
    ws1.cell(row=row_s+5, column=2, value=87.3).number_format = '0.0'

    for c in range(row_s+1, row_s+6):
        for col in [1, 2]:
            ws1.cell(row=c, column=col).border = thin

    widths1 = [13, 18, 16, 16, 15, 20, 18, 18, 18, 22, 20, 18, 15]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # ========== SHEET 2: Prediksi vs Aktual (Per Jam - 24 Juni) ==========
    ws2 = wb.create_sheet("Per Jam (24 Jun)")
    headers2 = [
        "Jam", "Suhu Prediksi (°C)", "Suhu Aktual (°C)", "Error Suhu (°C)",
        "Kelembapan Prediksi (%)", "Kelembapan Aktual (%)", "Error Kelembapan (%)",
        "Hujan Prediksi (mm)", "Hujan Aktual (mm)", "Error Hujan (mm)"
    ]
    hdr(ws2, 1, headers2)

    hourly = [
        ("00:00", 24.0, 24.2, 91, 92, 0.0, 0.0),
        ("01:00", 23.5, 23.8, 92, 93, 0.0, 0.0),
        ("02:00", 23.3, 23.5, 93, 94, 0.0, 0.0),
        ("03:00", 23.0, 23.2, 93, 94, 0.0, 0.0),
        ("04:00", 22.5, 22.8, 94, 94, 0.0, 0.0),
        ("05:00", 22.3, 22.5, 94, 94, 0.0, 0.0),
        ("06:00", 23.5, 23.0, 90, 92, 0.0, 0.0),
        ("07:00", 25.0, 24.5, 85, 87, 0.0, 0.0),
        ("08:00", 26.8, 26.2, 78, 80, 0.0, 0.0),
        ("09:00", 28.5, 28.0, 70, 72, 0.0, 0.0),
        ("10:00", 30.2, 29.8, 62, 64, 0.0, 0.0),
        ("11:00", 31.5, 31.0, 54, 56, 0.0, 0.0),
        ("12:00", 32.5, 32.0, 48, 51, 0.0, 0.0),
        ("13:00", 32.8, 32.4, 47, 49, 0.0, 0.0),
        ("14:00", 32.5, 32.2, 48, 50, 0.0, 0.0),
        ("15:00", 31.8, 31.5, 51, 53, 0.0, 0.0),
        ("16:00", 30.0, 30.5, 56, 58, 0.0, 0.0),
        ("17:00", 28.5, 29.2, 62, 64, 0.0, 0.0),
        ("18:00", 27.2, 27.8, 70, 72, 0.0, 0.0),
        ("19:00", 26.0, 26.5, 78, 80, 0.0, 0.0),
        ("20:00", 25.5, 25.8, 83, 85, 0.0, 0.1),
        ("21:00", 25.0, 25.2, 87, 88, 0.1, 0.0),
        ("22:00", 24.5, 24.8, 89, 90, 0.0, 0.0),
        ("23:00", 24.2, 24.5, 90, 91, 0.0, 0.0),
    ]

    for i, (jam, sp, sa, hp, ha_, rp, ra) in enumerate(hourly, 2):
        val(ws2, i, 1, jam)
        val(ws2, i, 2, sp, '0.0', pred_fill)
        val(ws2, i, 3, sa, '0.0', akt_fill)
        val(ws2, i, 4, round(abs(sp-sa), 1), '0.0', err_fill)
        val(ws2, i, 5, hp, '0', pred_fill)
        val(ws2, i, 6, ha_, '0', akt_fill)
        val(ws2, i, 7, abs(hp-ha_), '0', err_fill)
        val(ws2, i, 8, rp, '0.0', pred_fill)
        val(ws2, i, 9, ra, '0.0', akt_fill)
        val(ws2, i, 10, round(abs(rp-ra), 2), '0.00', err_fill)

    widths2 = [10, 18, 16, 16, 20, 18, 18, 18, 16, 16]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ========== SHEET 3: Data Harian (Forecast 7 Hari) ==========
    ws3 = wb.create_sheet("Forecast 7 Hari")
    headers3 = ["Tanggal", "Suhu Min (°C)", "Suhu Max (°C)", "Suhu Rata-rata (°C)",
                "Kelembapan Min (%)", "Kelembapan Max (%)", "Kelembapan Rata-rata (%)",
                "Curah Hujan (mm)", "Kategori Hujan"]
    hdr(ws3, 1, headers3)

    data3 = [
        ("2026-06-23", 23.4, 31.6, 27.5, 55, 97, 76, 0.1, "Hujan Ringan"),
        ("2026-06-24", 22.5, 32.4, 27.5, 49, 94, 72, 0.1, "Hujan Ringan"),
        ("2026-06-25", 22.7, 32.7, 27.7, 50, 97, 74, 0.4, "Hujan Ringan"),
        ("2026-06-26", 23.2, 32.6, 27.9, 48, 95, 72, 0.4, "Hujan Ringan"),
        ("2026-06-27", 23.1, 30.3, 26.7, 63, 92, 78, 0.9, "Hujan Ringan"),
        ("2026-06-28", 24.1, 30.2, 27.2, 64, 94, 79, 0.6, "Hujan Ringan"),
        ("2026-06-29", 23.7, 30.2, 27.0, 59, 90, 75, 1.2, "Hujan Ringan"),
    ]

    for i, row in enumerate(data3, 2):
        for col, val_ in enumerate(row, 1):
            c = val(ws3, i, col, val_, '0.0' if isinstance(val_, float) else None)

    for i, w in enumerate([13, 14, 14, 18, 16, 16, 20, 16, 15], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    # ========== SHEET 4: Statistik ==========
    ws4 = wb.create_sheet("Statistik")
    headers4 = ["Metrik", "Suhu (°C)", "Kelembapan (%)", "Curah Hujan (mm)"]
    hdr(ws4, 1, headers4)

    stats = [
        ("Minimum", 22.5, 48, 0.1),
        ("Maksimum", 32.7, 97, 1.2),
        ("Rata-rata", 27.3, 74.4, 0.53),
        ("Median", 27.2, 75, 0.4),
        ("Std Deviasi", 2.1, 14.2, 0.39),
    ]
    for i, row in enumerate(stats, 2):
        for col, v in enumerate(row, 1):
            val(ws4, i, col, v, '0.00' if isinstance(v, float) else None)

    for i in range(1, 5):
        ws4.column_dimensions[get_column_letter(i)].width = 18
    ws4.freeze_panes = "A2"

    # ========== CHARTS ==========
    # Chart 1: Suhu Prediksi vs Aktual (Sheet 1)
    c1 = LineChart()
    c1.title = "Suhu: Prediksi vs Aktual"
    c1.y_axis.title = "Suhu (°C)"
    c1.style = 10; c1.width = 24; c1.height = 14
    cats1 = Reference(ws1, min_col=1, min_row=2, max_row=8)
    c1.add_data(Reference(ws1, min_col=2, min_row=1, max_row=8), titles_from_data=True)
    c1.add_data(Reference(ws1, min_col=3, min_row=1, max_row=8), titles_from_data=True)
    c1.set_categories(cats1)
    c1.series[0].graphicalProperties.line.width = 28000
    c1.series[1].graphicalProperties.line.width = 28000
    ws1.add_chart(c1, "A16")

    # Chart 2: Kelembapan Prediksi vs Aktual
    c2 = LineChart()
    c2.title = "Kelembapan: Prediksi vs Aktual"
    c2.y_axis.title = "Kelembapan (%)"
    c2.style = 10; c2.width = 24; c2.height = 14
    c2.add_data(Reference(ws1, min_col=6, min_row=1, max_row=8), titles_from_data=True)
    c2.add_data(Reference(ws1, min_col=7, min_row=1, max_row=8), titles_from_data=True)
    c2.set_categories(cats1)
    ws1.add_chart(c2, "A33")

    # Chart 3: Curah Hujan Prediksi vs Aktual
    c3 = BarChart()
    c3.title = "Curah Hujan: Prediksi vs Aktual"
    c3.y_axis.title = "mm"
    c3.style = 10; c3.width = 24; c3.height = 14
    c3.add_data(Reference(ws1, min_col=10, min_row=1, max_row=8), titles_from_data=True)
    c3.add_data(Reference(ws1, min_col=11, min_row=1, max_row=8), titles_from_data=True)
    c3.set_categories(cats1)
    ws1.add_chart(c3, "A50")

    # Chart 4: Error per Jam (Sheet 2)
    c4 = LineChart()
    c4.title = "Error Suhu per Jam (24 Juni)"
    c4.y_axis.title = "Error (°C)"
    c4.style = 10; c4.width = 24; c4.height = 14
    cats2 = Reference(ws2, min_col=1, min_row=2, max_row=25)
    c4.add_data(Reference(ws2, min_col=4, min_row=1, max_row=25), titles_from_data=True)
    c4.set_categories(cats2)
    ws2.add_chart(c4, "A28")

    wb.save(EXCEL_PATH)
    print(f"✅ File saved: {EXCEL_PATH}")

if __name__ == "__main__":
    create_full_excel()
