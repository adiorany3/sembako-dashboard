#!/usr/bin/env python3
"""
Precompute AI Market Analysis — V4 COMPREHENSIVE
================================================
Reads ALL data → calculates technical indicators, correlations, anomaly detection,
seasonal patterns, confidence scoring → generates deep analysis.

Two modes:
  1. Groq API (VPS) — AI-enhanced with full context
  2. Rule-based (fallback) — algorithmic with all indicators

Flow: Excel files → indicators.py → prompt → Groq/rule → daily_analysis.json
"""
import os, json, sys, traceback, re
from pathlib import Path
from datetime import datetime
import openpyxl

# Add scripts dir to path (may run from ~/.hermes/scripts/ or ~/sembako/scripts/)
sys.path.insert(0, str(Path(__file__).parent))
sys.path.insert(0, os.path.expanduser("~/sembako/scripts"))
from indicators import (
    moving_average, ema, rate_of_change, volatility,
    find_support_resistance, trend_direction, momentum_score,
    detect_anomalies, pearson_correlation, confidence_score,
    get_seasonal_context, full_analysis,
    multi_period_trend, price_forecast, change_from_baseline, enhanced_confidence,
    real_interest_rate, currency_pressure_index, import_cost_pressure,
    food_inflation_proxy, stagflation_risk, risk_on_off_index, commodity_terms_of_trade
)

DATA_DIR = Path.home() / "sembako" / "data"
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")


def ts():
    return datetime.now().strftime("[%H:%M:%S]")


def read_sheet(fn, sheet=None):
    """Read Excel → list of dicts."""
    p = DATA_DIR / fn
    if not p.exists(): return []
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2: return []
        h = [re.sub(r'\s+', ' ', str(c).strip()) if c else f'c{i}' for i,c in enumerate(rows[0])]
        return [dict(zip(h, row)) for row in rows[1:]]
    except Exception: return []


def last_row(fn, sheet=None, n=0):
    d = read_sheet(fn, sheet)
    return d[-(n+1)] if len(d) > n else {}


def load_all_history():
    """Load ALL historical data for technical analysis."""
    print(f"{ts()} 📂 Loading historical data...")

    data = {}

    # Sembako (64 rows)
    rows = read_sheet("harga_sembako.xlsx", "Harga")
    if rows:
        keys = ["Beras Premium","Telur Ras","Gula Pasir","Cabai Merah","Minyak Goreng","Bawang Merah","Daging Sapi"]
        for k in keys:
            vals = [float(r.get(k, 0)) for r in rows if r.get(k) and float(str(r.get(k,0) or 0)) > 0]
            if len(vals) >= 2:
                data[f"sembako_{k}"] = vals
        print(f"  Sembako: {len(rows)} rows, {len([k for k in data if k.startswith('sembako')])} series")

    # Crypto (multiple rows)
    rows = read_sheet("crypto_monitor.xlsx", "Harga")
    if rows:
        col_map = {"BTC (USD)": "btc_usd", "ETH (USD)": "eth_usd", "SOL (USD)": "sol_usd"}
        for xl_col, key in col_map.items():
            vals = [float(r.get(xl_col, 0)) for r in rows if r.get(xl_col)]
            if len(vals) >= 2:
                data[f"crypto_{key}"] = vals
        print(f"  Crypto: {len(rows)} rows, {len([k for k in data if k.startswith('crypto')])} series")

    # Pertanian (1 row only)
    rows = read_sheet("harga_pertanian_ternak.xlsx", "Harga Komoditas")
    if rows:
        r = rows[-1]
        for k in ["Jagung Kering Pipil (Rp/kg)", "Kedelai Impor (Rp/kg)", "Pakan Broiler BR1 (Rp/kg)"]:
            v = r.get(k)
            if v:
                data[f"pertanian_{k.split()[0]}"] = [float(v)]  # single point
        print(f"  Pertanian: {len(rows)} rows")

    # Peternakan (3784 rows — rich!)
    rows = read_sheet("harga_peternakan_lengkap.xlsx")
    if rows:
        # Group by product, take latest price per day
        from collections import defaultdict
        prod_daily = defaultdict(dict)
        for r in rows:
            n = r.get("Produk","").strip()
            p = r.get("Harga (Rp)", 0)
            t = str(r.get("Tanggal", ""))[:10]
            if n and p and t:
                try: prod_daily[n][t] = float(p)
                except (ValueError, TypeError): pass
        # Get last 30 days per key product
        key_products = ["Telur Ayam Ras (Grosir)", "Ayam Broiler Hidup", "Daging Sapi Eceran",
                        "Daging Kambing Eceran", "Telur Ayam Kampung"]
        for pk in key_products:
            if pk in prod_daily:
                daily = sorted(prod_daily[pk].items())[-30:]
                vals = [v for _, v in daily]
                if len(vals) >= 2:
                    data[f"pet_{pk.split()[0]}_{pk.split()[-1]}"] = vals
        print(f"  Peternakan: {len(rows)} rows, {len([k for k in data if k.startswith('pet_')])} product series")

    # Kurs (1 row — but history from Excel)
    rows = read_sheet("kurs_valuta.xlsx", "Harian")
    if rows:
        for col in ["USD_IDR","EUR_IDR","SGD_IDR"]:
            vals = [float(r.get(col, 0)) for r in rows if r.get(col) and float(str(r.get(col,0) or 0)) > 1000]
            if len(vals) >= 2:
                data[f"kurs_{col}"] = vals
        print(f"  Kurs: {len(rows)} rows")

    # Minyak (1 row)
    rows = read_sheet("harga_minyak.xlsx", "Harian")
    if rows:
        b = [float(r.get("Brent_USD",0)) for r in rows if r.get("Brent_USD")]
        w = [float(r.get("WTI_USD",0)) for r in rows if r.get("WTI_USD")]
        if len(b) >= 2: data["oil_brent"] = b
        if len(w) >= 2: data["oil_wti"] = w
        print(f"  Minyak: {len(rows)} rows")

    # BI Rate (1 row)
    bi = last_row("bi_rate_inflasi.xlsx", "Harian")
    if bi:
        for k in ["BI_Rate","Inflasi_YoY","IHK"]:
            v = bi.get(k)
            if v: data[f"bi_{k}"] = [float(v)]
        print(f"  BI Rate: 1 snapshot")

    # CPO (1 row)
    cpo = last_row("harga_cpo.xlsx", "Harian")
    if cpo:
        myr = cpo.get("Harga_MYR")
        if myr: data["cpo_myr"] = [float(myr)]
        print(f"  CPO: 1 snapshot")

    # Sentimen (301 rows)
    rows = read_sheet("sentimen_berita.xlsx", "Detail")
    if rows:
        # Daily sentiment aggregate
        from collections import defaultdict
        daily_sent = defaultdict(lambda: {"pos":0,"neg":0,"neu":0,"total":0})
        for r in rows:
            t = str(r.get("Tanggal",""))[:10]
            s = str(r.get("Sentimen","")).upper()
            if t:
                daily_sent[t]["total"] += 1
                if s == "POSITIF": daily_sent[t]["pos"] += 1
                elif s == "NEGATIF": daily_sent[t]["neg"] += 1
                else: daily_sent[t]["neu"] += 1
        # Convert to score series
        dates_sorted = sorted(daily_sent.keys())
        scores = []
        for d in dates_sorted:
            ds = daily_sent[d]
            if ds["total"] > 0:
                score = (ds["pos"] - ds["neg"]) / ds["total"]
                scores.append(score)
        if scores:
            data["sentimen_score"] = scores
        print(f"  Sentimen: {len(rows)} articles, {len(scores)} daily scores")

    # Saham (31 rows)
    rows = read_sheet("harga_saham_ihsg.xlsx", "IHSG")
    if rows:
        vals = []
        for r in rows:
            v = r.get("IHSG") or r.get("Close")
            if v:
                try: vals.append(float(v))
                except (ValueError, TypeError): pass
        if len(vals) >= 2:
            data["saham_ihsg"] = vals
        print(f"  Saham: {len(rows)} rows")

    # Pakan (61 rows)
    rows = read_sheet("harga_pakan_ternak.xlsx", "Harga Pakan")
    if rows:
        for k in rows[-1].keys():
            if k and k != "Tanggal":
                vals = []
                for r in rows:
                    v = r.get(k)
                    if v:
                        try:
                            fv = float(v)
                            if fv > 0: vals.append(fv)
                        except (ValueError, TypeError):
                            pass
                if len(vals) >= 2:
                    data[f"pakan_{k[:20]}"] = vals
        print(f"  Pakan: {len(rows)} rows, {len([k for k in data if k.startswith('pakan')])} series")

    # Emas (from JSON)
    jp = DATA_DIR / "emas_history.json"
    if jp.exists():
        try:
            d = json.loads(jp.read_text())
            if isinstance(d, list) and len(d) > 1:
                vals = [float(e.get("antam_beli", 0)) for e in d if e.get("antam_beli")]
                if len(vals) >= 2:
                    data["emas_antam"] = vals
                print(f"  Emas: {len(d)} days history")
        except (Exception,): pass

    total = sum(len(v) for v in data.values())
    print(f"{ts()} ✅ {len(data)} series, {total} data points total")
    return data


def compute_all_indicators(history):
    """Compute technical indicators for all series — with multi-period trends, forecasts, enhanced confidence."""
    print(f"{ts()} 📐 Computing indicators...")
    indicators = {}

    for name, values in history.items():
        if len(values) >= 2:
            result = full_analysis(name, values)
            # Add multi-period trend (3/7/14 days)
            result["multi_period"] = multi_period_trend(values)
            # Add % change from baselines (1/7/14/30 days)
            result["changes"] = change_from_baseline(values)
            # Add 7-day price forecast (if enough data)
            if len(values) >= 5:
                result["forecast_7d"] = price_forecast(values, horizon=7)
            # Enhanced confidence score
            result["enhanced_confidence"] = enhanced_confidence(
                values, result["trend"], result.get("volatility"), result.get("forecast_7d")
            )
            indicators[name] = result

    # ── EXPANDED CROSS-CORRELATIONS (12 pairs) ──
    corr_pairs = [
        # Macro
        ("kurs_USD_IDR", "emas_antam", "USD↔Emas"),
        ("oil_brent", "cpo_myr", "Minyak↔CPO"),
        # Sembako internal
        ("sembako_Telur Ras", "sembako_Daging Sapi", "Telur↔Daging"),
        ("sembako_Gula Pasir", "sembako_Ber", "Gula↔Beras"),
        ("sembako_Cabai Merah", "sembako_Bawang Merah", "Cabai↔Bawang"),
        # Pakan chain (feed → livestock)
        ("kurs_USD_IDR", "sembako_Daging Sapi", "Kurs↔Daging Sapi"),
        ("oil_brent", "sembako_Minyak Goreng", "Minyak↔Minyak Goreng"),
        # Cross-sector
        ("saham_ihsg", "sembako_Ber", "IHSG↔Beras"),
        ("sentimen_score", "saham_ihsg", "Sentimen↔IHSG"),
        ("sentimen_score", "sembako_Telur Ras", "Sentimen↔Telur"),
        ("kurs_USD_IDR", "saham_ihsg", "Kurs↔IHSG"),
        ("emas_antam", "saham_ihsg", "Emas↔IHSG"),
    ]

    correlations = {}
    for a, b, label in corr_pairs:
        if a in history and b in history:
            n = min(len(history[a]), len(history[b]))
            if n >= 3:
                corr = pearson_correlation(history[a][-n:], history[b][-n:])
                if corr is not None:
                    correlations[label] = corr

    # Anomalies — only on genuine time-series
    anomalies = []
    time_series_prefixes = ("sembako_", "crypto_", "pet_", "saham_", "sentimen_", "kurs_", "oil_")
    for name, values in history.items():
        if not any(name.startswith(p) for p in time_series_prefixes):
            continue
        if len(values) >= 3:
            detected = detect_anomalies(values, threshold_pct=8)
            for a in detected:
                # Filter: skip anomali palsu dari values kecil (sentimen score ~0)
                if abs(a['from']) < 1:
                    continue
                anomalies.append({"series": name, **a})

    # Seasonal
    seasonal = get_seasonal_context()

    # ── MACRO ECONOMIC INDICATORS ──
    macro = {}
    print(f"{ts()} 📐 Computing macro indicators...")
    # 1. Real Interest Rate
    bi_rows = read_sheet("bi_rate_inflasi.xlsx", "Harian")
    if bi_rows:
        br = bi_rows[-1].get("BI_Rate")
        inf = bi_rows[-1].get("Inflasi_YoY")
        if br and inf:
            macro["real_interest_rate"] = real_interest_rate(float(br), float(inf))
    # 2. Currency Pressure Index
    if "kurs_USD_IDR" in history and len(history["kurs_USD_IDR"]) >= 5:
        macro["currency_pressure"] = currency_pressure_index(history["kurs_USD_IDR"])
    # 3. Import Cost Pressure
    oil_brent = history.get("oil_brent", [None])[-1]
    usd_idr = history.get("kurs_USD_IDR", [None])[-1]
    if oil_brent and usd_idr:
        macro["import_cost_pressure"] = import_cost_pressure(float(oil_brent), float(usd_idr))
    # 4. Food Inflation Proxy
    sembako_vals = []
    for k in ["sembako_Ber", "sembako_Telur Ras", "sembako_Gula Pasir",
              "sembako_Cabai Merah", "sembako_Minyak Goreng", "sembako_Bawang Merah"]:
        if k in history and history[k]:
            sembako_vals.append(history[k][-1])
    if sembako_vals:
        basket_avg = sum(sembako_vals) / len(sembako_vals)
        # Build basket history from available sembako series
        basket_history = []
        available_keys = [k for k in ["sembako_Ber", "sembako_Telur Ras"] if k in history and history[k]]
        if available_keys:
            max_len = max(len(history[k]) for k in available_keys)
            for i in range(max(max_len - 35, 0), max_len):
                day_vals = []
                for k in ["sembako_Ber", "sembako_Telur Ras", "sembako_Gula Pasir",
                           "sembako_Cabai Merah", "sembako_Minyak Goreng", "sembako_Bawang Merah"]:
                    v = history.get(k, [])
                    if i < len(v):
                        day_vals.append(v[i])
                if day_vals:
                    basket_history.append(sum(day_vals) / len(day_vals))
        if basket_history:
            macro["food_inflation_proxy"] = food_inflation_proxy(basket_history)
    # 5. Stagflation Risk
    ihsg_trend = indicators.get("saham_ihsg", {}).get("trend", "sideways")
    food_sig = macro.get("food_inflation_proxy", {}).get("signal", "stable") if macro.get("food_inflation_proxy") else "stable"
    rir = macro.get("real_interest_rate")
    if rir:
        # Extract actual BI rate and inflation from the latest data
        _bi_rows = read_sheet("bi_rate_inflasi.xlsx", "Harian")
        _br = float(_bi_rows[-1].get("BI_Rate")) if _bi_rows and _bi_rows[-1].get("BI_Rate") else None
        _inf = float(_bi_rows[-1].get("Inflasi_YoY")) if _bi_rows and _bi_rows[-1].get("Inflasi_YoY") else None
        macro["stagflation_risk"] = stagflation_risk(
            _br, _inf, ihsg_trend, food_sig
        )
    # 6. Risk-On/Off Index
    btc_chg = None
    if "crypto_btc_usd" in history and len(history["crypto_btc_usd"]) >= 5:
        btc_chg = (history["crypto_btc_usd"][-1] - history["crypto_btc_usd"][-5]) / history["crypto_btc_usd"][-5] * 100
    gold_chg = None
    if "emas_antam" in history and len(history["emas_antam"]) >= 5:
        gold_chg = (history["emas_antam"][-1] - history["emas_antam"][-5]) / history["emas_antam"][-5] * 100
    ihsg_chg = None
    if "saham_ihsg" in history and len(history["saham_ihsg"]) >= 5:
        ihsg_chg = (history["saham_ihsg"][-1] - history["saham_ihsg"][-5]) / history["saham_ihsg"][-5] * 100
    if btc_chg is not None or gold_chg is not None or ihsg_chg is not None:
        macro["risk_on_off"] = risk_on_off_index(btc_chg, gold_chg, ihsg_chg)
    # 7. Commodity Terms of Trade
    cpo_myr = history.get("cpo_myr", [None])[-1]
    if cpo_myr and oil_brent:
        macro["commodity_terms"] = commodity_terms_of_trade(float(cpo_myr), float(oil_brent))

    print(f"{ts()} ✅ {len(macro)} macro indicators computed")
    indicators["_macro"] = macro

    # Count stats
    with_mpt = sum(1 for v in indicators.values() if isinstance(v, dict) and v.get("multi_period"))
    with_fc = sum(1 for v in indicators.values() if isinstance(v, dict) and v.get("forecast_7d"))
    print(f"{ts()} ✅ {len([v for v in indicators if v != '_macro'])} indicators, {with_mpt} multi-period, {with_fc} forecasts, {len(correlations)} correlations, {len(anomalies)} anomalies, {len(macro)} macro")
    return indicators, correlations, anomalies, seasonal


def build_prompt(history, indicators, correlations, anomalies, seasonal):
    """Build comprehensive prompt for Groq API."""
    lines = []
    lines.append("ANALISA KOMPREHENSIF PASAR INDONESIA")
    lines.append(f"Tanggal: {datetime.now().strftime('%d %B %Y, %H:%M')}")
    lines.append(f"Data points: {sum(len(v) for v in history.values())}")
    lines.append("")

    # Current prices
    lines.append("=== HARGA TERKINI ===")
    snapshot_keys = {
        "Ber (Beras)": history.get("sembako_Ber", [0])[-1],
        "Telur Ras": history.get("sembako_Telur Ras", [0])[-1],
        "Gula Pasir": history.get("sembako_Gula Pasir", [0])[-1],
        "Cabai Merah": history.get("sembako_Cabai Merah", [0])[-1],
        "Minyak Goreng": history.get("sembako_Minyak Goreng", [0])[-1],
        "Daging Sapi": history.get("sembako_Daging Sapi", [0])[-1],
        "BTC (USD)": history.get("crypto_btc_usd", [0])[-1],
        "ETH (USD)": history.get("crypto_eth_usd", [0])[-1],
    }
    for k, v in snapshot_keys.items():
        if v:
            lines.append(f"  {k}: Rp{v:,.0f}" if v > 1000 else f"  {k}: ${v:,.2f}")

    lines.append("")

    # Technical indicators summary
    lines.append("=== INDIKATOR TEKNIS ===")
    for name, ind in sorted(indicators.items()):
        if ind.get("error"): continue
        trend = ind.get("trend", "n/a")
        mom = ind.get("momentum", 0)
        vol = ind.get("volatility")
        conf = ind.get("confidence", 0)
        ma5 = ind.get("ma5")
        ma20 = ind.get("ma20")
        roc = ind.get("roc_5d")
        support = ind.get("support", [])
        resistance = ind.get("resistance", [])

        trend_icon = "🟢" if trend == "uptrend" else "🔴" if trend == "downtrend" else "🟡"

        lines.append(f"  {name}:")
        lines.append(f"    Harga: Rp{ind['current']:,.0f} ({ind['change_pct']:+.1f}%) | {trend_icon} {trend}")
        if ma5: lines.append(f"    MA5: Rp{ma5:,.0f} | MA20: {'Rp'+f'{ma20:,.0f}' if ma20 else '-'}")
        if vol is not None: lines.append(f"    Volatility: {vol:.2f}% | Momentum: {mom}/100")
        if roc is not None: lines.append(f"    ROC(5): {roc:+.1f}%")
        if support: lines.append(f"    Support: {', '.join(f'Rp{s:,.0f}' for s in support[:2])}")
        if resistance: lines.append(f"    Resistance: {', '.join(f'Rp{r:,.0f}' for r in resistance[:2])}")
        lines.append(f"    Confidence: {'⭐'*conf} ({conf}/5)")

    lines.append("")

    # Correlations
    if correlations:
        lines.append("=== KORELASI LINTAS SEKTOR ===")
        for label, corr in correlations.items():
            strength = "kuat" if abs(corr) > 0.7 else "sedang" if abs(corr) > 0.4 else "lemah"
            direction = "positif" if corr > 0 else "negatif"
            lines.append(f"  {label}: {corr:+.3f} (korelasi {strength} {direction})")
        lines.append("")

    # Anomalies
    if anomalies:
        lines.append("=== ANOMALI TERDETEKSI ===")
        for a in anomalies[:10]:
            lines.append(f"  ⚠️ {a['series']}: {a['type']} {a['change_pct']:+.1f}% ({a['from']:,.0f} → {a['to']:,.0f})")
        lines.append("")

    # Seasonal
    if seasonal:
        lines.append("=== KONTEKS MUSIMAN ===")
        for s in seasonal:
            lines.append(f"  {s['pattern']}: {s['effect']} (dampak: {s['impact']})")
        lines.append("")

    return "\n".join(lines)


def build_rule_analysis(history, indicators, correlations, anomalies, seasonal, macro=None):
    """Generate comprehensive rule-based analysis from computed indicators."""
    date_str = datetime.now().strftime("%d %B %Y")
    lines = []

    # ── EXECUTIVE SUMMARY ──
    lines.append("## 📊 Executive Summary")

    # Count trends
    trends = [v.get("trend","n/a") for v in indicators.values() if v.get("trend")]
    up = trends.count("uptrend")
    down = trends.count("downtrend")
    side = trends.count("sideways")

    # Use enhanced confidence scores
    total_conf = [v.get("enhanced_confidence", v.get("confidence", 0)) for v in indicators.values() if v.get("trend") != "insufficient_data"]
    avg_conf = sum(total_conf) / max(len(total_conf), 1)

    # Overall market direction
    if up > down * 1.5:
        market_dir = "**POSITIF** 🟢 — mayoritas komoditas uptrend"
    elif down > up * 1.5:
        market_dir = "**NEGATIF** 🔴 — mayoritas komoditas downtrend"
    else:
        market_dir = "**CAMPURAN** 🟡 — tren mixed antar sektor"

    lines.append(f"Per {date_str}, kondisi pasar {market_dir}.")
    lines.append(f"- {up} komoditas uptrend, {down} downtrend, {side} sideways")
    lines.append(f"- Rata-rata confidence level: {avg_conf:.1f}/5 {'⭐' * min(5, round(avg_conf))}")

    if anomalies:
        high_anomaly = [a for a in anomalies if abs(a['change_pct']) > 10]
        if high_anomaly:
            lines.append(f"- ⚠️ **{len(high_anomaly)} anomali signifikan** terdeteksi (pergerakan >10%)")

    if seasonal:
        active = [s for s in seasonal if s['impact'] == 'high']
        if active:
            lines.append(f"- 🌙 **Konteks musiman**: {active[0]['effect']}")
    lines.append("")

    # ── TOP 3 HAL PENTING ──
    highlights = []
    # Biggest movers
    biggest_up = [(k, v) for k, v in indicators.items()
                  if v.get('change_pct', 0) > 3 and not v.get('error') and v.get('trend') == 'uptrend']
    biggest_down = [(k, v) for k, v in indicators.items()
                    if v.get('change_pct', 0) < -3 and not v.get('error') and v.get('trend') == 'downtrend']
    if biggest_up:
        best = max(biggest_up, key=lambda x: x[1]['change_pct'])
        highlights.append(f"🟢 **{best[0].replace('_',' ').title()}** naik {best[1]['change_pct']:+.1f}% — momentum kuat")
    if biggest_down:
        worst = min(biggest_down, key=lambda x: x[1]['change_pct'])
        highlights.append(f"🔴 **{worst[0].replace('_',' ').title()}** turun {worst[1]['change_pct']:+.1f}% — tekanan jual")
    # Strongest correlation
    if correlations:
        top_corr = max(correlations.items(), key=lambda x: abs(x[1]))
        if abs(top_corr[1]) > 0.4:
            highlights.append(f"🔗 **{top_corr[0]}**: korelasi {top_corr[1]:+.3f} — {'searah' if top_corr[1] > 0 else 'berlawanan'}")
    if highlights:
        lines.append("**3 Hal Penting:**")
        for h in highlights[:3]:
            lines.append(f"- {h}")
    lines.append("")

    # ── SEKTOR SEMBAKO ──
    lines.append("## 🛒 Sembako — Analisa Teknis")
    sembako_keys = [k for k in indicators if k.startswith("sembako_")]
    if sembako_keys:
        lines.append("| Komoditas | Harga | MA5 | MA20 | Trend | Momentum | Volatility | Conf |")
        lines.append("|-----------|-------|-----|------|-------|----------|-----------|------|")
        for k in sorted(sembako_keys):
            ind = indicators[k]
            name = k.replace("sembako_", "")
            cur = f"Rp{ind['current']:,.0f}"
            ma5 = f"Rp{ind['ma5']:,.0f}" if ind.get('ma5') else "-"
            ma20 = f"Rp{ind['ma20']:,.0f}" if ind.get('ma20') else "-"
            t = "🟢↑" if ind['trend'] == 'uptrend' else "🔴↓" if ind['trend'] == 'downtrend' else "🟡→"
            mom = f"{ind['momentum']:+d}"
            vol = f"{ind['volatility']:.1f}%" if ind.get('volatility') else "-"
            conf = "⭐" * ind['confidence']
            # Enhanced confidence
            ec = ind.get('enhanced_confidence', ind['confidence'])
            conf_e = "⭐" * ec
            lines.append(f"| {name} | {cur} | {ma5} | {ma20} | {t} | {mom} | {vol} | {conf_e} |")
    lines.append("")

    # ── MULTI-PERIOD TRENDS + FORECAST ──
    trend_key_items = ["Beras Premium", "Telur Ras", "Daging Sapi", "Cabai Keriting", "Bawang Merah", "Gula Pasir", "Minyak Goreng"]
    trend_data = [(f"sembako_{t}", t) for t in trend_key_items if f"sembako_{t}" in indicators]
    # Also include pakan & peternakan
    trend_data += [(f"pakan_{k}", k) for k in ["Jagung Pipilan", "Bungkil Kedelai", "Bekatul"] if f"pakan_{k}" in indicators]
    trend_data += [(f"pet_{k}", k) for k in ["Ayam_Hidup", "Daging_Eceran", "Telur_(Grosir)"] if f"pet_{k}" in indicators]

    if trend_data:
        lines.append("## 📈 Tren Multi-Periode + Prediksi 7 Hari")
        lines.append("| Komoditas | Saat Ini | vs 1 hari | vs 7 hari | vs 14 hari | Prediksi 7h | Arah Prediksi |")
        lines.append("|-----------|----------|-----------|-----------|-----------|------------|---------------|")
        for key, label in trend_data:
            ind = indicators[key]
            cur = ind['current']
            chg = ind.get('changes', {})
            fc = ind.get('forecast_7d', [])
            chg1 = f"{chg.get(1, 0) or 0:+.1f}%" if chg.get(1) is not None else "-"
            chg7 = f"{chg.get(7, 0) or 0:+.1f}%" if chg.get(7) is not None else "-"
            chg14 = f"{chg.get(14, 0) or 0:+.1f}%" if chg.get(14) is not None else "-"
            if fc:
                fc_avg = sum(fc) / len(fc)
                fc_dir = "🟢↑" if fc_avg > cur * 1.01 else "🔴↓" if fc_avg < cur * 0.99 else "🟡→"
                fc_str = f"Rp{fc_avg:,.0f}" if cur > 100 else f"${fc_avg:,.0f}"
            else:
                fc_str = "-"
                fc_dir = "-"
            lines.append(f"| {label} | Rp{cur:,.0f} | {chg1} | {chg7} | {chg14} | {fc_str} | {fc_dir} |")
        lines.append("")

    # ── SEKTOR CRYPTO ──
    lines.append("## ₿ Crypto — Analisa Teknis")
    crypto_keys = [k for k in indicators if k.startswith("crypto_")]
    if crypto_keys:
        for k in sorted(crypto_keys):
            ind = indicators[k]
            name = k.replace("crypto_", "").upper()
            t = "🟢 Bullish" if ind['trend'] == 'uptrend' else "🔴 Bearish" if ind['trend'] == 'downtrend' else "🟡 Sideways"
            lines.append(f"**{name}**: ${ind['current']:,.2f} ({ind['change_pct']:+.1f}%) {t}")
            if ind.get('ma5'): lines.append(f"  MA5: ${ind['ma5']:,.2f}")
            if ind.get('volatility'): lines.append(f"  Volatility: {ind['volatility']:.2f}%")
            sr = ind
            if sr.get('support'): lines.append(f"  Support: {', '.join(f'${s:,.2f}' for s in sr['support'][:2])}")
            if sr.get('resistance'): lines.append(f"  Resistance: {', '.join(f'${r:,.2f}' for r in sr['resistance'][:2])}")
            lines.append(f"  Confidence: {'⭐'*ind['confidence']} ({ind['confidence']}/5)")
            lines.append("")
    else:
        lines.append("- Data historis crypto belum cukup untuk analisa teknis")
        lines.append("")

    # ── PETERNAKAN — Margin Analysis ──
    lines.append("## 🐄 Peternakan — Margin & Chain")
    pet_keys = [k for k in indicators if k.startswith("pet_")]
    if pet_keys:
        lines.append("| Produk | Harga | Trend | Momentum | Volatility |")
        lines.append("|--------|-------|-------|----------|-----------|")
        for k in sorted(pet_keys):
            ind = indicators[k]
            name = k.replace("pet_", "")
            t = "🟢↑" if ind['trend'] == 'uptrend' else "🔴↓" if ind['trend'] == 'downtrend' else "🟡→"
            lines.append(f"| {name} | Rp{ind['current']:,.0f} | {t} | {ind['momentum']:+d} | {ind.get('volatility','-')} |")
        lines.append("")

    # ── CROSS-SECTOR CORRELATIONS ──
    if correlations:
        lines.append("## 🔗 Korelasi Lintas Sektor")
        for label, corr in sorted(correlations.items(), key=lambda x: abs(x[1]), reverse=True):
            strength = "🔴 Kuat" if abs(corr) > 0.7 else "🟡 Sedang" if abs(corr) > 0.4 else "⚪ Lemah"
            direction = "positif" if corr > 0 else "negatif"
            lines.append(f"- **{label}**: {corr:+.3f} — korelasi {strength} {direction}")
            if corr > 0.7:
                lines.append(f"  → Keduanya bergerak searah — jika {label.split('↔')[0]} naik, {label.split('↔')[1]} cenderung naik juga")
            elif corr < -0.7:
                lines.append(f"  → Bergerak berlawanan — diversifikasi otomatis")
        lines.append("")

    # ── ANOMALI ──
    if anomalies:
        lines.append("## ⚠️ Anomali & Alert")
        for a in sorted(anomalies, key=lambda x: abs(x['change_pct']), reverse=True)[:8]:
            icon = "🔴" if a['type'] == 'crash' else "🟢"
            lines.append(f"- {icon} **{a['series']}**: {a['change_pct']:+.1f}% ({a['from']:,.0f} → {a['to']:,.0f})")
        lines.append("")
    else:
        lines.append("## ✅ Stabilitas Harga")
        lines.append("- Tidak ada anomali signifikan terdeteksi")
        lines.append("")

    # ── SEASONAL CONTEXT ──
    if seasonal:
        lines.append("## 🌙 Konteks Musiman")
        for s in seasonal:
            impact_icon = "🔴" if s['impact'] == 'high' else "🟡" if s['impact'] == 'medium' else "⚪"
            lines.append(f"- {impact_icon} **{s['pattern'].replace('_',' ').title()}**: {s['effect']}")
        lines.append("")

    # ── MACRO ECONOMIC DASHBOARD ──
    macro = macro or {}
    if macro:
        lines.append("## 🏛️ Indikator Makro Ekonomi")

        # Real Interest Rate
        rir = macro.get("real_interest_rate")
        if rir:
            sig_icon = {"hawkish": "🔴", "neutral": "🟢", "dovish": "🟡", "danger": "⚠️"}.get(rir["signal"], "⚪")
            lines.append(f"- {sig_icon} **Real Interest Rate**: {rir['value']:+.1f}% — {rir['interpretation']}")

        # Currency Pressure
        cp = macro.get("currency_pressure")
        if cp:
            sig_icon = {"danger": "🔴", "warning": "🟡", "cautious": "🟠", "stable": "🟢"}.get(cp["signal"], "⚪")
            lines.append(f"- {sig_icon} **Currency Pressure Index**: {cp['score']}/100 (5d: {cp['trend_5d_pct']:+.1f}%) — {cp['interpretation']}")

        # Import Cost Pressure
        icp = macro.get("import_cost_pressure")
        if icp:
            sig_icon = {"danger": "🔴", "warning": "🟡", "neutral": "🟢", "favorable": "💚"}.get(icp["signal"], "⚪")
            lines.append(f"- {sig_icon} **Import Cost Pressure**: ratio {icp['ratio_vs_baseline']:.2f}x baseline — {icp['interpretation']}")

        # Food Inflation Proxy
        fip = macro.get("food_inflation_proxy")
        if fip:
            sig_icon = {"high_inflation": "🔴", "moderate": "🟡", "stable": "🟢", "deflation": "🔵"}.get(fip["signal"], "⚪")
            lines.append(f"- {sig_icon} **Food Inflation Proxy**: 30d {fip['change_30d']:+.1f}%, 7d {fip['change_7d']:+.1f}% — {fip['interpretation']}")

        # Stagflation Risk
        sr = macro.get("stagflation_risk")
        if sr:
            sig_icon = {"high_risk": "🔴", "moderate": "🟡", "low": "🟢", "minimal": "✅"}.get(sr["signal"], "⚪")
            lines.append(f"- {sig_icon} **Stagflation Risk**: {sr['score']}/100 — {sr['interpretation']}")
            if sr["factors"]:
                for f in sr["factors"][:3]:
                    lines.append(f"  - ⮑ {f}")

        # Risk-On/Off
        roo = macro.get("risk_on_off")
        if roo:
            sig_icon = {"risk_on": "🟢", "mild_on": "🟢", "neutral": "⚪", "mild_off": "🟡", "risk_off": "🔴"}.get(roo["signal"], "⚪")
            lines.append(f"- {sig_icon} **Risk-On/Off Index**: {roo['score']:+d}/100 — {roo['interpretation']}")

        # Commodity Terms of Trade
        ctt = macro.get("commodity_terms")
        if ctt:
            sig_icon = {"favorable": "💚", "positive": "🟢", "neutral": "⚪", "warning": "🟡", "danger": "🔴"}.get(ctt["signal"], "⚪")
            lines.append(f"- {sig_icon} **Commodity Terms of Trade**: CPO/Oil {ctt['ratio']:.2f} ({ctt['pct_vs_baseline']:+.1f}% vs baseline) — {ctt['interpretation']}")

        # Summary
        danger_count = sum(1 for v in macro.values() if isinstance(v, dict) and v.get("signal") in ("danger", "high_risk", "risk_off"))
        warning_count = sum(1 for v in macro.values() if isinstance(v, dict) and v.get("signal") in ("warning", "moderate", "cautious", "mild_off"))
        lines.append("")
        if danger_count >= 2:
            lines.append(f"⚠️ **{danger_count} indikator dalam zona BAHAYA** — ekonomi butuh perhatian serius.")
        elif danger_count >= 1:
            lines.append(f"⚠️ **{danger_count} indikator zona bahaya** — waspada terhadap potensi tekanan.")
        elif warning_count >= 3:
            lines.append(f"🟡 **{warning_count} indikator waspada** — monitor perkembangan lebih lanjut.")
        else:
            lines.append("✅ **Kondisi makro stabil** — tidak ada zona bahaya terdeteksi.")
        lines.append("")

    # ── RISK ASSESSMENT ──
    lines.append("## ⚠️ Risk Assessment")
    risks = []

    # Check high volatility items
    high_vol = [(k, v) for k, v in indicators.items()
                if v.get('volatility') and v['volatility'] > 5 and not v.get('error')]
    if high_vol:
        names = ", ".join(k.split("_",1)[-1] for k,_ in high_vol[:3])
        risks.append(f"🔴 **Volatil**: {names} — volatility >5%, berisiko fluktuasi tajam")

    # Check downtrend
    downtrends = [(k, v) for k, v in indicators.items()
                  if v.get('trend') == 'downtrend' and not v.get('error')]
    if len(downtrends) >= 3:
        names = ", ".join(k.split("_",1)[-1] for k,_ in downtrends[:3])
        risks.append(f"🔴 **Downtrend**: {names} — tekanan jual berkelanjutan")

    # Check anomalies
    crashes = [a for a in anomalies if a['type'] == 'crash' and abs(a['change_pct']) > 10]
    if crashes:
        risks.append(f"🔴 **Crash detected**: {len(crashes)} komoditas turun >10%")

    # Low confidence
    low_conf = [(k, v) for k, v in indicators.items()
                if v.get('confidence', 5) <= 2 and not v.get('error')]
    if low_conf:
        names = ", ".join(k.split("_",1)[-1] for k,_ in low_conf[:3])
        risks.append(f"🟡 **Confidence rendah**: {names} — data kurang, prediksi kurang reliable")

    # Strong correlation risk
    strong_neg = [(l, c) for l, c in correlations.items() if c < -0.7]
    if strong_neg:
        lines.append(f"- ⚠️ Korelasi negatif kuat: {', '.join(l for l,_ in strong_neg)}")

    if not risks:
        risks.append("- ✅ Tidak ada risiko signifikan terdeteksi")

    for r in risks:
        lines.append(f"- {r}")
    lines.append("")

    # ── REKOMENDASI with CONFIDENCE ──
    lines.append("## 💡 Rekomendasi Berbasis Data")

    # Petani/ternak
    lines.append("**🌾 Petani & Peternak:**")
    pakan_vol = indicators.get("pakan_Bungkil Kelapa", {})
    if pakan_vol.get("trend") == "downtrend":
        lines.append("- ⭐⭐⭐⭐ Pakan bahan mentah downtrend — **stok sekarang** (confidence tinggi)")
    elif pakan_vol.get("trend") == "uptrend":
        lines.append("- ⭐⭐⭐ Pakan bahan mentah uptrend — **tunda stok, beli bertahap**")

    telur_ind = indicators.get("sembako_Telur Ras", {})
    if telur_ind.get("trend") == "uptrend" and telur_ind.get("momentum", 0) > 50:
        lines.append("- ⭐⭐⭐⭐ Telur uptrend kuat — **perluasan produksi menguntungkan**")
    elif telur_ind.get("trend") == "downtrend":
        lines.append("- ⭐⭐⭐ Telur downtrend — **jaga volume, kurangi stok berlebih**")

    sapi_ind = indicators.get("sembako_Daging Sapi", {})
    if sapi_ind.get("trend") == "uptrend":
        lines.append("- ⭐⭐⭐ Daging sapi naik — **peluang jual ternak lebih menguntungkan**")

    lines.append("")

    # Konsumen
    lines.append("**🛒 Konsumen:**")
    hot_items = [(k.replace("sembako_",""), v) for k, v in indicators.items()
                 if k.startswith("sembako_") and v.get("change_pct", 0) > 3 and not v.get("error")]
    if hot_items:
        names = ", ".join(n for n,_ in hot_items[:3])
        lines.append(f"- ⭐⭐⭐ **Harga naik signifikan**: {names} — beli seperlunya")
    cold_items = [(k.replace("sembako_",""), v) for k, v in indicators.items()
                  if k.startswith("sembako_") and v.get("change_pct", 0) < -2 and not v.get("error")]
    if cold_items:
        names = ", ".join(n for n,_ in cold_items[:3])
        lines.append(f"- ⭐⭐⭐ **Harga turun**: {names} — moment tepat untuk stok")
    lines.append("")

    # Investor
    lines.append("**📈 Investor:**")
    btc_ind = indicators.get("crypto_btc_usd", {})
    if btc_ind.get("trend") == "uptrend":
        lines.append(f"- ⭐⭐⭐⭐ BTC ${btc_ind.get('current',0):,.0f} uptrend — **DCA aktif**")
    elif btc_ind.get("trend") == "downtrend":
        lines.append(f"- ⭐⭐⭐ BTC ${btc_ind.get('current',0):,.0f} downtrend — **wait & see, support: ${btc_ind.get('support',['?'])[0] if btc_ind.get('support') else '?'}**")

    if correlations.get("USD↔Emas", 0) < -0.5:
        lines.append("- ⭐⭐⭐ USD↔Emas negatif — **lindungi portofolio dengan emas saat USD lemah**")

    lines.append("- Diversifikasi: 40% pasar uang, 30% ekuitas, 20% komoditas, 10% crypto")
    lines.append("")

    # ── PREDIKSI ──
    lines.append("## 🎯 Prediksi 7 Hari ke Depan")

    for sector_name, prefix, icon in [
        ("Sembako", "sembako_", "🛒"),
        ("Crypto", "crypto_", "₿"),
    ]:
        sector_keys = [k for k in indicators if k.startswith(prefix) and not indicators[k].get("error")]
        if not sector_keys: continue

        up_count = sum(1 for k in sector_keys if indicators[k].get("trend") == "uptrend")
        down_count = sum(1 for k in sector_keys if indicators[k].get("trend") == "downtrend")
        avg_roc = sum(indicators[k].get("roc_5d", 0) or 0 for k in sector_keys) / max(len(sector_keys), 1)

        if up_count > down_count * 2:
            pred = "🟢 **Cenderung Naik**"
        elif down_count > up_count * 2:
            pred = "🔴 **Cenderung Turun**"
        else:
            pred = "🟡 **Sideways/Mixed**"

        lines.append(f"{icon} **{sector_name}**: {pred} (↑{up_count} ↓{down_count} | ROC rata-rata: {avg_roc:+.1f}%)")
    lines.append("")

    lines.append("**Faktor yang perlu dipantau:**")
    lines.append("- Kebijakan BI Rate berikutnya (apotek moneter)")
    lines.append("- Sentimen global: Fed rate decision, crypto regulation")
    lines.append("- Cuaca: potensi El Niño affect pertanian")
    if seasonal:
        lines.append(f"- Musiman: {seasonal[0]['effect']}")
    lines.append("")
    lines.append("---")
    lines.append(f"_Analisis komprehensif | {date_str} | {sum(len(v) for v in history.values())} data points | {len(indicators)} indikator | {len(correlations)} korelasi_")

    return "\n".join(lines)


def call_groq(prompt, data_text):
    """Call Groq API with comprehensive data context."""
    import urllib.request
    if not GROQ_API_KEY:
        return ""

    SYSTEM_PROMPT = """Anda adalah analis pasar komoditas Indonesia profesional. Analisa berdasarkan data teknis dan fundamental.

Output dalam Bahasa Indonesia dengan format Markdown:
## 📊 Executive Summary
Ringkasan kondisi pasar 2-3 kalimat. Sebutkan angka spesifik.

## 📈 Analisa Sektoral
Untuk setiap sektor (Sembako, Crypto, Emas, Pertanian, Peternakan, Energi, Kurs, Moneter):
- Sebutkan harga aktual, MA5/MA20, trend (uptrend/downtrend/sideways)
- Sebutkan momentum (skor -100 to +100)
- Sebutkan support/resistance jika ada
- Berikan confidence level ⭐ (1-5)

## 🔗 Korelasi
Jelaskan korelasi antar sektor dan implikasinya.

## ⚠️ Risiko & Anomali
Sebutkan anomali, volatilitas tinggi, atau risiko yang terdeteksi.

## 💡 Rekomendasi
Rekomendasi SPESIFIK dengan confidence score:
- Petani/Peternak: kapan stok pakan, kapan jual ternak
- Konsumen: kapan beli, komoditas mana yang perlu diwaspadai
- Investor: strategi DCA, diversifikasi, hedging

## 🎯 Prediksi 7 Hari
Berdasarkan momentum dan trend, prediksi arah pasar.

Penting: Gunakan angka spesifik dari data. Jangan generalisasi."""

    payload = json.dumps({
        "model": "llama-3.1-8b-instant",
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": f"DATA PASAR:\n{data_text}\n\n{prompt}"}
        ],
        "temperature": 0.3,
        "max_tokens": 2000,
    })

    req = urllib.request.Request(
        "https://api.groq.com/openai/v1/chat/completions",
        data=payload.encode(),
        headers={
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json",
        },
    )

    with urllib.request.urlopen(req, timeout=30) as resp:
        result = json.loads(resp.read())
        return result["choices"][0]["message"]["content"]


def main():
    print(f"{ts()} 🧠 AI Market Analysis V4 — COMPREHENSIVE")
    print(f"{'=' * 50}")

    # 1. Load all historical data
    history = load_all_history()
    if not history:
        print(f"{ts()} ❌ No data available")
        return

    # 2. Compute indicators
    indicators, correlations, anomalies, seasonal = compute_all_indicators(history)

    # Extract macro indicators — don't let _macro leak into indicator iteration loops
    macro_indicators = indicators.pop("_macro", {})

    # 3. Build prompt
    prompt = build_prompt(history, indicators, correlations, anomalies, seasonal)
    print(f"{ts()} 📝 Prompt: {len(prompt)} chars")

    # 4. Try Groq first
    analysis = ""
    source = "groq_v4"
    try:
        print(f"{ts()} 🤖 Calling Groq (llama-3.1-8b-instant)...")
        analysis = call_groq(
            "Buat analisa komprehensif berdasarkan data di atas. "
            "Sertakan angka spesifik, support/resistance, confidence score.",
            prompt
        )
        if analysis and len(analysis) > 200:
            print(f"{ts()} ✅ Groq response: {len(analysis)} chars")
        else:
            analysis = ""
            source = "rule_based_v4"
    except Exception as e:
        print(f"{ts()} ⚠️ Groq failed: {e}")
        source = "rule_based_v4"

    # 5. Fallback to rule-based
    if not analysis:
        print(f"{ts()} 🔄 Generating comprehensive rule-based analysis...")
        analysis = build_rule_analysis(history, indicators, correlations, anomalies, seasonal, macro=macro_indicators)
        source = "rule_based_v4"

    print(f"{ts()} ✅ Analysis: {len(analysis)} chars ({source})")
    print(f"{ts()} Preview: {analysis[:200]}...")

    # 6. Save
    output = {
        "analysis": analysis,
        "source": source,
        "model": "llama-3.1-8b-instant" if "groq" in source else "rule_based_v4",
        "generated_at": datetime.now().strftime("%d %b %Y, %H:%M"),
        "timestamp": datetime.now().isoformat(),
        "cached": False,
        "refresh_in_seconds": 28800,
        "indicators_count": len(indicators),
        "correlations_count": len(correlations),
        "anomalies_count": len(anomalies),
        "data_points": sum(len(v) for v in history.values()),
    }

    out_path = DATA_DIR / "daily_analysis.json"
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"{ts()} 💾 Saved to {out_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"{ts()} ❌ Error: {e}")
        traceback.print_exc()
