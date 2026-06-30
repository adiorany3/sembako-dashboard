#!/usr/bin/env python3
"""
Deduplicate Excel files — remove duplicate entries per composite key, keep most complete row.
Run manually or as post-processing after updater scripts.

Usage: python3 dedup_excel.py [--dry-run]
"""
import os
import sys
import openpyxl
from collections import Counter

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")

# (filename, sheet_name, key_columns)
# key_columns = indices that form unique identity; duplicates = same values in ALL key columns
DEDUP_TARGETS = [
    # Simple date-only dedup (1 row per date)
    ("harga_sembako.xlsx", "Harga", [0]),
    ("cuaca_yogyakarta.xlsx", "Cuaca", [0]),
    ("harga_emas.xlsx", "Harian", [0]),
    ("harga_saham_ihsg.xlsx", "IHSG", [0]),
    ("kurs_valuta.xlsx", "Harian", [0]),
    ("harga_minyak.xlsx", "Harian", [0]),
    ("bi_rate_inflasi.xlsx", "Harian", [0]),
    ("harga_cpo.xlsx", "Harian", [0]),
    ("harga_pertanian_ternak.xlsx", "Harga Komoditas", [0]),

    # Composite key — multiple valid rows per date
    # Peternakan: Tanggal + Produk (each date has ~86 products)
    ("harga_peternakan_lengkap.xlsx", "Data Utama", [0, 3]),  # Tanggal + Produk
    # Sentimen: Tanggal + Headline (many articles per date)
    ("sentimen_berita.xlsx", "Detail", [0, 3]),  # Tanggal + Headline
]


def make_key(row, key_cols):
    """Create composite key from specified column indices."""
    return tuple(str(row[i].value).strip() if row[i].value is not None else "" for i in key_cols)


def row_completeness(row):
    """Score: how many non-empty cells."""
    return sum(1 for c in row if c.value is not None and str(c.value).strip() != "")


def dedup_sheet(filepath, sheet_name, key_cols):
    """Remove duplicate rows by composite key, keeping most complete entry."""
    if not os.path.exists(filepath):
        return 0, 0

    wb = openpyxl.load_workbook(filepath)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return 0, 0

    ws = wb[sheet_name]
    rows = list(ws.iter_rows())
    if len(rows) <= 1:
        wb.close()
        return 0, 0

    header = rows[0]
    data_rows = rows[1:]

    # Group by composite key
    key_groups = {}
    for row in data_rows:
        key = make_key(row, key_cols)
        if key not in key_groups:
            key_groups[key] = []
        key_groups[key].append(row)

    # Dedup: keep most complete row per key
    deduped_rows = []
    removed = 0

    for key, group in key_groups.items():
        if len(group) == 1:
            deduped_rows.append(group[0])
        else:
            # Most complete wins; ties broken by last (most recent run)
            best = max(group, key=lambda r: (row_completeness(r), group.index(r)))
            deduped_rows.append(best)
            removed += len(group) - 1

    if removed == 0:
        wb.close()
        return 0, 0

    # Rewrite
    for row_idx in range(ws.max_row, 1, -1):
        ws.delete_rows(row_idx)

    for row_data in deduped_rows:
        ws.append([cell.value for cell in row_data])

    after = ws.max_row - 1
    wb.save(filepath)
    wb.close()
    return removed, after


def main():
    dry_run = "--dry-run" in sys.argv
    total_before = 0
    total_removed = 0

    print("=" * 60)
    print("DEDUP EXCEL — Hapus Entry Ganda Per Composite Key")
    print(f"Mode: {'DRY RUN' if dry_run else 'EXECUTE'}")
    print("=" * 60)

    for filename, sheet_name, key_cols in DEDUP_TARGETS:
        filepath = os.path.join(DATA_DIR, filename)
        if not os.path.exists(filepath):
            print(f"⏭️  SKIP {filename} — file not found")
            continue

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                print(f"⏭️  SKIP {filename} — sheet '{sheet_name}' not found")
                continue
            ws = wb[sheet_name]
            rows_data = list(ws.iter_rows(values_only=True))
            wb.close()

            before = len(rows_data) - 1

            # Count unique keys
            keys = [tuple(str(r[c]).strip() if r[c] else "" for c in key_cols) for r in rows_data[1:]]
            unique = len(set(keys))
            dupes = len(keys) - unique

            key_label = "+".join([str(rows_data[0][c]) if rows_data[0][c] else f"col{c}" for c in key_cols])

            if dupes > 0:
                print(f"🔧 {filename} / {sheet_name}: {before} rows, unique keys ({key_label}): {unique}, duplicates: {dupes}")
            else:
                print(f"✅ {filename} / {sheet_name}: {before} rows — no duplicates")

            if not dry_run and dupes > 0:
                removed, after = dedup_sheet(filepath, sheet_name, key_cols)
                print(f"   → {after} rows (removed {removed})")
                total_before += before
                total_removed += removed
        except Exception as e:
            print(f"❌ {filename}: ERROR: {e}")

    if not dry_run and total_removed > 0:
        print()
        print("=" * 60)
        print(f"TOTAL: {total_before} → {total_before - total_removed} rows ({total_removed} removed)")


if __name__ == "__main__":
    main()
