#!/usr/bin/env python3
"""
Update kurs valuta asing harian (USD, EUR, SGD, MYR) ke IDR.
Sumber utama: Frankfurter API (gratis, no key)
Fallback: exchangerate-api.com
Fallback 2: BI via Jina
"""
import os
import sys
import json
import re
from datetime import datetime
from urllib.request import Request, urlopen
from urllib.error import URLError

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Try importing jina from scripts dir, continue if missing
try:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from cache_lib import jina as jina_read
    HAS_JINA = True
except ImportError:
    HAS_JINA = False
    jina_read = None

EXCEL_PATH = os.path.expanduser("~/sembako/data/kurs_valuta.xlsx")
SHEET_NAME = "Harian"
HEADERS = ["Tanggal", "USD_IDR", "EUR_IDR", "SGD_IDR", "MYR_IDR", "Sumber"]
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FREEZE_PANES = "A2"

# Column widths
COL_WIDTHS = {"A": 14, "B": 14, "C": 14, "D": 14, "E": 14, "F": 22}


def create_or_load_workbook():
    """Create workbook with headers if not exists, else load existing."""
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if SHEET_NAME in wb.sheetnames:
            return wb, wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Freeze panes & column widths
    ws.freeze_panes = FREEZE_PANES
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    return wb, ws


def fetch_exchange_rates_api():
    """Try exchangerate-api.com free endpoint (no key needed)."""
    url = "https://open.er-api.com/v6/latest/USD"
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"})
        with urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))

        if data.get("result") == "success":
            rates = data.get("rates", {})
            idr = rates.get("IDR")
            if not idr:
                return None

            usd_idr = round(idr, 2)
            eur_idr = round(idr / rates.get("EUR", 1), 2) if rates.get("EUR") else None
            sgd_idr = round(idr / rates.get("SGD", 1), 2) if rates.get("SGD") else None
            myr_idr = round(idr / rates.get("MYR", 1), 2) if rates.get("MYR") else None

            return {
                "USD_IDR": usd_idr,
                "EUR_IDR": eur_idr,
                "SGD_IDR": sgd_idr,
                "MYR_IDR": myr_idr,
                "Sumber": "exchangerate-api.com",
            }
    except Exception as e:
        print(f"  ⚠️ exchangerate-api error: {e}")
    return None


def fetch_from_bi_jina():
    """Fallback: scrape Bank Indonesia kurs page via Jina."""
    url = "https://www.bi.go.id/id/data/ekonomi-dan-keuangan/kurs-default.aspx"
    text = jina_read(url)
    if not text or len(text) < 200:
        print("  ⚠️ Jina scraping BI gagal")
        return None

    rates = {}
    lines = text.split("\n")

    # Try to find currency pairs with IDR rates
    patterns = {
        "USD_IDR": [r"USD.*?IDR.*?([\d.,]+)", r"US\s*Dollar.*?([\d.,]+)"],
        "EUR_IDR": [r"EUR.*?IDR.*?([\d.,]+)", r"Euro.*?([\d.,]+)"],
        "SGD_IDR": [r"SGD.*?IDR.*?([\d.,]+)", r"Singapore\s*Dollar.*?([\d.,]+)"],
        "MYR_IDR": [r"MYR.*?IDR.*?([\d.,]+)", r"Malaysian\s*Ringgit.*?([\d.,]+)"],
    }

    full_text = " ".join(lines)
    for key, regex_list in patterns.items():
        for pattern in regex_list:
            m = re.search(pattern, full_text, re.IGNORECASE)
            if m:
                raw = m.group(1).replace(",", "").replace(".", "")
                # Handle decimals: if last 2 digits are decimal
                if len(raw) > 4:
                    val = float(raw[:-2]) if len(raw) % 2 == 0 else float(raw)
                    if val > 100:
                        rates[key] = val
                        break

    if rates.get("USD_IDR"):
        rates["Sumber"] = "bi.go.id (jina)"
        # Fill missing rates with None
        for k in ["USD_IDR", "EUR_IDR", "SGD_IDR", "MYR_IDR"]:
            if k not in rates:
                rates[k] = None
        return rates

    return None


def add_daily_row(ws, data):
    """Add today's row to worksheet if not already present."""
    today = datetime.now().strftime("%Y-%m-%d")

    # Check if today already exists
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == today:
            # Update existing row
            ws.cell(row=row, column=2, value=data["USD_IDR"])
            ws.cell(row=row, column=3, value=data["EUR_IDR"])
            ws.cell(row=row, column=4, value=data["SGD_IDR"])
            ws.cell(row=row, column=5, value=data["MYR_IDR"])
            ws.cell(row=row, column=6, value=data["Sumber"])
            return False  # Updated, not new

    # Append new row
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=today)
    ws.cell(row=next_row, column=2, value=data["USD_IDR"])
    ws.cell(row=next_row, column=3, value=data["EUR_IDR"])
    ws.cell(row=next_row, column=4, value=data["SGD_IDR"])
    ws.cell(row=next_row, column=5, value=data["MYR_IDR"])
    ws.cell(row=next_row, column=6, value=data["Sumber"])
    return True  # New row added


def main():
    today = datetime.now().strftime("%Y-%m-%d")
    print("=" * 50)
    print("💱 KURS VALUTA ASING HARIAN")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 50)

    # Try primary source
    data = fetch_exchange_rates_api()

    # Fallback to BI via Jina
    if not data:
        print("  🔄 Fallback ke Bank Indonesia via Jina...")
        data = fetch_from_bi_jina()

    if not data:
        print("\n❌ Gagal mengambil data kurs dari semua sumber")
        return

    # Save to Excel
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb, ws = create_or_load_workbook()
    is_new = add_daily_row(ws, data)
    wb.save(EXCEL_PATH)

    # Print results
    usd = data.get("USD_IDR")
    eur = data.get("EUR_IDR")
    sgd = data.get("SGD_IDR")
    myr = data.get("MYR_IDR")

    print(f"\n💰 Kurs USD/IDR: Rp {usd:,.0f}" if usd else "\n💰 Kurs USD/IDR: N/A")
    if eur:
        print(f"💶 Kurs EUR/IDR: Rp {eur:,.0f}")
    if sgd:
        print(f"🇸🇬 Kurs SGD/IDR: Rp {sgd:,.0f}")
    if myr:
        print(f"🇲🇾 Kurs MYR/IDR: Rp {myr:,.0f}")
    print(f"📊 Sumber: {data['Sumber']}")
    print(f"{'✅ Baris baru ditambahkan' if is_new else '🔄 Data hari ini di-update'}")
    print(f"💾 Tersimpan: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
