#!/usr/bin/env python3
"""
WordPress Article Generator - Data-Driven Version
Generates articles based on actual market data from Excel files.
Each article uses real prices, trends, and analysis.
"""
import json
import random
import os
import sys
from datetime import datetime
from pathlib import Path
import openpyxl

WP_DIR = Path(__file__).resolve().parent.parent / "data" / "wp_articles"
WP_DIR.mkdir(exist_ok=True)
DATA_DIR = Path(__file__).resolve().parent.parent / "data"

# ============ DATA READERS ============

def load_excel(filename, sheet=None):
    """Load Excel data, return list of dicts with header mapping."""
    path = DATA_DIR / filename
    if not path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(h) if h else f"col{i}" for i, h in enumerate(rows[0])]
        data = []
        for row in rows[1:]:
            if row[0]:
                d = dict(zip(headers, row))
                # Add lowercase keys for compatibility (separate dict)
                d.update({k.lower(): v for k, v in d.items()})
                data.append(d)
        return data
    except Exception as e:
        print(f"  ⚠️ Error loading {filename}: {e}")
        return []

def load_peternakan():
    """Load peternakan data grouped by kategori."""
    data = load_excel("harga_peternakan_lengkap.xlsx", "Data Utama")
    if not data:
        return {}
    grouped = {}
    for row in data:
        cat = row.get("Kategori", "LAIN")
        if cat not in grouped:
            grouped[cat] = []
        grouped[cat].append(row)
    return grouped

def load_pakan():
    """Load feed price data."""
    data = load_excel("harga_pakan_ternak.xlsx")
    return data

def load_sembako():
    """Load sembako data."""
    data = load_excel("harga_sembako.xlsx", "Harga")
    return data

# ============ TREND ANALYSIS ============

def trend_7h(data, field):
    """Compare last 7 days vs previous 7 days."""
    vals = [row.get(field) for row in data[-14:] if row.get(field) and row.get(field) != '-']
    nums = []
    for v in vals:
        try:
            nums.append(float(str(v).replace('.', '').replace(',', '')))
        except:
            pass
    if len(nums) < 4:
        return None, None, None
    mid = len(nums) // 2
    avg_old = sum(nums[:mid]) / mid if mid else 0
    avg_new = sum(nums[mid:]) / (len(nums) - mid) if (len(nums) - mid) else 0
    if avg_old == 0:
        return avg_new, 0, "stabil"
    pct = ((avg_new - avg_old) / avg_old) * 100
    if pct > 3:
        label = "naik"
    elif pct < -3:
        label = "turun"
    else:
        label = "stabil"
    return avg_new, pct, label

def price_str(val):
    """Format price to Rp string."""
    try:
        return f"Rp {int(float(str(val).replace('.', '').replace(',', ''))):,}"
    except:
        return "-"

# ============ ARTICLE GENERATORS ============

def gen_peternakan_price_report(peternakan_data):
    """Generate article about peternakan price movements."""
    if not peternakan_data:
        return None

    # Pick a random category
    cats = list(peternakan_data.keys())
    if not cats:
        return None
    cat = random.choice(cats)
    items = peternakan_data[cat][:10]  # max 10 items

    cat_label = cat.capitalize()
    rows_html = []
    for item in items:
        produk = item.get("Produk", item.get("produk", "-"))
        harga = price_str(item.get("Harga (Rp)", item.get("Harga", item.get("harga", "-"))))
        sumber = item.get("Sumber", item.get("sumber", "-"))
        rows_html.append(f"<tr><td>{produk}</td><td>{harga}</td><td>{sumber}</td></tr>")

    title = f"Update Harga {cat_label} Terkini: {datetime.now().strftime('%d %B %Y')}"
    body = f"""
<h4>Data Harga {cat_label}</h4>
<p>Berikut data harga terbaru dari kategori <strong>{cat_label}</strong> yang tercatat di sistem:</p>

<table style="width:100%; border-collapse:collapse; margin:16px 0;">
  <thead>
    <tr style="background:#f0f0f0;">
      <th style="border:1px solid #ddd; padding:8px; text-align:left;">Produk</th>
      <th style="border:1px solid #ddd; padding:8px; text-align:left;">Harga</th>
      <th style="border:1px solid #ddd; padding:8px; text-align:left;">Satuan</th>
      <th style="border:1px solid #ddd; padding:8px; text-align:left;">Sumber</th>
    </tr>
  </thead>
  <tbody>
    {"".join(rows_html)}
  </tbody>
</table>

<h4>Catatan</h4>
<p>Data diambil dari monitoring dashboard secara real-time. Harga bisa berbeda antar daerah karena faktor transportasi, ketersediaan, dan musim.</p>

<h4>Saran</h4>
<ul>
  <li>Pantau harga secara berkala sebelum melakukan transaksi besar</li>
  <li>Bandingkan harga beberapa sumber sebelum membeli</li>
  <li>Perhatikan tren naik/turun sebelum stok dalam jumlah besar</li>
</ul>

<p style="font-style:italic; color:#666;">Data diupdate otomatis setiap hari. Untuk data real-time, kunjungi dashboard di http://43.153.196.161:5000</p>
"""
    return {"title": title, "category": "Harga Peternakan", "body": body}

def gen_pakan_analysis(pakan_data):
    """Generate article about feed prices."""
    if not pakan_data:
        return None

    items = pakan_data[-8:]  # last 8 items
    rows_html = []
    for item in items:
        produk = item.get("Produk", item.get("Nama", "-"))
        harga = price_str(item.get("Harga", item.get("Harga/kg")))
        rows_html.append(f"<tr><td>{produk}</td><td>{harga}</td></tr>")

    title = f"Analisis Harga Bahan Pakan Ternak: {datetime.now().strftime('%d %B %Y')}"
    body = f"""
<h4>Data Harga Bahan Pakan</h4>
<p>Monitoring harga bahan pakan ternak terkini untuk membantu peternak dalam perencanaan biaya:</p>

<table style="width:100%; border-collapse:collapse; margin:16px 0;">
  <thead>
    <tr style="background:#f0f0f0;">
      <th style="border:1px solid #ddd; padding:8px; text-align:left;">Bahan Pakan</th>
      <th style="border:1px solid #ddd; padding:8px; text-align:left;">Harga</th>
    </tr>
  </thead>
  <tbody>
    {"".join(rows_html)}
  </tbody>
</table>

<h4>Tips Hemat Biaya Pakan</h4>
<ul>
  <li><strong>Beli dalam jumlah besar</strong> saat harga stabil untuk menghemat</li>
  <li><strong>Gunakan bahan lokal</strong> yang tersedia di sekitar peternakan</li>
  <li><strong>Fermentasi pakan</strong> untuk meningkatkan nilai nutrisi</li>
  <li><strong>Pantau tren harga</strong> sebelum melakukan stok besar</li>
</ul>

<h4>Rekomendasi</h4>
<p>Dengan fluktuasi harga bahan pakan, peternak disarankan untuk melakukan analisis cost-benefit secara berkala dan mempertimbangkan formulasi pakan sendiri jika harga pakan jadi terlalu tinggi.</p>

<p style="font-style:italic; color:#666;">Data diupdate otomatis. Dashboard: http://43.153.196.161:5000</p>
"""
    return {"title": title, "category": "Analisis Pakan", "body": body}

def gen_sembako_market_report(sembako_data):
    """Generate article about sembako market trends."""
    if not sembako_data or len(sembako_data) < 2:
        return None

    # Pick random items to report
    items_to_report = random.sample([
        ("beras_premium", "Beras Premium"),
        ("minyak_goreng", "Minyak Goreng"),
        ("gula_pasir", "Gula Pasir"),
        ("telur_ras", "Telur Ayam Ras"),
        ("ayam_ras", "Daging Ayam Ras"),
        ("cabai_merah", "Cabai Merah"),
        ("bawang_merah", "Bawang Merah"),
        ("daging_sapi", "Daging Sapi"),
    ], min(5, 8))

    rows_html = []
    for key, label in items_to_report:
        latest = sembako_data[-1].get(key, "-")
        rows_html.append(f"<tr><td>{label}</td><td>{price_str(latest)}</td></tr>")

    title = f"Update Harga Sembako Hari Ini: {datetime.now().strftime('%d %B %Y')}"
    body = f"""
<h4>Data Harga Sembako Terkini</h4>
<p>Harga sembako yang tercatat di sistem monitoring pada hari ini:</p>

<table style="width:100%; border-collapse:collapse; margin:16px 0;">
  <thead>
    <tr style="background:#f0f0f0;">
      <th style="border:1px solid #ddd; padding:8px; text-align:left;">Komoditas</th>
      <th style="border:1px solid #ddd; padding:8px; text-align:left;">Harga</th>
    </tr>
  </thead>
  <tbody>
    {"".join(rows_html)}
  </tbody>
</table>

<h4>Analisis Pasar</h4>
<p>Harga sembakofluktuatif dipengaruhi oleh pasokan, cuaca, dan musim. Peternak dan konsumen perlu memantau perkembangan harga untuk pengambilan keputusan.</p>

<h4>Saran Konsumsi</h4>
<ul>
  <li>Beli sesuai kebutuhan untuk menghindari pemborosan</li>
  <li>Manfaatkan diskon atau promo di toko modern</li>
  <li>Stok bahan tahan lama (beras, minyak, gula) saat harga stabil</li>
</ul>

<p style="font-style:italic; color:#666;">Data diupdate otomatis dari Siskaperbapo Jatim dan sumber lainnya. Dashboard: http://43.153.196.161:5000</p>
"""
    return {"title": title, "category": "Harga Sembako", "body": body}

def gen_peternakan_tips():
    """Generate a tips article (data-light, knowledge-based)."""
    tips_pool = [
        {
            "title": "Tips Mengelola Kandang Ayam Broiler untuk Hasil Maksimal",
            "category": "Tips Ternak",
            "body": """
<h4>1. Ventilasi dan Sirkulasi Udara</h4>
<p>Kandang yang baik memiliki sirkulasi udara yang cukup. Suhu ideal untuk broiler adalah 25-28°C. Di atas 30°C, ayam mulai stres dan FCR memburuk.</p>

<h4>2. Manajemen Air Minum</h4>
<ul>
  <li>Gunakan air bersih, ganti minimal 2x sehari</li>
  <li>Perhatikan konsumsi air sebagai indikator kesehatan</li>
  <li>Tambahkan vitamin saat cuaca panas</li>
</ul>

<h4>3. Program Vaksinasi</h4>
<p>Ikuti jadwal vaksin dari supplier DOC. Biasanya: ND (Hari 1, 7, 14) → IB → IBD → ND booster.</p>

<h4>4. Monitoring Berat Badan</h4>
<p>Timbang sampel mingguan. Target berat 2kg dalam 35 hari dengan FCR 1.6-1.7.</p>

<h4>5. Biosekuriti</h4>
<p>Ganti sepatu karet saat masuk kandang, semprot desinfektan di pintu masuk, batasi pengunjung.</p>

<p style="font-style:italic; color:#666;">Tips ini berdasarkan praktik umum di lapangan. Konsultasikan dengan dokter hewan untuk kondisi spesifik.</p>
"""
        },
        {
            "category": "Tips Ternak",
            "title": "Cara Menghitung Keuntungan Beternak Ayam Broiler",
            "body": """
<h4>Rumus Dasar</h4>
<p><strong>Keuntungan = Total Pendapatan - Total Biaya</strong></p>

<h4>Komponen Biaya</h4>
<ul>
  <li><strong>DOC</strong>: Rp 5.000-7.000/ekor (tergantung lokasi)</li>
  <li><strong>Pakan</strong>: ±60% dari total biaya</li>
  <li><strong>Obat & vaksin</strong>: Rp 1.500-2.500/ekor</li>
  <li><strong>Listrik & BBM</strong>: tergantung skala</li>
  <li><strong>Sewa kandang/jamah</strong>: Rp 2.000-4.000/ekor</li>
</ul>

<h4>Contoh Hitungan (1.000 ekor)</h4>
<table style="width:100%; border-collapse:collapse; margin:16px 0;">
  <tr style="background:#f0f0f0;"><th style="border:1px solid #ddd; padding:8px;">Item</th><th style="border:1px solid #ddd; padding:8px;">Biaya</th></tr>
  <tr><td style="border:1px solid #ddd; padding:8px;">DOC 1.000 ekor</td><td style="border:1px solid #ddd; padding:8px;">Rp 6.000.000</td></tr>
  <tr><td style="border:1px solid #ddd; padding:8px;">Pakan (33 kg/ekor × Rp 6.500)</td><td style="border:1px solid #ddd; padding:8px;">Rp 214.500.000</td></tr>
  <tr><td style="border:1px solid #ddd; padding:8px;">Obat & vaksin</td><td style="border:1px solid #ddd; padding:8px;">Rp 2.000.000</td></tr>
  <tr><td style="border:1px solid #ddd; padding:8px;">Lain-lain</td><td style="border:1px solid #ddd; padding:8px;">Rp 8.000.000</td></tr>
  <tr style="background:#f0f0f0;"><td style="border:1px solid #ddd; padding:8px;"><strong>Total Biaya</strong></td><td style="border:1px solid #ddd; padding:8px;"><strong>Rp 230.500.000</strong></td></tr>
</table>

<h4>Revenue</h4>
<p>Jual 1.000 ekor × 2kg × Rp 33.000 = Rp 66.000.000.000</p>
<p style="color:red;"><strong>Catatan: Harga jual bervariasi. Gunakan harga aktual saat menghitung.</strong></p>

<h4>Kunci Profitabilitas</h4>
<ul>
  <li>FCR rendah (pakan efisien)</li>
  <li>Mortalitas rendah (&lt;5%)</li>
  <li>Pembelian pakan tepat waktu (tidak kehabisan/terlalu banyak)</li>
  <li>Harga jual ≥ HPP + margin</li>
</ul>
"""
        },
        {
            "category": "Tips Ternak",
            "title": "Memahami Nutrisi Pakan Ayam: Protein, Energi, dan Serat",
            "body": """
<h4>Kenapa Nutrisi Penting</h4>
<p>Pakan adalah biaya terbesar (±60%) dalam beternak unggas. Memahami komposisi nutrisi membantu peternak memilih bahan pakan yang tepat dan mengoptimalkan formulasi.</p>

<h4>Tiga Komponen Utama</h4>
<ul>
  <li><strong>Protein (CP)</strong>: untuk pertumbuhan otot. Ayam broiler butuh 21-23% di fase starter, 19-21% di finisher.</li>
  <li><strong>Energi (TDN/ME)</strong>: untuk metabolisme. Jagung adalah sumber utama energi.</li>
  <li><strong>Serat (CF)</strong>: terbatas untuk unggas. Maksimal 5-7% dalam ransum.</li>
</ul>

<h4>Sumber Protein Umum</h4>
<ul>
  <li>Bungkil kedelai (44% CP) — paling umum</li>
  <li>Tepung ikan (60-65% CP) — mahal tapi berkualitas</li>
  <li>Bungkil kelapa (18-20% CP) — alternatif murah</li>
  <li>Ampas tahu (25-30% CP) — tersedia lokal</li>
</ul>

<h4>Red Flags</h4>
<ul>
  <li>Harga bungkil kedelai naik → pertimbangkan campuran bahan lain</li>
  <li>Serat terlalu tinggi → FCR memburuk</li>
  <li>Protein rendah → pertumbuhan melambat</li>
</ul>

<p style="font-style:italic; color:#666;">Untuk formulasi detail, gunakan tools optimasi pakan di dashboard: http://43.153.196.161:5000</p>
"""
        },
    ]
    return random.choice(tips_pool)

# ============ MAIN ============

def pick_generator():
    """Pick article generator based on data availability and variety."""
    # Weight: prefer data-driven (peternakan/pakan/sembako) over tips
    weighted = [
        (gen_peternakan_price_report, load_peternakan(), 3),
        (gen_pakan_analysis, load_pakan(), 2),
        (gen_sembako_market_report, load_sembako(), 2),
        (gen_peternakan_tips, None, 1),  # always available
    ]
    
    # Filter generators that have data
    available = [(fn, data, w) for fn, data, w in weighted if data is not None or fn == gen_peternakan_tips]
    if not available:
        available = [(gen_peternakan_tips, None, 1)]
    
    # Weighted random selection
    total_weight = sum(w for _, _, w in available)
    pick = random.uniform(0, total_weight)
    cumulative = 0
    for fn, data, w in available:
        cumulative += w
        if pick <= cumulative:
            return fn, data
    return available[-1][0], available[-1][1]

def main():
    now = datetime.now()
    ds = now.strftime("%d %B %Y")

    generator, data = pick_generator()
    if data is not None:
        article = generator(data)
    else:
        article = generator()
    
    if not article:
        # Fallback to tips
        article = gen_peternakan_tips()
        print("  ⚠️ Data tidak tersedia, menggunakan tips")

    print("=" * 60)
    print("WordPress Article Generator - Data-Driven")
    print("=" * 60)
    print(f"Title: {article['title']}")
    print(f"Category: {article['category']}")
    print(f"Date: {ds}")
    print("=" * 60)

    # Generate HTML
    html_content = f"""<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <title>{article['title']} - Catatan Insani</title>
    <style>
        body {{ font-family: -apple-system, sans-serif; max-width: 800px; margin: 0 auto; padding: 20px; background: #f5f5f5; }}
        .card {{ background: #fff; border-radius: 8px; padding: 24px; margin-bottom: 16px; box-shadow: 0 1px 3px rgba(0,0,0,.1); }}
        h1 {{ color: #1a1a2e; font-size: 20px; }}
        pre {{ white-space: pre-wrap; word-wrap: break-word; background: #f8f8f8; padding: 12px; border-radius: 6px; font-size: 13px; max-height: 400px; overflow-y: auto; border: 1px solid #eee; }}
        .copy-btn {{ background: #0073aa; color: #fff; border: none; padding: 10px 20px; border-radius: 6px; cursor: pointer; font-size: 14px; margin-top: 8px; }}
        .copy-btn:hover {{ background: #005a87; }}
        .label {{ font-weight: bold; color: #555; margin-bottom: 4px; }}
        .success {{ color: green; font-weight: bold; display: none; margin-left: 10px; }}
        .preview {{ border: 1px solid #ddd; border-radius: 8px; padding: 16px; background: #fff; }}
    </style>
</head>
<body>
    <h1>Artikel WordPress</h1>
    <div class="card">
        <div class="label">Title:</div>
        <h1 id="title">{article['title']}</h1>
        <button class="copy-btn" onclick="copyText('title')">Copy Title</button>
        <span class="success" id="title-copied">Copied!</span>
    </div>
    <div class="card">
        <div class="label">Category: <strong>{article['category']}</strong> | Date: <strong>{ds}</strong></div>
    </div>
    <div class="card">
        <div class="label">HTML Source (paste this into WordPress editor HTML mode):</div>
        <pre id="raw">{article['body']}</pre>
        <button class="copy-btn" onclick="copyHtml()">📋 Copy HTML Source</button>
        <span class="success" id="success-text">Copied!</span>
    </div>
    <div class="card preview">
        <div class="label">Preview:</div>
        <div style="max-height: 300px; overflow-y: auto; border: 1px solid #ddd; padding: 12px; background: #fafafa;">{article['body']}</div>
    </div>
    <div class="card">
        <a href="/article?new=1" class="copy-btn" style="background:#28a745;text-decoration:none;display:inline-block;">Generate Artikel Baru</a>
    </div>
    <script>
        function copyText(id) {{
            var el = document.getElementById(id);
            var text = el.innerText || el.textContent;
            var ta = document.createElement("textarea");
            ta.value = text;
            ta.style.position = "fixed";
            ta.style.opacity = "0";
            document.body.appendChild(ta);
            ta.select();
            document.execCommand("copy");
            document.body.removeChild(ta);
            document.getElementById(id + '-copied').style.display = 'inline';
            setTimeout(function(){{ document.getElementById(id + '-copied').style.display = 'none'; }}, 2000);
        }}
        function copyHtml() {{
            var raw = document.getElementById("raw").textContent;
            var ta = document.createElement("textarea");
            ta.value = raw;
            ta.style.position = "fixed";
            ta.style.opacity = "0";
            document.body.appendChild(ta);
            ta.select();
            document.execCommand("copy");
            document.body.removeChild(ta);
            document.getElementById('success-text').style.display = 'inline';
            setTimeout(function(){{ document.getElementById('success-text').style.display = 'none'; }}, 2000);
        }}
    </script>
</body>
</html>"""

    # Save files
    date_str = now.strftime('%Y-%m-%d')
    meta_file = WP_DIR / f"{date_str}.json"
    html_file = WP_DIR / f"{date_str}.html"
    
    with open(meta_file, 'w', encoding='utf-8') as f:
        json.dump({
            'title': article['title'],
            'category': article['category'],
            'body': article['body'],
            'date': ds,
            'status': 'generated',
            'generated_at': now.isoformat(),
        }, f, ensure_ascii=False, indent=2)
    
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"\nFiles created:")
    print(f"  Metadata: {meta_file}")
    print(f"  HTML: {html_file}")

if __name__ == "__main__":
    main()
