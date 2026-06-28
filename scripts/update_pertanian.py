#!/usr/bin/env python3
"""
Update harga jagung, kedelai, pakan ternak dari berita online via Jina Reader.
Uses Kompas search + direct article reading for commodity prices.
"""
import re
import os
import sys
import json
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from create_pertanian_excel import add_row
from cache_lib import jina as jina_read_cached

EXCEL_PATH = os.path.expanduser("~/sembako/data/harga_pertanian_ternak.xlsx")
HISTORY_PATH = os.path.expanduser("~/sembako/data/pertanian_history.json")


def jina_read(url):
    """Fetch URL via Jina Reader with 1h cache."""
    text = jina_read_cached(url)
    return text if text else ""


def search_kompas(query, limit=5):
    """Search Kompas.com for articles matching query."""
    url = f"https://search.kompas.com/search/?q={query}"
    text = jina_read(url)
    if not text:
        return []
    # Extract kompas article URLs (skip search/asset/otomotif)
    urls = re.findall(r'https://[^\s<>"]+\.kompas\.com/read/[^\s<>")]+', text)
    unique = list(dict.fromkeys(urls))[:limit]
    return unique


def search_detik(keyword, limit=5):
    """Search detik.com finance for articles matching keyword."""
    url = f"https://www.detik.com/search/searchnews?query={keyword}&result_type=latest"
    text = jina_read(url)
    if not text:
        return []
    urls = re.findall(r'https://finance\.detik\.com/berita-ekonomi-bisnis/d-\d+/[^\s<>")]+', text)
    unique = list(dict.fromkeys(urls))[:limit]
    return unique


def extract_prices_from_text(text):
    """Extract commodity prices from article text."""
    results = []
    text_lower = text.lower()
    
    # Patterns: "Rp X.XXX", "RpX.XXX", "Rp X.XXX/kg"
    for match in re.finditer(r'[Rr][Pp]\s*?(\d[\d.,]*)', text):
        raw = match.group(1).replace('.', '').replace(',', '')
        try:
            val = int(raw)
            if 1000 < val < 100000:
                results.append(val)
        except ValueError:
            pass
    
    # Context extraction - find price near commodity words
    commodity_prices = {}
    for commodity in ['jagung', 'kedelai', 'pakan', 'broiler', 'layer', 'bebek', 'bungkil', 'ampas']:
        if commodity in text_lower:
            # Find prices near this commodity mention
            idx = text_lower.find(commodity)
            nearby = text_lower[max(0, idx-100):idx+300]
            for match in re.finditer(r'[Rr][Pp]\s*?(\d[\d.,]*)', nearby):
                raw = match.group(1).replace('.', '').replace(',', '')
                try:
                    val = int(raw)
                    if 1000 < val < 100000:
                        if commodity not in commodity_prices:
                            commodity_prices[commodity] = []
                        commodity_prices[commodity].append(val)
                except ValueError:
                    pass
    
    return commodity_prices


def main():
    today = datetime.now().strftime("%Y-%m-%d")
    print(f"🔍 Scraping harga pertanian & ternak - {today}\n")

    # Search queries
    queries_kompas = [
        ("harga jagung pipil 2026", "jagung"),
        ("harga kedelai impor 2026", "kedelai"),
        ("harga pakan ternak ayam broiler 2026", "pakan"),
    ]
    queries_detik = [
        ("harga+jagung+pipil", "jagung"),
        ("harga+kedelai+impor", "kedelai"),
        ("harga+pakan+ternak", "pakan"),
    ]

    all_commodity_prices = {}
    sources = []

    # Filter articles that actually contain commodity keywords in their titles
    commodity_keywords = ['jagung', 'kedelai', 'pakan', 'ternak', 'ayam', 'broiler', 'layer', 'bebek']
    
    # Try Kompas first (better article quality)
    for query, category in queries_kompas:
        print(f"🔍 Kompas: {category}")
        urls = search_kompas(query, limit=3)
        for url in urls:
            text = jina_read(url)
            if not text or len(text) < 200:
                continue
            # Get title
            lines = [l.strip() for l in text.split('\n') if l.strip() and not l.startswith('!')]
            title = lines[0][:80] if lines else url[-50:]
            
            # Skip articles that don't actually mention the commodity
            if not any(k in title.lower() for k in commodity_keywords):
                continue
                
            print(f"  📰 {title}")
            sources.append(title)

            prices = extract_prices_from_text(text)
            for commodity, vals in prices.items():
                if commodity not in all_commodity_prices:
                    all_commodity_prices[commodity] = []
                all_commodity_prices[commodity].extend(vals)

    # Fallback to detik if Kompas empty
    if len(all_commodity_prices) < 3:
        print("\n⚠️ Kompas kurang, coba Detik...")
        for query, category in queries_detik:
            urls = search_detik(query, limit=3)
            for url in urls:
                text = jina_read(url)
                if not text or len(text) < 200:
                    continue
                lines = [l.strip() for l in text.split('\n') if l.strip() and not l.startswith('!')]
                title = lines[0][:80] if lines else url[-50:]
                
                # Skip articles that don't actually mention the commodity
                if not any(k in title.lower() for k in commodity_keywords):
                    continue
                    
                print(f"  📰 {title}")
                sources.append(title)

                prices = extract_prices_from_text(text)
                for commodity, vals in prices.items():
                    if commodity not in all_commodity_prices:
                        all_commodity_prices[commodity] = []
                    all_commodity_prices[commodity].extend(vals)

    # Map commodity names to Excel columns
    def avg_price(prices_list):
        if not prices_list:
            return None
        # Remove outliers (below 30% or above 300% of median)
        sorted_p = sorted(prices_list)
        median = sorted_p[len(sorted_p) // 2]
        filtered = [p for p in prices_list if median * 0.3 <= p <= median * 3]
        return round(sum(filtered) / len(filtered)) if filtered else round(sum(prices_list) / len(prices_list))

    final = {}
    final['jagung_pipil'] = avg_price(all_commodity_prices.get('jagung', []))
    final['jagung_pakan'] = avg_price([p for p in all_commodity_prices.get('jagung', []) if p > 6000])
    final['kedelai_impor'] = avg_price(all_commodity_prices.get('kedelai', []))
    final['kedelai_lokal'] = avg_price([p for p in all_commodity_prices.get('kedelai', []) if p < 11000])
    final['pakan_broiler'] = avg_price(all_commodity_prices.get('broiler', all_commodity_prices.get('pakan', [])))
    final['pakan_layer'] = None
    final['pakan_bebek'] = None
    final['bungkil_kedelai'] = None
    final['jagung_giling'] = None

    # Print results
    print(f"\n📊 Hasil Scraping:")
    has_data = False
    for key, val in final.items():
        if val:
            print(f"  {key}: Rp {val:,}")
            has_data = True

    if not has_data:
        # Use last known data as fallback
        print("⚠️ Tidak ada data baru, menggunakan data terakhir...")
        if os.path.exists(EXCEL_PATH):
            import openpyxl
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
            ws = wb.active
            if ws.max_row > 1:
                last_row = [ws.cell(ws.max_row, c).value for c in range(1, ws.max_column + 1)]
                # Re-add last row with today's date
                data = {
                    'jagung_pipil': last_row[1],
                    'jagung_pakan': last_row[2],
                    'kedelai_impor': last_row[3],
                    'kedelai_lokal': last_row[4],
                    'pakan_broiler': last_row[5],
                    'pakan_layer': last_row[6],
                    'pakan_bebek': last_row[7],
                    'bungkil_kedelai': last_row[8],
                    'jagung_giling': last_row[9],
                }
                print(f"✅ Data fallback disimpan")
                fb_dict = {
                    'jagung_pipil': data.get('jagung_pipil', 0),
                    'jagung_pakan': data.get('jagung_pakan', 0),
                    'kedelai_impor': data.get('kedelai_impor', 0),
                    'kedelai_lokal': data.get('kedelai_lokal', 0),
                    'pakan_broiler': data.get('pakan_broiler', 0),
                    'pakan_layer': data.get('pakan_layer', 0),
                    'pakan_bebek': data.get('pakan_bebek', 0),
                    'bungkil_kedelai': data.get('bungkil_kedelai', 0),
                    'jagung_giling': data.get('jagung_giling', 0),
                }
                add_row(today, fb_dict, sumber='fallback')
                save_history(data, ['fallback data terakhir'])
                return
        print("❌ Tidak ada data sama sekali!")
        return

    data_dict = {
        'jagung_pipil': final.get('jagung_pipil') or 0,
        'jagung_pakan': final.get('jagung_pakan') or 0,
        'kedelai_impor': final.get('kedelai_impor') or 0,
        'kedelai_lokal': final.get('kedelai_lokal') or 0,
        'pakan_broiler': final.get('pakan_broiler') or 0,
        'pakan_layer': final.get('pakan_layer') or 0,
        'pakan_bebek': final.get('pakan_bebek') or 0,
        'bungkil_kedelai': final.get('bungkil_kedelai') or 0,
        'jagung_giling': final.get('jagung_giling') or 0,
    }
    add_row(today, data_dict, sumber='web scraping')
    
    save_history(final, sources[:5])
    print(f"\n✅ Data tersimpan ke {EXCEL_PATH}")


def save_history(data, sources):
    history = []
    if os.path.exists(HISTORY_PATH):
        try:
            with open(HISTORY_PATH) as f:
                history = json.load(f)
        except:
            pass
    history.append({
        "date": datetime.now().isoformat(),
        "prices": data,
        "sources": sources
    })
    history = history[-100:]
    with open(HISTORY_PATH, "w") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


if __name__ == "__main__":
    main()
