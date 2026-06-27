import os
import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(os.path.expanduser('~/sembako/data/harga_sembako.xlsx'))
ws = wb['Harga']

# Data dari detik.com - Harga Sembako Jawa Timur 24 Juni 2026
# Kolom: Tanggal, Beras Premium, Beras Medium, Minyak Goreng, Gula Pasir, Garam,
#        Tepung Terigu, Cabai Merah, Cabai Rawit, Bawang Merah, Bawang Putih,
#        Minyak Tanah, Telur Ras, Telur Kampung, Ayam Ras, Ayam Kampung, Daging Sapi, Sumber

new_row = [
    datetime(2026, 6, 24, 9, 55),   # Tanggal (dari artikel detikJatim)
    15030,    # Beras Premium
    12937,    # Beras Medium
    20229,    # Minyak Goreng (curah - yang paling umum dirujuk)
    17230,    # Gula Pasir (kristal putih)
    1912,     # Garam (bata)
    3200,     # Tepung Terigu (dari data terakhir)
    36458,    # Cabai Merah (keriting)
    47035,    # Cabai Rawit (merah)
    41480,    # Bawang Merah
    35413,    # Bawang Putih
    2200,     # Minyak Tanah (dari data terakhir)
    25008,    # Telur Ras
    45151,    # Telur Kampung
    32020,    # Ayam Ras
    70069,    # Ayam Kampung
    124293,   # Daging Sapi
    'DetikJatim',
]

ws.append(new_row)

wb.save(os.path.expanduser('~/sembako/data/harga_sembako.xlsx'))
print("Excel updated successfully!")

# Verify the last row
wb2 = openpyxl.load_workbook(os.path.expanduser('~/sembako/data/harga_sembako.xlsx'))
ws2 = wb2['Harga']
print(f"Total rows (including header): {ws2.max_row}")
print("Last row:")
for cell in ws2[ws2.max_row]:
    print(f"  {cell.column_letter}: {cell.value}")
