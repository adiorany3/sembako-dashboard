import os
import openpyxl
wb = openpyxl.load_workbook(os.path.expanduser('~/sembako/data/harga_sembako.xlsx'))
print('Sheets:', wb.sheetnames)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'\n--- Sheet: {sheet} ---')
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        print(row)
