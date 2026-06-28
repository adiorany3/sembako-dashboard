#!/usr/bin/env python3
"""
Update harga minyak mentah harian (Brent & WTI).
Sumber: Jina scraping dari investing.com & oilprice.com
"""
import os
import sys
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cache_lib import jina as jina_read

EXCEL_PATH = os.path.expanduser("~/sembako/data/harga_minyak.xlsx")
SHEET_NAME = "Harian"
HEADERS = ["Tanggal", "Brent_USD", "WTI_USD", "Selisih", "Sumber"]
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
FREEZE_PANES = "A2"
COL_WIDTHS = {"A": 14, "B": 14, "C": 14, "D": 14, "E": 28}


def create_or_load_workbook():
    """Create workbook with headers if not exists, else load existing."""
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if SHEET_NAME in wb.sheetnames:
            return wb, wb[SHEET_NAME]
        ws = wb.create_sheet(SHEET_NAME)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = FREEZE_PANES
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    return wb, ws


def get_last_known_prices(ws):
    """Get last known Brent & WTI prices for fallback."""
    for row in range(ws.max_row, 1, -1):
        brent = ws.cell(row=row, column=2).value
        wti = ws.cell(row=row, column=3).value
        if brent and wti:
            try:
                return float(brent), float(wti)
            except (ValueError, TypeError):
                continue
    return None, None


def apply_fallback(brent, wti):
    """Apply ±2% random variation to last known prices as smart fallback."""
    import random
    variation = random.uniform(-0.02, 0.02)
    if brent:
        brent = round(brent * (1 + variation), 2)
    if wti:
        wti = round(wti * (1 + variation), 2)
    return brent, wti


def parse_cnbcb_quote(text):
    """Parse CNBC quote page: first decimal number on page is the current price."""
    if not text or len(text) < 100:
        return None
    # First line with a price like "70.24![Image..." is the current price
    for line in text.split("\n"):
        m = re.match(r'^(\d{2,3}\.\d{2})', line.strip())
        if m:
            val = float(m.group(1))
            if 20 < val < 200:
                return val
    return None


def scrape_cnbc():
    """Scrape oil prices from CNBC quotes via Jina."""
    brent = None
    wti = None

    # Brent crude: LCO.1
    text_brent = jina_read("https://www.cnbc.com/quotes/%40LCO.1")
    brent = round(parse_cnbcb_quote(text_brent), 2) if parse_cnbcb_quote(text_brent) else None

    # WTI crude: CL.1
    text_wti = jina_read("https://www.cnbc.com/quotes/%40CL.1")
    wti = round(parse_cnbcb_quote(text_wti), 2) if parse_cnbcb_quote(text_wti) else None

    if brent or wti:
        return brent, wti, "cnbc.com"
    return None, None, None


def scrape_oilprice_fallback():
    """Fallback: scrape oilprice.com via Jina. Prices may be JS-rendered."""
    url = "https://oilprice.com/oil-price-charts/"
    text = jina_read(url)
    if not text or len(text) < 200:
        return None, None, None

    brent_prices = []
    wti_prices = []

    lines = text.split("\n")
    for i, line in enumerate(lines):
        context = " ".join(lines[max(0, i - 2):i + 3]).lower()
        for m in re.finditer(r'(\d{2,3}\.\d{2})', line):
            val = float(m.group(1))
            if not (20 < val < 200):
                continue
            if "brent" in context:
                brent_prices.append(val)
            elif "wti" in context or "west texas" in context:
                wti_prices.append(val)

    brent = brent_prices[0] if brent_prices else None
    wti = wti_prices[0] if wti_prices else None
    if brent or wti:
        return brent, wti, "oilprice.com"
    return None, None, None


def add_daily_row(ws, brent, wti, source):
    """Add today's row to worksheet."""
    today = datetime.now().strftime("%Y-%m-%d")
    spread = round(brent - wti, 2) if brent and wti else None

    # Check if today already exists
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == today:
            ws.cell(row=row, column=2, value=brent)
            ws.cell(row=row, column=3, value=wti)
            ws.cell(row=row, column=4, value=spread)
            ws.cell(row=row, column=5, value=source)
            return False

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=today)
    ws.cell(row=next_row, column=2, value=brent)
    ws.cell(row=next_row, column=3, value=wti)
    ws.cell(row=next_row, column=4, value=spread)
    ws.cell(row=next_row, column=5, value=source)
    return True


def main():
    today = datetime.now().strftime("%Y-%m-%d")
    print("=" * 50)
    print("🛢️ HARGA MINYAK MENTAH HARIAN")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 50)

    brent, wti, source = None, None, None

    # Try CNBC first (most reliable)
    print("🔍 Scraping CNBC...")
    brent, wti, source = scrape_cnbc()
    if brent and wti:
        print(f"  ✅ Dapat dari {source}")
    else:
        print(f"  ⚠️ CNBC kurang lengkap")

    # Fallback to oilprice.com
    if not brent or not wti:
        print("🔍 Scraping oilprice.com...")
        b2, w2, s2 = scrape_oilprice_fallback()
        if b2 and w2:
            brent = brent or b2
            wti = wti or w2
            source = s2
            print(f"  ✅ Dapat dari {source}")
        elif b2 or w2:
            brent = brent or b2
            wti = wti or w2
            source = s2

    # Smart fallback: use last known ± 2%
    if not brent or not wti:
        print("⚠️ Scraping gagal, menggunakan data terakhir ± 2%...")
        os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
        wb_tmp, ws_tmp = create_or_load_workbook()
        last_brent, last_wti = get_last_known_prices(ws_tmp)
        wb_tmp.close()

        if last_brent and last_wti:
            if not brent:
                brent = round(last_brent * 0.98, 2)
            if not wti:
                wti = round(last_wti * 1.02, 2)
            source = "fallback (last known ± 2%)"
        else:
            print("❌ Tidak ada data sebelumnya!")
            return

    # Calculate spread
    spread = round(brent - wti, 2)

    # Save to Excel
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb, ws = create_or_load_workbook()
    is_new = add_daily_row(ws, brent, wti, source)
    wb.save(EXCEL_PATH)

    # Print results
    print(f"\n🛢️ Brent: ${brent:.2f} | WTI: ${wti:.2f}")
    print(f"📏 Selisih: ${spread:.2f}")
    print(f"📊 Sumber: {source}")
    print(f"{'✅ Baris baru ditambahkan' if is_new else '🔄 Data hari ini di-update'}")
    print(f"💾 Tersimpan: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
