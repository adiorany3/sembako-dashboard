#!/usr/bin/env python3
"""Update sembako Excel with prices from detikJatim article (23 Juni 2026)."""
import openpyxl
from openpyxl.styles import Alignment, Border, Side

EXCEL_PATH = "/root/sembako/harga_sembako.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb.active

ca = Alignment(horizontal="center", vertical="center")
thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

# Prices from Siskaperbapo Jatim via detik.com (23 Juni 2026)
prices = {
    "beras_premium": 15004, "beras_medium": 12911,
    "gula": 17172, "minyak_curah": 20413, "minyak_kemasan": 21493,
    "telur_ras": 25032, "telur_kampung": 46201,
    "ayam_ras": 32081, "ayam_kampung": 69053,
    "sapi": 124778,
    "cabai_keriting": 35538, "cabai_rawit": 47354,
    "bawang_merah": 41274, "bawang_putih": 34946,
    "garam": 9184, "elpiji": 19909,
}

keys_order = [
    "beras_premium", "beras_medium", "gula", "minyak_curah", "minyak_kemasan",
    "telur_ras", "telur_kampung", "ayam_ras", "ayam_kampung", "sapi",
    "cabai_keriting", "cabai_rawit", "bawang_merah", "bawang_putih", "garam", "elpiji"
]

# Update existing 2026-06-23 row
target_row = None
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and "2026-06-23" in str(val):
        target_row = row
        break

if target_row:
    for i, key in enumerate(keys_order, 2):
        ws.cell(row=target_row, column=i, value=prices[key])
    ws.cell(row=target_row, column=18, value="Siskaperbapo Jatim via detik.com")
    print(f"Updated row {target_row} for 2026-06-23")

# Add 2026-06-24 row
has_24 = False
for row in range(2, ws.max_row + 1):
    val = ws.cell(row=row, column=1).value
    if val and "2026-06-24" in str(val):
        has_24 = True
        break

if not has_24:
    row = ws.max_row + 1
    c = ws.cell(row=row, column=1, value="2026-06-24")
    c.alignment = ca; c.border = thin
    for i, key in enumerate(keys_order, 2):
        c = ws.cell(row=row, column=i, value=prices[key])
        c.alignment = ca; c.border = thin
        if prices[key]:
            c.number_format = "#,##0"
    c = ws.cell(row=row, column=18, value="Siskaperbapo Jatim via detik.com")
    c.alignment = ca; c.border = thin
    print(f"Added row {row} for 2026-06-24")

wb.save(EXCEL_PATH)
print("Excel saved successfully!")
