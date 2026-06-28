#!/usr/bin/env python3
"""
Analisa sentimen berita Indonesia via Jina Reader.
Direct scrape dari homepage news tanpa Google Search.
"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cache_lib import jina as jina_read_cached

import urllib.request
import re
import json
from datetime import datetime
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer

EXCEL_PATH = os.path.expanduser("~/sembako/data/sentimen_berita.xlsx")
HISTORY_PATH = os.path.expanduser("~/sembako/data/sentimen_history.json")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Accept": "text/plain",
}

POSITIVE_ID = [
    "positif", "optimis", "berhasil", "sukses", "naik", "meningkat", "tumbuh",
    "untung", "profit", "bagus", "baik", "stabil", "aman", "sejahtera",
    "kemajuan", "berkah", "puas", "senang", "bangga", "pulih", "rebound",
    "bullish", "melonjak", "melambung", "cemerlang", "modal", "perkuat",
    "sentuh", "ciptakan", "bangkit", "kejutan"
]

NEGATIVE_ID = [
    "negatif", "gagal", "turun", "menurun", "anjlok", "merosot", "rugi",
    "krisis", "resesi", "inflasi", "korupsi", "bencana", "gempa", "banjir",
    "kecelakaan", "konflik", "perang", "demo", "protes", "ancaman",
    "bearish", "darurat", "wabah", "phk", "kemiskinan",
    "kelangkaan", "keluhan", "masalah", "buruk", "pelemahan", "melambat",
    "miskin", "ilegal", "ditutup", "dampak", "eskalasi"
]

OUTDATED_PATTERNS = [
    r'\bramadhan\b', r'\bramadan\b', r'\blebaran\b', r'\bmudik\b',
    r'\bidul\s*fitri\b', r'\bpuasa\b', r'\bnatal\b', r'\btahun\s*baru\b',
    r'\bimlek\b', r'\bkenaikan\s*gaji\b', r'\bthr\b', r'\bhari\s*raya\b',
    r'\bmenjelang\s*ramadhan\b', r'\bjelang\s*ramadhan\b',
    r'\bjelang\s*ramadan\b', r'\bjelang\s*lebaran\b'
]


def jina_read(url, timeout=15):
    """Fetch URL via Jina Reader with 1h cache (from cache_lib)."""
    text = jina_read_cached(url)
    return text if text else ""


def extract_headlines(text):
    """Extract article headlines and URLs from finance homepage markdown."""
    results = []
    lines = text.split("\n")
    current_title = ""
    current_url = ""

    noise_keywords = ["rekomendasi", "untuk anda", "diskon", "promo", "iklan", "terjual", "belanja"]

    for line in lines:
        line = line.strip()
        # URL lines
        url_match = re.search(r'https://[^\s<>)\"\']+', line)
        if url_match:
            url = url_match.group()
            if 'detik.com' in url or 'kompas.com' in url or 'cnbcindonesia' in url:
                if current_title and len(current_title) > 15:
                    # Check for noise
                    if any(k in current_title.lower() for k in noise_keywords):
                        current_title = ""
                        continue
                    results.append((current_title.strip(), url))
                current_url = url
                current_title = ""
        # Title lines (markdown headers or bold text)
        if line.startswith("#") or line.startswith("**"):
            title = re.sub(r'^#+\s*', '', line).replace("**", "").strip()
            if len(title) > 15:
                current_title = title

    return results[:20]


def fetch_article_headlines(limit_per_source=5):
    """Get headlines from multiple news homepages."""
    sources = [
        ("https://finance.detik.com/", "detik finance"),
        ("https://www.kompas.com/", "kompas"),
        ("https://www.cnbcindonesia.com/news/", "cnbc"),
    ]

    all_headlines = []
    seen = set()

    for url, name in sources:
        print(f"  🌐 Fetching {name}...")
        text = jina_read(url, timeout=15)
        if not text or len(text) < 200:
            print(f"    ⚠️ Gagal fetch {name}")
            continue

        headlines = extract_headlines(text)
        print(f"    Ditemukan {len(headlines)} headline")
        for title, article_url in headlines:
            if title not in seen and len(title) > 20:
                seen.add(title)
                all_headlines.append((title, article_url, name))

        if len(all_headlines) >= limit_per_source * len(sources):
            break

    return all_headlines[:limit_per_source * 2]


def is_outdated(text):
    return any(re.search(p, text.lower()) for p in OUTDATED_PATTERNS)


def analyze_sentiment_id(text):
    text_lower = text.lower()
    pos_count = sum(1 for w in POSITIVE_ID if w in text_lower)
    neg_count = sum(1 for w in NEGATIVE_ID if w in text_lower)
    analyzer = SentimentIntensityAnalyzer()
    vader = analyzer.polarity_scores(text)

    if pos_count > neg_count:
        label = "POSITIF"
        score = min(0.5 + (pos_count - neg_count) * 0.15, 1.0)
    elif neg_count > pos_count:
        label = "NEGATIF"
        score = max(-0.5 - (neg_count - pos_count) * 0.15, -1.0)
    else:
        compound = vader['compound']
        if compound > 0.05:
            label = "POSITIF"
            score = compound
        elif compound < -0.05:
            label = "NEGATIF"
            score = compound
        else:
            label = "NETRAL"
            score = compound

    return {"label": label, "score": round(score, 3)}


def analyze_news():
    """Fetch headlines and analyze sentiment."""
    headlines = fetch_article_headlines(limit_per_source=8)
    if not headlines:
        print("⚠️ Tidak ada headline ditemukan")
        return []

    results = []
    for title, url, source in headlines:
        if is_outdated(title):
            print(f"  SKIP (outdated): {title[:50]}...")
            continue

        sentiment = analyze_sentiment_id(title)
        emoji = {"POSITIF": "🟢", "NEGATIF": "🔴", "NETRAL": "⚪"}[sentiment["label"]]
        print(f"  {emoji} [{sentiment['label']:>7}] {title[:60]}")

        results.append({
            "keyword": "berita ekonomi",
            "headline": title,
            "sentiment": sentiment["label"],
            "score": sentiment["score"],
            "source": source,
            "timestamp": datetime.now().isoformat()
        })

    return results


def save_to_history(results):
    if not results:
        return
    history = []
    if os.path.exists(HISTORY_PATH):
        with open(HISTORY_PATH) as f:
            history = json.load(f)
    history.extend(results)
    history = history[-500:]
    with open(HISTORY_PATH, "w") as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def save_to_excel(results):
    if not results:
        return
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, Reference

    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
    la = Alignment(horizontal="left", vertical="center", wrap_text=True)

    pos_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    neg_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    net_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Detail"]
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detail"
        headers = ["Tanggal", "Waktu", "Keyword", "Headline", "Sentimen", "Skor", "Sumber"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
        ws.row_dimensions[1].height = 30
        widths = [13, 10, 20, 55, 12, 10, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    row = ws.max_row + 1
    for r in results:
        dt = datetime.fromisoformat(r["timestamp"])
        fill = {"POSITIF": pos_fill, "NEGATIF": neg_fill, "NETRAL": net_fill}[r["sentiment"]]
        data = [dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M"), r["keyword"],
                r["headline"], r["sentiment"], r["score"], r["source"]]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin; c.fill = fill
            c.alignment = la if col == 4 else ca
        row += 1

    # Summary sheet
    if "Ringkasan" in wb.sheetnames:
        del wb["Ringkasan"]
    ws_sum = wb.create_sheet("Ringkasan")

    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    pos = sum(1 for r in all_rows if r[4] == "POSITIF")
    neg = sum(1 for r in all_rows if r[4] == "NEGATIF")
    net = sum(1 for r in all_rows if r[4] == "NETRAL")
    total = pos + neg + net

    for col, h in enumerate(["Metrik", "Jumlah", "Persentase"], 1):
        c = ws_sum.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin

    scores = [r[5] for r in all_rows if r[5] is not None]
    avg_score = sum(scores) / len(scores) if scores else 0

    sum_data = [
        ("Total Berita", total, "100%"),
        ("🟢 Positif", pos, f"{pos/total*100:.1f}%" if total else "0%"),
        ("🔴 Negatif", neg, f"{neg/total*100:.1f}%" if total else "0%"),
        ("⚪ Netral", net, f"{net/total*100:.1f}%" if total else "0%"),
        ("Rata-rata Skor", round(avg_score, 3), ""),
    ]
    for i, rd in enumerate(sum_data, 2):
        for col, val in enumerate(rd, 1):
            c = ws_sum.cell(row=i, column=col, value=val)
            c.border = thin; c.alignment = ca

    for i in range(1, 4):
        ws_sum.column_dimensions[get_column_letter(i)].width = 18

    if total > 0:
        pie = PieChart()
        pie.title = "Distribusi Sentimen"
        pie.style = 10
        pie.add_data(Reference(ws_sum, min_col=2, min_row=1, max_row=4), titles_from_data=True)
        pie.set_categories(Reference(ws_sum, min_col=1, min_row=2, max_row=4))
        pie.width = 16; pie.height = 12
        ws_sum.add_chart(pie, "A9")

    wb.save(EXCEL_PATH)
    print(f"\n✅ Disimpan ke {EXCEL_PATH}")


def main():
    print("=" * 50)
    print("📰 ANALISA SENTIMEN BERITA INDONESIA")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M WIB')}")
    print("=" * 50)

    results = analyze_news()

    if results:
        save_to_history(results)
        save_to_excel(results)

        pos = sum(1 for r in results if r["sentiment"] == "POSITIF")
        neg = sum(1 for r in results if r["sentiment"] == "NEGATIF")
        net = sum(1 for r in results if r["sentiment"] == "NETRAL")
        total = len(results)
        avg = sum(r["score"] for r in results) / total if total else 0

        msg = f"📊 *Analisa Sentimen Hari Ini*\n"
        msg += f"📅 {datetime.now().strftime('%d %B %Y')}\n\n"
        msg += f"📰 Total berita: *{total}*\n"
        msg += f"🟢 Positif: *{pos}* ({pos/total*100:.0f}%)\n"
        msg += f"🔴 Negatif: *{neg}* ({neg/total*100:.0f}%)\n"
        msg += f"⚪ Netral: *{net}* ({net/total*100:.0f}%)\n"
        msg += f"📈 Skor rata-rata: *{avg:.2f}*\n\n"

        if neg > pos:
            msg += "⚠️ Sentimen cenderung *NEGATIF* hari ini"
        elif pos > neg:
            msg += "✅ Sentimen cenderung *POSITIF* hari ini"
        else:
            msg += "⚖️ Sentimen *SEIMBANG* hari ini"

        print(f"\n{msg}")
    else:
        print("⚠️ Tidak ada berita ditemukan - tidak ada perubahan pada data")


if __name__ == "__main__":
    main()
