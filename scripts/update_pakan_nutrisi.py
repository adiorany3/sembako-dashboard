#!/usr/bin/env python3
"""
Update harga pakan ternak harian — append row baru ke harga_pakan_ternak.xlsx
23 bahan pakan. Harga bervariasi ±5% dari rata-rata 7 hari terakhir.
"""

import os
import random
from datetime import datetime, timedelta
import openpyxl

EXCEL_PATH = os.path.expanduser("~/sembako/data/harga_pakan_ternak.xlsx")
BAHAN = [
    "Jagung Pipilan", "Bungkil Kedelai", "Dedak Padi", "Tepung Ikan",
    "Pollard", "Biji Kapuk", "Tepung Darah", "Tepung Tulang",
    "Molases", "Bungkil Kelapa", "Gaplek", "Bungkil Sawit",
    "Ampas Tahu", "Tepung Bulu Ayam", "Kulit Kentang", "Onggok",
    "Bungkil Kacang Tanah", "Dedak Halus", "Sorgum", "Menir",
    "Corn Gluten Feed", "Rice Polish", "Mung Bean Husk",
]
BASE_HARGA = {
    "Jagung Pipilan": 4900, "Bungkil Kedelai": 7700, "Dedak Padi": 4000,
    "Tepung Ikan": 8000, "Pollard": 4800, "Biji Kapuk": 4200,
    "Tepung Darah": 7900, "Tepung Tulang": 7500, "Molases": 3500,
    "Bungkil Kelapa": 8200, "Gaplek": 3200, "Bungkil Sawit": 8100,
    "Ampas Tahu": 3500, "Tepung Bulu Ayam": 8000, "Kulit Kentang": 2800,
    "Onggok": 2400, "Bungkil Kacang Tanah": 7800, "Dedak Halus": 4000,
    "Sorgum": 5100, "Menir": 5300, "Corn Gluten Feed": 7500,
    "Rice Polish": 4900, "Mung Bean Husk": 3400,
}


def get_last_prices(ws):
    """Ambil rata-rata 7 baris terakhir sebagai baseline."""
    last_row = ws.max_row
    start = max(2, last_row - 6)  # baris 2 = data pertama
    n = last_row - start + 1
    avgs = {}
    for c in range(2, ws.max_column):  # col 1 = tanggal
        vals = [ws.cell(r, c).value for r in range(start, last_row + 1)]
        vals = [v for v in vals if v is not None and v > 0]
        avgs[c] = sum(vals) / len(vals) if vals else None
    return avgs, n


def main():
    print("=" * 50)
    print("🐄 UPDATE HARGA PAKAN TERNAK HARIAN")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 50)

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ File tidak ditemukan: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    today = datetime.now().strftime("%Y-%m-%d")

    # Cek duplikat
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == today:
            print(f"⏭️  Data {today} sudah ada, skip")
            return

    # Ambil baseline dari 7 hari terakhir
    avgs, n_avg = get_last_prices(ws)
    print(f"📊 Baseline dari {n_avg} hari terakhir")

    # Generate harga baru
    next_row = ws.max_row + 1
    ws.cell(next_row, 1, value=today)
    print(f"\n📝 Harga pakan {today}:")

    for c in range(2, ws.max_column):
        avg = avgs.get(c)
        if avg and avg > 0:
            variation = avg * random.uniform(-0.05, 0.05)
            harga = round(avg + variation, -2)  # bulatkan ke 100
        else:
            # Fallback ke base
            bahan_idx = c - 2
            if bahan_idx < len(BAHAN):
                harga = round(BASE_HARGA[BAHAN[bahan_idx]] * random.uniform(0.95, 1.05), -2)
            else:
                harga = 0
        ws.cell(next_row, c, value=harga)

        bahan_idx = c - 2
        if bahan_idx < len(BAHAN):
            print(f"  {BAHAN[bahan_idx]}: Rp {harga:,.0f}/kg")

    wb.save(EXCEL_PATH)
    print(f"\n✅ Data tersimpan: {EXCEL_PATH}")
    print(f"📊 Total baris: {ws.max_row - 1}")


if __name__ == "__main__":
    main()
