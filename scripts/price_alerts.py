#!/usr/bin/env python3
"""
Price alert system for sembako dashboard.
Reads Excel data files, detects significant price changes,
outputs formatted Telegram alert text. Silent if no alerts.
"""
import os
import json
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed", file=sys.stderr)
    sys.exit(1)

SEMBAKO_DIR = os.path.expanduser("~/sembako")
DATA_DIR = os.path.join(SEMBAKO_DIR, "data")


def safe_load_workbook(filepath):
    """Load workbook, return None if missing."""
    if not os.path.exists(filepath):
        return None
    try:
        return openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return None


def get_sheet_rows(filepath, sheet_name):
    """Get all data rows (skip header) from a sheet. Returns list of tuples."""
    wb = safe_load_workbook(filepath)
    if wb is None:
        return []
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return []
    return rows  # includes header at index 0


def get_last_n_rows(filepath, sheet_name, n=2):
    """Get last n data rows (excluding header)."""
    all_rows = get_sheet_rows(filepath, sheet_name)
    if not all_rows:
        return []
    data_rows = all_rows[1:]  # skip header
    return data_rows[-n:] if len(data_rows) >= n else data_rows


def detect_bbm_alerts():
    """Check BBM price changes from bbm_history.json."""
    alerts = []
    history_path = os.path.join(DATA_DIR, "bbm_history.json")
    if not os.path.exists(history_path):
        return alerts

    try:
        with open(history_path) as f:
            history = json.load(f)
    except (json.JSONDecodeError, IOError):
        return alerts

    if len(history) < 2:
        return alerts

    latest = history[-1]
    prev = history[-2]

    # Check if fuel types changed or new entry appeared
    latest_types = set(latest.get("fuel_types", []))
    prev_types = set(prev.get("fuel_types", []))
    latest_status = latest.get("status", "").strip().upper()
    prev_status = prev.get("status", "").strip().upper()
    latest_ts = latest.get("timestamp", "")
    prev_ts = prev.get("timestamp", "")

    # Only alert on genuinely new data (different timestamp)
    if latest_ts == prev_ts:
        return alerts

    # BBM: any new entry with NAIK status = alert
    if "NAIK" in latest_status or "⬆" in latest_status:
        new_fuels = latest_types - prev_types
        if new_fuels:
            for fuel in sorted(new_fuels):
                alerts.append(f"⛽ BBM: {fuel} berpotensi naik")
        elif latest_types:
            # Same fuels but new news = possible alert
            headline = latest.get("headline", "")
            if headline:
                alerts.append(f"⛽ BBM: {headline[:80]}")
    elif "TURUN" in latest_status or "⬇" in latest_status:
        new_fuels = latest_types - prev_types
        if new_fuels:
            for fuel in sorted(new_fuels):
                alerts.append(f"⛽ BBM: {fuel} berpotensi turun")
        elif latest_types:
            headline = latest.get("headline", "")
            if headline:
                alerts.append(f"⛽ BBM: {headline[:80]}")

    return alerts


def detect_crypto_alerts():
    """Check crypto price changes. Drops >5% or gains >10% in 24h."""
    alerts = []
    filepath = os.path.join(DATA_DIR, "crypto_monitor.xlsx")

    rows = get_sheet_rows(filepath, "Harga")
    if not rows:
        return alerts

    header = rows[0]
    # Map column names to indices
    col_map = {}
    for i, h in enumerate(header):
        if h:
            col_map[str(h).strip()] = i

    data = rows[1:]
    if not data:
        return alerts

    latest = data[-1]
    prev = data[-2] if len(data) >= 2 else None

    # Crypto coins to monitor: (name, usd_col, idr_col, change_col)
    coins = [
        ("BTC", "BTC (USD)", "BTC (IDR)", "BTC 24h %"),
        ("ETH", "ETH (USD)", "ETH (IDR)", "ETH 24h %"),
        ("SOL", "SOL (USD)", "SOL (IDR)", "SOL 24h %"),
    ]

    for name, usd_col, idr_col, chg_col in coins:
        if chg_col not in col_map:
            continue
        chg_idx = col_map[chg_col]
        usd_idx = col_map.get(usd_col)
        idr_idx = col_map.get(idr_col)

        try:
            change = float(latest[chg_idx])
        except (TypeError, IndexError, ValueError):
            continue

        usd_val = None
        idr_val = None
        if usd_idx is not None:
            try:
                usd_val = float(latest[usd_idx])
            except (TypeError, IndexError, ValueError):
                pass
        if idr_idx is not None:
            try:
                idr_val = float(latest[idr_idx])
            except (TypeError, IndexError, ValueError):
                pass

        # Previous values for comparison
        prev_usd = None
        if prev and usd_idx is not None:
            try:
                prev_usd = float(prev[usd_idx])
            except (TypeError, IndexError, ValueError):
                pass

        # If we have prev row, calculate actual change from price comparison
        actual_change = change
        if prev_usd and usd_val:
            actual_change = ((usd_val - prev_usd) / prev_usd) * 100

        # Drop > 5%
        if actual_change < -5:
            arrow = "📉"
            usd_str = f"${usd_val:,.0f}" if usd_val else "?"
            prev_usd_str = f"${prev_usd:,.0f}" if prev_usd else "?"
            alerts.append(
                f"{arrow} {name}: turun {abs(actual_change):.1f}% "
                f"({prev_usd_str} → {usd_str})"
            )
        # Gain > 10%
        elif actual_change > 10:
            arrow = "🚀"
            usd_str = f"${usd_val:,.0f}" if usd_val else "?"
            prev_usd_str = f"${prev_usd:,.0f}" if prev_usd else "?"
            alerts.append(
                f"{arrow} {name}: naik {actual_change:.1f}% "
                f"({prev_usd_str} → {usd_str})"
            )

    return alerts


def detect_gold_alerts():
    """Check gold price: all-time high or drop >3%."""
    alerts = []
    filepath = os.path.join(DATA_DIR, "harga_emas.xlsx")

    rows = get_sheet_rows(filepath, "Harian")
    if not rows:
        return alerts

    header = rows[0]
    data = rows[1:]
    if not data:
        return alerts

    # Find Antam 1gr Beli column
    antam_idx = None
    for i, h in enumerate(header):
        if h and "antam" in str(h).lower() and "beli" in str(h).lower():
            antam_idx = i
            break

    if antam_idx is None:
        return alerts

    latest_price = None
    try:
        latest_price = float(data[-1][antam_idx])
    except (TypeError, IndexError, ValueError):
        return alerts

    if latest_price <= 0:
        return alerts

    # Calculate all-time high from all data
    all_prices = []
    for row in data:
        try:
            p = float(row[antam_idx])
            if p > 0:
                all_prices.append(p)
        except (TypeError, ValueError):
            continue

    if not all_prices:
        return alerts

    # Exclude current price from historical for high check
    historical = all_prices[:-1]
    previous_high = max(historical) if historical else 0

    # All-time high
    if previous_high > 0 and latest_price > previous_high:
        alerts.append(
            f"🥇 Emas: sentuh all-time high Rp {latest_price:,.0f}"
        )

    # Check drop > 3% from previous row
    if len(data) >= 2:
        try:
            prev_price = float(data[-2][antam_idx])
            if prev_price > 0:
                pct = ((latest_price - prev_price) / prev_price) * 100
                if pct < -3:
                    alerts.append(
                        f"🥇 Emas: turun {abs(pct):.1f}% "
                        f"(Rp {prev_price:,.0f} → Rp {latest_price:,.0f})"
                    )
                elif pct > 3:
                    alerts.append(
                        f"🥇 Emas: naik {pct:.1f}% "
                        f"(Rp {prev_price:,.0f} → Rp {latest_price:,.0f})"
                    )
        except (TypeError, ValueError):
            pass

    return alerts


def detect_sembako_alerts():
    """Check sembako items with price changes >10% between last 2 rows."""
    alerts = []

    # Check both sembako files
    for filepath, sheet_name in [
        (os.path.join(DATA_DIR, "harga_sembako.xlsx"), "Harga"),
        (os.path.join(DATA_DIR, "harga_sembako_desktop.xlsx"), "Harga Sembako"),
    ]:
        rows = get_sheet_rows(filepath, sheet_name)
        if not rows or len(rows) < 3:
            continue

        header = rows[0]
        data = rows[1:]
        if len(data) < 2:
            continue

        prev_row = data[-2]
        curr_row = data[-1]

        for i, col_name in enumerate(header):
            if col_name is None:
                continue
            col_str = str(col_name).strip()
            if col_str.lower() in ("tanggal", "sumber"):
                continue

            try:
                prev_val = float(prev_row[i])
                curr_val = float(curr_row[i])
            except (TypeError, IndexError, ValueError):
                continue

            if prev_val <= 0 or curr_val <= 0:
                continue

            pct = ((curr_val - prev_val) / prev_val) * 100

            if abs(pct) > 10:
                # Clean column name (remove newlines)
                item_name = col_str.split("\n")[0].strip()
                if pct > 0:
                    alerts.append(
                        f"🍚 {item_name}: naik {pct:.1f}% "
                        f"(Rp {prev_val:,.0f} → Rp {curr_val:,.0f})"
                    )
                else:
                    alerts.append(
                        f"🍚 {item_name}: turun {abs(pct):.1f}% "
                        f"(Rp {prev_val:,.0f} → Rp {curr_val:,.0f})"
                    )

        # Only need one working sembako file
        if alerts:
            break

    return alerts


def detect_usdidr_alerts():
    """Check USD/IDR moves >2%."""
    alerts = []
    filepath = os.path.join(DATA_DIR, "kurs_valuta.xlsx")

    rows = get_sheet_rows(filepath, "Harian")
    if not rows:
        return alerts

    header = rows[0]
    data = rows[1:]
    if len(data) < 2:
        return alerts

    # Find USD_IDR column
    usd_idx = None
    for i, h in enumerate(header):
        if h and "usd" in str(h).lower():
            usd_idx = i
            break

    if usd_idx is None:
        return alerts

    try:
        prev_val = float(data[-2][usd_idx])
        curr_val = float(data[-1][usd_idx])
    except (TypeError, IndexError, ValueError):
        return alerts

    if prev_val <= 0:
        return alerts

    pct = ((curr_val - prev_val) / prev_val) * 100

    if abs(pct) > 2:
        if pct > 0:
            alerts.append(
                f"💵 USD/IDR: naik {pct:.2f}% "
                f"Rp {prev_val:,.0f} → Rp {curr_val:,.0f}"
            )
        else:
            alerts.append(
                f"💵 USD/IDR: turun {abs(pct):.2f}% "
                f"Rp {prev_val:,.0f} → Rp {curr_val:,.0f}"
            )

    return alerts


def detect_cpo_alerts():
    """Check CPO price changes >5%."""
    # Look for CPO data in pertanian_ternak or dedicated file
    alerts = []

    # Check if there's a dedicated CPO file
    cpo_files = [
        (os.path.join(DATA_DIR, "harga_cpo.xlsx"), "Harga"),
        (os.path.join(DATA_DIR, "harga_commodity.xlsx"), "Harga"),
        (os.path.join(SEMBAKO_DIR, "harga_cpo.xlsx"), "Harga"),
    ]

    for filepath, sheet_name in cpo_files:
        rows = get_sheet_rows(filepath, sheet_name)
        if not rows or len(rows) < 3:
            continue

        header = rows[0]
        data = rows[1:]
        if len(data) < 2:
            continue

        # Find CPO column
        cpo_idx = None
        for i, h in enumerate(header):
            if h and "cpo" in str(h).lower():
                cpo_idx = i
                break

        if cpo_idx is None:
            continue

        try:
            prev_val = float(data[-2][cpo_idx])
            curr_val = float(data[-1][cpo_idx])
        except (TypeError, IndexError, ValueError):
            continue

        if prev_val <= 0:
            continue

        pct = ((curr_val - prev_val) / prev_val) * 100

        if abs(pct) > 5:
            if pct > 0:
                alerts.append(
                    f"🛢️ CPO: naik {pct:.1f}% "
                    f"(Rp {prev_val:,.0f} → Rp {curr_val:,.0f})"
                )
            else:
                alerts.append(
                    f"🛢️ CPO: turun {abs(pct):.1f}% "
                    f"(Rp {prev_val:,.0f} → Rp {curr_val:,.0f})"
                )
            break

    return alerts


def main():
    """Main entry: collect all alerts, output formatted text if any."""
    all_alerts = []

    # Collect alerts from each category
    all_alerts.extend(detect_bbm_alerts())
    all_alerts.extend(detect_crypto_alerts())
    all_alerts.extend(detect_gold_alerts())
    all_alerts.extend(detect_sembako_alerts())
    all_alerts.extend(detect_usdidr_alerts())
    all_alerts.extend(detect_cpo_alerts())

    # Silent if no alerts
    if not all_alerts:
        return

    # Format output for Telegram
    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    print(f"🚨 *ALERT HARGA* — {now}")
    print()
    for alert in all_alerts:
        print(alert)


if __name__ == "__main__":
    main()
