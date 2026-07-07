#!/usr/bin/env python3
"""Dynamic daily market report — reads ALL Excel data files."""
import os, sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed", file=sys.stderr)
    sys.exit(1)

DATA_DIR = os.path.expanduser("~/sembako/data")


def load_sheet(filepath, sheet_name="Harian"):
    if not os.path.exists(filepath):
        return [], []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet_name = sheet_name if sheet_name in wb.sheetnames else (wb.sheetnames[0] if wb.sheetnames else None)
        if not sheet_name:
            return [], []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return list(rows[0]), rows[1:] if len(rows) > 1 else []
    except Exception:
        return [], []


def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def fmt_rp(v):
    return f"Rp {v:,.0f}".replace(",", ".") if v else "N/A"


def fmt_pct(v):
    if v is None:
        return ""
    arrow = "📈" if v > 0 else "📉" if v < 0 else "➡️"
    return f"{arrow} {v:+.1f}%"


def pct_change(curr, prev):
    if not curr or not prev or prev == 0:
        return None
    return ((curr - prev) / prev) * 100


# ========== SEMBAKO ==========
def report_sembako():
    header, data = load_sheet(os.path.join(DATA_DIR, "harga_sembako.xlsx"), "Harga")
    if not data:
        return ""
    
    latest = data[-1]
    prev = data[-2] if len(data) >= 2 else None
    date_str = str(latest[0])[:10] if latest[0] else "?"
    
    lines = ["🍚 **Sembako**"]
    # Key commodities: Beras Premium (1), Beras Medium (2), Oils (3), Sugar (4)
    key_idx = {1: "Beras Premium", 2: "Beras Medium", 3: "Minyak Goreng", 4: "Gula Pasir", 
               11: "Telur Ras", 12: "Telur Kampung", 13: "Ayam Ras", 14: "Ayam Kampung"}
    
    for idx, name in key_idx.items():
        curr = safe_float(latest[idx])
        if curr and curr > 0:
            prev_val = safe_float(prev[idx]) if prev else None
            chg = pct_change(curr, prev_val) if prev_val else None
            chg_str = f" {fmt_pct(chg)}" if chg else ""
            lines.append(f"   {name}: {fmt_rp(curr)}{chg_str}")
    
    return "\n".join(lines)


# ========== CRYPTO ==========
def report_crypto():
    header, data = load_sheet(os.path.join(DATA_DIR, "crypto_monitor.xlsx"), "Harga")
    if not data:
        return ""
    
    col_map = {str(h).strip(): i for i, h in enumerate(header) if h}
    latest = data[-1]
    
    lines = ["🪙 **Crypto**"]
    coins = [("BTC", "BTC (USD)", "BTC 24h %"), ("ETH", "ETH (USD)", "ETH 24h %"), ("SOL", "SOL (USD)", "SOL 24h %")]
    
    for name, price_col, chg_col in coins:
        usd = safe_float(latest[col_map[price_col]]) if price_col in col_map else None
        chg = safe_float(latest[col_map[chg_col]]) if chg_col in col_map else None
        if usd:
            chg_str = f" ({chg:+.1f}% 24h)" if chg else ""
            lines.append(f"   {name}: ${usd:,.0f}{chg_str}")
    
    return "\n".join(lines) if len(lines) > 1 else ""


# ========== EMAS ==========
def report_gold():
    header, data = load_sheet(os.path.join(DATA_DIR, "harga_emas.xlsx"), "Harian")
    if not data:
        return ""
    
    antam_idx = next((i for i, h in enumerate(header) if h and "antam" in str(h).lower() and "beli" in str(h).lower()), None)
    if antam_idx is None:
        return ""
    
    curr = safe_float(data[-1][antam_idx])
    prev = safe_float(data[-2][antam_idx]) if len(data) >= 2 else None
    if not curr:
        return ""
    
    chg = pct_change(curr, prev) if prev else None
    date_str = str(data[-1][0])[:10] if data[-1][0] else "?"
    return f"💰 **Emas**\n   Antam 1gr: {fmt_rp(curr)}/gram ({date_str}) {fmt_pct(chg)}"


# ========== PERTANIAN ==========
def report_pertanian():
    header, data = load_sheet(os.path.join(DATA_DIR, "harga_pertanian_ternak.xlsx"), "Harga Komoditas")
    if not data:
        return ""
    
    col_map = {str(h).lower(): i for i, h in enumerate(header) if h}
    
    # Key commodities from actual headers
    candidates = [
        ("jagung kering", "Jagung Kering Pipil"),
        ("jagung pakan", "Jagung Pakan Ternak"),
        ("kedelai lokal", "Kedelai Lokal"),
    ]
    lines = ["🌾 **Pertanian**"]
    
    latest = data[-1]
    prev = data[-2] if len(data) >= 2 else None
    found = False
    
    for key, name in candidates:
        idx = col_map.get(key)
        if idx is not None:
            val = safe_float(latest[idx])
            if val and val > 0:
                prev_val = safe_float(prev[idx]) if prev else None
                chg = pct_change(val, prev_val) if prev_val else None
                chg_str = f" {fmt_pct(chg)}" if chg else ""
                lines.append(f"   {name}: {fmt_rp(val)}/kg{chg_str}")
                found = True
    
    return "\n".join(lines) if found else ""


# ========== PETERNAKAN ==========
def report_peternakan():
    header, data = load_sheet(os.path.join(DATA_DIR, "harga_peternakan_lengkap.xlsx"), "Data Utama")
    if not data:
        return ""
    
    lines = ["🐔 **Peternakan**"]
    found = False
    
    # Peternakan has: Kategori, Sub Kategori, Produk, Harga (Rp), Satuan
    for row in data[-3:]:  # Check last 3 rows
        try:
            produk = str(row[3]).strip() if row[3] else ""
            harga = safe_float(row[4])
            if harga and harga > 0 and produk:
                lines.append(f"   {produk}: {fmt_rp(harga)}")
                found = True
        except (IndexError, TypeError):
            continue
    
    return "\n".join(lines) if found else ""


# ========== PAKAN ==========
def report_pakan():
    header, data = load_sheet(os.path.join(DATA_DIR, "harga_pakan_ternak.xlsx"), "Harga Pakan")
    if not data:
        return ""
    
    col_map = {str(h).lower(): i for i, h in enumerate(header) if h}
    
    candidates = [
        ("jagung", "Jagung Pipilan"),
        ("bungkil kedelai", "Bungkil Kedelai"),
        ("dedak", "Dedak Padi"),
    ]
    lines = ["🌽 **Pakan**"]
    
    latest = data[-1]
    prev = data[-2] if len(data) >= 2 else None
    found = False
    
    for key, name in candidates:
        idx = col_map.get(key)
        if idx is not None:
            val = safe_float(latest[idx])
            if val and val > 0:
                prev_val = safe_float(prev[idx]) if prev else None
                chg = pct_change(val, prev_val) if prev_val else None
                chg_str = f" {fmt_pct(chg)}" if chg else ""
                lines.append(f"   {name}: {fmt_rp(val)}/kg{chg_str}")
                found = True
    
    return "\n".join(lines) if found else ""


# ========== KURS ==========
def report_usdidr():
    header, data = load_sheet(os.path.join(DATA_DIR, "kurs_valuta.xlsx"), "Harian")
    if not data:
        return ""
    
    usd_idx = next((i for i, h in enumerate(header) if h and "usd" in str(h).lower()), None)
    if usd_idx is None:
        return ""
    
    curr = safe_float(data[-1][usd_idx])
    prev = safe_float(data[-2][usd_idx]) if len(data) >= 2 else None
    if not curr:
        return ""
    
    chg = pct_change(curr, prev) if prev else None
    date_str = str(data[-1][0])[:10] if data[-1][0] else "?"
    return f"💵 **USD/IDR**\n   Rp {curr:,.0f} ({date_str}) {fmt_pct(chg)}"


# ========== INFLASI ==========
def report_bi_rate():
    header, data = load_sheet(os.path.join(DATA_DIR, "bi_rate_inflasi.xlsx"), "Harian")
    if not data:
        return ""
    
    bi_idx = next((i for i, h in enumerate(header) if h and ("bi_rate" in str(h).lower())), None)
    yoy_idx = next((i for i, h in enumerate(header) if h and "yoy" in str(h).lower()), None)
    
    curr = data[-1]
    bi = safe_float(curr[bi_idx]) if bi_idx is not None else None
    yoy = safe_float(curr[yoy_idx]) if yoy_idx is not None else None
    
    parts = []
    if bi is not None:
        parts.append(f"BI Rate: {bi:.2f}%")
    if yoy is not None:
        parts.append(f"Inflasi: {yoy:.2f}%")
    
    if parts:
        date_str = str(curr[0])[:10] if curr[0] else "?"
        return f"🏦 **Macro**\n   {', '.join(parts)} ({date_str})"
    return ""


# ========== CPO & OIL ==========
def report_cpo():
    header, data = load_sheet(os.path.join(DATA_DIR, "harga_cpo.xlsx"), "Harian")
    if not data:
        return ""
    
    cpo_idx = next((i for i, h in enumerate(header) if h and "cpo" in str(h).lower()), None)
    if cpo_idx is None:
        return ""
    
    curr = safe_float(data[-1][cpo_idx])
    prev = safe_float(data[-2][cpo_idx]) if len(data) >= 2 else None
    if not curr:
        return ""
    
    chg = pct_change(curr, prev) if prev else None
    date_str = str(data[-1][0])[:10] if data[-1][0] else "?"
    return f"🌴 **CPO**\n   Rp {curr:,.0f}/kg ({date_str}) {fmt_pct(chg)}"


def report_oil():
    header, data = load_sheet(os.path.join(DATA_DIR, "harga_minyak.xlsx"), "Harian")
    if not data:
        return ""
    
    price_idx = next((i for i, h in enumerate(header) if h and ("brent" in str(h).lower() or "wti" in str(h).lower())), None)
    if price_idx is None and len(header) > 1:
        price_idx = 1
    
    curr = safe_float(data[-1][price_idx]) if price_idx else None
    prev = safe_float(data[-2][price_idx]) if len(data) >= 2 and price_idx else None
    if not curr:
        return ""
    
    chg = pct_change(curr, prev) if prev else None
    date_str = str(data[-1][0])[:10] if data[-1][0] else "?"
    return f"🛢️ **Minyak**\n   ${curr:.2f}/barel ({date_str}) {fmt_pct(chg)}"


# ========== MAIN ==========
def main():
    now = datetime.now()
    print(f"📊 *Laporan Pasar Harian*")
    print(f"📅 {now.strftime('%d %B %Y')} | 🕐 {now.strftime('%H:%M')} WIB")
    print("=" * 40)
    print()
    
    sections = [
        report_sembako(),
        report_crypto(),
        report_gold(),
        report_pertanian(),
        report_peternakan(),
        report_pakan(),
        report_usdidr(),
        report_bi_rate(),
        report_cpo(),
        report_oil(),
    ]
    
    for s in sections:
        if s:
            print(s)
            print()
    
    print("=" * 40)
    print("📌 Data dari dashboard sembako Indonesia")


if __name__ == "__main__":
    main()