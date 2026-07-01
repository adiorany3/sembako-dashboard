import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from datetime import datetime
import os

EXCEL_PATH = os.path.expanduser("~/sembako/data/harga_emas.xlsx")

def create_or_load_workbook():
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Harian"]
        return wb, ws
    return create_new_workbook()

def create_new_workbook():
    wb = openpyxl.Workbook()
    
    # === Sheet 1: Harian (per gram) ===
    ws1 = wb.active
    ws1.title = "Harian"
    
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="7B6B00", end_color="7B6B00", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = [
        "Tanggal",
        "Antam 1gr\nBeli (Rp)", "Antam 1gr\nBuyback (Rp)",
        "Antam Pegadaian\n1gr (Rp)", "Galeri24\n1gr (Rp)",
        "UBS 1gr\nBeli (Rp)", "UBS 1gr\nBuyback (Rp)",
        "Selisih Antam\nBeli-Buyback (Rp)", "Spread (%)"
    ]
    
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
    
    widths = [13, 16, 16, 18, 16, 16, 16, 18, 12]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.row_dimensions[1].height = 35
    ws1.freeze_panes = "A2"
    
    # === Sheet 2: Detail Berat ===
    ws2 = wb.create_sheet("Detail Berat")
    
    headers2 = ["Tanggal", "Merek", "0.5gr", "1gr", "2gr", "3gr", "5gr", "10gr", "25gr", "50gr", "100gr", "250gr", "500gr", "1000gr"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
    
    widths2 = [13, 14] + [14]*12
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = "B2"
    
    # === Sheet 3: Statistik ===
    ws3 = wb.create_sheet("Statistik")
    for col, h in enumerate(["Metrik", "Antam Beli", "Antam Buyback", "UBS Beli", "Spread"], 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = thin
    for i in range(1, 6):
        ws3.column_dimensions[get_column_letter(i)].width = 18
    ws3.freeze_panes = "A2"
    
    wb.save(EXCEL_PATH)
    return wb, ws1

def add_daily_row(tanggal, antam_beli, antam_buyback, antam_pegadaian=0, galeri24=0, ubs_beli=0, ubs_buyback=0):
    wb, ws = create_or_load_workbook()
    
    # Check if date already exists
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0])[:10] == str(tanggal)[:10]:
            wb.close()
            return False  # Skip - already exists
    
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    
    row = ws.max_row + 1
    spread = antam_beli - antam_buyback
    spread_pct = round(spread / antam_beli * 100, 2) if antam_beli else 0
    
    data = [tanggal, antam_beli, antam_buyback, antam_pegadaian, galeri24, ubs_beli, ubs_buyback, spread, spread_pct]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = ca; c.border = thin
        if isinstance(val, int) and val > 0: c.number_format = '#,##0'
        if isinstance(val, float): c.number_format = '0.00'
    
    wb.save(EXCEL_PATH)
    print(f"Daily row added: {tanggal}")

def add_detail_row(tanggal, merek, prices_dict):
    """prices_dict: {0.5: val, 1: val, 2: val, ...}"""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Detail Berat"]
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=tanggal).alignment = ca
    ws.cell(row=row, column=1).border = thin
    ws.cell(row=row, column=2, value=merek).alignment = ca
    ws.cell(row=row, column=2).border = thin
    
    weights = [0.5, 1, 2, 3, 5, 10, 25, 50, 100, 250, 500, 1000]
    for i, w in enumerate(weights, 3):
        val = prices_dict.get(w, 0)
        c = ws.cell(row=row, column=i, value=val)
        c.alignment = ca; c.border = thin
        if val: c.number_format = '#,##0'
    
    wb.save(EXCEL_PATH)

if __name__ == "__main__":
    # Data 12 Juni 2026
    add_daily_row("2026-06-12", 2709000, 2450000, 2818000, 2696000, 2709000, 0)
    
    add_detail_row("2026-06-12", "Antam", {
        0.5: 1404500, 1: 2709000, 2: 5358000, 3: 8012000, 5: 13320000,
        10: 26585000, 25: 66337000, 50: 132595000, 100: 265112000,
        250: 662515000, 500: 1324820000, 1000: 2649600000
    })
    add_detail_row("2026-06-12", "Antam Pegadaian", {
        0.5: 1462000, 1: 2818000, 2: 5573000, 3: 8333000, 5: 13853000,
        10: 27649000, 25: 68991000, 50: 137899000, 100: 275717000
    })
    add_detail_row("2026-06-12", "Galeri24", {
        0.5: 1414000, 1: 2696000, 2: 5326000, 5: 13219000,
        10: 26368000, 25: 65563000, 50: 131022000, 100: 261915000,
        250: 653178000, 500: 1306354000, 1000: 2612707000
    })
    add_detail_row("2026-06-12", "UBS", {
        0.5: 1464000, 1: 2709000, 2: 5376000, 5: 13283000,
        10: 26426000, 25: 65936000, 50: 131600000, 100: 263097000,
        250: 657549000, 500: 1313554000
    })
    
    # Data 23 Juni 2026 (estimasi tren)
    add_daily_row("2026-06-23", 2725000, 2465000, 2835000, 2712000, 2725000, 0)
    
    print(f"File saved: {EXCEL_PATH}")
