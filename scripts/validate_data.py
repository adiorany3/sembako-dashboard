#!/usr/bin/env python3
"""
Validate all Excel data files.
Prints summary. Exits non-zero if critical failures.
"""
import os
import sys
from pathlib import Path
from datetime import datetime
import openpyxl

# Add parent to path
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from utils.validation import validate_price, validate_date, validate_percentage
except ImportError:
    # Fallback functions if validation module unavailable
    def validate_price(price, min_price=100, max_price=100000000):
        if price is None: return False, "Empty"
        try:
            p = int(price)
            return p > 0 and min_price <= p <= max_price, ""
        except (ValueError, TypeError, KeyError): return False, "Invalid"
    
    def validate_date(date_val, max_age_days=365):
        if not date_val: return False, "Empty"
        try:
            dt = datetime.strptime(str(date_val)[:10], "%Y-%m-%d")
            age = (datetime.now() - dt).days
            return 0 <= age <= max_age_days, ""
        except (ValueError, TypeError, KeyError): return False, "Invalid"
    
    def validate_percentage(pct, max_val=100):
        try:
            p = float(pct)
            return -max_val <= p <= max_val, ""
        except (ValueError, TypeError, KeyError): return False, "Invalid"

DATA_DIR = Path(__file__).parent.parent / "data"

# File-specific validation rules: (filename, sheet_name, key_cols)
# key_cols: indices that must be present - price is always col with numeric > 1000
VALIDATION_TARGETS = [
    ("harga_sembako.xlsx", "Harga", None),
    ("harga_emas.xlsx", "Harian", None),
    ("harga_minyak.xlsx", "Harian", None),
    ("bi_rate_inflasi.xlsx", "Harian", None),
    ("harga_cpo.xlsx", "Harian", None),
    ("kurs_valuta.xlsx", "Harian", None),
    ("harga_peternakan_lengkap.xlsx", "Data Utama", None),
    ("harga_pakan_ternak.xlsx", None, None),
    ("harga_saham_ihsg.xlsx", "IHSG", None),
    ("sentimen_berita.xlsx", "Detail", None),
    ("crypto_monitor.xlsx", "Sheet1", None),
]


def find_price_column(row):
    """Find column index containing price-like value (>1000 int/float)."""
    for i, val in enumerate(row):
        if val is not None:
            try:
                v = float(val)
                if v > 1000:
                    return i
            except (ValueError, TypeError):
                pass
    return None


def validate_file(filepath, sheet_name):
    """Validate single file. Returns (is_valid, errors, warnings)."""
    errors = []
    warnings = []
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return False, [f"Cannot open: {e}"], []
    
    if sheet_name and sheet_name not in wb.sheetnames:
        return False, [f"Sheet '{sheet_name}' not found"], []
    
    ws = wb[sheet_name] if sheet_name else wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        return False, ["No data rows"], []
    
    headers = rows[0]
    data_rows = rows[1:]
    
    # Check row count
    if len(data_rows) == 0:
        errors.append("Empty dataset")
        wb.close()
        return False, errors, warnings
    
    # Find price column
    price_col = None
    for row in data_rows[:10]:
        pc = find_price_column(row)
        if pc is not None:
            price_col = pc
            break
    
    # Validate each row
    valid_count = 0
    for idx, row in enumerate(data_rows, start=2):
        if not row or len(row) == 0:
            continue
        
        # Validate date (first column)
        date_valid, date_err = validate_date(row[0])
        if not date_valid:
            errors.append(f"Row {idx}: invalid date - {date_err}")
        
        # Validate price if found
        if price_col is not None and price_col < len(row):
            price_val = row[price_col]
            if price_val is not None:
                price_valid, price_err = validate_price(price_val)
                if not price_valid:
                    errors.append(f"Row {idx}: invalid price {price_val} - {price_err}")
                else:
                    valid_count += 1
            else:
                warnings.append(f"Row {idx}: empty price")
        else:
            valid_count += 1  # No price column found, assume valid
    
    wb.close()
    
    # Determine status
    is_valid = len(errors) == 0
    
    # Warnings for stale/simulated data
    if valid_count == 0:
        warnings.append("No valid prices found")
    
    return is_valid, errors, warnings


def main():
    print("=" * 60)
    print("VALIDATE EXCEL DATA FILES")
    print(f"Data dir: {DATA_DIR}")
    print("=" * 60)
    
    total_errors = 0
    total_warnings = 0
    files_checked = 0
    files_failed = 0
    
    for filename, sheet_name, _ in VALIDATION_TARGETS:
        filepath = DATA_DIR / filename
        
        if not filepath.exists():
            print(f"⏭️  SKIP {filename} — file not found")
            continue
        
        files_checked += 1
        is_valid, errors, warnings = validate_file(str(filepath), sheet_name)
        
        if is_valid:
            print(f"✅ {filename}: OK ({len(warnings)} warnings)")
            for w in warnings:
                print(f"   ⚠️  {w}")
        else:
            files_failed += 1
            total_errors += len(errors)
            print(f"❌ {filename}: FAILED")
            for e in errors:
                print(f"   ❌ {e}")
            for w in warnings:
                print(f"   ⚠️  {w}")
        
        total_warnings += len(warnings)
    
    # Check any other xlsx files not in list
    for f in DATA_DIR.glob("*.xlsx"):
        if f.name not in [t[0] for t in VALIDATION_TARGETS]:
            files_checked += 1
            is_valid, errors, warnings = validate_file(str(f), None)
            if not is_valid:
                files_failed += 1
                total_errors += len(errors)
                print(f"❌ {f.name}: FAILED")
                for e in errors:
                    print(f"   ❌ {e}")
            else:
                print(f"✅ {f.name}: OK")
    
    print()
    print("=" * 60)
    print(f"SUMMARY: {files_checked} files checked, {files_failed} failed")
    print(f"Errors: {total_errors}, Warnings: {total_warnings}")
    
    if files_failed > 0:
        print("❌ VALIDATION FAILED")
        sys.exit(1)
    else:
        print("✅ ALL VALIDATIONS PASSED")
        sys.exit(0)


if __name__ == "__main__":
    main()
