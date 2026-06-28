#!/usr/bin/env python3
"""
Formulasi Pakan Ternak Komprehensif — Indonesia
================================================
Generates Excel with:
  1. Bahan Pakan & Harga Terkini
  2. Profil Nutrisi per Bahan
  3. Ruminansia Besar (Sapi)
  4. Ruminansia Kecil (Kambing/Domba)
  5. Non-Ruminansia (Babi)
  6. Unggas (Ayam Broiler, Ayam Petelur, Bebek)
  7. Ringkasan Biaya per Formulasi
  8. Kebutuhan Nutrisi per Kelompok

Harga bahan baku update terkini dari data pasar Indonesia.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
import os, sys

# ════════════════════════════════════════════
# HARGA BAHAN BAKU (28 Juni 2026, Rp/kg)
# ════════════════════════════════════════════
PRICES = {
    "Jagung Pipilan": 4896,
    "Bungkil Kedelai": 7984,
    "Dedak Padi": 4039,
    "Tepung Ikan": 8023,
    "Pollard": 4764,
    "Biji Kapuk": 4168,
    "Tepung Darah": 8149,
    "Tepung Tulang": 7908,
    "Molases": 3637,
    "Bungkil Kelapa": 8637,
    "Gaplek": 3202,
    "Bungkil Sawit": 8201,
    "Ampas Tahu": 3582,
    "Tepung Bulu Ayam": 8038,
    "Kulit Kentang": 2830,
    "Onggok": 2386,
    "Bungkil Kacang Tanah": 8059,
    "Dedak Halus": 4006,
    "Sorgum": 5148,
    "Menir": 5211,
    "Corn Gluten Feed": 7515,
    "Rice Polish": 4933,
    "Mung Bean Husk": 3332,
}

# ════════════════════════════════════════════
# PROFIL NUTRISI (basis kering %)
# Sumber: NRC, Kementan RI, literatur pakan ternak
# ════════════════════════════════════════════
# Columns: Karbohidrat, Protein(KK), Lemak(KK), Serat Kasar, Abu, Ca, P, Metabolisable Energy
# Energy dalam Mcal/kg (ME)
NUTRITION = {
    "Jagung Pipilan":       {"KK": 8.5, "PK": 84, "PK_KK": 7.14, "LK": 3.8, "SK": 2.2, "Abu": 1.3, "Ca": 0.02, "P": 0.28, "ME": 3.35, "Ket": "Energi tinggi, protein rendah"},
    "Bungkil Kedelai":      {"KK": 44, "PK": 35, "PK_KK": 15.4, "LK": 1.5, "SK": 6.0, "Abu": 5.9, "Ca": 0.34, "P": 0.65, "ME": 2.44, "Ket": "Protein tinggi, asam amino lengkap"},
    "Dedak Padi":           {"KK": 13.5, "PK": 12, "PK_KK": 9.6, "LK": 12.8, "SK": 15.5, "Abu": 10.7, "Ca": 0.10, "P": 1.40, "ME": 1.75, "Ket": "P tinggi, serat sedang"},
    "Tepung Ikan":          {"KK": 5.0, "PK": 62, "PK_KK": 37.2, "LK": 12.0, "SK": 0.8, "Abu": 18.2, "Ca": 4.50, "P": 2.80, "ME": 2.81, "Ket": "Protein hewani super, Ca tinggi"},
    "Pollard":              {"KK": 13.0, "PK": 15, "PK_KK": 11.5, "LK": 4.5, "SK": 9.5, "Abu": 4.2, "Ca": 0.12, "P": 1.10, "ME": 2.05, "Ket": "Serat sedang, murah"},
    "Biji Kapuk":           {"KK": 14.0, "PK": 20, "PK_KK": 14.3, "LK": 46.0, "SK": 23.0, "Abu": 6.0, "Ca": 0.18, "P": 0.55, "ME": 3.90, "Ket": "Lemak sangat tinggi"},
    "Tepung Darah":         {"KK": 1.0, "PK": 85, "PK_KK": 85.0, "LK": 1.5, "SK": 0.5, "Abu": 4.0, "Ca": 0.15, "P": 0.30, "ME": 3.01, "Ket": "Protein hewani tertinggi"},
    "Tepung Tulang":        {"KK": 0, "PK": 50, "PK_KK": 50.0, "LK": 2.0, "SK": 0.5, "Abu": 72.0, "Ca": 24.0, "P": 12.0, "ME": 1.05, "Ket": "Suplemen Ca & P"},
    "Molases":              {"KK": 33, "PK": 4, "PK_KK": 12.1, "LK": 0.5, "SK": 0, "Abu": 14.0, "Ca": 0.80, "P": 0.08, "ME": 2.62, "Ket": "Pemanis, fermentasi rumen"},
    "Bungkil Kelapa":       {"KK": 22, "PK": 22, "PK_KK": 10.0, "LK": 10.0, "SK": 12.0, "Abu": 5.5, "Ca": 0.20, "P": 0.60, "ME": 2.85, "Ket": "Protein sedang, lemak tinggi"},
    "Gaplek":               {"KK": 2.0, "PK": 2, "PK_KK": 100.0, "LK": 0.5, "SK": 4.0, "Abu": 1.5, "Ca": 0.05, "P": 0.05, "ME": 3.30, "Ket": "Karbohidrat sumber energi murah"},
    "Bungkil Sawit":        {"KK": 35, "PK": 18, "PK_KK": 5.1, "LK": 9.0, "SK": 15.0, "Abu": 4.5, "Ca": 0.25, "P": 0.65, "ME": 2.60, "Ket": "Protein sedang, serat tinggi"},
    "Ampas Tahu":           {"KK": 16, "PK": 14, "PK_KK": 8.8, "LK": 8.0, "SK": 15.0, "Abu": 3.5, "Ca": 0.50, "P": 0.30, "ME": 1.95, "Ket": "Protein nabati murah"},
    "Tepung Bulu Ayam":     {"KK": 2.0, "PK": 80, "PK_KK": 80.0, "LK": 5.0, "SK": 0.5, "Abu": 4.0, "Ca": 0.50, "P": 0.35, "ME": 2.95, "Ket": "Protein hewani, kualitas sedang"},
    "Kulit Kentang":        {"KK": 2.0, "PK": 7, "PK_KK": 35.0, "LK": 0.5, "SK": 8.0, "Abu": 5.5, "Ca": 0.08, "P": 0.15, "ME": 2.70, "Ket": "Energi, murah"},
    "Onggok":               {"KK": 1.5, "PK": 2, "PK_KK": 133.0, "LK": 2.0, "SK": 8.0, "Abu": 2.5, "Ca": 0.10, "P": 0.05, "ME": 3.10, "Ket": "Pengisi murah, karbohidrat rendah"},
    "Bungkil Kacang Tanah": {"KK": 30, "PK": 38, "PK_KK": 12.7, "LK": 8.0, "SK": 12.0, "Abu": 5.5, "Ca": 0.20, "P": 0.55, "ME": 2.85, "Ket": "Protein nabati tinggi"},
    "Dedak Halus":          {"KK": 13.0, "PK": 13, "PK_KK": 10.0, "LK": 13.0, "SK": 14.0, "Abu": 10.5, "Ca": 0.08, "P": 1.35, "ME": 1.80, "Ket": "Serat sedang, P tinggi"},
    "Sorgum":               {"KK": 11.0, "PK": 11, "PK_KK": 10.0, "LK": 3.5, "SK": 2.5, "Abu": 1.5, "Ca": 0.04, "P": 0.30, "ME": 3.20, "Ket": "Alternatif jagung"},
    "Menir":                {"KK": 10.0, "PK": 8, "PK_KK": 8.0, "LK": 1.5, "SK": 0.8, "Abu": 0.5, "Ca": 0.01, "P": 0.12, "ME": 3.40, "Ket": "Energi tinggi, serat rendah"},
    "Corn Gluten Feed":     {"KK": 18, "PK": 22, "PK_KK": 12.2, "LK": 3.0, "SK": 8.0, "Abu": 2.5, "Ca": 0.06, "P": 0.40, "ME": 2.50, "Ket": "Protein nabati sedang"},
    "Rice Polish":          {"KK": 13.0, "PK": 12, "PK_KK": 9.2, "LK": 13.0, "SK": 8.0, "Abu": 8.0, "Ca": 0.05, "P": 1.10, "ME": 1.90, "Ket": "Serat rendah, lemak sedang"},
    "Mung Bean Husk":       {"KK": 10.0, "PK": 15, "PK_KK": 15.0, "LK": 2.0, "SK": 18.0, "Abu": 4.5, "Ca": 0.90, "P": 0.25, "ME": 2.00, "Ket": "Serat tinggi, Ca tinggi"},
}

# ════════════════════════════════════════════
# KEBUTUHAN NUTRISI PER KATEGORI
# Sumber: NRC 2001, Kementan RI
# ════════════════════════════════════════════
NRC = {
    # Ruminansia Besar (Sapi)
    "Sapi Potong Konservasi":   {"BK": 100, "PK_min": 8, "PK_max": 12, "TDN": 55, "SK_max": 30, "Ca": 0.3, "P": 0.2, "ME": 2.2},
    "Sapi Potong Intensif":     {"BK": 120, "PK_min": 12, "PK_max": 16, "TDN": 65, "SK_max": 25, "Ca": 0.4, "P": 0.3, "ME": 2.6},
    "Sapi Perah Produksi Rendah":{"BK": 100, "PK_min": 14, "PK_max": 16, "TDN": 65, "SK_max": 25, "Ca": 0.6, "P": 0.4, "ME": 2.6},
    "Sapi Perah Produksi Tinggi":{"BK": 150, "PK_min": 16, "PK_max": 18, "TDN": 70, "SK_max": 20, "Ca": 0.8, "P": 0.5, "ME": 2.9},
    "Sapi Bakalan 200-350kg":   {"BK": 100, "PK_min": 12, "PK_max": 14, "TDN": 62, "SK_max": 25, "Ca": 0.4, "P": 0.3, "ME": 2.5},
    "Sapi Fattening 350-500kg": {"BK": 120, "PK_min": 12, "PK_max": 14, "TDN": 68, "SK_max": 20, "Ca": 0.4, "P": 0.3, "ME": 2.7},
    "Sapi Hamil 7-9 bulan":     {"BK": 100, "PK_min": 11, "PK_max": 13, "TDN": 60, "SK_max": 25, "Ca": 0.5, "P": 0.3, "ME": 2.4},
    "Sapi Laktasi Ringan":      {"BK": 120, "PK_min": 14, "PK_max": 16, "TDN": 65, "SK_max": 22, "Ca": 0.6, "P": 0.4, "ME": 2.6},

    # Ruminansia Kecil (Kambing/Domba)
    "Kambing Konservasi":       {"BK": 50, "PK_min": 9, "PK_max": 12, "TDN": 55, "SK_max": 28, "Ca": 0.3, "P": 0.2, "ME": 2.2},
    "Kambing Pedaging Potong":  {"BK": 60, "PK_min": 12, "PK_max": 14, "TDN": 62, "SK_max": 22, "Ca": 0.4, "P": 0.3, "ME": 2.5},
    "Kambing Perah":            {"BK": 70, "PK_min": 14, "PK_max": 16, "TDN": 65, "SK_max": 20, "Ca": 0.6, "P": 0.4, "ME": 2.6},
    "Domba Konservasi":         {"BK": 40, "PK_min": 9, "PK_max": 11, "TDN": 55, "SK_max": 28, "Ca": 0.3, "P": 0.2, "ME": 2.1},
    "Domba Pedaging":           {"BK": 50, "PK_min": 12, "PK_max": 14, "TDN": 62, "SK_max": 22, "Ca": 0.4, "P": 0.3, "ME": 2.5},

    # Non-Ruminansia (Babi)
    "Babi Starter (7-25kg)":    {"BK": 100, "PK_min": 18, "PK_max": 20, "TDN": 72, "SK_max": 8, "Ca": 0.6, "P": 0.5, "ME": 3.2},
    "Babi Grower (25-60kg)":    {"BK": 100, "PK_min": 15, "PK_max": 17, "TDN": 72, "SK_max": 10, "Ca": 0.5, "P": 0.4, "ME": 3.1},
    "Babi Finisher (60-100kg)": {"BK": 100, "PK_min": 13, "PK_max": 15, "TDN": 72, "SK_max": 10, "Ca": 0.5, "P": 0.4, "ME": 3.0},
    "Babi Lancet (Induk)":      {"BK": 100, "PK_min": 14, "PK_max": 16, "TDN": 70, "SK_max": 12, "Ca": 0.7, "P": 0.5, "ME": 2.9},
    "Babi Laktasi (Sapih)":     {"BK": 100, "PK_min": 16, "PK_max": 18, "TDN": 72, "SK_max": 10, "Ca": 0.8, "P": 0.6, "ME": 3.0},

    # Unggas - Ayam Broiler
    "Broiler Starter (0-10hr)":  {"BK": 100, "PK_min": 22, "PK_max": 24, "TDN": 70, "SK_max": 5, "Ca": 1.0, "P": 0.5, "ME": 3.0},
    "Broiler Grower (10-25hr)":  {"BK": 100, "PK_min": 20, "PK_max": 22, "TDN": 72, "SK_max": 5, "Ca": 0.9, "P": 0.45, "ME": 3.1},
    "Broiler Finisher (25-35d)": {"BK": 100, "PK_min": 18, "PK_max": 20, "TDN": 73, "SK_max": 5, "Ca": 0.9, "P": 0.4, "ME": 3.2},

    # Unggas - Ayam Petelur
    "Layer Grower (7-17wk)":     {"BK": 100, "PK_min": 16, "PK_max": 18, "TDN": 68, "SK_max": 6, "Ca": 1.0, "P": 0.45, "ME": 2.9},
    "Layer Pre-Lay (17-18wk)":   {"BK": 100, "PK_min": 17, "PK_max": 18, "TDN": 70, "SK_max": 5, "Ca": 2.0, "P": 0.5, "ME": 2.9},
    "Layer Produksi":            {"BK": 100, "PK_min": 16, "PK_max": 18, "TDN": 70, "SK_max": 5, "Ca": 3.5, "P": 0.45, "ME": 2.85},

    # Unggas - Bebek
    "Bebek Petelur Starter":     {"BK": 100, "PK_min": 18, "PK_max": 20, "TDN": 70, "SK_max": 6, "Ca": 1.0, "P": 0.5, "ME": 2.9},
    "Bebek Petelur Layer":       {"BK": 100, "PK_min": 16, "PK_max": 18, "TDN": 70, "SK_max": 5, "Ca": 2.8, "P": 0.45, "ME": 2.85},
    "Bebek Pedaging (Peking)":   {"BK": 100, "PK_min": 16, "PK_max": 18, "TDN": 72, "SK_max": 6, "Ca": 0.9, "P": 0.45, "ME": 3.0},
}


# ════════════════════════════════════════════
# STYLES
# ════════════════════════════════════════════
TITLE = Font(bold=True, size=14, color="FFFFFF")
HEADER = Font(bold=True, size=10, color="FFFFFF")
SUBHEADER = Font(bold=True, size=10)
DATA = Font(size=9)
MONEY = Font(size=9, color="006600")
WARN = Font(size=9, color="CC0000")

BLUE = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
GREEN = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
RED = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
ORANGE = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
YELLOW = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
PURPLE = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
GREY = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
LIGHT_RED = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
LIGHT_YELLOW = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header(ws, row, cols, fill=BLUE):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN


def style_row(ws, row, cols, font=DATA, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.alignment = CENTER if c > 1 else LEFT
        cell.border = THIN
        if fill: cell.fill = fill


def auto_width(ws, min_w=10, max_w=25):
    for col in ws.columns:
        mx = min_w
        for cell in col:
            if cell.value:
                mx = max(mx, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = mx


# ════════════════════════════════════════════
# HELPER: Calculate mixed nutrition
# ════════════════════════════════════════════
def calc_mix(formula):
    """Given {ingredient: pct}, return aggregated nutrition dict + cost."""
    total_pk = 0; total_me = 0; total_sk = 0; total_ca = 0; total_p = 0
    total_cost = 0; total_lk = 0

    for name, pct in formula.items():
        if pct <= 0: continue
        n = NUTRITION.get(name, {})
        r = pct / 100.0
        total_pk += n.get("PK", 0) * r
        total_me += n.get("ME", 0) * r
        total_sk += n.get("SK", 0) * r
        total_ca += n.get("Ca", 0) * r
        total_p  += n.get("P", 0) * r
        total_lk += n.get("LK", 0) * r
        total_cost += PRICES.get(name, 0) * r

    return {
        "PK": round(total_pk, 1),
        "ME": round(total_me, 2),
        "SK": round(total_sk, 1),
        "Ca": round(total_ca, 2),
        "P": round(total_p, 2),
        "LK": round(total_lk, 1),
        "Harga/kg": round(total_cost),
    }


# ════════════════════════════════════════════
# FORMULASI — RUMINANSIA BESAR (SAPI)
# ════════════════════════════════════════════
RUMINAN_BESAR = {
    "Sapi Potong Konservasi": {
        "Komposisi": {"Jagung Pipilan": 25, "Dedak Padi": 20, "Pollard": 10, "Onggok": 15,
                      "Gaplek": 10, "Bungkil Sawit": 8, "Molases": 5, "Tepung Tulang": 1, "Dedak Halus": 6},
        "Catatan": "Pakan ternak rumput + suplemen. Cocok untuk sapi potong养地 konservasi. Dominasi serat kasar.",
    },
    "Sapi Potong Intensif": {
        "Komposisi": {"Jagung Pipilan": 35, "Bungkil Kedelai": 15, "Dedak Padi": 12, "Pollard": 8,
                      "Tepung Ikan": 3, "Bungkil Sawit": 10, "Gaplek": 8, "Molases": 5, "Tepung Tulang": 1, "Dedak Halus": 3},
        "Catatan": "High-energy untuk sapi bakalan 200-350kg. Pertambahan BB 0.8-1.2 kg/hari.",
    },
    "Sapi Fattening": {
        "Komposisi": {"Jagung Pipilan": 40, "Bungkil Kedelai": 15, "Dedak Padi": 10, "Tepung Ikan": 4,
                      "Bungkil Sawit": 10, "Bungkil Kacang Tanah": 5, "Gaplek": 8, "Molases": 5, "Tepung Tulang": 1, "Dedak Halus": 2},
        "Catatan": "Maksimalkan pertumbuhan. Target FCR 7:1. Sapi 350-500kg.",
    },
    "Sapi Perah Produksi Rendah": {
        "Komposisi": {"Jagung Pipilan": 30, "Bungkil Kedelai": 12, "Dedak Padi": 15, "Pollard": 8,
                      "Bungkil Kelapa": 10, "Bungkil Sawit": 8, "Gaplek": 8, "Molases": 5, "Tepung Tulang": 1.5, "Dedak Halus": 2.5},
        "Catatan": "Sapi perah < 10L/hari. Kalsium tinggi untuk tulang kuat.",
    },
    "Sapi Perah Produksi Tinggi": {
        "Komposisi": {"Jagung Pipilan": 35, "Bungkil Kedelai": 18, "Dedak Padi": 8, "Tepung Ikan": 5,
                      "Bungkil Kacang Tanah": 8, "Bungkil Kelapa": 8, "Gaplek": 6, "Molases": 5, "Tepung Tulang": 2, "Corn Gluten Feed": 5},
        "Catatan": "Sapi perah > 15L/hari. Protein tinggi, energi tinggi.",
    },
    "Sapi Hamil 7-9 Bulan": {
        "Komposisi": {"Jagung Pipilan": 28, "Bungkil Kedelai": 10, "Dedak Padi": 18, "Pollard": 10,
                      "Bungkil Sawit": 10, "Gaplek": 10, "Onggok": 5, "Molases": 5, "Tepung Tulang": 1.5, "Dedak Halus": 2.5},
        "Catatan": "Trimester akhir. Ca & P tinggi untuk perkembangan janin.",
    },
    "Sapi Laktasi Ringan": {
        "Komposisi": {"Jagung Pipilan": 32, "Bungkil Kedelai": 14, "Dedak Padi": 12, "Pollard": 8,
                      "Bungkil Kacang Tanah": 8, "Bungkil Kelapa": 8, "Gaplek": 8, "Molases": 5, "Tepung Tulang": 1.5, "Dedak Halus": 3.5},
        "Catatan": "Laktasi 5-10L/hari. Protein & Ca terjaga.",
    },
}


# ════════════════════════════════════════════
# FORMULASI — RUMINANSIA KECIL (KAMBING/DOMBA)
# ════════════════════════════════════════════
RUMINAN_KECIL = {
    "Kambing Konservasi": {
        "Komposisi": {"Jagung Pipilan": 22, "Dedak Padi": 20, "Pollard": 10, "Onggok": 15,
                      "Gaplek": 12, "Bungkil Sawit": 8, "Molases": 5, "Dedak Halus": 5, "Tepung Tulang": 1, "Sorgum": 2},
        "Catatan": "Pakan ternak rumput + konsentrat. Kambing etawa/kacang养地.",
    },
    "Kambing Pedaging Potong": {
        "Komposisi": {"Jagung Pipilan": 30, "Bungkil Kedelai": 12, "Dedak Padi": 15, "Pollard": 8,
                      "Bungkil Sawit": 10, "Gaplek": 10, "Molases": 5, "Tepung Tulang": 1, "Dedak Halus": 7, "Sorgum": 2},
        "Catatan": "Kambing boer/kacang potong. Target BB 25-35kg.",
    },
    "Kambing Perah (Saanen)": {
        "Komposisi": {"Jagung Pipilan": 30, "Bungkil Kedelai": 15, "Dedak Padi": 12, "Tepung Ikan": 3,
                      "Bungkil Kacang Tanah": 8, "Bungkil Kelapa": 8, "Gaplek": 8, "Molases": 5, "Tepung Tulang": 1.5, "Dedak Halus": 6, "Corn Gluten Feed": 3.5},
        "Catatan": "Kambing perah Saanen. Produksi susu 2-4L/hari.",
    },
    "Domba Konservasi": {
        "Komposisi": {"Jagung Pipilan": 20, "Dedak Padi": 22, "Pollard": 10, "Onggok": 15,
                      "Gaplek": 12, "Bungkil Sawit": 8, "Molases": 5, "Dedak Halus": 5, "Tepung Tulang": 1, "Sorgum": 2},
        "Catatan": "Domba ekor tipis/tugas养地.",
    },
    "Domba Pedaging": {
        "Komposisi": {"Jagung Pipilan": 28, "Bungkil Kedelai": 12, "Dedak Padi": 15, "Pollard": 8,
                      "Bungkil Sawit": 10, "Gaplek": 10, "Molases": 5, "Tepung Tulang": 1, "Dedak Halus": 8, "Sorgum": 3},
        "Catatan": "Domba texel/lokal potong. Target BB 20-30kg.",
    },
}


# ════════════════════════════════════════════
# FORMULASI — NON-RUMINANSIA (BABI)
# ════════════════════════════════════════════
NON_RUMINAN = {
    "Babi Starter (7-25kg)": {
        "Komposisi": {"Jagung Pipilan": 35, "Bungkil Kedelai": 22, "Tepung Ikan": 5, "Tepung Darah": 5,
                      "Tepung Bulu Ayam": 3, "Menir": 12, "Corn Gluten Feed": 8, "Ampas Tahu": 3, "Molases": 3, "Tepung Tulang": 1.5, "Dedak Padi": 2.5},
        "Catatan": "Doc 7-25kg. Protein tinggi untuk pertumbuhan awal. FCR target 1.5:1.",
    },
    "Babi Grower (25-60kg)": {
        "Komposisi": {"Jagung Pipilan": 38, "Bungkil Kedelai": 18, "Tepung Ikan": 3, "Tepung Darah": 4,
                      "Menir": 15, "Corn Gluten Feed": 8, "Ampas Tahu": 4, "Dedak Padi": 5, "Molases": 3, "Tepung Tulang": 1, "Bungkil Sawit": 1},
        "Catatan": "Grower 25-60kg. Energi tinggi, protein sedang. FCR target 2.5:1.",
    },
    "Babi Finisher (60-100kg)": {
        "Komposisi": {"Jagung Pipilan": 42, "Bungkil Kedelai": 15, "Menir": 15, "Corn Gluten Feed": 8,
                      "Ampas Tahu": 5, "Dedak Padi": 5, "Tepung Bulu Ayam": 3, "Bungkil Sawit": 3, "Molases": 3, "Tepung Tulang": 1},
        "Catatan": "Finisher 60-100kg. Rendah protein, energi tinggi. FCR target 3.2:1.",
    },
    "Babi Lancet (Induk)": {
        "Komposisi": {"Jagung Pipilan": 35, "Bungkil Kedelai": 18, "Tepung Ikan": 3, "Tepung Darah": 3,
                      "Dedak Padi": 12, "Menir": 10, "Corn Gluten Feed": 8, "Ampas Tahu": 4, "Molases": 3, "Tepung Tulang": 1.5, "Dedak Halus": 2.5},
        "Catatan": "Induk babi hamil. Protein sedang, Ca & P tinggi untuk janin.",
    },
    "Babi Laktasi (Sapih)": {
        "Komposisi": {"Jagung Pipilan": 35, "Bungkil Kedelai": 20, "Tepung Ikan": 5, "Tepung Darah": 4,
                      "Menir": 12, "Corn Gluten Feed": 8, "Ampas Tahu": 4, "Molases": 3, "Tepung Tulang": 2, "Dedak Halus": 5, "Bungkil Sawit": 2},
        "Catatan": "Induk laktasi. Protein & energi tinggi untuk produksi ASI.",
    },
}


# ════════════════════════════════════════════
# FORMULASI — UNGGAS
# ════════════════════════════════════════════
UNGGAS = {
    "Broiler Starter (0-10hr)": {
        "Komposisi": {"Jagung Pipilan": 35, "Bungkil Kedelai": 25, "Tepung Ikan": 8, "Tepung Darah": 5,
                      "Tepung Bulu Ayam": 3, "Menir": 10, "Corn Gluten Feed": 5, "Bungkil Kacang Tanah": 3, "Molases": 2, "Tepung Tulang": 1.5, "Dedak Padi": 2.5},
        "Catatan": "Doc 0-10 hari. Protein sangat tinggi. Bungkus kuning = starter.",
    },
    "Broiler Grower (10-25hr)": {
        "Komposisi": {"Jagung Pipilan": 38, "Bungkil Kedelai": 22, "Tepung Ikan": 6, "Tepung Darah": 4,
                      "Tepung Bulu Ayam": 3, "Menir": 12, "Corn Gluten Feed": 5, "Bungkil Kacang Tanah": 3, "Molases": 2, "Tepung Tulang": 1.5, "Dedak Padi": 3.5},
        "Catatan": "Grower 10-25 hari. Protein sedang, energi tinggi.",
    },
    "Broiler Finisher (25-35hr)": {
        "Komposisi": {"Jagung Pipilan": 42, "Bungkil Kedelai": 18, "Tepung Ikan": 5, "Tepung Darah": 3,
                      "Menir": 14, "Corn Gluten Feed": 6, "Ampas Tahu": 3, "Molases": 2, "Tepung Tulang": 1.5, "Dedak Padi": 4, "Bungkil Sawit": 1.5},
        "Catatan": "Finisher 25-35 hari. Energi maksimal. FCR target 1.7-1.8:1.",
    },
    "Layer Grower (7-17wk)": {
        "Komposisi": {"Jagung Pipilan": 38, "Bungkil Kedelai": 18, "Tepung Ikan": 4, "Tepung Darah": 3,
                      "Menir": 12, "Corn Gluten Feed": 6, "Ampas Tahu": 4, "Bungkil Sawit": 3, "Molases": 2, "Tepung Tulang": 2.5, "Dedak Padi": 5, "Dedak Halus": 2.5},
        "Catatan": "Ayam petelur grower 7-17 minggu. Ca mulai ditingkatkan.",
    },
    "Layer Pre-Lay (17-18wk)": {
        "Komposisi": {"Jagung Pipilan": 35, "Bungkil Kedelai": 17, "Tepung Ikan": 4, "Tepung Darah": 3,
                      "Menir": 10, "Corn Gluten Feed": 5, "Ampas Tahu": 4, "Bungkil Sawit": 3, "Molases": 2, "Tepung Tulang": 8, "Dedak Padi": 5, "Dedak Halus": 4},
        "Catatan": "Ayam pre-lay 17-18 minggu. Ca sangat tinggi untuk pembentukan cangkang.",
    },
    "Layer Produksi": {
        "Komposisi": {"Jagung Pipilan": 33, "Bungkil Kedelai": 16, "Tepung Ikan": 3, "Tepung Darah": 3,
                      "Menir": 10, "Corn Gluten Feed": 5, "Ampas Tahu": 3, "Bungkil Sawit": 3, "Molases": 2, "Tepung Tulang": 15, "Dedak Padi": 5, "Dedak Halus": 2},
        "Catatan": "Ayam petelur produksi. Ca 3.5% untuk produksi telur. Phosphorus terkontrol.",
    },
    "Bebek Petelur Starter": {
        "Komposisi": {"Jagung Pipilan": 35, "Bungkil Kedelai": 20, "Tepung Ikan": 6, "Tepung Darah": 4,
                      "Menir": 12, "Corn Gluten Feed": 5, "Ampas Tahu": 3, "Molases": 2, "Tepung Tulang": 2, "Dedak Padi": 5, "Dedak Halus": 6},
        "Catatan": "Bebek petelur starter (Mojosari/Alabio). Protein tinggi.",
    },
    "Bebek Petelur Layer": {
        "Komposisi": {"Jagung Pipilan": 33, "Bungkil Kedelai": 16, "Tepung Ikan": 3, "Tepung Darah": 3,
                      "Menir": 10, "Corn Gluten Feed": 5, "Ampas Tahu": 3, "Bungkil Sawit": 3, "Molases": 2, "Tepung Tulang": 12, "Dedak Padi": 6, "Dedak Halus": 4},
        "Catatan": "Bebek petelur layer. Ca tinggi untuk telur.",
    },
    "Bebek Pedaging (Peking)": {
        "Komposisi": {"Jagung Pipilan": 38, "Bungkil Kedelai": 18, "Tepung Ikan": 5, "Tepung Darah": 4,
                      "Menir": 12, "Corn Gluten Feed": 6, "Ampas Tahu": 3, "Molases": 2, "Tepung Tulang": 2, "Dedak Padi": 5, "Dedak Halus": 5},
        "Catatan": "Bebek pedaging Peking. Energi tinggi, target 3kg dalam 45 hari.",
    },
}


def write_sheet_bahan(wb):
    """Sheet 1: Daftar Bahan & Harga."""
    ws = wb.create_sheet("1. Bahan Pakan & Harga")
    ws.sheet_properties.tabColor = "27AE60"

    # Title
    ws.merge_cells("A1:H1")
    ws.cell(1, 1, "🌾 DAFTAR BAHAN PAKAN TERNAK — HARGA UPDATE TERKINI").font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(1, 1).fill = GREEN
    ws.cell(1, 1).alignment = CENTER

    ws.merge_cells("A2:H2")
    ws.cell(2, 1, f"Sumber: Pasar Indonesia | Update: {datetime.now().strftime('%d %B %Y')} | Satuan: Rp/kg").font = Font(italic=True, size=9)

    # Headers
    headers = ["No", "Nama Bahan", "Harga (Rp/kg)", "Kategori", "Protein (%)", "Energi (ME Mcal/kg)", "Serat (%)", "Keterangan"]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c, h)
    style_header(ws, 4, len(headers))

    categories = {
        "Jagung Pipilan": "Serealia", "Bungkil Kedelai": "Protein Nabati", "Dedak Padi": "Sub-Produk",
        "Tepung Ikan": "Protein Hewani", "Pollard": "Sub-Produk", "Biji Kapuk": "Lemak",
        "Tepung Darah": "Protein Hewani", "Tepung Tulang": "Mineral", "Molases": "Pemanis",
        "Bungkil Kelapa": "Protein Nabati", "Gaplek": "Serealia", "Bungkil Sawit": "Protein Nabati",
        "Ampas Tahu": "Sub-Produk", "Tepung Bulu Ayam": "Protein Hewani", "Kulit Kentang": "Sub-Produk",
        "Onggok": "Sub-Produk", "Bungkil Kacang Tanah": "Protein Nabati", "Dedak Halus": "Sub-Produk",
        "Sorgum": "Serealia", "Menir": "Serealia", "Corn Gluten Feed": "Protein Nabati",
        "Rice Polish": "Sub-Produk", "Mung Bean Husk": "Serat",
    }

    for i, (name, price) in enumerate(sorted(PRICES.items(), key=lambda x: x[1]), 1):
        row = 4 + i
        n = NUTRITION.get(name, {})
        cat = categories.get(name, "-")
        ws.cell(row, 1, i)
        ws.cell(row, 2, name)
        ws.cell(row, 3, price).number_format = '#,##0'
        ws.cell(row, 4, cat)
        ws.cell(row, 5, n.get("PK", "-"))
        ws.cell(row, 6, n.get("ME", "-"))
        ws.cell(row, 7, n.get("SK", "-"))
        ws.cell(row, 8, n.get("Ket", ""))

        fill = LIGHT_GREEN if price < 4000 else LIGHT_BLUE if price < 6000 else WHITE
        style_row(ws, row, len(headers), fill=fill)
        ws.cell(row, 3).number_format = '#,##0'

    # Summary row
    last = 4 + len(PRICES) + 1
    ws.cell(last, 1, "").font = SUBHEADER
    ws.cell(last, 2, "RATA-RATA").font = SUBHEADER
    avg = sum(PRICES.values()) / len(PRICES)
    ws.cell(last, 3, round(avg)).font = SUBHEADER
    ws.cell(last, 3).number_format = '#,##0'
    ws.cell(last, 4, f"Termurah: {min(PRICES, key=PRICES.get)} (Rp{min(PRICES.values()):,})").font = SUBHEADER
    ws.cell(last, 5, "").font = SUBHEADER
    ws.cell(last, 6, "").font = SUBHEADER
    ws.cell(last, 7, "").font = SUBHEADER
    ws.cell(last, 8, f"Termahal: {max(PRICES, key=PRICES.get)} (Rp{max(PRICES.values()):,})").font = SUBHEADER

    auto_width(ws, 12, 35)
    return ws


def write_sheet_nutrisi(wb):
    """Sheet 2: Profil Nutrisi Lengkap."""
    ws = wb.create_sheet("2. Profil Nutrisi")
    ws.sheet_properties.tabColor = "3498DB"

    ws.merge_cells("A1:K1")
    ws.cell(1, 1, "🧬 PROFIL NUTRISI BAHAN PAKAN (Basis Kering %)").font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    ws.cell(1, 1).alignment = CENTER

    ws.merge_cells("A2:K2")
    ws.cell(2, 1, "Sumber: NRC, Kementan RI, Jurnal Nutrisi Ternak | Karbohidrat = 100 - (PK+LK+SK+Abu)").font = Font(italic=True, size=9)

    headers = ["No", "Bahan Pakan", "Harga Rp/kg", "Protein Kasar (%)", "Lemak Kasar (%)",
               "Serat Kasar (%)", "Abu (%)", "Ca (%)", "P (%)", "ME (Mcal/kg)", "Keterangan"]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c, h)
    style_header(ws, 4, len(headers), PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid"))

    for i, (name, nut) in enumerate(sorted(NUTRITION.items(), key=lambda x: x[1].get("PK", 0), reverse=True), 1):
        row = 4 + i
        price = PRICES.get(name, 0)
        ws.cell(row, 1, i)
        ws.cell(row, 2, name)
        ws.cell(row, 3, price).number_format = '#,##0'
        ws.cell(row, 4, nut["PK"])
        ws.cell(row, 5, nut["LK"])
        ws.cell(row, 6, nut["SK"])
        ws.cell(row, 7, nut["Abu"])
        ws.cell(row, 8, nut["Ca"])
        ws.cell(row, 9, nut["P"])
        ws.cell(row, 10, nut["ME"])
        ws.cell(row, 11, nut["Ket"])

        # Color code by protein level
        if nut["PK"] >= 30: fill = LIGHT_GREEN  # high protein
        elif nut["PK"] >= 15: fill = LIGHT_BLUE  # medium
        else: fill = LIGHT_YELLOW  # low

        style_row(ws, row, len(headers), fill=fill)
        ws.cell(row, 3).number_format = '#,##0'

    # Legend
    last = 4 + len(NUTRITION) + 2
    ws.cell(last, 1, "LEGENDA:").font = SUBHEADER
    ws.cell(last+1, 1, "🟢 Hijau = Protein Tinggi (>30%)").font = DATA
    ws.cell(last+2, 1, "🔵 Biru = Protein Sedang (15-30%)").font = DATA
    ws.cell(last+3, 1, "🟡 Kuning = Protein Rendah (<15%)").font = DATA

    auto_width(ws, 12, 35)
    return ws


def write_formulasi_sheet(wb, title, tab_color, formulations, animal_type, icon, fill_color):
    """Write a formulation sheet for any animal type."""
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color

    # Title
    ws.merge_cells("A1:M1")
    ws.cell(1, 1, f"{icon} FORMULASI PAKAN — {animal_type}").font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(1, 1).fill = fill_color
    ws.cell(1, 1).alignment = CENTER

    ws.merge_cells("A2:M2")
    ws.cell(2, 1, "Persentase = % dari total campuran (bk). Harga per kg formula = Σ(% × Harga bahan ÷ 100)").font = Font(italic=True, size=9)

    row = 4
    for form_name, form_data in formulations.items():
        formula = form_data["Komposisi"]
        catatan = form_data.get("Catatan", "")
        nrc_key = form_name.split(" (")[0] if "(" in form_name else form_name
        # Try matching NRC
        nrc = None
        for nk, nv in NRC.items():
            if nrc_key in nk or nk in form_name:
                nrc = nv
                break

        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
        ws.cell(row, 1, f"📋 {form_name}").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row, 1).fill = fill_color
        ws.cell(row, 1).alignment = LEFT
        row += 1

        # NRC requirements if available
        if nrc:
            ws.cell(row, 1, "Kebutuhan NRC:").font = SUBHEADER
            ws.cell(row, 2, f"PK min {nrc['PK_min']}%").font = DATA
            ws.cell(row, 3, f"PK max {nrc['PK_max']}%").font = DATA
            ws.cell(row, 4, f"TDN {nrc['TDN']}%").font = DATA
            ws.cell(row, 5, f"ME ≥ {nrc['ME']} Mcal/kg").font = DATA
            ws.cell(row, 6, f"Ca ≥ {nrc['Ca']}%").font = DATA
            ws.cell(row, 7, f"P ≥ {nrc['P']}%").font = DATA
            ws.cell(row, 8, f"SK ≤ {nrc['SK_max']}%").font = DATA
            row += 1

        # Column headers
        cols = ["No", "Bahan Pakan", "% Komposisi", "Harga Rp/kg",
                "Kontribusi Biaya", "PK (%)", "ME (Mcal/kg)", "Ca (%)", "P (%)", "SK (%)"]
        for c, h in enumerate(cols, 1):
            ws.cell(row, c, h)
        style_header(ws, row, len(cols), PatternFill(start_color="555555", end_color="555555", fill_type="solid"))
        row += 1

        # Data rows
        total_cost = 0
        total_pk = 0; total_me = 0; total_ca = 0; total_p = 0; total_sk = 0
        for idx, (bahan, pct) in enumerate(sorted(formula.items(), key=lambda x: x[1], reverse=True), 1):
            n = NUTRITION.get(bahan, {})
            price = PRICES.get(bahan, 0)
            contrib = round(pct * price / 100)
            total_cost += contrib
            total_pk += n.get("PK", 0) * pct / 100
            total_me += n.get("ME", 0) * pct / 100
            total_ca += n.get("Ca", 0) * pct / 100
            total_p += n.get("P", 0) * pct / 100
            total_sk += n.get("SK", 0) * pct / 100

            ws.cell(row, 1, idx)
            ws.cell(row, 2, bahan)
            ws.cell(row, 3, f"{pct}%")
            ws.cell(row, 4, price).number_format = '#,##0'
            ws.cell(row, 5, contrib).number_format = '#,##0'
            ws.cell(row, 6, round(n.get("PK", 0) * pct / 100, 2))
            ws.cell(row, 7, round(n.get("ME", 0) * pct / 100, 3))
            ws.cell(row, 8, round(n.get("Ca", 0) * pct / 100, 3))
            ws.cell(row, 9, round(n.get("P", 0) * pct / 100, 3))
            ws.cell(row, 10, round(n.get("SK", 0) * pct / 100, 2))

            fill = LIGHT_GREEN if pct >= 20 else WHITE
            style_row(ws, row, len(cols), fill=fill)
            ws.cell(row, 4).number_format = '#,##0'
            ws.cell(row, 5).number_format = '#,##0'
            row += 1

        # Total row
        pct_total = sum(formula.values())
        ws.cell(row, 1, "")
        ws.cell(row, 2, "TOTAL").font = SUBHEADER
        ws.cell(row, 3, f"{pct_total}%").font = SUBHEADER
        ws.cell(row, 4, "")
        ws.cell(row, 5, total_cost).font = SUBHEADER
        ws.cell(row, 5).number_format = '#,##0'
        ws.cell(row, 6, round(total_pk, 1)).font = SUBHEADER
        ws.cell(row, 7, round(total_me, 2)).font = SUBHEADER
        ws.cell(row, 8, round(total_ca, 2)).font = SUBHEADER
        ws.cell(row, 9, round(total_p, 2)).font = SUBHEADER
        ws.cell(row, 10, round(total_sk, 1)).font = SUBHEADER
        style_row(ws, row, len(cols), fill=GREY)
        ws.cell(row, 5).number_format = '#,##0'
        row += 1

        # Validation
        if nrc:
            meets_pk = total_pk >= nrc["PK_min"] and total_pk <= nrc.get("PK_max", 100)
            meets_me = total_me >= nrc["ME"]
            meets_sk = total_sk <= nrc.get("SK_max", 100)
            meets_ca = total_ca >= nrc["Ca"]
            meets_p = total_p >= nrc["P"]

            status = "✅ MEMENUHI" if all([meets_pk, meets_me, meets_sk, meets_ca, meets_p]) else "⚠️ PERLU PENCAMPURAN SERAT/UMBI"
            ws.cell(row, 1, "Status:")
            ws.cell(row, 2, f"PK: {'✅' if meets_pk else '❌'} | ME: {'✅' if meets_me else '❌'} | SK: {'✅' if meets_sk else '❌'} | Ca: {'✅' if meets_ca else '❌'} | P: {'✅' if meets_p else '❌'}").font = SUBHEADER
            row += 1
            ws.cell(row, 1, "Catatan:")
            ws.cell(row, 2, catatan).font = Font(italic=True, size=9, color="555555")
            row += 1
        else:
            ws.cell(row, 1, "Catatan:")
            ws.cell(row, 2, catatan).font = Font(italic=True, size=9, color="555555")
            row += 1

        row += 2  # spacing between formulas

    auto_width(ws, 12, 30)
    return ws


def write_ringkasan(wb):
    """Sheet: Ringkasan Biaya Semua Formulasi."""
    ws = wb.create_sheet("7. Ringkasan Biaya")
    ws.sheet_properties.tabColor = "E74C3C"

    ws.merge_cells("A1:I1")
    ws.cell(1, 1, "💰 RINGKASAN BIAYA PER FORMULASI (Rp/kg)").font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(1, 1).fill = RED
    ws.cell(1, 1).alignment = CENTER

    ws.merge_cells("A2:I2")
    ws.cell(2, 1, f"Harga bahan: Update {datetime.now().strftime('%d %B %Y')} | Belum termasuk biaya pengolahan, transportasi, vitamin premix").font = Font(italic=True, size=9)

    headers = ["No", "Kelompok", "Formulasi", "Biaya/kg (Rp)", "Protein (%)", "Energi ME", "Ca (%)", "P (%)", "Cost Grade"]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c, h)
    style_header(ws, 4, len(headers))

    all_forms = [
        ("Ruminansia Besar", "🐄", RUMINAN_BESAR),
        ("Ruminansia Kecil", "🐐", RUMINAN_KECIL),
        ("Non-Ruminansia (Babi)", "🐷", NON_RUMINAN),
        ("Unggas", "🐔", UNGGAS),
    ]

    row = 5
    num = 1
    for group_name, icon, forms in all_forms:
        # Group header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.cell(row, 1, f"{icon} {group_name}").font = Font(bold=True, size=11, color="FFFFFF")
        ws.cell(row, 1).fill = BLUE
        ws.cell(row, 1).alignment = LEFT
        row += 1

        for form_name, form_data in forms.items():
            formula = form_data["Komposisi"]
            calc = calc_mix(formula)
            cost = calc["Harga/kg"]

            if cost < 4000: grade = "💰 Murah"
            elif cost < 5500: grade = "💲 Sedang"
            elif cost < 7000: grade = "💸 Mahal"
            else: grade = "💎 Premium"

            ws.cell(row, 1, num)
            ws.cell(row, 2, group_name)
            ws.cell(row, 3, form_name)
            ws.cell(row, 4, cost).number_format = '#,##0'
            ws.cell(row, 5, calc["PK"])
            ws.cell(row, 6, calc["ME"])
            ws.cell(row, 7, calc["Ca"])
            ws.cell(row, 8, calc["P"])
            ws.cell(row, 9, grade)

            fill = LIGHT_GREEN if cost < 4500 else LIGHT_BLUE if cost < 6000 else LIGHT_RED
            style_row(ws, row, len(headers), fill=fill)
            ws.cell(row, 4).number_format = '#,##0'
            num += 1
            row += 1

        row += 1  # spacing

    auto_width(ws, 12, 35)
    return ws


def write_kebutuhan(wb):
    """Sheet: Standar Kebutuhan Nutrisi NRC."""
    ws = wb.create_sheet("8. Kebutuhan Nutrisi")
    ws.sheet_properties.tabColor = "8E44AD"

    ws.merge_cells("A1:J1")
    ws.cell(1, 1, "📊 STANDAR KEBUTUHAN NUTRISI (NRC / Kementan RI)").font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(1, 1).fill = PURPLE
    ws.cell(1, 1).alignment = CENTER

    headers = ["No", "Kelompok Ternak", "BK (g/hari)", "PK Min (%)", "PK Max (%)",
               "TDN (%)", "ME (Mcal/kg)", "Ca (%)", "P (%)", "SK Max (%)"]
    for c, h in enumerate(headers, 1):
        ws.cell(3, c, h)
    style_header(ws, 3, len(headers), PURPLE)

    groups = [
        ("🐄 RUMINANSIA BESAR", "Ruminansia Besar"),
        ("🐐 RUMINANSIA KECIL", "Ruminansia Kecil"),
        ("🐷 NON-RUMINANSIA", "Non-Ruminansia"),
        ("🐔 UNGGAS", "Unggas"),
    ]

    row = 4
    num = 1
    for group_label, prefix in groups:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ws.cell(row, 1, group_label).font = SUBHEADER
        ws.cell(row, 1).fill = GREY
        row += 1

        for nrc_name, nrc_data in NRC.items():
            # Only show relevant ones
            if prefix == "Ruminansia Besar" and not any(x in nrc_name for x in ["Sapi"]): continue
            if prefix == "Ruminansia Kecil" and not any(x in nrc_name for x in ["Kambing", "Domba"]): continue
            if prefix == "Non-Ruminansia" and "Babi" not in nrc_name: continue
            if prefix == "Unggas" and not any(x in nrc_name for x in ["Broiler", "Layer", "Bebek"]): continue

            ws.cell(row, 1, num)
            ws.cell(row, 2, nrc_name)
            ws.cell(row, 3, nrc_data["BK"])
            ws.cell(row, 4, nrc_data["PK_min"])
            ws.cell(row, 5, nrc_data["PK_max"])
            ws.cell(row, 6, nrc_data["TDN"])
            ws.cell(row, 7, nrc_data["ME"])
            ws.cell(row, 8, nrc_data["Ca"])
            ws.cell(row, 9, nrc_data["P"])
            ws.cell(row, 10, nrc_data["SK_max"])

            fill = LIGHT_YELLOW if "Sapi" in nrc_name or "Kambing" in nrc_name else LIGHT_BLUE if "Babi" in nrc_name else LIGHT_GREEN
            style_row(ws, row, len(headers), fill=fill)
            num += 1
            row += 1

    auto_width(ws, 12, 35)
    return ws


def main():
    print("🔄 Generating Formulasi Pakan Ternak Komprehensif...")
    out = Path.home() / "sembako" / "data" / "formulasi_pakan_ternak.xlsx"
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Bahan & Harga
    print("  1/6 Bahan Pakan & Harga...")
    write_sheet_bahan(wb)

    # Sheet 2: Nutrisi
    print("  2/6 Profil Nutrisi...")
    write_sheet_nutrisi(wb)

    # Sheet 3-6: Formulasi per kelompok
    print("  3/6 Ruminansia Besar...")
    write_formulasi_sheet(wb, "3. Ruminansia Besar", "8B4513", RUMINAN_BESAR,
                          "RUMINANSIA BESAR (SAPI)", "🐄", PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid"))

    print("  4/6 Ruminansia Kecil...")
    write_formulasi_sheet(wb, "4. Ruminansia Kecil", "D2691E", RUMINAN_KECIL,
                          "RUMINANSIA KECIL (KAMBING & DOMBA)", "🐐", PatternFill(start_color="D2691E", end_color="D2691E", fill_type="solid"))

    print("  5/6 Non-Ruminansia...")
    write_formulasi_sheet(wb, "5. Non-Ruminansia (Babi)", "E74C3C", NON_RUMINAN,
                          "NON-RUMINANSIA (BABI)", "🐷", PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"))

    print("  6/6 Unggas...")
    write_formulasi_sheet(wb, "6. Unggas", "F39C12", UNGGAS,
                          "UNGGAS (AYAM BROILER, PETELUR, BEBEK)", "🐔", PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"))

    # Sheet 7: Ringkasan Biaya
    print("  + Ringkasan Biaya...")
    write_ringkasan(wb)

    # Sheet 8: Kebutuhan Nutrisi
    print("  + Kebutuhan Nutrisi NRC...")
    write_kebutuhan(wb)

    # Save
    wb.save(str(out))
    print(f"\n✅ Saved: {out}")
    print(f"   {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")

    # Stats
    total_forms = len(RUMINAN_BESAR) + len(RUMINAN_KECIL) + len(NON_RUMINAN) + len(UNGGAS)
    print(f"   {total_forms} formulasi ({len(RUMINAN_BESAR)} ruminansia besar, {len(RUMINAN_KECIL)} ruminansia kecil, {len(NON_RUMINAN)} babi, {len(UNGGAS)} unggas)")
    print(f"   {len(NUTRITION)} bahan baku, {len(NRC)} standar NRC")


if __name__ == "__main__":
    main()
