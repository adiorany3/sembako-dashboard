#!/usr/bin/env python3
"""
Update BI Rate & Inflasi (CPI) data for sembako dashboard.
Scrapes from Bank Indonesia and BPS via Jina Reader.
Saves to ~/sembako/data/bi_rate_inflasi.xlsx
"""
import re
import os
import sys
import json
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cache_lib import jina as jina_read_cached

EXCEL_PATH = os.path.expanduser("~/sembako/data/bi_rate_inflasi.xlsx")
CACHE_PATH = os.path.expanduser("~/sembako/data/bi_rate_cache.json")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Accept": "text/plain",
}


def jina_read(url, timeout=15):
    """Fetch URL via Jina Reader with cache."""
    text = jina_read_cached(url)
    return text if text else ""


def get_last_known():
    """Load last known BI rate + inflasi from cache."""
    if os.path.exists(CACHE_PATH):
        try:
            with open(CACHE_PATH) as f:
                d = json.load(f)
                return d
        except (json.JSONDecodeError, OSError):
            pass
    # Fallback defaults
    return {
        "bi_rate": 6.00,
        "inflasi_yoy": 2.31,
        "inflasi_mom": 0.15,
        "ihk": 108.50,
        "date": datetime.now().strftime("%Y-%m-%d"),
    }


def save_cache(data):
    """Persist BI rate + inflasi to cache."""
    with open(CACHE_PATH, "w") as f:
        json.dump(data, f)


def scrape_bi_rate():
    """
    Scrape BI Rate from Bank Indonesia website via Jina.
    Returns (rate_float, date_str) or None.
    """
    url = "https://www.bi.go.id/id/moneter/bi-rate/default.aspx"
    text = jina_read(url)

    if not text or len(text) < 100:
        print("  ⚠️ Gagal fetch BI rate dari BI.go.id")
        return None

    # Look for rate pattern: "6,00%" or "6.00%" near "BI Rate" or "Suku Bunga Acuan"
    # Try multiple patterns
    patterns = [
        r'(\d{1,2}[,\.]\d{2})\s*%',
        r'(?:rate|suku\s*bunga)\s*(?:acuan)?\s*[:=\s]*(\d{1,2}[,\.]\d{2})',
        r'(?:BI\s*Rate)\s*[:=\s]*(\d{1,2}[,\.]\d{2})',
    ]

    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            rate_str = m.group(1).replace(",", ".")
            try:
                rate = float(rate_str)
                if 1.0 <= rate <= 15.0:  # Sanity check
                    # Try to extract date
                    date_str = datetime.now().strftime("%Y-%m-%d")
                    date_m = re.search(
                        r'(\d{1,2})\s*(Januari|Februari|Maret|April|Mei|Juni|'
                        r'Juli|Agustus|September|Oktober|November|Desember)\s*(\d{4})',
                        text, re.I
                    )
                    if date_m:
                        month_map = {
                            "januari": 1, "februari": 2, "maret": 3, "april": 4,
                            "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
                            "september": 9, "oktober": 10, "november": 11, "desember": 12,
                        }
                        day = int(date_m.group(1))
                        month = month_map.get(date_m.group(2).lower(), datetime.now().month)
                        year = int(date_m.group(3))
                        date_str = f"{year}-{month:02d}-{day:02d}"
                    return (rate, date_str)
            except ValueError:
                continue

    print("  ⚠️ Tidak bisa parse BI Rate dari respons")
    return None


def scrape_inflasi():
    """
    Scrape CPI / Inflasi from BPS website via Jina.
    Returns dict with inflasi_yoy, inflasi_mom, ihk or None.
    """
    url = "https://www.bps.go.id/id/statistics-table/2/MTczMyMx/nilai-indeks-harga-konsumen--ikh--dan-laju-inflasi-bulanan--ihs-.html"
    text = jina_read(url)

    if not text or len(text) < 200:
        print("  ⚠️ Gagal fetch data inflasi dari BPS")
        return None

    result = {}

    # Try to parse inflation YoY: "X,XX%" pattern near "inflasi" or "year on year" or "yoy"
    # Look for "Laju Inflasi" sections
    joined = " ".join(text.split("\n"))

    # Inflasi YoY
    yoy_pat = re.search(
        r'(?:inflasi|year.{0,5}year|yoy)\s*[:=\s]*(\d{1,2}[,\.]\d{1,2})\s*%',
        joined, re.I
    )
    if yoy_pat:
        result["inflasi_yoy"] = float(yoy_pat.group(1).replace(",", "."))

    # Inflasi MoM
    mom_pat = re.search(
        r'(?:month.{0,5}month|mom|bulan\s*ke\s*bulan)\s*[:=\s]*(\d{1,2}[,\.]\d{1,2})\s*%',
        joined, re.I
    )
    if mom_pat:
        result["inflasi_mom"] = float(mom_pat.group(1).replace(",", "."))

    # IHK value - look for "indeks harga konsumen" near a number
    ihk_pat = re.search(
        r'(?:indeks\s*harga\s*konsumen|ikh)\s*[:=\s]*(\d{2,3}[,\.]\d{1,2})',
        joined, re.I
    )
    if ihk_pat:
        result["ihk"] = float(ihk_pat.group(1).replace(",", "."))

    # If no specific patterns found, try to grab any percentage that looks like inflation
    if not result.get("inflasi_yoy"):
        # Look for the first reasonable percentage in inflasi context
        all_pcts = re.findall(r'(\d{1,2}[,\.]\d{1,2})\s*%', joined)
        reasonable = [float(p.replace(",", ".")) for p in all_pcts if 0.5 <= float(p.replace(",", ".")) <= 10.0]
        if reasonable:
            result["inflasi_yoy"] = reasonable[0]

    if not result.get("inflasi_mom"):
        all_pcts = re.findall(r'(\d{1,2}[,\.]\d{1,2})\s*%', joined)
        reasonable = [float(p.replace(",", ".")) for p in all_pcts if -2.0 <= float(p.replace(",", ".")) <= 5.0]
        if len(reasonable) >= 2:
            result["inflasi_mom"] = reasonable[1]

    return result if result else None


def vary_pct(val, pct=0.05):
    """No variation - return base value."""
    return val


def save_to_excel(data):
    """
    Save BI rate & inflasi data to Excel.
    Sheet 'Harian': [Tanggal, BI_Rate, Inflasi_MoM, Inflasi_YoY, IHK, Sumber]
    """
    import openpyxl

    headers = ["Tanggal", "BI_Rate", "Inflasi_MoM", "Inflasi_YoY", "IHK", "Sumber"]

    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
    else:
        wb = openpyxl.Workbook()

    if "Harian" in wb.sheetnames:
        ws = wb["Harian"]
    else:
        ws = wb.active
        ws.title = "Harian"
        ws.append(headers)

    # Check if today already exists
    today = data["date"]
    existing_dates = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing_dates.append(str(row[0])[:10])

    row_data = [
        today,
        data["bi_rate"],
        data.get("inflasi_mom", ""),
        data.get("inflasi_yoy", ""),
        data.get("ihk", ""),
        data.get("sumber", "BI.go.id / BPS"),
    ]

    if today in existing_dates:
        # Update existing row
        for i, row in enumerate(ws.iter_rows(min_row=2)):
            if row[0].value and str(row[0].value)[:10] == today:
                for j, val in enumerate(row_data):
                    ws.cell(row=i + 2, column=j + 1, value=val)
                break
    else:
        ws.append(row_data)

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"  📁 Tersimpan ke {EXCEL_PATH}")


def main():
    """Update BI Rate & Inflasi data."""
    print("=" * 55)
    print("🏦 BI RATE & INFLASI (CPI)")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M WIB')}")
    print("=" * 55)

    last_known = get_last_known()
    today = datetime.now().strftime("%Y-%m-%d")
    data = {"date": today}

    # --- BI Rate ---
    print("\n🔍 Scraping BI Rate...")
    bi_result = scrape_bi_rate()
    if bi_result:
        data["bi_rate"] = bi_result[0]
        data["bi_date"] = bi_result[1]
        data["bi_source"] = "BI.go.id"
        print(f"  ✅ BI Rate ditemukan: {bi_result[0]}%")
    else:
        data["bi_rate"] = vary_pct(last_known["bi_rate"])
        data["bi_source"] = "fallback (last known)"
        print(f"  ⚡ Fallback BI Rate: {data['bi_rate']}%")

    # --- Inflasi ---
    print("\n🔍 Scraping data inflasi BPS...")
    inf_result = scrape_inflasi()
    if inf_result:
        data.update(inf_result)
        data["inflasi_source"] = "BPS"
        print(f"  ✅ Inflasi YoY: {inf_result.get('inflasi_yoy', 'N/A')}%")
        print(f"  ✅ Inflasi MoM: {inf_result.get('inflasi_mom', 'N/A')}%")
        if inf_result.get("ihk"):
            print(f"  ✅ IHK: {inf_result['ihk']}")
    else:
        data["inflasi_yoy"] = vary_pct(last_known["inflasi_yoy"])
        data["inflasi_mom"] = vary_pct(last_known["inflasi_mom"])
        data["ihk"] = vary_pct(last_known["ihk"])
        data["inflasi_source"] = "fallback (last known)"
        print(f"  ⚡ Fallback Inflasi YoY: {data['inflasi_yoy']}%")

    # Determine source
    if bi_result and inf_result:
        data["sumber"] = "BI.go.id + BPS"
    elif bi_result:
        data["sumber"] = "BI.go.id + fallback"
    elif inf_result:
        data["sumber"] = "fallback + BPS"
    else:
        data["sumber"] = "fallback (last known)"

    # --- Save ---
    save_to_excel(data)
    save_cache(data)

    # --- Summary ---
    print("\n" + "=" * 55)
    print(f"🏦 BI Rate: {data['bi_rate']:.2f}% | "
          f"Inflasi YoY: {data.get('inflasi_yoy', 'N/A')}%")
    if data.get("inflasi_mom"):
        print(f"   MoM: {data['inflasi_mom']}% | IHK: {data.get('ihk', 'N/A')}")
    print("=" * 55)


if __name__ == "__main__":
    main()
