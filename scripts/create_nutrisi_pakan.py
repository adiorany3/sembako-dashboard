#!/usr/bin/env python3
"""
Buat Excel Lengkap: Profil Nutrisi Bahan Pakan Ternak + Harga
==============================================================
Sheet 1: Profil Nutrisi Lengkap (protein, serat, lemak, energi, mineral, vitamin)
Sheet 2: Harga Bahan Pakan Hari Ini
Sheet 3: Kalkulator Formulasi Pakan
Sheet 4: Komposisi Pakan Siap Pakai (ayam broiler, layer, itik, sapi)
"""

import os
import sys
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

DATA_DIR = os.path.expanduser("~/sembako/data")
OUTPUT = os.path.join(DATA_DIR, "nutrisi_bahan_pakan_lengkap.xlsx")

# ==========================================
# DATA NUTRISI LENGKAP - 23 BAHAN PAKAN
# Sumber: NRC (National Research Council), IPB, Referensi Industri
# ==========================================
NUTRISI = [
    # (nama, protein%, serat%, lemak%, EM kkal/kg, Ca%, P%, Lys%, Met%, Treo%, VitA IU/kg, VitD3 IU/kg, VitE IU/kg, kalsiumfosfor_ratio)
    ("Jagung Pipilan",           8.5, 2.5, 3.8, 3350, 0.02, 0.28, 0.24, 0.18, 0.30, 400, 0, 11, "1:14"),
    ("Bungkil Kedelai",          44.0, 6.0, 1.5, 2400, 0.20, 0.60, 2.90, 0.60, 1.70, 0, 0, 3, "1:3"),
    ("Dedak Padi",               10.0, 10.0, 3.5, 2800, 0.05, 1.00, 0.40, 0.22, 0.45, 0, 0, 11, "1:20"),
    ("Tepung Ikan",              62.0, 1.0, 8.0, 2800, 3.50, 2.00, 4.80, 1.80, 2.70, 0, 0, 5, "7:4"),
    ("Pollard",                  13.5, 10.0, 4.0, 2500, 0.10, 1.20, 0.55, 0.25, 0.50, 0, 0, 14, "1:12"),
    ("Biji Kapuk",               20.0, 22.0, 20.0, 2100, 0.15, 0.40, 0.50, 0.20, 0.60, 0, 0, 20, "1:3"),
    ("Tepung Darah",             80.0, 1.0, 1.0, 2800, 0.20, 0.20, 7.50, 0.80, 3.50, 0, 0, 1, "1:1"),
    ("Tepung Tulang",            50.0, 2.0, 3.0, 1800, 24.0, 12.0, 2.00, 0.50, 1.50, 0, 0, 1, "2:1"),
    ("Molases",                  4.0, 0.5, 1.0, 2600, 0.80, 0.10, 0.20, 0.05, 0.15, 0, 0, 0, "8:1"),
    ("Bungkil Kelapa",           22.0, 12.0, 6.0, 2300, 0.15, 0.50, 0.70, 0.30, 0.80, 0, 0, 3, "1:3"),
    ("Gaplek",                   3.0, 4.0, 0.5, 3100, 0.05, 0.05, 0.10, 0.03, 0.15, 0, 0, 0, "1:1"),
    ("Bungkil Sawit",            18.0, 18.0, 8.0, 2200, 0.15, 0.40, 0.60, 0.25, 0.70, 0, 0, 10, "1:3"),
    ("Ampas Tahu",               12.0, 14.0, 4.0, 2000, 0.20, 0.30, 0.60, 0.15, 0.50, 0, 0, 0, "1:2"),
    ("Tepung Bulu Ayam",         85.0, 1.0, 3.0, 3000, 0.10, 0.10, 1.50, 0.30, 4.00, 0, 0, 1, "1:1"),
    ("Kulit Kentang",            10.0, 8.0, 0.5, 2500, 0.05, 0.15, 0.40, 0.10, 0.30, 0, 0, 0, "1:3"),
    ("Onggok",                   5.0, 20.0, 1.0, 1800, 0.10, 0.05, 0.15, 0.05, 0.20, 0, 0, 0, "2:1"),
    ("Bungkil Kacang Tanah",     45.0, 6.0, 8.0, 2800, 0.15, 0.55, 1.50, 0.45, 1.60, 0, 0, 5, "1:4"),
    ("Dedak Halus",              11.0, 8.0, 3.0, 2700, 0.05, 0.90, 0.45, 0.20, 0.40, 0, 0, 12, "1:18"),
    ("Sorgum",                   11.0, 3.0, 3.0, 3200, 0.03, 0.30, 0.30, 0.15, 0.35, 0, 0, 5, "1:10"),
    ("Menir",                    8.0, 1.0, 1.0, 3400, 0.02, 0.10, 0.25, 0.15, 0.30, 0, 0, 0, "1:5"),
    ("Corn Gluten Feed",         18.0, 8.0, 3.0, 2700, 0.05, 0.35, 0.50, 0.30, 0.60, 0, 0, 10, "1:7"),
    ("Rice Polish",              12.0, 3.0, 12.0, 3200, 0.05, 1.50, 0.45, 0.20, 0.40, 0, 0, 15, "1:30"),
    ("Mung Bean Husk",           14.0, 20.0, 2.0, 1900, 0.10, 0.20, 0.55, 0.15, 0.40, 0, 0, 5, "1:2"),
]

# ==========================================
# KATEGORI PAKAN SIAP PAKAI
# Sumber: Standar Komposisi Pakan Nasional, IPB, Kementan
# ==========================================
PAKAN_SIAP_PAKAI = {
    "Ayam Broiler Starter (0-21 h)": {
        "target_nutrisi": {"PK": "23%", "SK": "5%", "EM": "2950 kkal", "Ca": "1.0%", "P": "0.5%", "Lys": "1.30%", "Met": "0.55%"},
        "komposisi": {
            "Jagung Pipilan": 48, "Bungkil Kedelai": 28, "Tepung Ikan": 5,
            "Bungkil Sawit": 5, "Dedak Padi": 5, "Premix": 2, "CaCO3": 1, "DCP": 0.8,
            "Metionin": 0.15, "Lisin": 0.10, "Salt": 0.3, "Choline": 0.05,
        },
    },
    "Ayam Broiler Finisher (21-35 h)": {
        "target_nutrisi": {"PK": "21%", "SK": "4.5%", "EM": "3050 kkal", "Ca": "0.9%", "P": "0.45%", "Lys": "1.15%", "Met": "0.50%"},
        "komposisi": {
            "Jagung Pipilan": 55, "Bungkil Kedelai": 25, "Tepung Ikan": 3,
            "Bungkil Sawit": 5, "Dedak Padi": 5, "Premix": 2, "CaCO3": 0.9,
            "DCP": 0.8, "Metionin": 0.15, "Lisin": 0.10, "Salt": 0.3, "Choline": 0.05,
        },
    },
    "Ayam Layer (18-72 bln)": {
        "target_nutrisi": {"PK": "16.5%", "SK": "6%", "EM": "2700 kkal", "Ca": "3.5%", "P": "0.4%", "Lys": "0.75%", "Met": "0.35%"},
        "komposisi": {
            "Jagung Pipilan": 52, "Bungkil Kedelai": 18, "Dedak Padi": 10,
            "Bungkil Sawit": 5, "Tepung Tulang": 3, "CaCO3": 5, "Premix": 2,
            "DCP": 0.8, "Metionin": 0.15, "Lisin": 0.10, "Salt": 0.3, "Choline": 0.05,
        },
    },
    "Itik Pedaging (0-8 wk)": {
        "target_nutrisi": {"PK": "22%", "SK": "6%", "EM": "2850 kkal", "Ca": "0.9%", "P": "0.45%", "Lys": "1.20%", "Met": "0.50%"},
        "komposisi": {
            "Jagung Pipilan": 50, "Bungkil Kedelai": 26, "Tepung Ikan": 4,
            "Dedak Padi": 8, "Bungkil Sawit": 5, "Premix": 2, "CaCO3": 1,
            "DCP": 0.8, "Metionin": 0.12, "Lisin": 0.08, "Salt": 0.3,
        },
    },
    "Sapi Potong (growing)": {
        "target_nutrisi": {"PK": "14%", "SK": "25%", "EM": "2200 kkal", "Ca": "0.6%", "P": "0.3%"},
        "komposisi": {
            "Sorgum": 30, "Dedak Padi": 20, "Gaplek": 15, "Onggok": 10,
            "Bungkil Sawit": 10, "Molases": 8, "Urea": 2, "Premix": 1, "Salt": 0.5,
        },
    },
    "Sapi Perah (lactating)": {
        "target_nutrisi": {"PK": "17%", "SK": "20%", "EM": "2500 kkal", "Ca": "0.8%", "P": "0.4%"},
        "komposisi": {
            "Sorgum": 25, "Bungkil Kedelai": 15, "Dedak Padi": 18, "Gaplek": 10,
            "Bungkil Sawit": 12, "Molases": 8, "Tepung Ikan": 3, "Premix": 2,
            "CaCO3": 1, "DCP": 0.5, "Salt": 0.5,
        },
    },
}


def get_latest_prices():
    """Baca harga terbaru dari harga_pakan_ternak.xlsx"""
    path = os.path.join(DATA_DIR, "harga_pakan_ternak.xlsx")
    if not os.path.exists(path):
        return {}
    wb = load_workbook(path)
    ws = wb["Harga Pakan"]
    last_row = ws.max_row
    prices = {}
    for c in range(2, 25):
        name = ws.cell(1, c).value
        if name:
            name_clean = name.replace(chr(10), " ").strip()
            prices[name_clean] = ws.cell(last_row, c).value or 0
    return prices


def create_sheet_nutrisi(wb):
    """Sheet 1: Profil Nutrisi Lengkap"""
    ws = wb.active
    ws.title = "Profil Nutrisi"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2E7D32")
    cat_font = Font(bold=True, color="FFFFFF", size=10)
    cat_fill = PatternFill("solid", fgColor="558B2F")
    num_font = Font(size=10)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "🧬 PROFIL NUTRISI LENGKAP BAHAN PAKAN TERNAK"
    ws["A1"].font = Font(bold=True, size=16, color="1B5E20")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:P2")
    ws["A2"] = f"Sumber: NRC, IPB, Kementan RI | Updated: {date.today().strftime('%d %B %Y')}"
    ws["A2"].font = Font(size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers - Row 4
    headers = [
        "No", "Bahan Pakan", "Protein\nKasar (%)", "Serat\nKasar (%)", "Lemak\nKasar (%)",
        "Energi\nMetabolisme\n(Kkal/kg)", "Kalsium\n(Ca) (%)", "Fosfor\n(P) (%)",
        "Ca:P\nRatio", "Lisin\n(Lys) (%)", "Metionin\n(Met) (%)", "Treonin\n(%)",
        "Vitamin A\n(IU/kg)", "Vitamin D3\n(IU/kg)", "Vitamin E\n(IU/kg)", "Kategori"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Data rows
    kategori_map = {
        "Jagung Pipilan": "Sumber Energi",
        "Bungkil Kedelai": "Sumber Protein",
        "Dedak Padi": "Sumber Protein",
        "Tepung Ikan": "Protein Hewani",
        "Pollard": "Sumber Energi",
        "Biji Kapuk": "Protein Khusus",
        "Tepung Darah": "Protein Hewani",
        "Tepung Tulang": "Mineral",
        "Molases": "Pengikat & Energi",
        "Bungkil Kelapa": "Sumber Protein",
        "Gaplek": "Sumber Energi",
        "Bungkil Sawit": "Sumber Protein",
        "Ampas Tahu": "Sumber Protein",
        "Tepung Bulu Ayam": "Protein Hewani",
        "Kulit Kentang": "Sumber Energi",
        "Onggok": "Serat & Energi",
        "Bungkil Kacang Tanah": "Sumber Protein",
        "Dedak Halus": "Sumber Protein",
        "Sorgum": "Sumber Energi",
        "Menir": "Sumber Energi",
        "Corn Gluten Feed": "Sumber Protein",
        "Rice Polish": "Sumber Energi",
        "Mung Bean Husk": "Serat",
    }

    for i, row_data in enumerate(NUTRISI):
        r = 5 + i
        nama = row_data[0]
        ws.cell(r, 1, i + 1).font = num_font
        ws.cell(r, 1).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, nama).font = Font(bold=True, size=10)
        
        # Nutritional values
        for j, val in enumerate(row_data[1:], 3):
            cell = ws.cell(r, j, val)
            cell.font = num_font
            cell.alignment = Alignment(horizontal="center")
            cell.number_format = '#,##0.00' if isinstance(val, float) else '#,##0'
        
        # Ca:P ratio
        ws.cell(r, 8, row_data[11])  # already string
        ws.cell(r, 8).alignment = Alignment(horizontal="center")
        
        # Kategori
        kategori = kategori_map.get(nama, "-")
        ws.cell(r, 16, kategori)
        ws.cell(r, 16).font = Font(size=10, italic=True)
        
        # Borders
        for c in range(1, 17):
            ws.cell(r, c).border = border

    # Column widths
    widths = [5, 22, 10, 10, 10, 14, 10, 10, 8, 10, 10, 10, 10, 10, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes
    ws.freeze_panes = "C5"
    return ws


def create_sheet_harga(wb, prices):
    """Sheet 2: Harga Bahan Pakan Hari Ini + Nutrisi Ringkas"""
    ws = wb.create_sheet("Harga & Nutrisi")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1565C0")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "💰 HARGA BAHAN PAKAN + PROFIL NUTRISI"
    ws["A1"].font = Font(bold=True, size=14, color="0D47A1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Per: {prices.get('Tanggal', date.today().isoformat())} | Sumber: Referensi Industri"
    ws["A2"].font = Font(size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["No", "Bahan Pakan", "Harga\n(Rp/kg)", "Protein (%)", "Energi\n(Kkal/kg)",
               "Harga/kg Protein\n(Rp)", "Harga/100 Kkal\n(Rp)", "Kategori", "Efisiensi"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Build price lookup by normalized name
    price_lookup = {}
    for k, v in prices.items():
        if k != "Tanggal":
            price_lookup[k.lower()] = v

    for i, nd in enumerate(NUTRISI):
        r = 5 + i
        nama = nd[0]
        protein = nd[1]
        em = nd[4]
        
        # Find matching price
        harga = 0
        for pk, pv in price_lookup.items():
            if nama.lower() in pk or pk in nama.lower():
                harga = pv
                break
        
        # Cost efficiency
        harga_per_protein = round(harga / protein, 0) if protein > 0 else 0
        harga_per_100kcal = round(harga / (em / 100), 0) if em > 0 else 0

        # Rating
        rating = "⭐⭐⭐"
        if protein >= 40:
            rating = "⭐⭐⭐⭐⭐"
        elif protein >= 20:
            rating = "⭐⭐⭐⭐"
        elif em >= 3200:
            rating = "⭐⭐⭐⭐"
        elif protein < 8:
            rating = "⭐⭐"

        kategori_map = {
            8.5: "Energi", 44: "Protein", 10: "Protein", 62: "Protein HW",
            13.5: "Energi", 20: "Protein", 80: "Protein HW", 50: "Mineral",
            4: "Pengikat", 22: "Protein", 3: "Energi", 18: "Protein",
            12: "Protein", 85: "Protein HW", 10: "Energi", 5: "Serat",
            45: "Protein", 11: "Protein", 11: "Energi", 8: "Energi",
            18: "Protein", 12: "Energi", 14: "Serat",
        }
        kategori = ""
        for k_p, k_t in kategori_map.items():
            if abs(k_p - protein) < 0.01:
                kategori = k_t
                break
        if not kategori:
            if protein >= 40: kategori = "Protein"
            elif em >= 3200: kategori = "Energi"
            elif protein >= 15: kategori = "Protein"

        ws.cell(r, 1, i+1).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, nama).font = Font(bold=True, size=10)
        ws.cell(r, 3, harga).number_format = '#,##0'
        ws.cell(r, 4, protein).number_format = '#,##0.0'
        ws.cell(r, 5, em).number_format = '#,##0'
        ws.cell(r, 6, harga_per_protein).number_format = '#,##0'
        ws.cell(r, 7, harga_per_100kcal).number_format = '#,##0'
        ws.cell(r, 8, kategori)
        ws.cell(r, 9, rating).alignment = Alignment(horizontal="center")
        
        # Conditional coloring for efficiency
        if harga_per_protein > 0 and harga_per_protein < 200:
            ws.cell(r, 6).fill = PatternFill("solid", fgColor="C8E6C9")  # Green
        elif harga_per_protein > 500:
            ws.cell(r, 6).fill = PatternFill("solid", fgColor="FFCDD2")  # Red
        
        for c in range(1, 10):
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = Alignment(horizontal="center") if c != 2 else Alignment()

    # Summary row
    r_sum = 5 + len(NUTRISI) + 1
    ws.merge_cells(f"A{r_sum}:B{r_sum}")
    ws[f"A{r_sum}"] = "📊 REKOMENDASI Bahan Pak (Harga/Protein terbaik):"
    ws[f"A{r_sum}"].font = Font(bold=True, size=11, color="1B5E20")

    # Top 5 cheapest protein
    protein_costs = []
    for i, nd in enumerate(NUTRISI):
        nama = nd[0]
        protein = nd[1]
        harga = 0
        for pk, pv in price_lookup.items():
            if nama.lower() in pk or pk in nama.lower():
                harga = pv
                break
        if protein > 0 and harga > 0:
            protein_costs.append((nama, harga/protein, protein, harga))
    protein_costs.sort(key=lambda x: x[1])

    for j, (nama, cost, prot, hrg) in enumerate(protein_costs[:5]):
        ws.cell(r_sum + 1 + j, 1, j + 1)
        ws.cell(r_sum + 1 + j, 2, f"{nama}")
        ws.cell(r_sum + 1 + j, 3, f"Rp{hrg:,}/kg ({prot:.0f}% PK) = Rp{cost:,.0f}/kg protein")
        ws.cell(r_sum + 1 + j, 2).font = Font(bold=True, color="2E7D32")

    # Column widths
    for i, w in enumerate([5, 22, 12, 10, 12, 14, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "C5"

    # Bar chart - Harga per kg Protein
    chart = BarChart()
    chart.title = "Harga per kg Protein (Rp)"
    chart.style = 10
    chart.width = 30
    chart.height = 15
    
    data_ref = Reference(ws, min_col=6, min_row=4, max_row=4+len(NUTRISI))
    cats_ref = Reference(ws, min_col=2, min_row=5, max_row=4+len(NUTRISI))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    ws.add_chart(chart, f"A{r_sum + 8}")

    return ws


def create_sheet_formulasi(wb):
    """Sheet 3: Kalkulator Formulasi Pakan"""
    ws = wb.create_sheet("Formulasi Pakan")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="6A1B9A")

    ws.merge_cells("A1:L1")
    ws["A1"] = "🧪 FORMULASI PAKAN SIAP PAKAI"
    ws["A1"].font = Font(bold=True, size=14, color="4A148C")
    ws["A1"].alignment = Alignment(horizontal="center")

    r = 3
    for nama_pakan, data in PAKAN_SIAP_PAKAI.items():
        # Pakan name header
        ws.merge_cells(f"A{r}:L{r}")
        ws[f"A{r}"] = f"📋 {nama_pakan}"
        ws[f"A{r}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{r}"].fill = PatternFill("solid", fgColor="6A1B9A")
        r += 1

        # Target nutrisi
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"] = "🎯 Target:"
        ws[f"A{r}"].font = Font(bold=True, size=10)
        target_text = " | ".join([f"{k}: {v}" for k, v in data["target_nutrisi"].items()])
        ws.merge_cells(f"C{r}:L{r}")
        ws[f"C{r}"] = target_text
        ws[f"C{r}"].font = Font(size=10, italic=True)
        r += 1

        # Komposisi table
        sub_headers = ["Bahan Pakan", "Komposisi (%)", "Harga/kg (Rp)", "Bobot Harga"]
        for col, h in enumerate(sub_headers, 1):
            cell = ws.cell(r, col, h)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="8E24AA")
        r += 1

        total_pct = 0
        total_cost = 0
        for bahan, pct in data["komposisi"].items():
            ws.cell(r, 1, bahan)
            ws.cell(r, 2, pct)
            ws.cell(r, 2).number_format = '0.0'
            # Approximate cost
            harga = 0
            for k, v in data["komposisi"].items():
                if bahan == k:
                    pass
            ws.cell(r, 4, f"Rp{pct * 50:,.0f}")  # rough estimate
            total_pct += pct
            r += 1

        ws.cell(r, 1, "TOTAL").font = Font(bold=True)
        ws.cell(r, 2, total_pct).font = Font(bold=True)
        ws.cell(r, 2).number_format = '0.0'
        r += 2

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 15

    return ws


def create_sheet_perbandingan(wb):
    """Sheet 4: Perbandingan Bahan Pak per Kategori"""
    ws = wb.create_sheet("Perbandingan")

    ws.merge_cells("A1:G1")
    ws["A1"] = "📊 PERBANDINGAN BAHAN PAK PER KATEGORI"
    ws["A1"].font = Font(bold=True, size=14, color="E65100")
    ws["A1"].alignment = Alignment(horizontal="center")

    categories = {
        "🌾 Sumber Energi (EM > 3000)": ["Jagung Pipilan", "Menir", "Sorgum", "Gaplek", "Rice Polish"],
        "🥜 Sumber Protein (> 30%)": ["Bungkil Kedelai", "Tepung Ikan", "Bungkil Kacang Tanah"],
        "🥩 Protein Hewani (> 50%)": ["Tepung Darah", "Tepung Bulu Ayam", "Tepung Tulang"],
        "🌿 Bahan Serat (> 15%)": ["Onggok", "Biji Kapuk", "Mung Bean Husk", "Bungkil Sawit"],
    }

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="E65100")

    r = 3
    for cat_name, items in categories.items():
        ws.merge_cells(f"A{r}:G{r}")
        ws[f"A{r}"] = cat_name
        ws[f"A{r}"].font = Font(bold=True, size=12, color="BF360C")
        r += 1

        headers = ["Bahan", "Protein (%)", "Serat (%)", "Lemak (%)", "EM (Kkal)", "Ca (%)", "P (%)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(r, col, h)
            cell.font = header_font
            cell.fill = header_fill
        r += 1

        for item_name in items:
            for nd in NUTRISI:
                if nd[0] == item_name:
                    ws.cell(r, 1, nd[0]).font = Font(bold=True, size=10)
                    ws.cell(r, 2, nd[1])
                    ws.cell(r, 3, nd[2])
                    ws.cell(r, 4, nd[3])
                    ws.cell(r, 5, nd[4])
                    ws.cell(r, 6, nd[5])
                    ws.cell(r, 7, nd[6])
                    for c in range(1, 8):
                        ws.cell(r, c).border = Border(
                            left=Side(style="thin"), right=Side(style="thin"),
                            top=Side(style="thin"), bottom=Side(style="thin"),
                        )
                    r += 1
                    break
        r += 1

    for i, w in enumerate([22, 10, 10, 10, 12, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


def main():
    print("=" * 60)
    print("🧬 NUTRISI BAHAN PAKAN TERNAK - EXCEL GENERATOR")
    print(f"📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 60)

    # Get latest prices
    print("\n💰 Membaca harga bahan pakan...")
    prices = get_latest_prices()
    if prices:
        print(f"  Harga terbaru: {len(prices)-1} bahan")
    else:
        print("  ⚠️ Harga belum tersedia")

    # Create workbook
    print("\n📊 Membuat Excel...")
    wb = Workbook()

    ws1 = create_sheet_nutrisi(wb)
    print("  ✅ Sheet 1: Profil Nutrisi Lengkap")

    ws2 = create_sheet_harga(wb, prices)
    print("  ✅ Sheet 2: Harga & Nutrisi")

    ws3 = create_sheet_formulasi(wb)
    print("  ✅ Sheet 3: Formulasi Pakan")

    ws4 = create_sheet_perbandingan(wb)
    print("  ✅ Sheet 4: Perbandingan Kategori")

    # Save
    wb.save(OUTPUT)
    size = os.path.getsize(OUTPUT)
    print(f"\n💾 Tersimpan: {OUTPUT}")
    print(f"   Size: {size:,} bytes")
    print(f"\n{'='*60}")
    print("Ringkasan:")
    print(f"  • 23 bahan pakan + profil nutrisi lengkap")
    print(f"  • Harga terkini + biaya per kg protein")
    print(f"  • 6 formulasi pakan (ayam broiler/layer, itik, sapi)")
    print(f"  • Rekomendasi bahan terbaik berdasarkan harga/nutrisi")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
