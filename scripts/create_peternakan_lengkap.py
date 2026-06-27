#!/usr/bin/env python3
"""
Create comprehensive livestock product data (hulu to hilir)
Hulu (Input) → Industri (Process) → Hilir (Jual)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def create_peternakan_excel():
    wb = openpyxl.Workbook()
    
    # Style definitions
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    section_fill = PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid')
    section_font = Font(bold=True, color='FFFFFF', size=12)
    date_fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
    date_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============================================
    # SHEET 1: DATA UTAMA (Summary)
    # ============================================
    ws1 = wb.active
    ws1.title = "Data Utama"
    
    # Headers
    headers = ['Tanggal', 'Kategori', 'Sub Kategori', 'Produk', 'Harga (Rp)', 'Satuan', 'Sumber', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Sample data (hulu to hilir)
    today = datetime.now().strftime('%Y-%m-%d')
    sample_data = [
        # HULU - Bibit
        (today, 'HULU', 'Bibit', 'DOC Ayam Ras (Super)', 4500, 'Ekor', 'Peternakan Lokal', 'Aktif'),
        (today, 'HULU', 'Bibit', 'DOC Ayam Kampung', 7500, 'Ekor', 'Peternakan Lokal', 'Aktif'),
        (today, 'HULU', 'Bibit', 'Bibit Itik Manila', 12000, 'Ekor', 'Peternakan Lokal', 'Aktif'),
        (today, 'HULU', 'Bibit', 'Bibit Kambing PE', 1500000, 'Ekor', 'Peternakan Lokal', 'Aktif'),
        (today, 'HULU', 'Bibit', 'Bibit Sapi Simental', 25000000, 'Ekor', 'Peternakan Lokal', 'Aktif'),
        # HULU - Pakan
        (today, 'HULU', 'Pakan', 'Konsentrat Broiler', 280000, '50 Kg', 'Pabrik Pakan', 'Aktif'),
        (today, 'HULU', 'Pakan', 'Konsentrat Layer', 265000, '50 Kg', 'Pabrik Pakan', 'Aktif'),
        (today, 'HULU', 'Pakan', 'Dedak Padi', 4500, 'Kg', 'Penggilingan', 'Aktif'),
        (today, 'HULU', 'Pakan', 'Jagung Pipilan', 8500, 'Kg', 'Pasar Ternak', 'Aktif'),
        (today, 'HULU', 'Pakan', 'Bungkil Kedelai', 12000, 'Kg', 'Pabrik', 'Aktif'),
        # HULU - Obat & Vitamin
        (today, 'HULU', 'Obat', 'Vitamin Mix (1 Pack)', 85000, 'Pack', 'Apotek Veteriner', 'Aktif'),
        (today, 'HULU', 'Obat', 'Antibiotik Ayam (Vial)', 45000, 'Vial', 'Apotek Veteriner', 'Aktif'),
        (today, 'HULU', 'Obat', 'Vaccine ND (1000 Dosis)', 175000, 'Box', 'Apotek Veteriner', 'Aktif'),
        (today, 'HULU', 'Obat', 'Obat Cacing Kambing', 25000, 'Tablet', 'Apotek Veteriner', 'Aktif'),
        # INDUSTRI - Pemotongan
        (today, 'INDUSTRI', 'Pemotongan', 'Jasa Pemotongan Ayam', 3000, 'Ekor', 'Rumah Potong', 'Aktif'),
        (today, 'INDUSTRI', 'Pemotongan', 'Jasa Pemotongan Sapi', 350000, 'Ekor', 'Rumah Potong', 'Aktif'),
        (today, 'INDUSTRI', 'Pemotongan', 'Jasa Pemotongan Kambing', 75000, 'Ekor', 'Rumah Potong', 'Aktif'),
        # INDUSTRI - Cold Storage
        (today, 'INDUSTRI', 'Cold Storage', 'Penyimpanan Cold Storage', 5000, 'Kg/Hari', 'Gudang', 'Aktif'),
        (today, 'INDUSTRI', 'Cold Storage', 'Sewa Cold Box', 150000, 'Hari', 'Rental', 'Aktif'),
        # INDUSTRI - Pengemasan
        (today, 'INDUSTRI', 'Pengemasan', 'Kardus Ayam (50 Ekor)', 15000, 'Pcs', 'Supplier', 'Aktif'),
        (today, 'INDUSTRI', 'Pengemasan', 'Plastik Vakum (10 Kg)', 8000, 'Roll', 'Supplier', 'Aktif'),
        (today, 'INDUSTRI', 'Pengemasan', 'Label Harga (1000 Pcs)', 50000, 'Roll', 'Supplier', 'Aktif'),
        # HILIR - Harga Panen
        (today, 'HILIR', 'Harga Panen', 'Ayam Broiler Hidup', 32000, 'Kg', 'Peternak', 'Aktif'),
        (today, 'HILIR', 'Harga Panen', 'Ayam Kampung Hidup', 45000, 'Kg', 'Peternak', 'Aktif'),
        (today, 'HILIR', 'Harga Panen', 'Itik Hidup', 38000, 'Kg', 'Peternak', 'Aktif'),
        (today, 'HILIR', 'Harga Panen', 'Kambing Hidup', 65000, 'Kg', 'Peternak', 'Aktif'),
        (today, 'HILIR', 'Harga Panen', 'Sapi Potong Hidup', 55000, 'Kg', 'Peternak', 'Aktif'),
        # HILIR - Harga Grosir
        (today, 'HILIR', 'Grosir', 'Daging Ayam Segar', 38000, 'Kg', 'Pasar Grosir', 'Aktif'),
        (today, 'HILIR', 'Grosir', 'Daging Sapi Murni', 135000, 'Kg', 'Pasar Grosir', 'Aktif'),
        (today, 'HILIR', 'Grosir', 'Daging Kambing', 120000, 'Kg', 'Pasar Grosir', 'Aktif'),
        (today, 'HILIR', 'Grosir', 'Telur Ayam Ras', 28000, 'Kg', 'Pasar Grosir', 'Aktif'),
        (today, 'HILIR', 'Grosir', 'Telur Ayam Kampung', 42000, 'Kg', 'Pasar Grosir', 'Aktif'),
        # HILIR - Harga Eceran
        (today, 'HILIR', 'Eceran', 'Daging Ayam Eceran', 42000, 'Kg', 'Pasar Tradisional', 'Aktif'),
        (today, 'HILIR', 'Eceran', 'Daging Sapi Eceran', 150000, 'Kg', 'Pasar Tradisional', 'Aktif'),
        (today, 'HILIR', 'Eceran', 'Daging Kambing Eceran', 135000, 'Kg', 'Pasar Tradisional', 'Aktif'),
        (today, 'HILIR', 'Eceran', 'Telur Ayam Ras Eceran', 30000, 'Kg', 'Pasar Tradisional', 'Aktif'),
        (today, 'HILIR', 'Eceran', 'Telur Ayam Kampung Eceran', 48000, 'Kg', 'Pasar Tradisional', 'Aktif'),
        # HILIR - Modern Retail
        (today, 'HILIR', 'Modern Retail', 'Daging Ayam (Supermarket)', 48000, 'Kg', 'Superindo', 'Aktif'),
        (today, 'HILIR', 'Modern Retail', 'Daging Sapi (Supermarket)', 165000, 'Kg', 'Hypermart', 'Aktif'),
        (today, 'HILIR', 'Modern Retail', 'Telur Ayam Ras (Supermarket)', 32000, 'Kg', 'Carrefour', 'Aktif'),
        # ANALISIS - Margin
        (today, 'ANALISIS', 'Margin', 'Margin Ayam Broiler (Peternak→Pasar)', 6000, 'Kg', 'Kalkulasi', 'Aktif'),
        (today, 'ANALISIS', 'Margin', 'Margin Daging Sapi (Grosir→Eceran)', 15000, 'Kg', 'Kalkulasi', 'Aktif'),
        (today, 'ANALISIS', 'Margin', 'Margin Telur (Peternak→Eceran)', 2000, 'Kg', 'Kalkulasi', 'Aktif'),
        (today, 'ANALISIS', 'Margin', 'Selisih Grosir vs Modern Retail', 8000, 'Kg', 'Kalkulasi', 'Aktif'),
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 1:  # Date column
                cell.fill = date_fill
                cell.font = date_font
                cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 30
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 10
    ws1.column_dimensions['G'].width = 20
    ws1.column_dimensions['H'].width = 10
    
    # ============================================
    # SHEET 2: HULU (Input/Production)
    # ============================================
    ws2 = wb.create_sheet("Hulu")
    
    headers_hulu = ['Tanggal', 'Sub Kategori', 'Produk', 'Harga (Rp)', 'Satuan', 'Sumber', 'Notes']
    for col, header in enumerate(headers_hulu, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    hulu_data = [
        # Bibit
        (today, 'Bibit', 'DOC Ayam Ras (Super)', 4500, 'Ekor', 'Cibitung Farm', ' strain Lohmann'),
        (today, 'Bibit', 'DOC Ayam Ras (Medium)', 3500, 'Ekor', 'Cibitung Farm', ' strain ISA'),
        (today, 'Bibit', 'DOC Ayam Kampung', 7500, 'Ekor', 'Peternak Lokal', 'Umur 1 hari'),
        (today, 'Bibit', 'DOC Itik Mojosari', 8000, 'Ekor', 'BPPT', 'Produktif'),
        (today, 'Bibit', 'Bibit Itik Manila', 12000, 'Ekor', 'Peternak Lokal', 'Umur 2 bulan'),
        (today, 'Bibit', 'Bibit Kambing Boer', 1800000, 'Ekor', 'Peternak Kakao', 'Pejantan'),
        (today, 'Bibit', 'Bibit Kambing PE', 1500000, 'Ekor', 'BPTU Htips', 'Betina'),
        (today, 'Bibit', 'Bibit Sapi Limousin', 28000000, 'Ekor', 'BPTU Htips', 'Jantan'),
        (today, 'Bibit', 'Bibit Sapi Simental', 25000000, 'Ekor', 'BPTU Htips', 'Jantan'),
        (today, 'Bibit', 'Bibit Sapi Brahman', 30000000, 'Ekor', 'BPTU Htips', 'Jantan Cross'),
        # Pakan
        (today, 'Pakan', 'BR 1 (Broiler Starter)', 280000, '50 Kg', 'Charoen Pokphand', '23% protein'),
        (today, 'Pakan', 'BR 2 (Broiler Finisher)', 275000, '50 Kg', 'Charoen Pokphand', '21% protein'),
        (today, 'Pakan', 'L 1 (Layer Starter)', 265000, '50 Kg', 'Charoen Pokphand', '18% protein'),
        (today, 'Pakan', 'L 3 (Layer Finisher)', 260000, '50 Kg', 'Charoen Pokphand', '16% protein'),
        (today, 'Pakan', 'Dedak Padi Halus', 4500, 'Kg', 'Penggilingan Padi', 'Kadar air 13%'),
        (today, 'Pakan', 'Dedak Padi Kasar', 3500, 'Kg', 'Penggilingan Padi', 'Kadar air 14%'),
        (today, 'Pakan', 'Jagung Pipilan', 8500, 'Kg', 'Pasar Ternak', 'Kadar air 14%'),
        (today, 'Pakan', 'Jagung Halus', 9000, 'Kg', 'Pasar Ternak', ' untuk starter'),
        (today, 'Pakan', 'Bungkil Kelapa', 9500, 'Kg', 'Pabrik Kelapa', '22% protein'),
        (today, 'Pakan', 'Bungkil Kedelai', 12000, 'Kg', 'Pabrik', '44% protein'),
        (today, 'Pakan', 'Tepung Ikan', 35000, 'Kg', 'Pabrik', '60% protein'),
        (today, 'Pakan', 'Tepung Darah', 15000, 'Kg', 'Rumah Potong', '80% protein'),
        (today, 'Pakan', 'Palm Kernel Cake', 4500, 'Kg', 'Pabrik CPO', '15% protein'),
        # Obat & Vitamin
        (today, 'Obat', 'Vitamin Mix Poultry', 85000, 'Kg', 'Novell Pharma', '1 pack = 500g'),
        (today, 'Obat', 'Vitamin B Complex', 50000, '100 ml', 'Medical Est', 'Injeksi'),
        (today, 'Obat', 'Antibiotik Enroxy', 45000, '100 ml', 'Novell Pharma', 'Enrofloxacin'),
        (today, 'Obat', 'Antibiotik Colibosin', 55000, '100 ml', 'Medical Est', 'Colistin'),
        (today, 'Obat', 'Vaccine Newcastle (Clone 30)', 175000, '1000 Dosis', 'Medion', 'Tetes mata'),
        (today, 'Obat', 'Vaccine Gumboro', 225000, '1000 Dosis', 'Medion', ' Drinking water'),
        (today, 'Obat', 'Vaccine AI', 350000, '500 Dosis', 'Medion', 'H5N1'),
        (today, 'Obat', 'Obat Cacing Worminex', 25000, 'Strip', 'Medical Est', 'Untuk kambing'),
        (today, 'Obat', 'Obat Kutu/Sekrep', 35000, '100 ml', 'Medical Est', 'Spray'),
        (today, 'Obat', 'Desinfektan Sterilari', 125000, 'Liter', 'Bio Security', 'Farm sanitizer'),
    ]
    
    for row_idx, row_data in enumerate(hulu_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.fill = date_fill
                cell.font = date_font
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws2.column_dimensions[col].width = 18 if col in ['A', 'F'] else 25 if col == 'C' else 15
    
    # ============================================
    # SHEET 3: INDUSTRI (Processing)
    # ============================================
    ws3 = wb.create_sheet("Industri")
    
    headers_industri = ['Tanggal', 'Sub Kategori', 'Produk/Jasa', 'Harga (Rp)', 'Satuan', 'Lokasi', 'Kapasitas', 'Notes']
    for col, header in enumerate(headers_industri, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center')
    
    industri_data = [
        # Pemotongan
        (today, 'Pemotongan', 'Jasa Pemotongan Ayam Broiler', 3000, 'Ekor', 'RPH Surabaya', '5000 ekor/hari', ' termasuk bulu'),
        (today, 'Pemotongan', 'Jasa Pemotongan Ayam Kampung', 5000, 'Ekor', 'RPH Surabaya', '500 ekor/hari', ' termasuk bulu'),
        (today, 'Pemotongan', 'Jasa Pemotongan Itik', 4500, 'Ekor', 'RPH Surabaya', '300 ekor/hari', ' termasuk bulu'),
        (today, 'Pemotongan', 'Jasa Pemotongan Kambing', 75000, 'Ekor', 'RPH Surabaya', '100 ekor/hari', ' lengkap'),
        (today, 'Pemotongan', 'Jasa Pemotongan Sapi', 350000, 'Ekor', 'RPH Surabaya', '50 ekor/hari', ' lengkap'),
        (today, 'Pemotongan', 'Jasa Pemotongan Sapi (Kecil)', 250000, 'Ekor', 'RPH Sidoarjo', '30 ekor/hari', ' untuk兽医'),
        # Cold Chain
        (today, 'Cold Storage', 'Sewa Cold Storage (Short Term)', 5000, 'Kg/Hari', 'Gudang dingin Surabaya', '100 ton', 'Min 1 ton'),
        (today, 'Cold Storage', 'Sewa Cold Storage (Monthly)', 100000, 'Ton/Bulan', 'Gudang dingin Surabaya', '100 ton', 'Long term'),
        (today, 'Cold Storage', 'Sewa Cold Box (Portable)', 150000, 'Hari', 'Rental Alat', '500 L', ' termasuk es'),
        (today, 'Cold Storage', 'Biaya Transportasi Cold Chain', 500, 'Kg/Km', 'Logistik', '-', 'Min 100 kg'),
        # Pengolahan
        (today, 'Pengolahan', 'Marge (Boneless Ayam)', 18000, 'Kg', 'Processing Plant', 'Per 100 kg', ' tanpa tulang'),
        (today, 'Pengolahan', 'Fillet Ayam', 25000, 'Kg', 'Processing Plant', 'Per 100 kg', ' premium'),
        (today, 'Pengolahan', 'Chicken Liver', 5000, 'Kg', 'Processing Plant', '-', ' byproduct'),
        (today, 'Pengolahan', 'Chicken Feet', 8000, 'Kg', 'Processing Plant', '-', ' siap export'),
        (today, 'Pengolahan', 'Sosis Ayam (Premium)', 55000, 'Kg', 'Food Industry', '-', ' siap jual'),
        (today, 'Pengolahan', 'Nugget Ayam', 48000, 'Kg', 'Food Industry', '-', ' siap jual'),
        # Pengemasan
        (today, 'Pengemasan', 'Kardus Box (50 Ekor)', 15000, 'Pcs', 'Supplier Kemasan', '-', ' standar'),
        (today, 'Pengemasan', 'Kardus Box (20 Kg)', 10000, 'Pcs', 'Supplier Kemasan', '-', ' untuk daging'),
        (today, 'Pengemasan', 'Plastik Vakum (10 Kg)', 8000, 'Roll', 'Supplier Kemasan', '-', 'PA/PE'),
        (today, 'Pengemasan', 'Plastik Cling Wrap', 25000, 'Roll', 'Supplier Kemasan', '-', '50 m roll'),
        (today, 'Pengemasan', 'Label/stiker Harga', 50000, '1000 Pcs', 'Percetakan', '-', ' custom'),
        (today, 'Pengemasan', 'Code Date (Inkjet)', 100, 'Ekor/Pcs', 'Supplier', '-', ' termasuk tinta'),
    ]
    
    for row_idx, row_data in enumerate(industri_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.fill = date_fill
                cell.font = date_font
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws3.column_dimensions[col].width = 18 if col in ['A', 'E'] else 22 if col in ['C', 'F'] else 14
    
    # ============================================
    # SHEET 4: HILIR (Distribution/Retail)
    # ============================================
    ws4 = wb.create_sheet("Hilir")
    
    headers_hilir = ['Tanggal', 'Sub Kategori', 'Produk', 'Harga (Rp)', 'Satuan', 'Lokasi/Toko', 'Keterangan']
    for col, header in enumerate(headers_hilir, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center')
    
    hilir_data = [
        # Harga Panen (Farm Gate)
        (today, 'Farm Gate', 'Ayam Broiler Hidup', 32000, 'Kg', 'Peternak Blitar', 'Berat 1.2-1.5 kg'),
        (today, 'Farm Gate', 'Ayam Kampung Hidup', 45000, 'Kg', 'Peternak Banyuwangi', 'Berat 1.0-1.3 kg'),
        (today, 'Farm Gate', 'Itik Manila Hidup', 38000, 'Kg', 'Peternak Mojokerto', 'Berat 1.5-2 kg'),
        (today, 'Farm Gate', 'Kambing PE Hidup', 65000, 'Kg', 'Peternak Bojanegara', 'Berat 20-30 kg'),
        (today, 'Farm Gate', 'Kambing Kacang Hidup', 55000, 'Kg', 'Peternak Lokal', 'Berat 15-25 kg'),
        (today, 'Farm Gate', 'Sapi Simental Hidup', 55000, 'Kg', 'Peternak Bojanegara', 'Berat 300-400 kg'),
        (today, 'Farm Gate', 'Sapi Brahman Cross', 50000, 'Kg', 'Peternak Feedlot', 'Berat 350-450 kg'),
        # Harga Grosir
        (today, 'Grosir', 'Daging Ayam Segar (PS)', 38000, 'Kg', 'Pasar Turi Surabaya', 'Kualitas 1'),
        (today, 'Grosir', 'Daging Sapi Murni (KIM)', 135000, 'Kg', 'Pasar Turi Surabaya', 'Kualitas 1'),
        (today, 'Grosir', 'Daging Kambing (KIM)', 120000, 'Kg', 'Pasar Turi Surabaya', 'Kualitas 1'),
        (today, 'Grosir', 'Telur Ayam Ras (Grosir)', 28000, 'Kg', 'Pasar Turi Surabaya', 'Grade A'),
        (today, 'Grosir', 'Telur Ayam Kampung (Grosir)', 42000, 'Kg', 'Pasar Gentan', 'Grade A'),
        (today, 'Grosir', 'Daging Ayam Segar (KIM)', 38500, 'Kg', 'Pasar Kembang Keris', 'Kualitas 1'),
        (today, 'Grosir', 'Daging Sapi Has Dalam', 145000, 'Kg', 'Pasar Turi Surabaya', 'Premium'),
        # Harga Eceran Tradisional
        (today, 'Eceran Tradisional', 'Daging Ayam Eceran', 42000, 'Kg', 'Pasar tradisional', ' varies by location'),
        (today, 'Eceran Tradisional', 'Daging Sapi Eceran', 150000, 'Kg', 'Pasar tradisional', ' varies by location'),
        (today, 'Eceran Tradisional', 'Daging Kambing Eceran', 135000, 'Kg', 'Pasar tradisional', ' varies by location'),
        (today, 'Eceran Tradisional', 'Telur Ayam Ras Eceran', 30000, 'Kg', 'Pasar tradisional', ' isi 16-18'),
        (today, 'Eceran Tradisional', 'Telur Ayam Kampung Eceran', 48000, 'Kg', 'Pasar tradisional', ' isi 10-12'),
        (today, 'Eceran Tradisional', 'Ayam Potong Eceran', 38000, 'Kg', 'Pasar tradisional', 'live weight'),
        # Harga Modern Retail
        (today, 'Modern Retail', 'Daging Ayam Segar (SM)', 48000, 'Kg', 'Superindo', 'In store'),
        (today, 'Modern Retail', 'Daging Ayam Fillet (SM)', 65000, 'Kg', 'Superindo', 'Premium'),
        (today, 'Modern Retail', 'Daging Sapi Sirloin (SM)', 185000, 'Kg', 'Superindo', 'Import Australia'),
        (today, 'Modern Retail', 'Daging Sapi Has Dalam (SM)', 175000, 'Kg', 'Hypermart', 'Lokal'),
        (today, 'Modern Retail', 'Telur Ayam Ras (SM)', 32000, 'Kg', 'Carrefour', 'Farm fresh'),
        (today, 'Modern Retail', 'Telur Ayam Organic (SM)', 48000, 'Kg', ' Ranch Market', 'Free range'),
        (today, 'Modern Retail', 'Chicken Nugget (SM)', 55000, 'Kg', 'Giant', 'Brand utama'),
        (today, 'Modern Retail', 'Sosis Ayam (SM)', 58000, 'Kg', 'Giant', ' various brand'),
        # Harga Online
        (today, 'Online/E-commerce', 'Daging Ayam Segar (Tokopedia)', 42000, 'Kg', 'Tokopedia', 'Delivery 24h'),
        (today, 'Online/E-commerce', 'Daging Sapi Has Dalam (Tokopedia)', 155000, 'Kg', 'Tokped Fresh', 'chilled'),
        (today, 'Online/E-commerce', 'Telur Ayam Ras (Shopee)', 28500, 'Kg', 'Shopee Fresh', ' includes delivery'),
    ]
    
    for row_idx, row_data in enumerate(hilir_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.fill = date_fill
                cell.font = date_font
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws4.column_dimensions[col].width = 18 if col in ['A', 'E'] else 25 if col == 'C' else 20 if col == 'F' else 14
    
    # ============================================
    # SHEET 5: ANALISIS (Margin & Kalkulasi)
    # ============================================
    ws5 = wb.create_sheet("Analisis")
    
    headers_analisis = ['Tanggal', 'Produk', 'Harga Input (Rp)', 'Harga Output (Rp)', 'Margin (Rp)', 'Margin (%)', 'Keterangan']
    for col, header in enumerate(headers_analisis, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color='8b5cf6', end_color='8b5cf6', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center')
    
    analisis_data = [
        # Perhitungan margin ayam broiler
        (today, 'Ayam Broiler DOC→Panen', 4500, 32000, 27500, 611, 'DOC 4500 + Pakan 22000 + Obat 3000 + Others 2500'),
        (today, 'Ayam Broiler Panen→Grosir', 32000, 38000, 6000, 19, 'Biaya pemotongan 3000 + Transport 1000 + Margin 2000'),
        (today, 'Ayam Broiler Grosir→Eceran', 38000, 42000, 4000, 11, 'Biaya psikis + transport pasar'),
        (today, 'Ayam Broiler Grosir→Modern', 38000, 48000, 10000, 26, ' plus biaya display, kemas, service fee'),
        # Perhitungan margin telur
        (today, 'Telur Layer Pakan→Grosir', 24000, 28000, 4000, 17, 'Pakan per kg telur'),
        (today, 'Telur Grosir→Eceran', 28000, 30000, 2000, 7, 'Biaya pecah 1-2%'),
        (today, 'Telur Grosir→Modern', 28000, 32000, 4000, 14, '-service fee 5%'),
        # Perhitungan margin daging sapi
        (today, 'Sapi Potong Hidup→RPH', 55000, 125000, 70000, 127, 'Per kg karkas (killing 50%)'),
        (today, 'Karkas Sapi Grosir→Eceran', 135000, 150000, 15000, 11, 'Margin eceran'),
        (today, 'Karkas Sapi Eceran→Modern', 150000, 165000, 15000, 10, '-service fee 5%'),
        # Perhitungan margin kambing
        (today, 'Kambing Hidup→RPH', 65000, 120000, 55000, 85, 'Per kg karkas (killing 45%)'),
        (today, 'Karkas Ktg Eceran', 120000, 135000, 15000, 13, 'Margin eceran'),
        # Perbandingan harga
        (today, 'Selisih Grosir vs Eceran (Ayam)', 38000, 42000, 4000, 11, ''),
        (today, 'Selisih Grosir vs Eceran (Sapi)', 135000, 150000, 15000, 11, ''),
        (today, 'Selisih Eceran vs Modern (Ayam)', 42000, 48000, 6000, 14, ''),
        (today, 'Selisih Eceran vs Modern (Sapi)', 150000, 165000, 15000, 10, ''),
        # Break Even
        (today, 'BEP Ayam Broiler (per 1000)', 32000000, 35200000, 3200000, 10, 'DOC 4.5jt + Pakan 22jt + Obat 3jt + Others 2.5jt'),
        (today, 'BEP Layer Telur (per 1000)', 45000000, 58800000, 13800000, 31, 'Pullet 15jt + Pakan 25jt + Others 5jt per periode'),
    ]
    
    for row_idx, row_data in enumerate(analisis_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.fill = date_fill
                cell.font = date_font
            elif col_idx in [5, 6]:  # Margin columns
                cell.number_format = '#,##0' if col_idx == 5 else '0.0"%"'
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws5.column_dimensions[col].width = 18 if col == 'A' else 28 if col == 'B' else 14 if col in ['C', 'D', 'E'] else 10 if col == 'F' else 35
    
    # ============================================
    # SHEET 6: HISTORI PERUBAHAN HARGA
    # ============================================
    ws6 = wb.create_sheet("Histori Harga")
    
    headers_histori = ['Tanggal', 'Produk', 'Harga Lama (Rp)', 'Harga Baru (Rp)', 'Perubahan (Rp)', 'Perubahan (%)', 'Sumber']
    for col, header in enumerate(headers_histori, 1):
        cell = ws6.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color='6366f1', end_color='6366f1', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center')
    
    # Sample histori
    histori_data = [
        (today, 'DOC Ayam Ras', 4000, 4500, 500, 12.5, 'Cibitung Farm'),
        ('2026-06-25', 'Jagung Pipilan', 8000, 8500, 500, 6.3, 'Pasar Ternak'),
        ('2026-06-24', 'Daging Ayam Grosir', 36000, 38000, 2000, 5.6, 'Pasar Turi'),
        ('2026-06-23', 'Telur Ayam Ras', 26000, 28000, 2000, 7.7, 'Pasar Turi'),
        ('2026-06-22', 'Konsentrat Broiler', 270000, 280000, 10000, 3.7, 'Charoen Pokphand'),
    ]
    
    for row_idx, row_data in enumerate(histori_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws6.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.fill = date_fill
                cell.font = date_font
            elif col_idx in [5, 6]:
                cell.number_format = '#,##0' if col_idx == 5 else '0.0"%"'
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws6.column_dimensions[col].width = 18 if col == 'A' else 25 if col == 'B' else 14 if col in ['C', 'D', 'E'] else 10 if col == 'F' else 20
    
    # ============================================
    # SHEET 7: MONITORING HARGA
    # ============================================
    ws7 = wb.create_sheet("Monitoring")
    
    # Ringkasan monitoring per kategori
    monitoring_summary = [
        ['Kategori', 'Jumlah Produk', 'Rata-rata Harga', 'Min', 'Max', 'Trend'],
        ['HULU - Bibit', 12, '=AVERAGE(E2:E13)', 0, 0, '↑'],
        ['HULU - Pakan', 13, '=AVERAGE(E15:E27)', 0, 0, '→'],
        ['HULU - Obat', 11, '=AVERAGE(E29:E39)', 0, 0, '→'],
        ['INDUSTRI - Pemotongan', 6, '=AVERAGE(E41:E46)', 0, 0, '→'],
        ['INDUSTRI - Cold Storage', 4, '=AVERAGE(E48:E51)', 0, 0, '↑'],
        ['INDUSTRI - Pengolahan', 6, '=AVERAGE(E53:E58)', 0, 0, '→'],
        ['INDUSTRI - Pengemasan', 6, '=AVERAGE(E60:E65)', 0, 0, '→'],
        ['HILIR - Farm Gate', 7, '=AVERAGE(E67:E73)', 0, 0, '→'],
        ['HILIR - Grosir', 7, '=AVERAGE(E75:E81)', 0, 0, '↑'],
        ['HILIR - Eceran Tradisional', 6, '=AVERAGE(E83:E88)', 0, 0, '↑'],
        ['HILIR - Modern Retail', 8, '=AVERAGE(E90:E97)', 0, 0, '↑'],
        ['HILIR - Online', 3, '=AVERAGE(E99:E101)', 0, 0, '→'],
        ['ANALISIS - Margin', 14, '=AVERAGE(E103:E116)', 0, 0, '→'],
    ]
    
    for row_idx, row_data in enumerate(monitoring_summary, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws7.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws7.column_dimensions[col].width = 22 if col == 'A' else 15 if col in ['B', 'C', 'D', 'E'] else 10
    
    # Save workbook
    output_path = os.path.expanduser("~/sembako/data/harga_peternakan_lengkap.xlsx")
    wb.save(output_path)
    print(f"✅ File created: {output_path}")
    print(f"📊 Sheets: {[ws.title for ws in wb.worksheets]}")
    return output_path

if __name__ == '__main__':
    create_peternakan_excel()
