#!/usr/bin/env python3
"""Generate optimalisasi_pakan.xlsx from cached optimization results."""
import numpy as np, openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

INGREDIENTS=["Jagung Pipilan","Bungkil Kedelai","Dedak Padi","Tepung Ikan","Pollard","Biji Kapuk","Tepung Darah","Tepung Tulang","Molases","Bungkil Kelapa","Gaplek","Bungkil Sawit","Ampas Tahu","Tepung Bulu Ayam","Kulit Kentang","Onggok","Bungkil Kacang Tanah","Dedak Halus","Sorgum","Menir","Corn Gluten Feed","Rice Polish","Mung Bean Husk"]
N=len(INGREDIENTS)
PRICES={"Jagung Pipilan":4896,"Bungkil Kedelai":7984,"Dedak Padi":4039,"Tepung Ikan":8023,"Pollard":4764,"Biji Kapuk":4168,"Tepung Darah":8149,"Tepung Tulang":7908,"Molases":3637,"Bungkil Kelapa":8637,"Gaplek":3202,"Bungkil Sawit":8201,"Ampas Tahu":3582,"Tepung Bulu Ayam":8038,"Kulit Kentang":2830,"Onggok":2386,"Bungkil Kacang Tanah":8059,"Dedak Halus":4006,"Sorgum":5148,"Menir":5211,"Corn Gluten Feed":7515,"Rice Polish":4933,"Mung Bean Husk":3332}
c_arr=np.array([PRICES[i] for i in INGREDIENTS],dtype=float)
NR={"Jagung Pipilan":[8.5,3.35,0.02,0.28,2.2,3.8],"Bungkil Kedelai":[35.0,2.44,0.34,0.65,6.0,1.5],"Dedak Padi":[12.0,1.75,0.10,1.40,15.5,12.8],"Tepung Ikan":[62.0,2.81,4.50,2.80,0.8,12.0],"Pollard":[15.0,2.05,0.12,1.10,9.5,4.5],"Biji Kapuk":[20.0,3.90,0.18,0.55,23.0,46.0],"Tepung Darah":[85.0,3.01,0.15,0.30,0.5,1.5],"Tepung Tulang":[50.0,1.05,24.0,12.0,0.5,2.0],"Molases":[4.0,2.62,0.80,0.08,0.0,0.5],"Bungkil Kelapa":[22.0,2.85,0.20,0.60,12.0,10.0],"Gaplek":[2.0,3.30,0.05,0.05,4.0,0.5],"Bungkil Sawit":[18.0,2.60,0.25,0.65,15.0,9.0],"Ampas Tahu":[14.0,1.95,0.50,0.30,15.0,8.0],"Tepung Bulu Ayam":[80.0,2.95,0.50,0.35,0.5,5.0],"Kulit Kentang":[7.0,2.70,0.08,0.15,8.0,0.5],"Onggok":[2.0,3.10,0.10,0.05,8.0,2.0],"Bungkil Kacang Tanah":[38.0,2.85,0.20,0.55,12.0,8.0],"Dedak Halus":[13.0,1.80,0.08,1.35,14.0,13.0],"Sorgum":[11.0,3.20,0.04,0.30,2.5,3.5],"Menir":[8.0,3.40,0.01,0.12,0.8,1.5],"Corn Gluten Feed":[22.0,2.50,0.06,0.40,8.0,3.0],"Rice Polish":[12.0,1.90,0.05,1.10,8.0,13.0],"Mung Bean Husk":[15.0,2.00,0.90,0.25,18.0,2.0]}
NP=np.array([NR[i][0] for i in INGREDIENTS]);NM=np.array([NR[i][1] for i in INGREDIENTS])
NC=np.array([NR[i][2] for i in INGREDIENTS]);NPo=np.array([NR[i][3] for i in INGREDIENTS]);NS=np.array([NR[i][4] for i in INGREDIENTS])
MB={"Molases":8,"Tepung Tulang":12,"Tepung Darah":10,"Tepung Bulu Ayam":8,"Biji Kapuk":12,"Mung Bean Husk":15,"Tepung Ikan":12,"Gaplek":25,"Onggok":25,"Kulit Kentang":25,"Ampas Tahu":20}
bds=[(0,MB.get(INGREDIENTS[i],45)) for i in range(N)]; W=10000

NRC={"Sapi Potong Konservasi":(8,12,2.2,0.15,0.15,30),"Sapi Potong Intensif":(12,16,2.6,0.20,0.20,25),"Sapi Fattening":(12,14,2.7,0.20,0.20,22),"Sapi Perah Produksi Rendah":(14,16,2.6,0.25,0.25,25),"Sapi Perah Produksi Tinggi":(16,18,2.9,0.30,0.30,22),"Sapi Hamil 7-9 Bulan":(11,13,2.4,0.25,0.25,25),"Sapi Laktasi Ringan":(14,16,2.6,0.25,0.25,22),"Kambing Konservasi":(9,12,2.2,0.15,0.15,28),"Kambing Pedaging Potong":(12,14,2.5,0.20,0.20,22),"Kambing Perah (Saanen)":(14,16,2.6,0.25,0.25,20),"Domba Konservasi":(9,11,2.1,0.15,0.15,28),"Domba Pedaging":(12,14,2.5,0.20,0.20,22),"Babi Starter (7-25kg)":(18,20,3.0,0.20,0.25,10),"Babi Grower (25-60kg)":(15,17,2.9,0.15,0.20,12),"Babi Finisher (60-100kg)":(13,15,2.8,0.15,0.20,12),"Babi Lancet (Induk)":(14,16,2.7,0.20,0.25,14),"Babi Laktasi (Sapih)":(16,18,2.8,0.25,0.30,12),"Broiler Starter (0-10hr)":(22,24,3.0,0.40,0.35,10),"Broiler Grower (10-25hr)":(20,22,3.1,0.35,0.30,10),"Broiler Finisher (25-35d)":(18,20,3.2,0.30,0.25,10),"Layer Grower (7-17wk)":(16,18,2.9,0.40,0.30,10),"Layer Pre-Lay (17-18wk)":(17,18,2.9,0.60,0.35,10),"Layer Produksi":(16,18,2.85,0.80,0.30,10),"Bebek Petelur Starter":(18,20,2.9,0.40,0.30,10),"Bebek Petelur Layer":(16,18,2.85,0.70,0.30,10),"Bebek Pedaging (Peking)":(16,18,3.0,0.35,0.30,10)}

ORIG={"Sapi Potong Konservasi":{"Jagung Pipilan":25,"Dedak Padi":20,"Pollard":10,"Onggok":15,"Gaplek":10,"Bungkil Sawit":8,"Molases":5,"Tepung Tulang":1,"Dedak Halus":6},"Sapi Potong Intensif":{"Jagung Pipilan":35,"Bungkil Kedelai":15,"Dedak Padi":12,"Pollard":8,"Tepung Ikan":3,"Bungkil Sawit":10,"Gaplek":8,"Molases":5,"Tepung Tulang":1,"Dedak Halus":3},"Sapi Fattening":{"Jagung Pipilan":40,"Bungkil Kedelai":15,"Dedak Padi":10,"Tepung Ikan":4,"Bungkil Sawit":10,"Bungkil Kacang Tanah":5,"Gaplek":8,"Molases":5,"Tepung Tulang":1,"Dedak Halus":2},"Sapi Perah Produksi Rendah":{"Jagung Pipilan":30,"Bungkil Kedelai":12,"Dedak Padi":15,"Pollard":8,"Bungkil Kelapa":10,"Bungkil Sawit":8,"Gaplek":8,"Molases":5,"Tepung Tulang":1.5,"Dedak Halus":2.5},"Sapi Perah Produksi Tinggi":{"Jagung Pipilan":35,"Bungkil Kedelai":18,"Dedak Padi":8,"Tepung Ikan":5,"Bungkil Kacang Tanah":8,"Bungkil Kelapa":8,"Gaplek":6,"Molases":5,"Tepung Tulang":2,"Corn Gluten Feed":5},"Sapi Hamil 7-9 Bulan":{"Jagung Pipilan":28,"Bungkil Kedelai":10,"Dedak Padi":18,"Pollard":10,"Bungkil Sawit":10,"Gaplek":10,"Onggok":5,"Molases":5,"Tepung Tulang":1.5,"Dedak Halus":2.5},"Sapi Laktasi Ringan":{"Jagung Pipilan":32,"Bungkil Kedelai":14,"Dedak Padi":12,"Pollard":8,"Bungkil Kacang Tanah":8,"Bungkil Kelapa":8,"Gaplek":8,"Molases":5,"Tepung Tulang":1.5,"Dedak Halus":3.5},"Kambing Konservasi":{"Jagung Pipilan":22,"Dedak Padi":20,"Pollard":10,"Onggok":15,"Gaplek":12,"Bungkil Sawit":8,"Molases":5,"Dedak Halus":5,"Tepung Tulang":1,"Sorgum":2},"Kambing Pedaging Potong":{"Jagung Pipilan":30,"Bungkil Kedelai":12,"Dedak Padi":15,"Pollard":8,"Bungkil Sawit":10,"Gaplek":10,"Molases":5,"Tepung Tulang":1,"Dedak Halus":7,"Sorgum":2},"Kambing Perah (Saanen)":{"Jagung Pipilan":30,"Bungkil Kedelai":15,"Dedak Padi":12,"Tepung Ikan":3,"Bungkil Kacang Tanah":8,"Bungkil Kelapa":8,"Gaplek":8,"Molases":5,"Tepung Tulang":1.5,"Dedak Halus":6,"Corn Gluten Feed":3.5},"Domba Konservasi":{"Jagung Pipilan":20,"Dedak Padi":22,"Pollard":10,"Onggok":15,"Gaplek":12,"Bungkil Sawit":8,"Molases":5,"Dedak Halus":5,"Tepung Tulang":1,"Sorgum":2},"Domba Pedaging":{"Jagung Pipilan":28,"Bungkil Kedelai":12,"Dedak Padi":15,"Pollard":8,"Bungkil Sawit":10,"Gaplek":10,"Molases":5,"Tepung Tulang":1,"Dedak Halus":8,"Sorgum":3},"Babi Starter (7-25kg)":{"Jagung Pipilan":35,"Bungkil Kedelai":22,"Tepung Ikan":5,"Tepung Darah":5,"Tepung Bulu Ayam":3,"Menir":12,"Corn Gluten Feed":8,"Ampas Tahu":3,"Molases":3,"Tepung Tulang":1.5,"Dedak Padi":2.5},"Babi Grower (25-60kg)":{"Jagung Pipilan":38,"Bungkil Kedelai":18,"Tepung Ikan":3,"Tepung Darah":4,"Menir":15,"Corn Gluten Feed":8,"Ampas Tahu":4,"Dedak Padi":5,"Molases":3,"Tepung Tulang":1,"Bungkil Sawit":1},"Babi Finisher (60-100kg)":{"Jagung Pipilan":42,"Bungkil Kedelai":15,"Menir":15,"Corn Gluten Feed":8,"Ampas Tahu":5,"Dedak Padi":5,"Tepung Bulu Ayam":3,"Bungkil Sawit":3,"Molases":3,"Tepung Tulang":1},"Babi Lancet (Induk)":{"Jagung Pipilan":35,"Bungkil Kedelai":18,"Tepung Ikan":3,"Tepung Darah":3,"Dedak Padi":12,"Menir":10,"Corn Gluten Feed":8,"Ampas Tahu":4,"Molases":3,"Tepung Tulang":1.5,"Dedak Halus":2.5},"Babi Laktasi (Sapih)":{"Jagung Pipilan":35,"Bungkil Kedelai":20,"Tepung Ikan":5,"Tepung Darah":4,"Menir":12,"Corn Gluten Feed":8,"Ampas Tahu":4,"Molases":3,"Tepung Tulang":2,"Dedak Halus":5,"Bungkil Sawit":2},"Broiler Starter (0-10hr)":{"Jagung Pipilan":35,"Bungkil Kedelai":25,"Tepung Ikan":8,"Tepung Darah":5,"Tepung Bulu Ayam":3,"Menir":10,"Corn Gluten Feed":5,"Bungkil Kacang Tanah":3,"Molases":2,"Tepung Tulang":1.5,"Dedak Padi":2.5},"Broiler Grower (10-25hr)":{"Jagung Pipilan":38,"Bungkil Kedelai":22,"Tepung Ikan":6,"Tepung Darah":4,"Tepung Bulu Ayam":3,"Menir":12,"Corn Gluten Feed":5,"Bungkil Kacang Tanah":3,"Molases":2,"Tepung Tulang":1.5,"Dedak Padi":3.5},"Broiler Finisher (25-35d)":{"Jagung Pipilan":42,"Bungkil Kedelai":18,"Tepung Ikan":5,"Tepung Darah":3,"Menir":14,"Corn Gluten Feed":6,"Ampas Tahu":3,"Molases":2,"Tepung Tulang":1.5,"Dedak Padi":4,"Bungkil Sawit":1.5},"Layer Grower (7-17wk)":{"Jagung Pipilan":38,"Bungkil Kedelai":18,"Tepung Ikan":4,"Tepung Darah":3,"Menir":12,"Corn Gluten Feed":6,"Ampas Tahu":4,"Bungkil Sawit":3,"Molases":2,"Tepung Tulang":2.5,"Dedak Padi":5,"Dedak Halus":2.5},"Layer Pre-Lay (17-18wk)":{"Jagung Pipilan":35,"Bungkil Kedelai":17,"Tepung Ikan":4,"Tepung Darah":3,"Menir":10,"Corn Gluten Feed":5,"Ampas Tahu":4,"Bungkil Sawit":3,"Molases":2,"Tepung Tulang":8,"Dedak Padi":5,"Dedak Halus":4},"Layer Produksi":{"Jagung Pipilan":33,"Bungkil Kedelai":16,"Tepung Ikan":3,"Tepung Darah":3,"Menir":10,"Corn Gluten Feed":5,"Ampas Tahu":3,"Bungkil Sawit":3,"Molases":2,"Tepung Tulang":15,"Dedak Padi":5,"Dedak Halus":2},"Bebek Petelur Starter":{"Jagung Pipilan":35,"Bungkil Kedelai":20,"Tepung Ikan":6,"Tepung Darah":4,"Menir":12,"Corn Gluten Feed":5,"Ampas Tahu":3,"Molases":2,"Tepung Tulang":2,"Dedak Padi":5,"Dedak Halus":6},"Bebek Petelur Layer":{"Jagung Pipilan":33,"Bungkil Kedelai":16,"Tepung Ikan":3,"Tepung Darah":3,"Menir":10,"Corn Gluten Feed":5,"Ampas Tahu":3,"Bungkil Sawit":3,"Molases":2,"Tepung Tulang":12,"Dedak Padi":6,"Dedak Halus":4},"Bebek Pedaging (Peking)":{"Jagung Pipilan":38,"Bungkil Kedelai":18,"Tepung Ikan":5,"Tepung Darah":4,"Menir":12,"Corn Gluten Feed":6,"Ampas Tahu":3,"Molases":2,"Tepung Tulang":2,"Dedak Padi":5,"Dedak Halus":5}}

GROUPS={"Rum.Sapi": ["Sapi Potong Konservasi","Sapi Potong Intensif","Sapi Fattening","Sapi Perah Produksi Rendah","Sapi Perah Produksi Tinggi","Sapi Hamil 7-9 Bulan","Sapi Laktasi Ringan"],"Rum.Kecil": ["Kambing Konservasi","Kambing Pedaging Potong","Kambing Perah (Saanen)","Domba Konservasi","Domba Pedaging"],"Babi": ["Babi Starter (7-25kg)","Babi Grower (25-60kg)","Babi Finisher (60-100kg)","Babi Lancet (Induk)","Babi Laktasi (Sapih)"],"Unggas": ["Broiler Starter (0-10hr)","Broiler Grower (10-25hr)","Broiler Finisher (25-35d)","Layer Grower (7-17wk)","Layer Pre-Lay (17-18wk)","Layer Produksi","Bebek Petelur Starter","Bebek Petelur Layer","Bebek Pedaging (Peking)"]}
GC={"Rum.Sapi":("8B4513",PatternFill(start_color="8B4513",end_color="8B4513",fill_type="solid")),"Rum.Kecil":("D2691E",PatternFill(start_color="D2691E",end_color="D2691E",fill_type="solid")),"Babi":("E74C3C",PatternFill(start_color="E74C3C",end_color="E74C3C",fill_type="solid")),"Unggas":("F39C12",PatternFill(start_color="F39C12",end_color="F39C12",fill_type="solid"))}
HF=Font(bold=True,size=10,color="FFFFFF");DF=Font(size=9);BF=Font(bold=True,size=9);TF=Font(bold=True,size=14,color="FFFFFF")
THIN=Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
CTR=Alignment(horizontal="center",vertical="center",wrap_text=True);LFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
GREY=PatternFill(start_color="ECF0F1",end_color="ECF0F1",fill_type="solid")
GRN=PatternFill(start_color="E8F5E9",end_color="E8F5E9",fill_type="solid")
RED_BG=PatternFill(start_color="FFEBEE",end_color="FFEBEE",fill_type="solid")
YLW=PatternFill(start_color="FFF8E1",end_color="FFF8E1",fill_type="solid")
BLU=PatternFill(start_color="E3F2FD",end_color="E3F2FD",fill_type="solid")

def sh(ws,r,cols,fill=None):
    if fill is None:fill=PatternFill(start_color="2C3E50",end_color="2C3E50",fill_type="solid")
    for c in range(1,cols+1):cl=ws.cell(r,c);cl.font=HF;cl.fill=fill;cl.alignment=CTR;cl.border=THIN
def sr(ws,r,cols,fill=None):
    for c in range(1,cols+1):cl=ws.cell(r,c);cl.font=DF;cl.alignment=CTR if c>1 else LFT;cl.border=THIN
    if fill:
        for c in range(1,cols+1):ws.cell(r,c).fill=fill

def dopt(fname):
    from scipy.optimize import minimize
    pk_min,pk_max,me_min,ca_min,p_min,sk_max=NRC[fname]
    def obj(x):
        pv=np.dot(c_arr,x)/100;pk=np.dot(NP,x)/100;me=np.dot(NM,x)/100
        ca=np.dot(NC,x)/100;p=np.dot(NPo,x)/100;sk=np.dot(NS,x)/100
        pen=W*(np.sum(x)-100)**2
        if pk<pk_min:pen+=W*(pk_min-pk)**2
        if pk>pk_max:pen+=W*(pk-pk_max)**2
        if me<me_min:pen+=W*(me_min-me)**2
        if ca<ca_min:pen+=W*(ca_min-ca)**2
        if p<p_min:pen+=W*(p_min-p)**2
        if sk>sk_max:pen+=W*(sk-sk_max)**2
        return pv+pen
    best=None
    for _ in range(4):
        x0=np.random.dirichlet(np.ones(N))*100
        r=minimize(obj,x0,method='L-BFGS-B',bounds=bds,options={'maxiter':5000,'ftol':1e-14})
        x=np.where(r.x<0.1,0,r.x);s=np.sum(x)
        if s>0:x=x/s*100
        pk=np.dot(NP,x)/100;me=np.dot(NM,x)/100;ca=np.dot(NC,x)/100;p=np.dot(NPo,x)/100;sk=np.dot(NS,x)/100
        ok=(pk_min<=pk<=pk_max)and(me>=me_min)and(ca>=ca_min)and(p>=p_min)and(sk<=sk_max)
        if ok:
            cv=round(np.dot(c_arr,x)/100)
            if best is None or cv<best[0]:best=(cv,x,pk,me,ca,p,sk)
    if best:
        cv,x,pk,me,ca,p,sk=best
        f={INGREDIENTS[i]:round(x[i],1) for i in range(N) if x[i]>0.1}
        return f,cv,{"PK":pk,"ME":me,"Ca":ca,"P":p,"SK":sk},"optimal"
    return None,0,{},"infeasible"

def cc(d):return round(sum(PRICES.get(i,0)*p/100 for i,p in d.items()))
def cn(d):
    n={"PK":0,"ME":0,"Ca":0,"P":0,"SK":0,"LK":0};ks=["PK","ME","Ca","P","SK","LK"]
    for i,p in d.items():nr=NR.get(i,[0]*6)
    for ki,k in enumerate(ks):n[k]+=nr[ki]*p/100
    return {k:round(v,2) for k,v in n.items()}

# Optimize all
ao={}
for fn in sorted(NRC.keys()):
    of,oc,on,st=dopt(fn);oc_orig=cc(ORIG.get(fn,{}));sav=oc_orig-oc if oc>0 else 0;sp=round(sav/oc_orig*100,1) if oc_orig>0 else 0
    ao[fn]=(of,oc,on,st,oc_orig,sav,sp)
    icon="+" if sav>0 else "!"
    print(f"  {icon} {fn:35s} Rp{oc_orig:>6,}->{oc:>6,} Hemat Rp{sav:>5,} ({sp}%)")

# Excel — CORRECT PATH
out = Path.home() / "sembako" / "data" / "optimalisasi_pakan.xlsx"
wb=openpyxl.Workbook();wb.remove(wb.active)
for gn,fns in GROUPS.items():
    ws=wb.create_sheet(gn);ws.sheet_properties.tabColor=GC[gn][0]
    ws.merge_cells("A1:K1");ws.cell(1,1,f"OPTIMALISASI - {gn.upper()}").font=TF;ws.cell(1,1).fill=GC[gn][1];ws.cell(1,1).alignment=CTR
    ws.merge_cells("A2:K2");ws.cell(2,1,f"SLSQP minimize | {datetime.now().strftime('%d %B %Y')} | Premix mineral/vitamin terpisah").font=Font(italic=True,size=9)
    row=4
    for fn in fns:
        if fn not in NRC or fn not in ORIG:continue
        of,oc,on,st,oc_orig,sav,sp=ao[fn];orig=ORIG[fn];orig_n=cn(orig)
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=10)
        ws.cell(row,1,fn).font=Font(bold=True,size=11,color="FFFFFF");ws.cell(row,1).fill=GC[gn][1];ws.cell(row,1).alignment=LFT;row+=1
        cols=["No","Bahan","Harga","Awal%","Biaya","Opt%","Biaya","Hemat","Status",""]
        for ci,h in enumerate(cols,1):ws.cell(row,ci,h)
        sh(ws,row,10,PatternFill(start_color="555555",end_color="555555",fill_type="solid"));row+=1
        ings=sorted(set(list(orig.keys())+list(of.keys()) if of else list(orig.keys())),key=lambda x:-max(orig.get(x,0),(of or {}).get(x,0)))
        for ni,ing in enumerate(ings,1):
            op=orig.get(ing,0);np2=(of or {}).get(ing,0);pr=PRICES.get(ing,0)
            oc2=round(op*pr/100);nc2=round(np2*pr/100)
            ws.cell(row,1,ni);ws.cell(row,2,ing);ws.cell(row,3,pr)
            ws.cell(row,4,f"{op}%" if op>0 else "-");ws.cell(row,5,oc2)
            ws.cell(row,6,f"{np2}%" if np2>0 else "-");ws.cell(row,7,nc2)
            ws.cell(row,8,nc2-oc2)
            stx="BARU" if op==0 and np2>0 else "HAPUS" if op>0 and np2==0 else "UBAH" if abs(op-np2)>0.5 else "OK"
            ws.cell(row,9,stx)
            fill=YLW if abs(op-np2)>0.5 else BLU if op==0 and np2>0 else RED_BG if op>0 and np2==0 else GRN
            sr(ws,row,10,fill)
            for c in [3,5,7,8]:ws.cell(row,c).number_format='#,##0'
            row+=1
        ws.cell(row,2,"TOTAL").font=BF;ws.cell(row,4,"100%").font=BF
        ws.cell(row,5,oc_orig).font=BF;ws.cell(row,6,"100%").font=BF;ws.cell(row,7,oc).font=BF
        ws.cell(row,8,sav).font=Font(bold=True,color="006600" if sav>0 else "CC0000")
        sr(ws,row,10,GREY)
        for c in [3,5,7,8]:ws.cell(row,c).number_format='#,##0'
        row+=1
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=10)
        ws.cell(row,1,f"Rp{oc_orig:,}/kg -> Rp{oc:,}/kg | HEMAT Rp{sav:,}/kg ({sp}%)").font=Font(bold=True,size=10,color="006600" if sav>0 else "CC0000")
        row+=2
    for col in ws.columns:mx=max(10,min(28,max(len(str(cell.value or "")) for cell in col)+2));ws.column_dimensions[get_column_letter(col[0].column)].width=mx

# Summary sheet
ws2=wb.create_sheet("Ringkasan");ws2.sheet_properties.tabColor="2ECC71"
ws2.merge_cells("A1:L1");ws2.cell(1,1,"RINGKASAN OPTIMALISASI").font=TF;ws2.cell(1,1).fill=PatternFill(start_color="2ECC71",end_color="2ECC71",fill_type="solid");ws2.cell(1,1).alignment=CTR
hdr=["No","Grp","Formulasi","Awal","Opt","Hemat","Hemat%","PKo","PKn","MEo","MEn","Status"]
for ci,h in enumerate(hdr,1):ws2.cell(3,ci,h)
sh(ws2,3,12,PatternFill(start_color="27AE60",end_color="27AE60",fill_type="solid"))
row=4;num=1;to=0;tn=0
for gn,fns in GROUPS.items():
    ws2.merge_cells(start_row=row,start_column=1,end_row=row,end_column=12);ws2.cell(row,1,gn).font=Font(bold=True,color="FFFFFF");ws2.cell(row,1).fill=GC[gn][1];ws2.cell(row,1).alignment=LFT;row+=1
    for fn in fns:
        if fn not in NRC or fn not in ORIG:continue
        of,oc,on,st,oc_orig,sav,sp=ao[fn];orig_n=cn(ORIG[fn]);to+=oc_orig;tn+=oc
        ws2.cell(row,1,num);ws2.cell(row,2,gn);ws2.cell(row,3,fn)
        ws2.cell(row,4,oc_orig);ws2.cell(row,5,oc);ws2.cell(row,6,sav);ws2.cell(row,7,f"{sp}%")
        ws2.cell(row,8,orig_n.get("PK","-"));ws2.cell(row,9,on.get("PK","-"))
        ws2.cell(row,10,orig_n.get("ME","-"));ws2.cell(row,11,on.get("ME","-"))
        ws2.cell(row,12,f"{'OK' if sav>=0 else 'X'} {st}")
        sr(ws2,row,12,GRN if sav>0 else RED_BG)
        for c in [4,5,6]:ws2.cell(row,c).number_format='#,##0'
        num+=1;row+=1
    row+=1
ws2.cell(row,1,"AVG").font=BF
ws2.cell(row,4,round(to/len(NRC))).number_format='#,##0'
ws2.cell(row,5,round(tn/len(NRC))).number_format='#,##0'
ws2.cell(row,6,round((to-tn)/len(NRC))).number_format='#,##0'
sr(ws2,row,12,GREY)
for col in ws2.columns:mx=max(10,min(28,max(len(str(cell.value or "")) for cell in col)+2));ws2.column_dimensions[get_column_letter(col[0].column)].width=mx

wb.save(str(out))
ok=sum(1 for v in ao.values() if v[3]=="optimal" and v[5]>0)
ts=sum(v[5] for v in ao.values() if v[5]>0)
print(f"\nDONE {out} | {len(wb.sheetnames)} sheets | {ok}/{len(NRC)} optimal | Total hemat Rp{ts:,}")
