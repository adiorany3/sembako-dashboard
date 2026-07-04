#!/usr/bin/env python3
"""Update CPO price from external source. Saves to harga_cpo.xlsx."""

import datetime as dt
from pathlib import Path

import re
import requests
import openpyxl

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)
OUTPUT = DATA_DIR / "harga_cpo.xlsx"

URL = "https://tradingeconomics.com/commodity/palm-oil"

def fetch_cpo() -> tuple[float, str]:
    """Return (myr_per_tonne, source). Raises on failure."""
    r = requests.get(URL, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
    r.raise_for_status()
    m = re.search(r"([\d,\.]+)\s*MYR/T", r.text)
    if m:
        return float(m.group(1).replace(",", "")), URL
    raise RuntimeError("Could not parse CPO price from tradingeconomics")


def main():
    today = dt.date.today().isoformat()
    myr, source = fetch_cpo()
    myr_per_kg = myr / 1000  # tonne to kg

    # Load or create workbook
    if OUTPUT.exists():
        wb = openpyxl.load_workbook(OUTPUT)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CPO"
        ws.append(["Tanggal", "Harga_MYR", "Harga_IDR", "Perubahan_Persen", "Sumber"])

    # Get last row for change calc — find most recent date
    last_my = None
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row, 2).value is not None:
            last_my = ws.cell(row, 2).value
            # Delete today's entry if exists (re-run safety)
            if ws.cell(row, 1).value == today:
                ws.delete_rows(row)
                last_my = ws.cell(row - 1, 2).value if row > 2 else None
            break
    pct_change = round((myr - last_my) / last_my * 100, 2) if last_my else 0.0
    idr = round(myr_per_kg * 18000, 0) if myr_per_kg else 0  # MYR/kg * 18000

    ws.append([today, myr, idr, pct_change, source])
    wb.save(OUTPUT)
    print(f"✅ CPO price updated: MYR {myr}/tonne ({pct_change:+.2f}%)")

if __name__ == "__main__":
    main()
