#!/usr/bin/env python3
"""
Expand Peternakan Data - More complete Hulu & Industri
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def expand_peternakan_data():
    filepath = '/root/sembako/harga_peternakan_lengkap.xlsx'
    wb = openpyxl.load_workbook(filepath)
    
    header_fill = PatternFill(start_color='1e3c72', end_color='2a5298', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    
    today = datetime.now().strftime('%Y-%m-%d')
    
    # ============================================
    # EXPAND HULU SHEET
    # ============================================
    ws = wb['Hulu']
    ws.delete_rows(2, ws.max_row)  # Clear existing data
    
    hulu_data = [
        # Bibit Unggas (15 items)
        ('Bibit', 'DOC Ayam Ras (Super)', 4600, 'Ekor', 'Cibitung Farm', 'Umur 1 hari'),
        ('Bibit', 'DOC Ayam Ras (Medium)', 3400, 'Ekor', 'Peternakan Lokal', 'Umur 1 hari'),
        ('Bibit', 'DOC Ayam Ras (Pullet)', 5200, 'Ekor', 'Cibitung Farm', 'Umur 1 hari'),
        ('Bibit', 'DOC Ayam Kampung', 7800, 'Ekor', 'Peternakan Lokal', 'Umur 1 hari'),
        ('Bibit', 'DOC Ayam Petelur (Hyline)', 8500, 'Ekor', 'Breeding Farm', 'Umur 1 hari'),
        ('Bibit', 'DOC Itik Mojosari', 8100, 'Ekor', 'Cibitung Farm', 'Umur 1 hari'),
        ('Bibit', 'Bibit Itik Manila ( Adult)', 11800, 'Ekor', 'Peternakan Lokal', 'Umur 4 bulan'),
        ('Bibit', 'DOC Puyuh (Quail)', 5500, 'Ekor', 'Breeding Farm', 'Umur 1 hari'),
        ('Bibit', 'Bibit Angsa (Chinese)', 35000, 'Ekor', 'Peternakan Lokal', 'Umur 2 bulan'),
        ('Bibit', 'Bibit Merpati (Racing)', 25000, 'Ekor', 'Breeder', 'Umur 3 bulan'),
        ('Bibit', 'DOC Bebek Peking (Muscovy)', 7500, 'Ekor', 'Farm', 'Umur 1 hari'),
        ('Bibit', 'Bibit Kelinci (New Zealand)', 85000, 'Ekor', 'Rabbit Farm', 'Umur 2 bulan'),
        ('Bibit', 'Bibit Kelinci (Rex)', 75000, 'Ekor', 'Rabbit Farm', 'Umur 2 bulan'),
        ('Bibit', 'DOC Ayam Arab', 6200, 'Ekor', 'Farm', 'Umur 1 hari'),
        ('Bibit', 'Bibit Turkey (Bronze)', 180000, 'Ekor', 'Import', 'Umur 1 hari'),
        
        # Bibit Ternak Besar (10 items)
        ('Bibit', 'Bibit Kambing Boer', 1824800, 'Ekor', 'Boer Farm', 'Umur 6 bulan'),
        ('Bibit', 'Bibit Kambing PE (Persilangan)', 1505700, 'Ekor', 'PE Farm', 'Umur 6 bulan'),
        ('Bibit', 'Bibit Kamping Kacang', 890000, 'Ekor', 'Lokal', 'Umur 6 bulan'),
        ('Bibit', 'Bibit Kambing Etawa', 2100000, 'Ekor', 'Etawa Farm', 'Umur 6 bulan'),
        ('Bibit', 'Bibit Sapi Limousin', 27571100, 'Ekor', 'BPTU', 'Umur 8 bulan'),
        ('Bibit', 'Bibit Sapi Simental', 25605100, 'Ekor', 'BPTU', 'Umur 8 bulan'),
        ('Bibit', 'Bibit Sapi Brahman', 28500000, 'Ekor', 'Import', 'Umur 8 bulan'),
        ('Bibit', 'Bibit Sapi Angus', 30000000, 'Ekor', 'Import', 'Umur 8 bulan'),
        ('Bibit', 'Bibit Sapi PO (Perah Ongole)', 15000000, 'Ekor', 'BPTU', 'Umur 8 bulan'),
        ('Bibit', 'Bibit Domba Merino', 1200000, 'Ekor', 'BPPT', 'Umur 6 bulan'),
        
        # Pakan Unggas (15 items)
        ('Pakan', 'BR1 (Bogasari) 50kg', 285000, 'Sak', 'Pabrik Pakan', 'Pakan layer fase starter'),
        ('Pakan', 'BR2 (Bogasari) 50kg', 275000, 'Sak', 'Pabrik Pakan', 'Pakan layer fase grower'),
        ('Pakan', 'BJ1 (Jati Sari) 50kg', 265000, 'Sak', 'Pabrik Pakan', 'Pakan broiler fase finisher'),
        ('Pakan', 'BJ2 (Jati Sari) 50kg', 255000, 'Sak', 'Pabrik Pakan', 'Pakan broiler fase grower'),
        ('Pakan', 'Konsentrat Layer 30%', 425000, 'Sak', 'Charoen Pokphand', 'Supplement telur'),
        ('Pakan', 'Konsentrat Broiler 25%', 395000, 'Sak', 'Charoen Pokphand', 'Supplement daging'),
        ('Pakan', 'Dedak Halus (Bekatul)', 180000, '100kg', 'Pasar Grosir', 'Sumber energi'),
        ('Pakan', 'Dedak Kasar', 150000, '100kg', 'Pasar Grosir', 'Sumber energi'),
        ('Pakan', 'Jagung Pipilan (Feed Grade)', 380000, '100kg', 'Pasar Grosir', 'Bahan utama energi'),
        ('Pakan', 'Bungkil Kedelai', 680000, '100kg', 'Import', 'Sumber protein tinggi'),
        ('Pakan', 'Tepung Ikan (Imported)', 1850000, '100kg', 'Import', 'Sumber protein & mineral'),
        ('Pakan', 'Tepung Darah (Blood Meal)', 950000, '50kg', 'Lokal', 'Sumber protein'),
        ('Pakan', 'Minyak Kelapa Sawit (CPO)', 1850000, 'Drum', 'Pabrik Sawit', 'Sumber energi'),
        ('Pakan', 'Premix Vitamin (1kg)', 85000, 'Pouch', 'Vitamin Supplier', 'Suplemen vitamin'),
        ('Pakan', 'Tepung Cangkang Telur', 120000, '50kg', 'Farm', 'Sumber kalsium'),
        
        # Pakan Ternak Besar (10 items)
        ('Pakan', 'Rumput Raja (King Grass) - 1 Ha', 450000, 'Ha', 'Perkebunan', 'Hijauan segar'),
        ('Pakan', 'Rumput Gajah (Napier)', 350000, 'Ha', 'Peternakan', 'Hijauan segar'),
        ('Pakan', 'Hay (Rumput Kering)', 180000, 'Bal', 'Peternakan', 'Hijauan kering'),
        ('Pakan', 'Silase Jagung', 220000, '100kg', 'Farm', 'Hijauan fermentasi'),
        ('Pakan', 'Pakan jadi Sapi (Total Mixed Ration)', 450000, '100kg', 'Pabrik Pakan', 'Pakan komplit'),
        ('Pakan', 'Ampas Tahu (Bungkil Tahu)', 85000, '100kg', 'Industri Tahu', 'Pakan samping'),
        ('Pakan', 'Ampas Bir (Wet Grain)', 65000, '100kg', 'Pabrik Bir', 'Pakan samping'),
        ('Pakan', 'Onggok (Singkong)', 55000, '100kg', 'Industri Tapioka', 'Sumber energi'),
        ('Pakan', 'Kulit Kopi (Coffee Pulp)', 45000, '100kg', 'Pabrik Kopi', 'Sumber serat'),
        ('Pakan', 'Urea-MOL (Probiotik)', 125000, 'Liter', 'Lab Peternakan', 'Suplemen'),
        
        # Obat & Veteriner (20 items)
        ('Obat', 'Vitamin B Complex (100ml)', 45000, 'Botol', 'Kimia Farma', 'Vitamin neurotropik'),
        ('Obat', 'Vitamin ADE (100ml)', 55000, 'Botol', 'Vet Pharma', 'Vitamin larut lemak'),
        ('Obat', 'Vitamin C (100g)', 35000, 'Pouch', 'Vet Pharma', 'Anti stress'),
        ('Obat', 'Liquid Glucose (1L)', 28000, 'Botol', 'Farm', 'Energi cepat'),
        ('Obat', 'Electrolyte (100g)', 25000, 'Pouch', 'Vet Pharma', 'Rehidrasi'),
        ('Obat', 'Antibiotik Oxytetracycline 20% (100ml)', 85000, 'Botol', 'Vet Pharma', 'Infeksi bakteri'),
        ('Obat', 'Antibiotik Enrofloxacin 10% (100ml)', 95000, 'Botol', 'Vet Pharma', 'Infeksi saluran napas'),
        ('Obat', 'Antibiotik Neomycin (100g)', 75000, 'Pouch', 'Vet Pharma', 'Diare'),
        ('Obat', 'Coccidiostat (100g)', 120000, 'Pouch', 'Vet Pharma', 'Koksidiosis'),
        ('Obat', 'Anthelmintik (Cacing) - Albendazole (10 tablet)', 35000, 'Strip', 'Kimia Farma', 'Cacingan'),
        ('Obat', 'Anthelmintik - Levamisole (100ml)', 65000, 'Botol', 'Vet Pharma', 'Cacingan'),
        ('Obat', 'Vaccine AI (Avian Influenza) - 100 doses', 175000, 'Vial', 'Vet Pharma', 'Flu burung'),
        ('Obat', 'Vaccine ND (Newcastle Disease) - 100 doses', 125000, 'Vial', 'Vet Pharma', 'Tetelo'),
        ('Obat', 'Vaccine Gumboro - 100 doses', 95000, 'Vial', 'Vet Pharma', 'Gumboro'),
        ('Obat', 'Desinfektan Virkon S (1kg)', 185000, 'Pouch', 'Vet Pharma', 'Biosecurity'),
        ('Obat', 'Desinfektan Formalin 37% (5L)', 95000, 'Jerigen', 'Kimia Farma', 'Biosecurity'),
        ('Obat', 'Obat Anti Parasit (Kutu/Ringworm) - 100ml', 55000, 'Botol', 'Vet Pharma', 'Ektoparasit'),
        ('Obat', 'Salep Mata Vet (20g)', 25000, 'Tube', 'Vet Pharma', 'Infeksi mata'),
        ('Obat', 'Antiseptik Betadine (100ml)', 28000, 'Botol', 'Kimia Farma', 'Luka luar'),
        ('Obat', 'ACR (Anti Colibacillosis) - 100ml', 85000, 'Botol', 'Vet Pharma', 'Coli infeksi'),
    ]
    
    for row_data in hulu_data:
        ws.append([today] + list(row_data))
    
    # Format header
    for col in range(1, 8):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
    
    print(f"✅ Hulu: {len(hulu_data)} items added")
    
    # ============================================
    # EXPAND INDUSTRI SHEET
    # ============================================
    ws = wb['Industri']
    ws.delete_rows(2, ws.max_row)
    
    industri_data = [
        # Pemotongan (RPH) - 10 items
        ('Pemotongan', 'Jasa Pemotongan Ayam Broiler', 3000, 'Ekor', 'RPH Sidoarjo', '-', 'Kapasitas 5000/hari'),
        ('Pemotongan', 'Jasa Pemotongan Ayam Kampung', 4900, 'Ekor', 'RPH Sidoarjo', '-', 'Kapasitas 1000/hari'),
        ('Pemotongan', 'Jasa Pemotongan Itik/Entok', 4300, 'Ekor', 'RPH Surabaya', '-', 'Kapasitas 500/hari'),
        ('Pemotongan', 'Jasa Pemotongan Puyuh', 1500, 'Ekor', 'RPH Sidoarjo', '-', 'Kapasitas 2000/hari'),
        ('Pemotongan', 'Jasa Pemotongan Kambing', 76700, 'Ekor', 'RPH Surabaya', '-', 'Kapasitas 100/hari'),
        ('Pemotongan', 'Jasa Pemotongan Domba', 65000, 'Ekor', 'RPH Surabaya', '-', 'Kapasitas 100/hari'),
        ('Pemotongan', 'Jasa Pemotongan Sapi', 336000, 'Ekor', 'RPH Surabaya', '-', 'Kapasitas 50/hari'),
        ('Pemotongan', 'Jasa Pemotongan Sapi (Kecil)', 241300, 'Ekor', 'RPH Sidoarjo', '-', 'Kapasitas 30/hari'),
        ('Pemotongan', 'Jasa Pemotongan Kerbau', 385000, 'Ekor', 'RPH Kediri', '-', 'Kapasitas 20/hari'),
        ('Pemotongan', 'Sertifikasi兽医 (Vet Certificate)', 45000, 'Ekor', 'PUSDAKTAN', '-', 'Wajib الحجر الصحي'),
        
        # Cold Storage & Logistik - 10 items
        ('Cold Storage', 'Sewa Cold Storage Short Term', 5100, 'Kg/Hari', 'Sidoarjo', '500 Ton', 'Min 1 hari'),
        ('Cold Storage', 'Sewa Cold Storage Monthly', 99100, 'Kg/Bulan', 'Sidoarjo', '500 Ton', 'Contract 6 bln'),
        ('Cold Storage', 'Sewa Cold Box (Portable) 10FT', 154200, 'Unit/Hari', 'Rental', '-', 'Termasuk refrigasi'),
        ('Cold Storage', 'Cold Chain Transport (Colt Diesel)', 850000, 'Trip', 'Sidoarjo-Jakarta', '-', '8 ton kapasitas'),
        ('Cold Storage', 'Cold Chain Transport (CDD Bak)', 550000, 'Trip', 'Sidoarjo-Surabaya', '-', '3 ton kapasitas'),
        ('Cold Storage', 'Ice Block (Es Batu) untuk Transport', 18000, 'Block 25kg', 'Pabrik Es', '-', 'Pendingin darurat'),
        ('Cold Storage', 'Thermal Blanket (Jas Hangat)', 35000, 'Pcs', 'Supplier', '-', 'Insulasi pengiriman'),
        ('Cold Storage', 'Cooling Pad (Media Pendingin)', 125000, 'Set', 'Kandang Closed House', '-', 'Evaporative cooling'),
        ('Cold Storage', 'Freezer Storage 20FT (Container)', 2500000, 'Month', 'Depo Container', '-', 'Kap 15 ton'),
        ('Cold Storage', 'Quality Control Cool Room', 350000, 'Day', 'Lab', '-', 'QC temperatura'),
        
        # Pengolahan Daging - 15 items
        ('Pengolahan', 'Jasa Pemotongan & Pembersihan (屠宰)', 8500, 'Kg', 'RPH Plus', '-', 'SIK dan halal'),
        ('Pengolahan', 'Jasa Marination (Bumbu Injeksi)', 12000, 'Kg', 'Food Processor', '-', '100kg min order'),
        ('Pengolahan', 'Jasa Fillet & Deboning', 15000, 'Kg', 'Food Processor', '-', 'Ayam & ikan'),
        ('Pengolahan', 'Jasa Cutting (Potong Kotak)', 8000, 'Kg', 'Food Processor', '-', 'Daging segar'),
        ('Pengolahan', 'Jasa Grinding (Giling)', 6000, 'Kg', 'Food Processor', '-', 'Daging giling'),
        ('Pengolahan', 'Sosis Sapi (Casings alami)', 45000, 'Kg', 'Bakery/Food', '-', '50kg min'),
        ('Pengolahan', 'Nuggets Ayam (Frozen)', 32000, 'Kg', 'Food Factory', '-', '50kg min'),
        ('Pengolahan', 'Bakso Sapi (Medium)', 35000, 'Kg', 'Bakso Merek', '-', '20kg min'),
        ('Pengolahan', 'Abon Sapi (250g)', 55000, 'Kg', 'Food Processor', '-', 'Kering, 6 bulan shelf life'),
        ('Pengolahan', 'Daging asap (Smoked Beef)', 65000, 'Kg', 'Food Processor', '-', 'Cold smoked'),
        ('Pengolahan', 'Dendeng Sapi (Balado)', 58000, 'Kg', 'Food Processor', '-', 'Kering, 3 bulan'),
        ('Pengolahan', 'Rendang Pack (200g)', 75000, 'Kg', 'Home Industry', '-', 'Ready to eat'),
        ('Pengolahan', 'Kaldu Sapi Bubuk (100g)', 28000, 'Kg', 'Food Factory', '-', 'Stock powder'),
        ('Pengolahan', 'Keripik Kulit Sapi', 48000, 'Kg', 'Snack Maker', '-', 'Camilan'),
        ('Pengolahan', 'Mie Basah Ayam (Kwetiaw)', 18000, 'Kg', 'Food Processor', '-', '3 hari shelf life'),
        
        # Pengolahan Telur & Susu - 10 items
        ('Pengolahan', 'Telur Pecah Liquid (Pasteurized)', 22000, 'Kg', 'Egg Processor', '-', '25kg per carton'),
        ('Pengolahan', 'Telur Bubuk (Whole Egg Powder)', 85000, 'Kg', 'Export Quality', '-', 'Shelf life 1 tahun'),
        ('Pengolahan', 'Telur Asin (Traditional)', 3500, 'Pcs', 'Home Industry', '-', 'Umur 14 hari'),
        ('Pengolahan', 'Telur Asin (Modern/Pressurized)', 4500, 'Pcs', 'Factory', '-', 'Umur 7 hari'),
        ('Pengolahan', 'Kue Lebaran dari Telur (Nastar, dll)', 120000, 'Kg', 'Bakery', '-', 'Premium packaging'),
        ('Pengolahan', 'Susu Segar Sapi (Perah)', 8500, 'Liter', 'Koperasi Susu', '-', 'Siap minum'),
        ('Pengolahan', 'Susu Kental Manis (SKM) 1L', 18000, 'Liter', 'Factory', '-', 'Homogenized'),
        ('Pengolahan', 'Yoghurt Plain (500ml)', 15000, 'Botol', 'Dairy Farm', '-', 'Shelf life 14 hari'),
        ('Pengolahan', 'Keju Cottage (250g)', 35000, 'Pcs', 'Dairy Processor', '-', 'Fresh cheese'),
        ('Pengolahan', 'Butter (250g)', 45000, 'Pcs', 'Dairy Factory', '-', 'Salted butter'),
        
        # Pengemasan - 10 items
        ('Pengemasan', 'Kardus Box 30x30x20 (Ayam)', 2500, 'Pcs', 'Carton Maker', '-', '5kg kapasitas'),
        ('Pengemasan', 'Kardus Box 40x30x25 (Daging)', 3500, 'Pcs', 'Carton Maker', '-', '10kg kapasitas'),
        ('Pengemasan', 'Plastik Vakum (Vacuum Bag) 30cm', 850, 'Pcs', 'Packaging Supplier', '-', '25 micron'),
        ('Pengemasan', 'Plastik Cling Wrap (Fresh Meat)', 15000, 'Roll', 'Packaging Supplier', '-', '45cm x 300m'),
        ('Pengemasan', 'Styrofoam Tray (Egg Tray) 30 cells', 3500, 'Pcs', 'Foam Factory', '-', 'Untuk 30 telur'),
        ('Pengemasan', 'Label/stiker Price (Roll 1000)', 75000, 'Roll', 'Printing', '-', 'Custom design'),
        ('Pengemasan', 'Barcode Sticker (Rolos 1000)', 45000, 'Roll', 'Printing', '-', 'Standar 30x10mm'),
        ('Pengemasan', 'Seng/Bungkus Karung (50kg)', 5500, 'Pcs', 'Packaging', '-', 'Untuk pakan & grain'),
        ('Pengemasan', 'Zip Lock Bag (1kg)', 1200, 'Pcs', 'Packaging', '-', 'Food grade'),
        ('Pengemasan', 'Vacuum Pouch (Stand Up Pouch) 500g', 2500, 'Pcs', 'Packaging', '-', 'With zipper'),
    ]
    
    for row_data in industri_data:
        ws.append([today] + list(row_data))
    
    # Format header
    for col in range(1, 9):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
    
    print(f"✅ Industri: {len(industri_data)} items added")
    
    # Save
    wb.save(filepath)
    print(f"💾 Saved: {filepath}")
    
    # Summary
    print(f"\n📊 Summary:")
    print(f"   Hulu: {len(hulu_data)} items (Bibit+Pakan+Obat)")
    print(f"   Industri: {len(industri_data)} items (Pemotongan+ColdStorage+Pengolahan+Pengemasan)")

if __name__ == '__main__':
    expand_peternakan_data()
