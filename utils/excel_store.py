"""
Excel data store with metadata support.
Provides consistent Excel read/write with validation and deduplication.
"""
import os
import json
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import from config
try:
    from utils.config import (
        DATA_DIR, STATUS_VALID, STATUS_STALE, STATUS_FAILED, STATUS_ESTIMATED
    )
except ImportError:
    # Fallback if config not available
    DATA_DIR = Path(__file__).resolve().parent.parent / 'data'
    STATUS_VALID = 'valid'
    STATUS_STALE = 'stale'
    STATUS_FAILED = 'failed'
    STATUS_ESTIMATED = 'estimated'


def get_excel_path(filename: str) -> Path:
    """Get full path for data file."""
    return DATA_DIR / filename


def get_metadata_sheet_name() -> str:
    """Name for metadata sheet."""
    return "Metadata"


def read_excel_with_metadata(filename: str, sheet_name: Optional[str] = None) -> Tuple[List, Dict]:
    """
    Read Excel file with metadata.
    Returns (data_rows, metadata_dict).
    """
    path = get_excel_path(filename)
    if not path.exists():
        return [], _empty_metadata()
    
    try:
        wb = load_workbook(path, data_only=True)
        
        # Get metadata if exists
        meta = _empty_metadata()
        if get_metadata_sheet_name() in wb.sheetnames:
            ws_meta = wb[get_metadata_sheet_name()]
            for row in ws_meta.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    meta[row[0]] = row[1] if len(row) > 1 else ''
        
        # Get data sheet
        if sheet_name:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        else:
            ws = wb.active
        
        if ws is None:
            wb.close()
            return [], meta

        # Read data rows
        rows = list(ws.iter_rows(values_only=True))
        data = rows[1:] if len(rows) > 1 else []  # Skip header
        
        # Update last_attempt_at
        meta['last_attempt_at'] = datetime.now().isoformat()
        
        wb.close()
        return data, meta
    
    except Exception as e:
        meta = _empty_metadata()
        meta['status'] = STATUS_FAILED
        meta['validation_errors'].append(str(e))
        return [], meta


def _empty_metadata() -> Dict:
    """Create empty metadata dict."""
    return {
        'source': '',
        'source_url': '',
        'last_success_at': None,
        'last_attempt_at': datetime.now().isoformat(),
        'status': STATUS_VALID,
        'validation_errors': [],
        'row_count': 0
    }


def write_excel_with_metadata(
    filename: str,
    data: List[List],
    metadata: Dict,
    sheet_name: Optional[str] = None,
    headers: Optional[List] = None
) -> bool:
    """
    Write data to Excel with metadata sheet.
    Returns True if successful.
    """
    path = get_excel_path(filename)
    
    try:
        wb = load_workbook(path) if path.exists() else openpyxl.Workbook()
        
        # Create or get data sheet
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
        else:
            ws = wb.active
            if ws is not None:
                ws.title = sheet_name or "Data"
        
        # Write headers
        if headers:
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
        
        # Write data rows
        for row_idx, row in enumerate(data, 2):  # Start from row 2
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        
        # Write metadata sheet
        meta_ws = wb[get_metadata_sheet_name()] if get_metadata_sheet_name() in wb.sheetnames else wb.create_sheet(get_metadata_sheet_name())
        meta_ws.title = get_metadata_sheet_name()
        
        # Metadata headers
        meta_ws.cell(row=1, column=1, value="key")
        meta_ws.cell(row=1, column=2, value="value")
        
        # Write metadata rows
        meta_keys = ['source', 'source_url', 'last_success_at', 'last_attempt_at', 
                   'status', 'validation_errors', 'row_count']
        for idx, key in enumerate(meta_keys, 2):
            meta_ws.cell(row=idx, column=1, value=key)
            val = metadata.get(key, '')
            if key == 'validation_errors' and isinstance(val, list):
                val = '; '.join(str(v) for v in val)
            meta_ws.cell(row=idx, column=2, value=str(val) if val else '')
        
        wb.save(path)
        wb.close()
        return True
    
    except Exception as e:
        print(f"Error writing Excel: {e}")
        return False


def add_row_to_excel(
    filename: str,
    row_data: List,
    sheet_name: Optional[str] = None,
    headers: Optional[List] = None
) -> bool:
    """Append a single row to Excel file."""
    path = get_excel_path(filename)
    
    try:
        # Create new if not exists
        if not path.exists():
            return write_excel_with_metadata(filename, [row_data], {}, sheet_name, headers)
        
        wb = load_workbook(path)
        
        # Get or create sheet
        if sheet_name:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        else:
            ws = wb.active
        
        # Write headers if needed
        if headers and ws.max_row == 0:
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h)
        
        # Append row
        row_num = ws.max_row + 1
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col, value=val)
        
        wb.save(path)
        wb.close()
        return True
    
    except Exception as e:
        print(f"Error adding row: {e}")
        return False


def get_data_status(metadata: Dict) -> str:
    """Determine data status from metadata."""
    if not metadata.get('last_success_at'):
        return STATUS_FAILED
    
    last_success = metadata.get('last_success_at')
    if isinstance(last_success, str):
        try:
            last_success = datetime.fromisoformat(last_success.replace('Z', '+00:00'))
        except:
            return STATUS_FAILED

    if not isinstance(last_success, datetime):
        return STATUS_FAILED
    
    now = datetime.now()
    hours_old = (now - last_success).total_seconds() / 3600
    
    # Import thresholds
    try:
        from utils.config import STALE_THRESHOLDS
        threshold = STALE_THRESHOLDS.get('sembako', 24)
    except:
        threshold = 24
    
    if hours_old > threshold:
        return STATUS_STALE
    
    return metadata.get('status', STATUS_VALID)


def content_hash(data: List) -> str:
    """Generate hash for content."""
    content = json.dumps(data, sort_keys=True, default=str)
    return hashlib.md5(content.encode()).hexdigest()[:12]