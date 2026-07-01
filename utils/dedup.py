"""
Deduplication utility for all Excel data files.
Provides consistent dedup with unique key rules per dataset.
"""
import os
import sys
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from typing import Dict, List, Tuple, Optional, Callable

import openpyxl

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from utils.config import DATA_DIR, STATUS_VALID
except ImportError:
    DATA_DIR = Path(__file__).parent.parent / 'data'
    STATUS_VALID = 'valid'


# Unique key rules per dataset
DEDUP_KEYS = {
    'harga_sembako.xlsx':           lambda row: (str(row[0])[:10] if row[0] else '', str(row[24]) if len(row) > 24 else ''),  # Tanggal + Sumber
    'crypto_monitor.xlsx':           lambda row: (str(row[0])[:10] if row[0] else '',),  # Tanggal only (keep latest per day)
    'harga_emas.xlsx':              lambda row: (str(row[0])[:10] if row[0] else '', 'harian'),  # Tanggal
    'harga_pertanian_ternak.xlsx':  lambda row: (str(row[0])[:10] if row[0] else '',),  # Tanggal
    'harga_peternakan_lengkap.xlsx': lambda row: (str(row[0])[:10] if row[0] else '', str(row[1]) if len(row) > 1 else '', str(row[3]) if len(row) > 3 else ''),  # Tanggal + Kategori + Produk
    'harga_pakan_ternak.xlsx':      lambda row: (str(row[0])[:10] if row[0] else '',),  # Tanggal
    'kurs_valuta.xlsx':             lambda row: (str(row[0])[:10] if row[0] else '', str(row[1]) if len(row) > 1 else ''),  # Tanggal + Mata uang
    'harga_minyak.xlsx':            lambda row: (str(row[0])[:10] if row[0] else '',),  # Tanggal
    'bi_rate_inflasi.xlsx':         lambda row: (str(row[0])[:10] if row[0] else '',),  # Tanggal
    'harga_cpo.xlsx':               lambda row: (str(row[0])[:10] if row[0] else '',),  # Tanggal
    'harga_saham_ihsg.xlsx':        lambda row: (str(row[0])[:10] if row[0] else '',),  # Tanggal
    'sentimen_berita.xlsx':         lambda row: (str(row[0])[:10] if row[0] else '', str(row[3]) if len(row) > 3 else ''),  # Tanggal + Headline
    'cuaca_yogyakarta.xlsx':        lambda row: (str(row[0])[:10] if row[0] else '', str(row[1]) if len(row) > 1 else ''),  # Tanggal + Lokasi
    'harga_sembako_desktop.xlsx':   lambda row: (str(row[0])[:10] if row[0] else '',),  # Tanggal
    'formulasi_pakan_ternak.xlsx':  None,  # Reference data - row hash dedup
    'nutrisi_bahan_pakan_lengkap.xlsx': None,  # Reference data - row hash dedup
    'optimalisasi_pakan.xlsx':      None,  # Reference data - row hash dedup
}


def get_dedup_key(filename: str) -> Optional[Callable]:
    """Get dedup key function for filename."""
    return DEDUP_KEYS.get(filename)


def dedup_excel_file(filepath: str, dry_run: bool = False, verbose: bool = True) -> Dict:
    """
    Deduplicate a single Excel file.
    Returns report dict.
    """
    filename = Path(filepath).name
    key_func = get_dedup_key(filename)
    
    report = {
        'file': filename,
        'filepath': filepath,
        'rows_before': 0,
        'rows_after': 0,
        'duplicates_removed': 0,
        'status': 'ok',
        'error': None,
    }
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        # Process all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                continue  # Empty or header only
            
            headers = list(rows[0])
            data_rows = rows[1:]
            report['rows_before'] += len(data_rows)
            
            if not data_rows:
                continue
            
            # Group by key
            if key_func:
                groups = defaultdict(list)
                for r in data_rows:
                    try:
                        key = key_func(r)
                        groups[key].append(r)
                    except:
                        # If key func fails, keep the row
                        key = ('__fallback__',)
                        groups[key].append(r)
            else:
                # No dedup rule - group by entire row hash
                groups = defaultdict(list)
                for r in data_rows:
                    key = tuple(str(v) for v in r if v)
                    groups[key].append(r)
            
            # Keep best row per key (prioritize valid status)
            deduped = []
            for key, rows_group in groups.items():
                if len(rows_group) == 1:
                    deduped.append(rows_group[0])
                else:
                    # Multiple rows - pick best
                    best = rows_group[-1]  # Default: last
                    
                    # Try to find row with status='valid'
                    valid_rows = [r for r in rows_group 
                                  if len(r) > 1 and str(r[-1]).lower() == STATUS_VALID.lower()]
                    if valid_rows:
                        best = valid_rows[-1]
                    
                    deduped.append(best)
            
            report['duplicates_removed'] += len(data_rows) - len(deduped)
            report['rows_after'] += len(deduped)
            
            if not dry_run:
                # Rewrite sheet
                ws.delete_rows(2, ws.max_row)
                for row_idx, row in enumerate(deduped, 2):
                    for col_idx, val in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=val)
        
        if not dry_run:
            wb.save(filepath)
        
        wb.close()
        
        if verbose:
            print(f"  {filename}: {report['rows_before']} -> {report['rows_after']} rows | removed {report['duplicates_removed']} duplicates")
    
    except Exception as e:
        report['status'] = 'error'
        report['error'] = str(e)
        if verbose:
            print(f"  {filename}: ERROR - {e}")
    
    return report


def dedup_all(dry_run: bool = False, verbose: bool = True) -> List[Dict]:
    """
    Deduplicate all Excel files in data directory.
    Returns list of reports.
    """
    if verbose:
        print(f"\n{'='*60}")
        print(f"DEDUPLICATION {'DRY RUN' if dry_run else 'RUN'}")
        print(f"Data dir: {DATA_DIR}")
        print(f"{'='*60}")
    
    reports = []
    total_before = 0
    total_removed = 0
    
    excel_files = sorted(DATA_DIR.glob('*.xlsx'))
    
    for filepath in excel_files:
        report = dedup_excel_file(str(filepath), dry_run=dry_run, verbose=verbose)
        reports.append(report)
        total_before += report['rows_before']
        total_removed += report['duplicates_removed']
    
    if verbose:
        print(f"\n{'='*60}")
        print(f"SUMMARY: {len(reports)} files | {total_before} total rows before | {total_removed} duplicates removed")
        if dry_run:
            print("(DRY RUN - no files written)")
        print(f"{'='*60}\n")
    
    return reports


if __name__ == '__main__':
    import argparse
    
    parser = argparse.ArgumentParser(description='Deduplicate Excel data files')
    parser.add_argument('--dry-run', action='store_true', help='Show what would be done without making changes')
    parser.add_argument('--file', help='Process specific file only')
    args = parser.parse_args()
    
    if args.file:
        report = dedup_excel_file(str(DATA_DIR / args.file), dry_run=args.dry_run)
    else:
        reports = dedup_all(dry_run=args.dry_run)