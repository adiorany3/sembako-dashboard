import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart
from datetime import datetime
import os

EXCEL_PATH = os.path.expanduser("~/sembako/harga_pertanian_ternak.xlsx")

def create_or_load_workbook():
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        return wb, ws
    return create_new_workbook()

def create_new_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Harga Komoditas"
    
    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    headers = [
        "Tanggal",
        "Jagung Kering\nPipil (Rp/kg)", "Jagung Pakan\nTernak (Rp/kg)",
        "Kedelai Impor\n(Rp/kg)", "Kedelai Lokal\n(Rp/kg)",
        "Pakan Broiler\nBR1 (Rp/kg)", "Pakan Layer\n(Rp/kg)",
        "Pakan Bebek\n(Rp/kg)", "Bungkil Kedelai\n(Rp/kg)",
        "Jagung Giling\nPakan (Rp/kg)",
        "Sumber"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hf; cell.fill = hfill; cell.alignment = ha; cell.border = thin
    
    # Category colors
    cat_colors = {
        range(2, 4): "548235",    # Jagung - hijau
        range(4, 6): "7030A0",    # Kedelai - ungu
        range(6, 10): "BF8F00",   # Pakan - kuning
    }
    for cols, color in cat_colors.items():
        for col in cols:
            ws.cell(row=1, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    widths = [13, 18, 20, 18, 18, 18, 16, 16, 18, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "B2"
    return wb, ws

def add_row(tanggal, data_dict, sumber=""):
    """
    data_dict keys: jagung_pipil, jagung_pakan, kedelai_impor, kedelai_lokal,
    pakan_broiler, pakan_layer, pakan_bebek, bungkil_kedelai, jagung_giling
    """
    wb, ws = create_or_load_workbook()
    ca = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    row = ws.max_row + 1
    keys_order = [
        "jagung_pipil", "jagung_pakan", "kedelai_impor", "kedelai_lokal",
        "pakan_broiler", "pakan_layer", "pakan_bebek", "bungkil_kedelai", "jagung_giling"
    ]
    
    c = ws.cell(row=row, column=1, value=tanggal)
    c.alignment = ca; c.border = thin
    
    for i, key in enumerate(keys_order, 2):
        val = data_dict.get(key, 0)
        c = ws.cell(row=row, column=i, value=val)
        c.alignment = ca; c.border = thin
        if val: c.number_format = '#,##0'
    
    c = ws.cell(row=row, column=11, value=sumber)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = thin
    
    wb.save(EXCEL_PATH)
    print(f"Row added: {tanggal}")
    return EXCEL_PATH

if __name__ == "__main__":
    # Data 14 Juni 2026 (dari berita detik)
    add_row("2026-06-14", {
        "jagung_pipil": 6000, "jagung_pakan": 6800,
        "kedelai_impor": 11200, "kedelai_lokal": 10000,
        "pakan_broiler": 8200, "pakan_layer": 7600,
        "pakan_bebek": 7000, "bungkil_kedelai": 9500,
        "jagung_giling": 6300
    }, "detik.com (kedelai naik Rp 11.200)")
    
    # Data 23 Juni 2026 (estimasi pasar terkini)
    add_row("2026-06-23", {
        "jagung_pipil": 6200, "jagung_pakan": 7000,
        "kedelai_impor": 11500, "kedelai_lokal": 10200,
        "pakan_broiler": 8300, "pakan_layer": 7700,
        "pakan_bebek": 7100, "bungkil_kedelai": 9800,
        "jagung_giling": 6500
    }, "estimasi pasar + berita")
    
    print(f"File saved: {EXCEL_PATH}")
