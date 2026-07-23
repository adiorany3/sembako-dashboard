"""
Data pipeline: collect, validate, normalize, score, format Excel data for agents.

ponytail: Dataset is a plain class. Upgrade to dataclass/pydantic when validation
complexity grows or serialization needed.
"""

from enum import Enum
from openpyxl import load_workbook
from uuid import uuid4
from datetime import datetime, timedelta
from pathlib import Path
import os
import json

DATA_DIR = Path(os.environ.get("SEMBAKO_DATA_DIR", "/root/sembako/data"))

# File -> (sheet_name, required_columns) mapping
# Handles actual filenames on disk; fallback names tried if primary missing.
DATASET_CONFIGS = {
    "harga_sembako": {
        "files": ["harga_sembako.xlsx"],
        "sheet": "Harga",
        "required": ["Tanggal"],
        "date_col": "Tanggal",
    },
    "crypto_monitor": {
        "files": ["crypto_monitor.xlsx"],
        "sheet": "Harga",
        "required": ["Tanggal"],
        "date_col": "Tanggal",
    },
    "harga_emas": {
        "files": ["harga_emas.xlsx"],
        "sheet": "Harian",
        "required": ["Tanggal"],
        "date_col": "Tanggal",
    },
    "harga_saham_ihsg": {
        "files": ["harga_saham_ihsg.xlsx"],
        "sheet": "IHSG",
        "required": ["Tanggal", "IHSG"],
        "date_col": "Tanggal",
    },
    "kurs": {
        "files": ["kurs.xlsx", "kurs_valuta.xlsx"],
        "sheet": "Harian",
        "required": ["Tanggal"],
        "date_col": "Tanggal",
    },
    "harga_komoditas": {
        # Combined from two files
        "files": ["harga_komoditas.xlsx", "harga_minyak.xlsx"],
        "sheets": ["Oil", "Harian"],
        "required": ["Tanggal"],
        "date_col": "Tanggal",
    },
    "harga_cpo": {
        "files": ["harga_cpo.xlsx"],
        "sheet": "Harian",
        "required": ["Tanggal"],
        "date_col": "Tanggal",
    },
    "bi_rate": {
        "files": ["bi_rate.xlsx", "bi_rate_inflasi.xlsx"],
        "sheet": "Harian",
        "required": ["Tanggal"],
        "date_col": "Tanggal",
    },
    "keuangan": {
        "files": ["keuangan.xlsx"],
        "sheet": None,  # auto-detect first sheet
        "required": ["Tanggal"],
        "date_col": "Tanggal",
    },
}

# Status weights for quality scoring
STATUS_SCORES = {
    "VALID": 100,
    "PARTIAL": 70,
    "STALE": 50,
    "INVALID": 20,
    "UNAVAILABLE": 0,
}

STALE_HOURS = 7 * 24  # 7 days


class DatasetStatus(Enum):
    VALID = "VALID"
    PARTIAL = "PARTIAL"
    STALE = "STALE"
    INVALID = "INVALID"
    UNAVAILABLE = "UNAVAILABLE"


class Dataset:
    """Container for a single dataset's metadata and rows."""

    def __init__(
        self,
        name: str,
        status: DatasetStatus,
        last_update,
        row_count: int,
        columns: list,
        null_fields: list,
        freshness_hours: float,
        source_file: str,
        rows: list = None,
        sheet_name: str = "",
    ):
        self.name = name
        self.status = status
        self.last_update = last_update
        self.row_count = row_count
        self.columns = columns
        self.null_fields = null_fields
        self.freshness_hours = freshness_hours
        self.source_file = source_file
        self.rows = rows or []
        self.sheet_name = sheet_name

    def to_dict(self):
        return {
            "name": self.name,
            "status": self.status.value,
            "last_update": self.last_update.isoformat() if self.last_update else None,
            "row_count": self.row_count,
            "columns": self.columns,
            "null_fields": self.null_fields,
            "freshness_hours": self.freshness_hours,
            "source_file": self.source_file,
            "sheet_name": self.sheet_name,
        }


def _find_file(file_candidates: list) -> str | None:
    """Return first existing file path from candidates."""
    for f in file_candidates:
        p = DATA_DIR / f
        if p.exists():
            return str(p)
    return None


def _parse_date(val):
    """Try to parse a value as a date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%B %d, %Y"):
            try:
                return datetime.strptime(val.strip(), fmt)
            except ValueError:
                continue
    return None


def _try_numeric(val):
    """Convert to float if possible, else return original."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = val.strip().replace(",", "").replace(".", "", 1)
        try:
            return float(cleaned)
        except ValueError:
            return val
    return val


def load_dataset(filepath, sheet_name, required_columns=None):
    """
    Load a single Excel sheet. Returns Dataset with status, rows, columns, null analysis.

    Status logic:
      - UNAVAILABLE: file/sheet not found
      - INVALID: required columns missing or empty sheet
      - STALE: last data older than STALE_HOURS
      - PARTIAL: has data but >20% null in non-date columns
      - VALID: all good
    """
    if not os.path.exists(filepath):
        return Dataset(
            name=os.path.basename(filepath),
            status=DatasetStatus.UNAVAILABLE,
            last_update=None,
            row_count=0,
            columns=[],
            null_fields=[],
            freshness_hours=0,
            source_file=filepath,
        )

    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return Dataset(
            name=os.path.basename(filepath),
            status=DatasetStatus.INVALID,
            last_update=None,
            row_count=0,
            columns=[],
            null_fields=[],
            freshness_hours=0,
            source_file=filepath,
        )

    # Resolve sheet
    actual_sheet = sheet_name
    if sheet_name is None:
        # auto-detect first sheet
        actual_sheet = wb.sheetnames[0] if wb.sheetnames else None

    if actual_sheet not in wb.sheetnames:
        wb.close()
        return Dataset(
            name=os.path.basename(filepath),
            status=DatasetStatus.UNAVAILABLE,
            last_update=None,
            row_count=0,
            columns=[],
            null_fields=[],
            freshness_hours=0,
            source_file=filepath,
        )

    ws = wb[actual_sheet]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return Dataset(
            name=os.path.basename(filepath),
            status=DatasetStatus.INVALID,
            last_update=None,
            row_count=0,
            columns=[],
            null_fields=[],
            freshness_hours=0,
            source_file=filepath,
        )

    # First row = headers; flatten multi-line headers
    raw_headers = all_rows[0]
    headers = [str(h).replace("\n", " ").strip() if h else f"Col_{i}" for i, h in enumerate(raw_headers)]
    data_rows = all_rows[1:]

    # Filter out completely empty rows
    data_rows = [r for r in data_rows if any(v is not None for v in r)]

    # Required column check
    if required_columns:
        headers_lower = [h.lower() for h in headers]
        for req in required_columns:
            if not any(req.lower() in h for h in headers_lower):
                return Dataset(
                    name=os.path.basename(filepath),
                    status=DatasetStatus.INVALID,
                    last_update=None,
                    row_count=0,
                    columns=headers,
                    null_fields=[],
                    freshness_hours=0,
                    source_file=filepath,
                    sheet_name=actual_sheet,
                )

    if not data_rows:
        return Dataset(
            name=os.path.basename(filepath),
            status=DatasetStatus.INVALID,
            last_update=None,
            row_count=0,
            columns=headers,
            null_fields=[],
            freshness_hours=0,
            source_file=filepath,
            sheet_name=actual_sheet,
        )

    # Find date column index
    date_col_idx = None
    for i, h in enumerate(headers):
        if "tanggal" in h.lower() or "date" in h.lower():
            date_col_idx = i
            break

    # Null analysis
    null_counts = {}
    for i, h in enumerate(headers):
        if i == date_col_idx:
            continue
        total = 0
        nulls = 0
        for row in data_rows:
            if i < len(row):
                total += 1
                if row[i] is None or (isinstance(row[i], str) and row[i].strip() in ("", "-", "N/A", "n/a")):  # type: ignore[union-attr]
                    nulls += 1
        if total > 0 and nulls > 0:
            null_pct = nulls / total
            if null_pct > 0:
                null_counts[h] = {"count": nulls, "total": total, "pct": round(null_pct, 3)}

    # Determine freshness
    last_date = None
    freshness_hours = 0
    if date_col_idx is not None:
        for row in reversed(data_rows):
            if date_col_idx < len(row):
                d = _parse_date(row[date_col_idx])
                if d:
                    last_date = d
                    break
        if last_date:
            freshness_hours = (datetime.now() - last_date).total_seconds() / 3600

    # Status determination
    has_nulls = len(null_counts) > 0
    high_null = any(v["pct"] > 0.2 for v in null_counts.values())
    is_stale = freshness_hours > STALE_HOURS

    if is_stale:
        status = DatasetStatus.STALE
    elif high_null:
        status = DatasetStatus.PARTIAL
    elif has_nulls:
        status = DatasetStatus.PARTIAL
    else:
        status = DatasetStatus.VALID

    null_fields = list(null_counts.keys())

    return Dataset(
        name=os.path.basename(filepath),
        status=status,
        last_update=last_date,
        row_count=len(data_rows),
        columns=headers,
        null_fields=null_fields,
        freshness_hours=round(freshness_hours, 1),
        source_file=filepath,
        rows=data_rows,
        sheet_name=actual_sheet,
    )


def _load_commodity_combined():
    """Load harga_komoditas from oil + cpo files, return merged Dataset."""
    oil_file = _find_file(["harga_komoditas.xlsx", "harga_minyak.xlsx"])
    cpo_file = _find_file(["harga_cpo.xlsx"])

    oil_ds = load_dataset(oil_file, "Harian") if oil_file else Dataset(
        "harga_minyak.xlsx", DatasetStatus.UNAVAILABLE, None, 0, [], [], 0, "")
    cpo_ds = load_dataset(cpo_file, "Harian") if cpo_file else Dataset(
        "harga_cpo.xlsx", DatasetStatus.UNAVAILABLE, None, 0, [], [], 0, "")

    # Merge: use oil as base, add CPO columns
    merged_rows = []
    merged_cols = list(oil_ds.columns) if oil_ds.status != DatasetStatus.UNAVAILABLE else []
    cpo_data = {}
    cpo_date_idx = None
    if cpo_ds.status != DatasetStatus.UNAVAILABLE:
        for i, h in enumerate(cpo_ds.columns):
            if "tanggal" in h.lower():
                cpo_date_idx = i
                break
    if cpo_date_idx is not None:
        for row in cpo_ds.rows:
            d = _parse_date(row[cpo_date_idx])
            if d:
                key = d.strftime("%Y-%m-%d") if isinstance(d, datetime) else str(d)
                cpo_data[key] = row

    if oil_ds.status != DatasetStatus.UNAVAILABLE:
        oil_date_idx = None
        for i, h in enumerate(oil_ds.columns):
            if "tanggal" in h.lower():
                oil_date_idx = i
                break

        # Add CPO columns
        if cpo_ds.status != DatasetStatus.UNAVAILABLE:
            cpo_non_date = [c for i, c in enumerate(cpo_ds.columns) if i != cpo_date_idx]
            merged_cols = list(oil_ds.columns) + cpo_non_date

        for row in oil_ds.rows:
            new_row = list(row)
            if oil_date_idx is not None:
                d = _parse_date(row[oil_date_idx])
                if d:
                    key = d.strftime("%Y-%m-%d")
                    cpo_row = cpo_data.get(key)
                    if cpo_row:
                        new_row += [v for i, v in enumerate(cpo_row) if i != cpo_date_idx]
                    else:
                        new_row += [None] * (len(merged_cols) - len(new_row))
            merged_rows.append(tuple(new_row))

    # Use more recent last_update
    last_update = oil_ds.last_update
    if cpo_ds.last_update and (not last_update or cpo_ds.last_update > last_update):
        last_update = cpo_ds.last_update

    freshness = 0
    if last_update:
        freshness = (datetime.now() - last_update).total_seconds() / 3600

    # Determine status from both
    statuses = [oil_ds.status, cpo_ds.status]
    if DatasetStatus.UNAVAILABLE in statuses:
        status = DatasetStatus.PARTIAL
    elif all(s == DatasetStatus.UNAVAILABLE for s in statuses):
        status = DatasetStatus.UNAVAILABLE
    else:
        status = max(statuses, key=lambda s: STATUS_SCORES.get(s.value, 0))

    return Dataset(
        name="harga_komoditas",
        status=status,
        last_update=last_update,
        row_count=len(merged_rows),
        columns=merged_cols,
        null_fields=oil_ds.null_fields + cpo_ds.null_fields,
        freshness_hours=round(freshness, 1),
        source_file=f"{oil_file or 'N/A'} + {cpo_file or 'N/A'}",
        rows=merged_rows,
        sheet_name="merged(Oil+Harian,Harian)",
    )


def collect_all_data() -> dict:
    """
    Load all datasets. Returns snapshot dict with quality_score.

    Structure:
      {
        snapshot_id: str,
        timestamp: str (ISO),
        datasets: {name: Dataset},
        quality_score: float 0-100
      }
    """
    datasets = {}

    for name, cfg in DATASET_CONFIGS.items():
        if name == "harga_komoditas":
            datasets[name] = _load_commodity_combined()
            continue

        filepath = _find_file(cfg["files"])
        if not filepath:
            datasets[name] = Dataset(
                name=name,
                status=DatasetStatus.UNAVAILABLE,
                last_update=None,
                row_count=0,
                columns=[],
                null_fields=[],
                freshness_hours=0,
                source_file=", ".join(cfg["files"]),
            )
        else:
            datasets[name] = load_dataset(
                filepath, cfg["sheet"], cfg.get("required")
            )

    # Weighted quality score: equal weight per dataset
    n = len(datasets)
    if n == 0:
        quality_score = 0.0
    else:
        total = sum(STATUS_SCORES.get(d.status.value, 0) for d in datasets.values())
        quality_score = round(total / n, 1)

    return {
        "snapshot_id": str(uuid4()),
        "timestamp": datetime.now().isoformat(),
        "datasets": datasets,
        "quality_score": quality_score,
    }


def validate_data(snapshot: dict) -> dict:
    """
    Per-dataset validation. Returns:
      {
        per_dataset: {name: {status, issues: [], warnings: []}},
        overall: {score, status}
      }
    """
    per = {}
    for name, ds in snapshot["datasets"].items():
        issues = []
        warnings = []

        # Check freshness
        if ds.freshness_hours > STALE_HOURS:
            issues.append(f"Data {round(ds.freshness_hours / 24, 1)} hari lama (>7 hari)")
        elif ds.freshness_hours > 3 * 24:
            warnings.append(f"Data {round(ds.freshness_hours / 24, 1)} hari lalu, mungkin kurang fresh")

        # Check nulls
        if ds.null_fields:
            warnings.append(f"Kolom dengan null: {', '.join(ds.null_fields[:5])}")

        # Check row count
        if ds.row_count == 0:
            issues.append("Tidak ada data baris")

        # Check status
        if ds.status == DatasetStatus.UNAVAILABLE:
            issues.append(f"File tidak ditemukan: {ds.source_file}")
        elif ds.status == DatasetStatus.INVALID:
            issues.append("Data tidak valid (format/sheet error)")

        per[name] = {
            "status": ds.status.value,
            "issues": issues,
            "warnings": warnings,
        }

    # Overall
    score = snapshot["quality_score"]
    if score >= 80:
        overall_status = "GOOD"
    elif score >= 50:
        overall_status = "FAIR"
    else:
        overall_status = "POOR"

    return {
        "per_dataset": per,
        "overall": {"score": score, "status": overall_status},
    }


def normalize_snapshot(snapshot: dict) -> dict:
    """
    Normalize all data in-place:
      - Dates -> ISO format strings
      - Numbers -> float
      - Empty/placeholder -> None
      - Currency markers preserved as-is
    """
    for name, ds in snapshot["datasets"].items():
        if not ds.rows:
            continue

        # Find date column
        date_col = None
        for i, h in enumerate(ds.columns):
            if "tanggal" in h.lower() or "date" in h.lower():
                date_col = i
                break

        normalized_rows = []
        for row in ds.rows:
            norm_row = list(row)
            for i in range(len(norm_row)):
                val = norm_row[i]

                # Date column -> ISO string
                if i == date_col:
                    d = _parse_date(val)
                    norm_row[i] = d.isoformat() if d else None
                    continue

                # Null/placeholder -> None
                if val is None or (isinstance(val, str) and val.strip() in ("", "-", "N/A", "n/a", "N/A (use update_saham.py)")):
                    norm_row[i] = None
                    continue

                # Try numeric conversion for price/value columns
                if isinstance(val, (int, float)):
                    norm_row[i] = float(val)
                elif isinstance(val, str):
                    # Preserve currency markers (Rp, $, USD, IDR) — just clean the number
                    stripped = val.strip()
                    # If it's purely numeric-like
                    cleaned = stripped.replace(",", "").replace(".", "", 1).lstrip("-")
                    if cleaned and cleaned.replace(".", "").isdigit():
                        try:
                            norm_row[i] = float(stripped.replace(",", ""))
                        except ValueError:
                            pass  # keep original (likely text with currency marker)

            normalized_rows.append(tuple(norm_row))

        ds.rows = normalized_rows

    return snapshot


def _get_val(row, col_idx):
    """Safe value extraction from row."""
    if col_idx is not None and col_idx < len(row):
        return row[col_idx]
    return None


def format_for_agents(snapshot: dict, mode: str = "full") -> str:
    """
    Format snapshot as structured text for LLM prompts.

    Modes:
      - 'full': last 30 days + latest per dataset (default)
      - 'quick': latest 1 row per dataset
      - 'macro': bi_rate, kurs, komoditas focus
      - 'risk': all with volatility/change indicators
    """
    lines = []
    ts = snapshot.get("timestamp", "")
    sid = snapshot.get("snapshot_id", "")[:8]
    score = snapshot.get("quality_score", 0)

    status_tag = {100: "🟢", 70: "🟡", 50: "🟠", 20: "🔴", 0: "⚫"}
    tag_for = lambda s: status_tag.get(STATUS_SCORES.get(s.value, 0), "⚫")

    lines.append(f"=== SNAPSHOT {sid} | {ts} | Quality: {score}/100 ===\n")

    # Determine which datasets to show
    if mode == "macro":
        show = ["bi_rate", "kurs", "harga_komoditas", "harga_cpo"]
    else:
        show = list(snapshot["datasets"].keys())

    for name in show:
        ds = snapshot["datasets"].get(name)
        if ds is None:
            continue

        tag = tag_for(ds.status)
        status_str = ds.status.value

        if ds.status == DatasetStatus.UNAVAILABLE:
            lines.append(f"{tag} [{status_str}] {name} — TIDAK TERSEDIA")
            lines.append(f"  File: {ds.source_file}")
            lines.append("")
            continue

        freshness = f"{ds.freshness_hours:.0f}h lalu" if ds.freshness_hours else "unknown"
        lines.append(f"{tag} [{status_str}] {name} — {ds.row_count} baris, {freshness}")

        if ds.null_fields:
            lines.append(f"  ⚠ Null: {', '.join(ds.null_fields[:8])}")

        if ds.rows:
            # Date column index
            date_idx = None
            for i, h in enumerate(ds.columns):
                if "tanggal" in h.lower():
                    date_idx = i
                    break

            if mode == "quick":
                # Latest 1 row
                last_row = ds.rows[-1]
                lines.append(f"  Latest ({_get_val(last_row, date_idx)}):")
                for i, col in enumerate(ds.columns):
                    v = _get_val(last_row, i)
                    if i == date_idx:
                        continue
                    if v is not None:
                        lines.append(f"    {col}: {v}")

            elif mode == "full":
                # Last 7 rows for compact view
                recent = ds.rows[-7:]
                lines.append(f"  Columns: {', '.join(ds.columns[:10])}{'...' if len(ds.columns) > 10 else ''}")
                for row in recent[-3:]:
                    date_val = _get_val(row, date_idx) if date_idx is not None else ""
                    vals = [str(v)[:12] for i, v in enumerate(row) if i != date_idx and v is not None][:5]
                    lines.append(f"  {date_val}: {', '.join(vals)}")
                if len(recent) > 3:
                    lines.append(f"  ... ({len(ds.rows) - 3} more rows)")

            elif mode == "risk":
                # All data with change indicators
                lines.append(f"  Columns: {', '.join(ds.columns)}")
                if len(ds.rows) > 1:
                    last = ds.rows[-1]
                    prev = ds.rows[-2]
                    lines.append("  Latest vs Previous:")
                    for i, col in enumerate(ds.columns):
                        if i == date_idx:
                            continue
                        lv = _get_val(last, i)
                        pv = _get_val(prev, i)
                        if lv is not None and pv is not None and isinstance(lv, (int, float)) and isinstance(pv, (int, float)):
                            chg = lv - pv
                            pct = (chg / pv * 100) if pv != 0 else 0
                            arrow = "↑" if chg > 0 else "↓" if chg < 0 else "="
                            lines.append(f"    {col}: {lv} {arrow} ({pct:+.1f}%)")
                lines.append(f"  Total rows: {ds.row_count}")

            else:  # fallback
                lines.append(f"  Rows: {ds.row_count}, Cols: {len(ds.columns)}")

        lines.append("")

    # Validation summary
    validation = validate_data(snapshot)
    overall = validation["overall"]
    lines.append(f"=== OVERALL: {overall['score']}/100 ({overall['status']}) ===")

    # Flag any datasets with issues
    for dname, info in validation["per_dataset"].items():
        if info["issues"]:
            lines.append(f"  ❌ {dname}: {'; '.join(info['issues'])}")

    return "\n".join(lines)


def get_latest_values(snapshot: dict) -> dict:
    """
    Quick access: latest value per numeric column per dataset.
    Returns {dataset_name: {col: val, ...}, ...}
    """
    result = {}
    for name, ds in snapshot["datasets"].items():
        if ds.status == DatasetStatus.UNAVAILABLE or not ds.rows:
            continue

        date_idx = None
        for i, h in enumerate(ds.columns):
            if "tanggal" in h.lower():
                date_idx = i
                break

        last_row = ds.rows[-1]
        latest = {}
        for i, col in enumerate(ds.columns):
            if i == date_idx:
                latest["Tanggal"] = _get_val(last_row, i)
                continue
            val = _get_val(last_row, i)
            if val is not None:
                latest[col] = val
        result[name] = latest

    return result


# Alias for backwards compat with __init__.py (must be before __main__ check)
MarketDataPipeline = collect_all_data


# ---- Self-check ----
if __name__ == "__main__":
    snap = collect_all_data()
    ns = normalize_snapshot(snap)
    vals = get_latest_values(ns)
    fmt = format_for_agents(ns, mode="quick")

    # Assertions
    assert isinstance(snap["quality_score"], float), "quality_score must be float"
    assert len(snap["datasets"]) >= 5, f"Expected >=5 datasets, got {len(snap['datasets'])}"
    assert "datasets" in snap
    assert isinstance(fmt, str) and len(fmt) > 50, "format_for_agents output too short"
    assert isinstance(vals, dict)

    # Verify UNAVAILABLE reported for missing files
    unavail = [d for d in snap["datasets"].values() if d.status == DatasetStatus.UNAVAILABLE]
    print(f"✓ {len(snap['datasets'])} datasets loaded, score={snap['quality_score']}")
    print(f"  {len(unavail)} unavailable, {len(snap['datasets'])-len(unavail)} loaded")
    print(f"  format_for_agents length: {len(fmt)} chars")
    print(fmt[:500])
