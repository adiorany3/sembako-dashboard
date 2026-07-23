#!/usr/bin/env python3
"""
Dashboard Backend - Flask API
Aggregates all monitoring data (sembako, crypto, cuaca, emas, keuangan)
"""

import os
import json
import time
import openpyxl
import hashlib
from datetime import datetime
from flask import Flask, render_template, jsonify, send_from_directory, request
from flask_cors import CORS
import sys
import os

# Fix: Add parent directory to sys.path for 'core' module imports
# When running as `python3 /root/sembako/core/app.py`, sys.path[0] = /root/sembako/core
# We need /root/sembako in the path to import 'core' module
_script_dir = os.path.dirname(os.path.abspath(__file__))
_parent_dir = os.path.dirname(_script_dir)
if _parent_dir not in sys.path:
    sys.path.insert(0, _parent_dir)

CACHE_DIR = os.path.expanduser("~/.hermes/cache")
os.makedirs(CACHE_DIR, exist_ok=True)

# Load GROQ_API_KEY from config.py into os.environ so providers.py can read it
# Check both parent dir (/root/sembako) and app.py dir (/root/sembako/core)
if not os.environ.get("GROQ_API_KEY"):
    for _config_dir in ["/root/sembako", "/root/sembako/core"]:
        _config_path = os.path.join(_config_dir, "config.py")
        if os.path.exists(_config_path):
            import importlib.util
            spec = importlib.util.spec_from_file_location("config", _config_path)
            _config = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(_config)
            if hasattr(_config, "GROQ_API_KEY") and _config.GROQ_API_KEY and len(_config.GROQ_API_KEY) > 10:
                os.environ["GROQ_API_KEY"] = _config.GROQ_API_KEY
                break


def _cache_path(key: str) -> str:
    safe = hashlib.sha256(key.encode()).hexdigest()[:16]
    return os.path.join(CACHE_DIR, f"{safe}.json")


def get_cached(key: str, ttl: int = 28800):
    path = _cache_path(key)
    if not os.path.exists(path):
        return None
    try:
        with open(path) as f:
            entry = json.load(f)
        if time.time() - entry.get("ts", 0) > ttl:
            return None
        return entry.get("data")
    except (json.JSONDecodeError, OSError):
        return None


def set_cache(key: str, data) -> None:
    path = _cache_path(key)
    try:
        with open(path, "w") as f:
            json.dump({"ts": time.time(), "data": data}, f, ensure_ascii=False)
    except OSError:
        pass



app = Flask(__name__, template_folder='../web/templates', static_folder='../web/static')
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
CORS(app)

DATA_DIR = os.path.expanduser("~/sembako/data")


# ============ Helper Functions ============


def load_excel_data(filename, sheet_name=None):
    """Load data from Excel file."""
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        return None, None

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                data.append(row)

        # Get file modification time
        file_mtime = os.path.getmtime(path)
        file_date = datetime.fromtimestamp(file_mtime).strftime("%d %b %Y, %H:%M")

        return data if data else [], file_date
    except Exception as e:
        print(f"Error loading {filename}: {e}")
        return [], None


def get_file_last_update(filename):
    """Get last update time for a file."""
    path = os.path.join(DATA_DIR, filename)
    if os.path.exists(path):
        mtime = os.path.getmtime(path)
        return datetime.fromtimestamp(mtime).strftime("%d %b %Y, %H:%M")
    return None


def get_file_mtime_iso(filename):
    """Get file modification time as ISO string for meta."""
    path = os.path.join(DATA_DIR, filename)
    if os.path.exists(path):
        mtime = os.path.getmtime(path)
        return datetime.fromtimestamp(mtime).isoformat()
    return None


def build_meta(filename, source, data, error=None):
    """Build meta dict for API responses."""
    status = "valid" if not error else "failed"
    last_attempt_at = get_file_mtime_iso(filename)
    validation_errors = [error] if error else []
    
    return {
        "status": status,
        "last_success_at": last_attempt_at,
        "last_attempt_at": last_attempt_at,
        "row_count": len(data) if data else 0,
        "source": source,
        "validation_errors": validation_errors
    }


def load_json_data(filename):
    """Load JSON data."""
    path = os.path.join(DATA_DIR, filename)
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading JSON data: {e}")
            return {}
    return {}


# ============ API Routes ============


@app.route("/")
def index():
    """Serve dashboard."""
    return render_template("index.html")


@app.route("/api/sembako")
def get_sembako():
    """Get latest sembako prices."""
    try:
        data, last_update = load_excel_data("harga_sembako.xlsx", "Harga")
    except Exception as e:
        meta = build_meta("harga_sembako.xlsx", "detik.com/Jina Reader", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("harga_sembako.xlsx", "detik.com/Jina Reader", [], error="Data not found")
        return jsonify({"data": [], "meta": meta}), 404

    result = []
    for row in data:  # All data
        if row[0]:
            result.append(
                {
                    "tanggal": str(row[0]),
                    "beras_premium": row[1],
                    "beras_medium": row[2],
                    "minyak_goreng": row[3],
                    "gula_pasir": row[4],
                    "garam": row[5],
                    "tepung_terigu": row[6],
                    "cabai_merah": row[7],
                    "cabai_rawit": row[8],
                    "bawang_merah": row[9],
                    "bawang_putih": row[10],
                    "minyak_tanah": row[11],
                    "telur_ras": row[12],
                    "telur_kampung": row[13],
                    "ayam_ras": row[14],
                    "ayam_kampung": row[15],
                    "daging_sapi": row[16],
                    "gas_elpiji": row[17] if len(row) > 17 else None,
                    "garam_bata": row[18] if len(row) > 18 else None,
                    "garam_halus": row[19] if len(row) > 19 else None,
                    "susu_km_bendera": row[20] if len(row) > 20 else None,
                    "susu_km_indomilk": row[21] if len(row) > 21 else None,
                    "susu_bubuk_bendera": row[22] if len(row) > 22 else None,
                    "susu_bubuk_indomilk": row[23] if len(row) > 23 else None,
                }
            )

    meta = build_meta("harga_sembako.xlsx", "detik.com/Jina Reader", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/crypto")
def get_crypto():
    """Get latest crypto prices."""
    try:
        data, last_update = load_excel_data("crypto_monitor.xlsx", "Harga")
    except Exception as e:
        meta = build_meta("crypto_monitor.xlsx", "CoinGecko API", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("crypto_monitor.xlsx", "CoinGecko API", [], error="Data not found")
        return jsonify({"data": [], "meta": meta}), 404

    result = []
    for row in data:
        if row[0]:
            result.append(
                {
                    "tanggal": str(row[0]),
                    "waktu": str(row[1]) if row[1] else "",
                    "btc_usd": row[2],
                    "btc_idr": row[3],
                    "btc_24h": row[4],
                    "eth_usd": row[5],
                    "eth_idr": row[6],
                    "eth_24h": row[7],
                    "sol_usd": row[8],
                    "sol_idr": row[9],
                    "sol_24h": row[10],
                    "ada_usd": row[11] if len(row) > 11 else 0,
                    "ada_idr": row[12] if len(row) > 12 else 0,
                    "ada_24h": row[13] if len(row) > 13 else 0,
                    "doge_usd": row[14] if len(row) > 14 else 0,
                    "doge_idr": row[15] if len(row) > 15 else 0,
                    "doge_24h": row[16] if len(row) > 16 else 0,
                    "xrp_usd": row[17] if len(row) > 17 else 0,
                    "xrp_idr": row[18] if len(row) > 18 else 0,
                    "xrp_24h": row[19] if len(row) > 19 else 0,
                    "market_cap": row[20] if len(row) > 20 else 0,
                    "sentimen": row[21] if len(row) > 21 else "",
                }
            )

    meta = build_meta("crypto_monitor.xlsx", "CoinGecko API", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/emas")
def get_emas():
    """Get latest gold prices."""
    try:
        data, last_update = load_excel_data("harga_emas.xlsx", "Harian")
    except Exception as e:
        meta = build_meta("harga_emas.xlsx", "detik.com (Antam/UBS)", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("harga_emas.xlsx", "detik.com (Antam/UBS)", [], error="Data not found")
        return jsonify({"data": [], "meta": meta}), 404

    result = []
    for row in data:
        if row[0]:
            result.append(
                {
                    "tanggal": str(row[0]),
                    "antam_beli": row[1],
                    "antam_buyback": row[2],
                    "antam_pegadaian": row[3],
                    "galeri24": row[4],
                    "ubs_beli": row[5],
                    "selisih": row[7],
                    "spread_persen": row[8],
                }
            )

    meta = build_meta("harga_emas.xlsx", "detik.com (Antam/UBS)", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/pertanian")
def get_pertanian():
    """Get agriculture prices."""
    try:
        data, last_update = load_excel_data("harga_pertanian_ternak.xlsx")
    except Exception as e:
        meta = build_meta("harga_pertanian_ternak.xlsx", "kompas.com/detik.com", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("harga_pertanian_ternak.xlsx", "kompas.com/detik.com", [], error="Data not found")
        return jsonify({"data": [], "meta": meta}), 404

    result = []
    for row in data:
        if row[0]:
            result.append(
                {
                    "tanggal": str(row[0]),
                    "jagung_pipil": row[1],
                    "jagung_pakan": row[2],
                    "kedelai_impor": row[3],
                    "kedelai_lokal": row[4],
                    "pakan_broiler": row[5],
                    "pakan_layer": row[6],
                    "pakan_bebek": row[7],
                    "bungkil_kedelai": row[8],
                    "jagung_giling": row[9],
                }
            )

    meta = build_meta("harga_pertanian_ternak.xlsx", "kompas.com/detik.com", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/keuangan")
def get_keuangan():
    """Get financial data."""
    try:
        data, last_update = load_excel_data("keuangan.xlsx")
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    if not data:
        return jsonify({"error": "Data not found"}), 404

    result = []
    for row in data[-20:]:  # Last 20 transactions
        if row[0]:
            result.append(
                {
                    "tanggal": str(row[0]),
                    "jenis": row[1],
                    "kategori": row[2],
                    "deskripsi": row[3],
                    "jumlah": row[4],
                    "metode": row[5] or "",
                }
            )

    return jsonify({"data": result, "last_update": last_update})


@app.route("/api/peternakan")
def get_peternakan():
    """Get comprehensive livestock data (hulu to hilir)."""
    try:
        data, last_update = load_excel_data("harga_peternakan_lengkap.xlsx", "Data Utama")
    except Exception as e:
        meta = build_meta("harga_peternakan_lengkap.xlsx", "data generator", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("harga_peternakan_lengkap.xlsx", "data generator", [], error="Data not found")
        return jsonify({"data": [], "meta": meta}), 404

    result = []
    for row in data[-50:]:
        if row[0]:
            result.append(
                {
                    "tanggal": str(row[0]),
                    "kategori": row[1],
                    "sub_kategori": row[2],
                    "produk": row[3],
                    "harga": row[4],
                    "satuan": row[5],
                    "sumber": row[6],
                }
            )

    meta = build_meta("harga_peternakan_lengkap.xlsx", "data generator", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/peternakan/<kategori>")
def get_peternakan_by_kategori(kategori):
    """Get peternakan data filtered by kategori (hulu, industri, hilir, analisis)."""
    try:
        data, last_update = load_excel_data("harga_peternakan_lengkap.xlsx", "Data Utama")
    except Exception as e:
        meta = build_meta("harga_peternakan_lengkap.xlsx", "data generator", [], error=str(e))
        return jsonify({"data": [], "kategori": kategori.upper(), "meta": meta}), 404

    if not data:
        meta = build_meta("harga_peternakan_lengkap.xlsx", "data generator", [], error="Data not found")
        return jsonify({"data": [], "kategori": kategori.upper(), "meta": meta}), 404

    result = []
    for row in data:
        if row[0] and row[1] and row[1].upper() == kategori.upper():
            result.append(
                {
                    "tanggal": str(row[0]),
                    "kategori": row[1],
                    "sub_kategori": row[2],
                    "produk": row[3],
                    "harga": row[4],
                    "satuan": row[5],
                    "sumber": row[6],
                }
            )

    meta = build_meta("harga_peternakan_lengkap.xlsx", "data generator", result)
    return jsonify(
        {"data": result, "last_update": last_update, "kategori": kategori.upper(), "meta": meta}
    )


@app.route("/keuangan")
def keuangan():
    """Serve dedicated keuangan page."""
    return render_template("keuangan.html")


@app.route("/api/summary")
def get_summary():
    """Get overall summary."""
    sources = {
        "sembako": "detik.com/Jina Reader",
        "crypto": "CoinGecko API",
        "emas": "detik.com (Antam/UBS)",
        "pertanian": "kompas.com/detik.com",
        "keuangan": "manual entry",
        "peternakan": "data generator",
    }
    
    summary = {
        "timestamp": datetime.now().isoformat(),
        "sembako": {
            "status": "OK" if load_excel_data("harga_sembako.xlsx")[0] else "ERROR",
            "source": sources["sembako"]
        },
        "crypto": {
            "status": "OK" if load_excel_data("crypto_monitor.xlsx")[0] else "ERROR",
            "source": sources["crypto"]
        },
        "emas": {"status": "OK" if load_excel_data("harga_emas.xlsx")[0] else "ERROR", "source": sources["emas"]},
        "pertanian": {
            "status": (
                "OK" if load_excel_data("harga_pertanian_ternak.xlsx")[0] else "ERROR"
            ),
            "source": sources["pertanian"]
        },
        "keuangan": {
            "status": "OK" if load_excel_data("keuangan.xlsx")[0] else "ERROR",
            "source": sources["keuangan"]
        },
        "peternakan": {
            "status": (
                "OK" if load_excel_data("harga_peternakan_lengkap.xlsx")[0] else "ERROR"
            ),
            "source": sources["peternakan"]
        },
    }
    
    # Build meta combining all sources
    combined_data = []
    meta = {
        "status": "valid" if all(v.get("status") == "OK" for v in summary.values() if isinstance(v, dict)) else "stale",
        "last_attempt_at": datetime.now().isoformat(),
        "row_count": len(combined_data),
        "source": "combined",
        "validation_errors": []
    }
    
    return jsonify({"data": summary, "meta": meta})


@app.route("/api/saham")
def get_saham():
    """Get IHSG and bluechip stock data."""
    excel_path = os.path.join(DATA_DIR, "harga_saham_ihsg.xlsx")
    if not os.path.exists(excel_path):
        meta = build_meta("harga_saham_ihsg.xlsx", "Yahoo Finance/investing.com", [], error="File not found")
        return jsonify({"data": {}, "meta": meta}), 404

    result = {}
    wb = None
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Sheet 1: IHSG
        if "IHSG" in wb.sheetnames:
            ws_ihsg = wb["IHSG"]
            ihsg_data = []
            for row in ws_ihsg.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    ihsg_data.append({
                        "tanggal": str(row[0]),
                        "ihsg": row[1], "change": row[2], "change_pct": row[3],
                        "high": row[4], "low": row[5], "volume": row[6], "sentimen": row[7],
                    })
            result["ihsg"] = ihsg_data[-30:]

        # Sheet 2: Sektor
        if "Sektor" in wb.sheetnames:
            ws_sektor = wb["Sektor"]
            sektor_data = []
            for row in ws_sektor.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    sektor_data.append({
                        "sektor": row[0], "value": row[1], "daily_change": row[2],
                        "change_pct": row[3], "weekly_trend": row[4], "recommendation": row[8] if len(row) > 8 else "",
                    })
            result["sektor"] = sektor_data

        # Sheet 3: Bluechip
        if "Bluechip" in wb.sheetnames:
            ws_blue = wb["Bluechip"]
            bluechip_data = []
            for row in ws_blue.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    bluechip_data.append({
                        "symbol": row[0], "nama": row[1], "sektor": row[2], "harga": row[3],
                        "daily_chg": row[4], "daily_chg_pct": row[5], "rsi": row[6],
                        "trend": row[7], "signal": row[8], "recommendation": row[9] if len(row) > 9 else "",
                    })
            result["bluechip"] = bluechip_data

        # Sheet: Watchlist
        if "Watchlist" in wb.sheetnames:
            ws_watch = wb["Watchlist"]
            watchlist_data = []
            for row in ws_watch.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    watchlist_data.append({
                        "symbol": row[0], "nama": row[1], "harga": row[3],
                        "daily_chg": row[4], "rsi": row[5], "reason": row[6],
                        "potential": row[7], "recommendation": row[9] if len(row) > 9 else "",
                    })
            result["watchlist"] = watchlist_data

    except Exception as e:
        print(f"Error loading saham: {e}")
        meta = build_meta("harga_saham_ihsg.xlsx", "Yahoo Finance/investing.com", [], error=str(e))
        return jsonify({"data": {}, "meta": meta}), 404
    finally:
        if wb:
            wb.close()

    last_update = get_file_last_update("harga_saham_ihsg.xlsx")
    meta = build_meta("harga_saham_ihsg.xlsx", "Yahoo Finance/investing.com", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/sentimen")
def get_sentimen():
    """Get latest news sentiment data."""
    try:
        data, last_update = load_excel_data("sentimen_berita.xlsx", "Detail")
    except Exception as e:
        meta = build_meta("sentimen_berita.xlsx", "news scraping", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("sentimen_berita.xlsx", "news scraping", [], error="Data not found")
        return jsonify({"data": [], "meta": meta}), 404

    result = []
    for row in data:
        if row[0]:
            result.append(
                {
                    "tanggal": str(row[0]),
                    "waktu": str(row[1]),
                    "keyword": row[2],
                    "headline": row[3],
                    "sentimen": row[4],
                    "score": row[5],
                    "source": row[6],
                }
            )

    meta = build_meta("sentimen_berita.xlsx", "news scraping", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/pakan")
def get_pakan():
    """Get latest pakan/feed prices."""
    try:
        data, last_update = load_excel_data("harga_pakan_ternak.xlsx")
    except Exception as e:
        meta = build_meta("harga_pakan_ternak.xlsx", "detik.com/kompas.com", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("harga_pakan_ternak.xlsx", "detik.com/kompas.com", [], error="Data not found")
        return jsonify({"data": [], "meta": meta}), 404

    result = []
    for row in data:
        if row[0]:
            result.append(
                {
                    "tanggal": str(row[0]),
                    "jagung_pipilan": row[1],
                    "bungkil_kedelai": row[2],
                    "dedak_padi": row[3],
                    "tepung_ikan": row[4],
                    "pollard": row[5],
                    "biji_kapuk": row[6],
                    "tepung_darah": row[7],
                    "tepung_tulang": row[8],
                    "molases": row[9],
                    "bungkil_kelapa": row[10],
                    "gaplek": row[11],
                    "bungkil_sawit": row[12],
                    "ampas_tahu": row[13],
                    "tepung_bulu_ayam": row[14],
                    "kulit_kentang": row[15],
                    "onggok": row[16],
                    "bungkil_kacang_tanah": row[17],
                    "dedak_halus": row[18],
                    "sorgum": row[19],
                    "menir": row[20],
                    "corn_gluten_feed": row[21],
                    "rice_polish": row[22],
                    "mung_bean_husk": row[23],
                }
            )

    meta = build_meta("harga_pakan_ternak.xlsx", "detik.com/kompas.com", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/kurs")
def get_kurs():
    """Get exchange rates (USD, EUR, SGD, MYR vs IDR)."""
    try:
        data, last_update = load_excel_data("kurs_valuta.xlsx", "Harian")
    except Exception as e:
        meta = build_meta("kurs_valuta.xlsx", "currency API", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("kurs_valuta.xlsx", "currency API", [], error="Data not found")
        return jsonify({"data": [], "last_update": None, "meta": meta}), 404

    result = []
    for row in data:
        if row[0]:
            result.append({
                "tanggal": str(row[0]),
                "usd_idr": row[1],
                "eur_idr": row[2],
                "sgd_idr": row[3],
                "myr_idr": row[4],
            })
    meta = build_meta("kurs_valuta.xlsx", "currency API", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/minyak")
def get_minyak():
    """Get crude oil prices (Brent & WTI)."""
    try:
        data, last_update = load_excel_data("harga_minyak.xlsx", "Harian")
    except Exception as e:
        meta = build_meta("harga_minyak.xlsx", "oil price API", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("harga_minyak.xlsx", "oil price API", [], error="Data not found")
        return jsonify({"data": [], "last_update": None, "meta": meta}), 404

    result = []
    for row in data:
        if row[0]:
            result.append({
                "tanggal": str(row[0]),
                "brent": row[1],
                "wti": row[2],
                "selisih": row[3],
            })
    meta = build_meta("harga_minyak.xlsx", "oil price API", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/bi-rate")
def get_bi_rate():
    """Get BI Rate & Inflasi (CPI)."""
    try:
        data, last_update = load_excel_data("bi_rate_inflasi.xlsx", "Harian")
    except Exception as e:
        meta = build_meta("bi_rate_inflasi.xlsx", "BI.go.id", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("bi_rate_inflasi.xlsx", "BI.go.id", [], error="Data not found")
        return jsonify({"data": [], "last_update": None, "meta": meta}), 404

    result = []
    for row in data:
        if row[0]:
            result.append({
                "tanggal": str(row[0]),
                "bi_rate": row[1],
                "inflasi_mom": row[2],
                "inflasi_yoy": row[3],
                "ihk": row[4],
            })
    meta = build_meta("bi_rate_inflasi.xlsx", "BI.go.id", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/cpo")
def get_cpo():
    """Get Crude Palm Oil (CPO) / Kelapa Sawit price."""
    try:
        data, last_update = load_excel_data("harga_cpo.xlsx", "Harian")
    except Exception as e:
        meta = build_meta("harga_cpo.xlsx", "CPO market data", [], error=str(e))
        return jsonify({"data": [], "meta": meta}), 404

    if not data:
        meta = build_meta("harga_cpo.xlsx", "CPO market data", [], error="Data not found")
        return jsonify({"data": [], "last_update": None, "meta": meta}), 404

    result = []
    for row in data:
        if row[0]:
            result.append({
                "tanggal": str(row[0]),
                "harga_myr": row[1],
                "harga_idr": row[2],
                "perubahan_persen": row[3],
            })
    meta = build_meta("harga_cpo.xlsx", "CPO market data", result)
    return jsonify({"data": result, "last_update": last_update, "meta": meta})


@app.route("/api/alerts")
def get_alerts():
    """Get recent price alerts (significant changes)."""
    import json as _json
    alert_file = os.path.join(DATA_DIR, "price_alerts.json")
    if not os.path.exists(alert_file):
        meta = build_meta("price_alerts.json", "price alerts", [], error="File not found")
        return jsonify({"alerts": [], "last_check": None, "meta": meta})
    try:
        with open(alert_file) as f:
            data = _json.load(f)
        alerts = data.get("alerts", [])
        meta = build_meta("price_alerts.json", "price alerts", alerts)
        return jsonify({"alerts": alerts, "last_check": data.get("last_check"), "meta": meta})
    except Exception as e:
        meta = build_meta("price_alerts.json", "price alerts", [], error=str(e))
        return jsonify({"alerts": [], "last_check": None, "meta": meta})


@app.route("/test-article")
def test_article():
    return "<h1>TEST ROUTE WORKS</h1>", 200, {"Content-Type": "text/html"}

@app.route("/api/health")
def health():
    """Health check."""
    meta = {
        "status": "valid",
        "last_attempt_at": datetime.now().isoformat(),
        "row_count": 0,
        "source": "system",
        "validation_errors": []
    }
    return jsonify({"status": "healthy", "timestamp": datetime.now().isoformat(), "version": "cd2cec2", "meta": meta})


# ============ AI-Powered Analysis using Groq ============


# Try to import config (local only, not on GitHub)
try:
    
    from config import GROQ_API_KEY
except ImportError:
    # Fallback to environment variable
    GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")

GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"


@app.route("/api/ai-analysis")
def get_ai_analysis():
    """
    Read pre-computed analysis from daily cron job.
    ZERO tokens - just reads JSON file.
    Cron job precompute_analysis.py runs every 8h and calls Groq once.
    """
    import time
    import os

    cache_file = os.path.expanduser("~/sembako/data/daily_analysis.json")

    # Try to read pre-computed analysis
    if os.path.exists(cache_file):
        try:
            with open(cache_file) as f:
                entry = json.load(f)
            generated = entry.get("generated_at", "unknown")
            next_up = entry.get("next_update", 0)
            remaining = max(0, int(next_up - time.time()))
            return jsonify({
                "status": "success",
                "analysis": entry.get("analysis", ""),
                "timestamp": generated,
                "model": "Groq Llama 3.1 8B (precomputed)",
                "cached": True,
                "refresh_in_seconds": remaining,
            })
        except (json.JSONDecodeError, OSError):
            pass

    # Fallback: compute on-demand if cache missing (rare)
    return jsonify({
        "status": "error",
        "error": "No cached analysis. Run: python3 scripts/precompute_analysis.py",
        "refresh_in_seconds": None,
    }), 503


def generate_analysis_prompt():
    """Generate analysis prompt from all available data."""
    prompt_parts = []
    today = datetime.now().strftime("%d %b %Y")

    prompt_parts.append(f"# ANALISIS MARKER {today}\\n")

    # 1. IHSG & Saham
    try:
        ihsg_path = os.path.join(DATA_DIR, "harga_saham_ihsg.xlsx")
        if os.path.exists(ihsg_path):
            wb = openpyxl.load_workbook(ihsg_path, data_only=True)

            # IHSG data
            ws_ihsg = wb["IHSG"]
            ihsg_rows = list(ws_ihsg.iter_rows(min_row=2, values_only=True))
            if ihsg_rows:
                latest = ihsg_rows[-1]
                prompt_parts.append("\\n## IHSG (Indeks Saham)\\n")
                prompt_parts.append(f"- Level: {latest[1]}\\n")
                prompt_parts.append(f"- Change: {latest[2]} ({latest[3]}%)\\n")
                prompt_parts.append(f"- High: {latest[4]}, Low: {latest[5]}\\n")

            # Sektor
            ws_sektor = wb["Sektor"]
            sektor_rows = list(ws_sektor.iter_rows(min_row=2, values_only=True))
            if sektor_rows:
                prompt_parts.append("\\n## Sektor Performance\\n")
                for row in sektor_rows[:8]:
                    if row[0]:
                        prompt_parts.append(f"- {row[0]}: {row[1]} ({row[3]})\\n")

            # Bluechip
            ws_blue = wb["Bluechip"]
            blue_rows = list(ws_blue.iter_rows(min_row=2, values_only=True))
            if blue_rows:
                prompt_parts.append("\\n## Saham Bluechip LQ45 (Top 10)\\n")
                prompt_parts.append("| Symbol | Harga | RSI | Trend | Rec |\\n")
                prompt_parts.append("|--------|-------|-----|-------|-----|\\n")
                for row in blue_rows[:10]:
                    if row[0]:
                        prompt_parts.append(
                            f"| {row[0]} | {row[3]} | {row[6]} | {row[7]} | {row[9]} |\\n"
                        )

            # Watchlist
            ws_watch = wb["Watchlist"]
            watch_rows = list(ws_watch.iter_rows(min_row=2, values_only=True))
            if watch_rows:
                prompt_parts.append("\\n## Watchlist (Turun + Potensi)\\n")
                for row in watch_rows[:5]:
                    if row[0]:
                        prompt_parts.append(
                            f"- {row[0]} ({row[1]}): RSI={row[5]}, Potential={row[7]}, Rec={row[9]}\\n"
                        )
    except Exception as e:
        prompt_parts.append(f"\\n## IHSG/Saham: Data tidak tersedia ({e})\\n")
        pass

    # 2. Crypto
    try:
        crypto_path = os.path.join(DATA_DIR, "crypto_monitor.xlsx")
        if os.path.exists(crypto_path):
            wb = openpyxl.load_workbook(crypto_path, data_only=True)
            ws = wb["Harga"]
            crypto_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if crypto_rows:
                prompt_parts.append("\\n## Crypto\\n")
                for row in crypto_rows[:6]:
                    if row[0]:
                        prompt_parts.append(f"- {row[0]}: ${row[1]} ({row[4]}%)\\n")
    except Exception as e:
        print(f"Error generating Crypto data for prompt: {e}")
        pass

    # 3. Emas
    try:
        emas_path = os.path.join(DATA_DIR, "harga_emas.xlsx")
        if os.path.exists(emas_path):
            wb = openpyxl.load_workbook(emas_path, data_only=True)
            ws = wb["Harian"]
            emas_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if emas_rows:
                prompt_parts.append("\\n## Emas\\n")
                for row in emas_rows[:3]:
                    if row[0]:
                        prompt_parts.append(
                            f"- {row[0]}: Rp {row[1]} (Buyback: Rp {row[2]})\\n"
                        )
    except Exception as e:
        print(f"Error generating Emas data for prompt: {e}")
        pass

    # 4. Sembako (key items)
    try:
        sembako_path = os.path.join(DATA_DIR, "harga_sembako.xlsx")
        if os.path.exists(sembako_path):
            wb = openpyxl.load_workbook(sembako_path, data_only=True)
            ws = wb["Harga"]
            sembako_rows = list(ws.iter_rows(min_row=2, values_only=True))
            if sembako_rows:
                prompt_parts.append("\\n## Sembako (Key Items)\\n")
                for row in sembako_rows[:8]:
                    if row[0]:
                        prompt_parts.append(f"- {row[0]}: Rp {row[1]}\\n")
    except Exception as e:
        print(f"Error generating Sembako data for prompt: {e}")
        pass

    prompt_parts.append(
        "\\n---\\nBerikan analisis dalam Bahasa Indonesia dengan format yang rapi."
    )
    return "\\n".join(prompt_parts)


# ============ Error Handlers ============


@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Not found"}), 404


@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": "Server error"}), 500


@app.route("/download/<filename>")
def download_file(filename):
    """Download Excel/data files. Generate on-the-fly if missing."""
    import re as _re
    import subprocess
    if not _re.match(r'^[\w\-\.]+$', filename):
        return jsonify({"error": "Invalid filename"}), 400
    
    filepath = os.path.join(DATA_DIR, filename)
    
    # If file missing, try to generate it
    if not os.path.exists(filepath):
        generators = {
            "nutrisi_bahan_pakan_lengkap.xlsx": "../scripts/create_nutrisi_pakan.py",
            "formulasi_pakan_ternak.xlsx": "../scripts/create_formulasi_pakan.py",
        }
        if filename in generators:
            script = os.path.join(os.path.dirname(__file__), generators[filename])
            try:
                subprocess.run(["python3", script], timeout=30, capture_output=True)
            except Exception:
                pass
    
    if not os.path.exists(filepath):
        return jsonify({"error": "File not found"}), 404
    
    try:
        return send_from_directory(DATA_DIR, filename, as_attachment=True)
    except Exception:
        return jsonify({"error": "File not found"}), 404



@app.route("/api/generate-article", methods=["POST"])
def api_generate_article():
    """Generate article, save as JSON + HTML for copy-paste."""
    import subprocess as _sub
    try:
        result = _sub.run(
            ["python3", os.path.join(os.path.dirname(__file__), "../scripts/generate_wp_article.py")],
            capture_output=True, text=True, timeout=30
        )
        # Read generated file
        today = datetime.now().strftime("%Y-%m-%d")
        meta_file = os.path.join(DATA_DIR, "wp_articles", f"{today}.json")
        if os.path.exists(meta_file):
            with open(meta_file) as f:
                meta = json.load(f)
            return jsonify({
                "title": meta["title"],
                "content": meta["content"],
                "category": meta.get("category", ""),
                "date": meta.get("date", ""),
                "html_file": f"/download/wp_articles/{today}.html",
            })
        return jsonify({"error": "Generation failed", "output": result.stdout}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/article-new")
def api_article_new():
    """Generate new article - call via browser or curl."""
    meta_file = os.path.join(DATA_DIR, "wp_articles", datetime.now().strftime("%Y-%m-%d") + ".json")
    if os.path.exists(meta_file):
        os.remove(meta_file)
    # Trigger by redirecting to /article which auto-generates
    return jsonify({"status": "ok", "message": "New article generated", "url": "/article"})


@app.route("/article")
def article_page():
    """Serve latest generated article — title and content always synced."""
    articles_dir = os.path.join(DATA_DIR, "wp_articles")
    today = datetime.now().strftime("%Y-%m-%d")
    meta_file = os.path.join(articles_dir, f"{today}.json")

    # Force regenerate if ?new param
    if request.args.get("new") and os.path.exists(meta_file):
        os.remove(meta_file)

    # Generate if missing
    if not os.path.exists(meta_file):
        import subprocess as _sub
        try:
            script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "../scripts/generate_wp_article.py")
            _sub.run(["python3", script], timeout=30, capture_output=True)
        except Exception:
            pass

    if os.path.exists(meta_file):
        with open(meta_file) as fp:
            meta = json.load(fp)
        t = meta["title"]
        ct = meta["body"]
        cat = meta.get("category", "")
        dt = meta.get("date", "")
        import html as _html
        ct_esc = _html.escape(ct)
        return f"""<!DOCTYPE html>
<html lang="id"><head><meta charset="UTF-8">
<title>WordPress Article</title>
<style>
body{{font-family:-apple-system,sans-serif;max-width:800px;margin:0 auto;padding:20px;background:#f5f5f5}}
.card{{background:#fff;border-radius:8px;padding:24px;margin-bottom:16px;box-shadow:0 1px 3px rgba(0,0,0,.1)}}
h1{{color:#1a1a2e;font-size:20px}}
pre{{white-space:pre-wrap;word-wrap:break-word;background:#f8f8f8;padding:12px;border-radius:6px;font-size:13px;max-height:400px;overflow-y:auto;border:1px solid #eee}}
.copy-btn{{background:#0073aa;color:#fff;border:none;padding:10px 20px;border-radius:6px;cursor:pointer;font-size:14px;margin-top:8px}}
.copy-btn:hover{{background:#005a87}}
.label{{font-weight:bold;color:#555;margin-bottom:4px}}
.success{{color:green;font-weight:bold;display:none;margin-left:10px}}
.preview{{border:1px solid #ddd;border-radius:8px;padding:16px;background:#fff}}
</style></head><body>
<h1>Artikel WordPress</h1>
<div class="card">
<div class="label">Title:</div>
<h1 id="title">{t}</h1>
<button class="copy-btn" onclick="copyText('title')">Copy Title</button>
<span class="success" id="title-copied">Copied!</span>
</div>
<div class="card">
<div class="label">Category: <strong>{cat}</strong> | Date: <strong>{dt}</strong></div>
</div>
<div class="card">
<div class="label">HTML Source (copy ini, paste ke WordPress <strong>HTML/Code Editor</strong>):</div>
<pre id="raw" style="display:none">{ct_esc}</pre>
<button class="copy-btn" onclick="copyRaw()">📋 Copy HTML Source</button>
<span class="success" id="content-copied">Copied!</span>
<div id="readable" style="margin-top:12px;padding:12px;background:#f8f8f8;border-radius:6px;font-size:13px;max-height:300px;overflow-y:auto;border:1px solid #eee"></div>
<div class="card preview"><div class="label">Preview:</div>{ct}</div>
<div class="card">
<a href="/article?new=1" class="copy-btn" style="background:#28a745;text-decoration:none;display:inline-block">🔄 Generate Artikel Baru</a>
</div>
<script>
function copyRaw(){{
  var raw=document.getElementById("raw").textContent;
  var ta=document.createElement("textarea");
  ta.value=raw;ta.style.position="fixed";ta.style.opacity="0";
  document.body.appendChild(ta);ta.select();
  document.execCommand("copy");
  document.body.removeChild(ta);
  document.getElementById("content-copied").style.display="inline";
  setTimeout(function(){{document.getElementById("content-copied").style.display="none"}},2000);
}}
</script></body></html>"""
    return "No article yet. Click Generate.", 204, {"Content-Type": "text/plain"}
# ============ Multi-Agent AI Analysis ============

import uuid

@app.route("/api/ai/analyze", methods=["POST"])
def ai_analyze():
    """Start persistent analysis job. Body: {mode: quick|full|macro|risk}"""
    from core.ai.job_runner import get_runner
    from core.ai.analysis_store import get_store

    # Rate limit check
    client_ip = request.remote_addr or "unknown"
    if not _check_ai_rate_limit(client_ip):
        return jsonify({
            "error": "Terlalu banyak permintaan. Coba lagi dalam 1 menit.",
            "error_code": "AI_RATE_LIMITED",
        }), 429

    store = get_store()
    mode = (request.json or {}).get("mode", "full")
    if mode not in ("quick", "full", "macro", "risk"):
        return jsonify({
            "error": "mode tidak valid",
            "error_code": "AI_CONFIG_MISSING",
        }), 400

    # Idempotency: check for active job (not cancelled/failed)
    existing = store.get_active_job()
    if existing and existing["status"] in ("RUNNING", "PARTIAL"):
        return jsonify({"active_job_id": existing["id"], "status": existing["status"],
                         "message": "Job aktif sedang berjalan"})

    # Check provider availability
    try:
        from core.ai.providers import get_provider_manager
        mgr = get_provider_manager()
        if not mgr.is_available():
            return jsonify({
                "error": mgr._router_manager._init_error or "AI provider tidak dikonfigurasi",
                "error_code": "AI_CONFIG_MISSING",
            }), 503
    except Exception:
        pass

    job = store.create_job(mode=mode)
    job_id = job["id"]
    runner = get_runner()
    runner.start_job(job_id)
    return jsonify({"job_id": job_id, "status": "QUEUED"})


@app.route("/api/ai/jobs/<job_id>")
def ai_job_detail(job_id):
    """Get job status with agent tasks."""
    from core.ai.analysis_store import get_store
    store = get_store()
    job = store.get_job(job_id)
    if not job:
        return jsonify({"error": "Job tidak ditemukan"}), 404
    agents = store.get_agent_tasks(job_id)
    return jsonify({"job": job, "agents": agents})


@app.route("/api/ai/jobs/<job_id>/cancel", methods=["POST"])
def ai_job_cancel(job_id):
    """Cancel a running job."""
    from core.ai.job_runner import get_runner
    runner = get_runner()
    runner.cancel_job(job_id)
    return jsonify({"status": "cancelled"})


@app.route("/api/ai/jobs/<job_id>/retry-failed", methods=["POST"])
def ai_job_retry_failed(job_id):
    """Retry failed/skipped agents in a job."""
    from core.ai.job_runner import get_runner
    runner = get_runner()
    runner.retry_failed(job_id)
    return jsonify({"status": "retrying"})


@app.route("/api/ai/history")
def ai_history():
    """List past analyses (newest first)."""
    from core.ai.analysis_store import get_store
    store = get_store()
    items = store.get_history(limit=30)
    return jsonify(items)


@app.route("/api/ai/history/<job_id>")
def ai_history_detail(job_id):
    """Get full job result by ID."""
    from core.ai.analysis_store import get_store
    store = get_store()
    job = store.get_job(job_id)
    if not job:
        return jsonify({"error": "Analisis tidak ditemukan"}), 404
    agents = store.get_agent_tasks(job_id)
    return jsonify({"job": job, "agents": agents})


@app.route("/api/ai/provider-status")
def ai_provider_status():
    """Check all AI router health via 9Router system."""
    import threading as _threading

    try:
        from core.ai.providers import get_provider_manager
    except Exception as e:
        return jsonify({
            "available": False,
            "error": f"AI module load error: {e}",
            "error_code": "AI_SERVICE_INTERNAL_ERROR",
            "routers": {},
            "healthy_count": 0,
            "total_count": 0,
        })

    mgr = get_provider_manager()

    if not mgr.is_available():
        return jsonify({
            "available": False,
            "error": mgr._router_manager._init_error or "AI not configured",
            "error_code": "AI_CONFIG_MISSING",
            "routers": {},
            "healthy_count": 0,
            "total_count": 9,
            "config": {
                "has_api_key": bool(os.environ.get("AI_ROUTER_API_KEY") or os.environ.get("GROQ_API_KEY")),
                "base_url": os.environ.get("AI_ROUTER_BASE_URL", "(not set)"),
            },
        })

    # Health check with timeout
    result = [None]

    def do_check():
        try:
            result[0] = mgr.health_check_all()
        except Exception as e:
            result[0] = {"error": str(e), "error_code": "AI_SERVICE_INTERNAL_ERROR"}

    t = _threading.Thread(target=do_check, daemon=True)
    t.start()
    t.join(20)

    if t.is_alive() or result[0] is None:
        return jsonify({
            "available": False,
            "error": "Health check timeout (20s)",
            "error_code": "AI_TIMEOUT",
            "routers": {},
            "healthy_count": 0,
            "total_count": 9,
        })

    return jsonify(result[0])


# --- Rate Limiting for AI endpoints ---
_ai_rate_limit = {}  # ip -> [timestamps]
AI_RATE_LIMIT_MAX = 5  # max requests
AI_RATE_LIMIT_WINDOW = 60  # seconds

def _check_ai_rate_limit(ip):
    now = time.time()
    timestamps = _ai_rate_limit.get(ip, [])
    timestamps = [t for t in timestamps if now - t < AI_RATE_LIMIT_WINDOW]
    if len(timestamps) >= AI_RATE_LIMIT_MAX:
        return False
    timestamps.append(now)
    _ai_rate_limit[ip] = timestamps
    return True


if __name__ == "__main__":
    # Production mode - disable debug for stability
    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)
